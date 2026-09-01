"""
VERSA INVENTORY EXPORT API v3
- Dropbox as primary inventory source
- STYLE+OVERRIDES image lookup (matches frontend)
- Size-suffix stripping for correct image URLs
"""

# Gevent monkey-patch MUST happen before any other imports
# to avoid SSL recursion errors with boto3/urllib3
try:
    from gevent import monkey
    monkey.patch_all()
except ImportError:
    pass  # gevent not installed — run with sync workers

import os
import re
import hmac
import json
import time
import uuid
import random
import threading
import concurrent.futures
from datetime import datetime, timedelta
from io import BytesIO

import boto3
from botocore.exceptions import ClientError, NoCredentialsError
from flask import Flask, request, jsonify, send_file, Response, make_response, g
from flask_cors import CORS
import xlsxwriter
import requests as http_requests
import openpyxl
from PIL import Image as PilImage
from PIL import ImageOps
try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.colors import Color, HexColor
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.lib.utils import ImageReader
    HAS_REPORTLAB = True
except ImportError:
    HAS_REPORTLAB = False
    print("⚠ reportlab not installed — PDF export will be unavailable. Add 'reportlab' to requirements.txt")

app = Flask(__name__)
CORS(app, resources={
    r"/*": {
        "origins": "*",
        "methods": ["GET", "POST", "DELETE", "OPTIONS"],
        # Authorization is required by /factory-view and by the staff-session
        # gate (authz_gate); X-Api-Key is the machine-caller credential.
        # Without them here the browser's preflight never grants the header
        # and the request is blocked before it is sent.
        "allow_headers": ["Content-Type", "Authorization", "X-Api-Key"]
    }
})
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

AWS_REGION       = os.environ.get('AWS_REGION', 'us-east-2')
S3_BUCKET        = os.environ.get('S3_BUCKET', 'nauticaslimfit')
S3_INVENTORY_KEY = os.environ.get('S3_INVENTORY_KEY', 'inventory/daily_inventory.xlsx')
S3_EXPORT_PREFIX = os.environ.get('S3_EXPORT_PREFIX', 'exports/').rstrip('/') + '/'
S3_PHOTOS_PREFIX = os.environ.get('S3_PHOTOS_PREFIX',
                                   'ALL+INVENTORY+Photos/PHOTOS+INVENTORY')

S3_OVERRIDES_KEY = os.environ.get('S3_OVERRIDES_KEY', 'inventory/style_overrides.json')
S3_MANUAL_ALLOC_KEY = os.environ.get('S3_MANUAL_ALLOC_KEY', 'inventory/manual_allocations.json')
S3_DEDUCTION_ASSIGN_KEY = os.environ.get('S3_DEDUCTION_ASSIGN_KEY', 'inventory/deduction_assignments.json')
S3_SAVED_CATALOGS_KEY = 'inventory/saved_catalogs.json'
S3_PREPACK_DEFAULTS_KEY = os.environ.get('S3_PREPACK_DEFAULTS_KEY', 'inventory/prepack_defaults.json')
S3_SUPPRESSION_OVERRIDES_KEY = os.environ.get('S3_SUPPRESSION_OVERRIDES_KEY', 'inventory/suppression_overrides.json')
S3_BANNER_RULES_KEY = os.environ.get('S3_BANNER_RULES_KEY', 'inventory/banner_rules.json')

S3_PHOTOS_URL = f"https://{S3_BUCKET}.s3.{AWS_REGION}.amazonaws.com/{S3_PHOTOS_PREFIX}"

# STYLE+OVERRIDES — primary image source (matches frontend logic)
S3_OVERRIDES_IMG_URL = f"https://{S3_BUCKET}.s3.{AWS_REGION}.amazonaws.com/ALL+INVENTORY+Photos/STYLE+OVERRIDES"
S3_DROPBOX_SYNC_PREFIX = 'ALL INVENTORY Photos/DROPBOX_SYNC'
S3_DROPBOX_SYNC_URL = f"https://{S3_BUCKET}.s3.{AWS_REGION}.amazonaws.com/ALL+INVENTORY+Photos/DROPBOX_SYNC"
CLOUDFRONT_URL = os.environ.get('CLOUDFRONT_URL', 'https://duv8yzuad6dsp.cloudfront.net')
# Derive CloudFront equivalents of S3 prefixes
CLOUDFRONT_OVERRIDES_URL = f"{CLOUDFRONT_URL}/ALL+INVENTORY+Photos/STYLE+OVERRIDES"
CLOUDFRONT_DROPBOX_SYNC_URL = f"{CLOUDFRONT_URL}/ALL+INVENTORY+Photos/DROPBOX_SYNC"
CLOUDFRONT_PHOTOS_URL = f"{CLOUDFRONT_URL}/{S3_PHOTOS_PREFIX}"
# Bottom-priority fallback swatches (extracted from the brand style-number
# files, keyed by XX_NNN image code). Tried ONLY after every other image
# source fails — mirrors the frontend's swatch-fallback rung.
CLOUDFRONT_SWATCH_FALLBACK_URL = f"{CLOUDFRONT_URL}/ALL+INVENTORY+Photos/SWATCH+FALLBACK" 

# Dropbox direct download URL for hourly inventory file
# Dropbox shared link — will be converted to direct download at runtime
DROPBOX_URL = os.environ.get('DROPBOX_URL',
    'https://www.dropbox.com/scl/fi/de3nzb66mx41un0j69kyk/Inventory_ATS.xlsx?rlkey=ihoxu4s7gpb5ei14w2cl9dvyw&dl=1')

# Dropbox shared folder link for PHOTOS INVENTORY (daily image sync)
DROPBOX_PHOTOS_URL = os.environ.get('DROPBOX_PHOTOS_URL', '')
DROPBOX_PHOTOS_TOKEN = re.sub(r'\s+', '', os.environ.get('DROPBOX_PHOTOS_TOKEN', ''))
DROPBOX_PHOTOS_PATH = os.environ.get('DROPBOX_PHOTOS_PATH', '/Versa Share Files/PHOTOS INVENTORY')
DROPBOX_PHOTOS_SYNC_HOURS = int(os.environ.get('DROPBOX_PHOTOS_SYNC_HOURS', 8))

# Dropbox OAuth2 refresh token (never expires — auto-refreshes access tokens)
DROPBOX_APP_KEY = os.environ.get('DROPBOX_APP_KEY', '')
DROPBOX_APP_SECRET = os.environ.get('DROPBOX_APP_SECRET', '')
DROPBOX_REFRESH_TOKEN = os.environ.get('DROPBOX_REFRESH_TOKEN', '')

# Selling-data folder paths (per-customer files synced daily from Dropbox)
# Folder layout: /Versa Share Files/David - Dropbox/Selling Data/<Customer>/<file>.xlsx
DROPBOX_SELLING_BASE = os.environ.get('DROPBOX_SELLING_BASE',
    '/Versa Share Files/David - Dropbox/Selling Data')
# Daily refresh hour (UTC). Default 06:00 UTC ≈ 1-2am US East.
SELLING_REFRESH_HOUR_UTC = int(os.environ.get('SELLING_REFRESH_HOUR_UTC', 6))
SELLING_REFRESH_MIN_UTC  = int(os.environ.get('SELLING_REFRESH_MIN_UTC', 0))

# Auto-refresh state
_dropbox_access_token = DROPBOX_PHOTOS_TOKEN  # fallback to legacy token
_dropbox_token_expires = 0

def get_dropbox_token():
    """Get a valid Dropbox access token, auto-refreshing if needed."""
    global _dropbox_access_token, _dropbox_token_expires
    
    # If we have refresh token config, use auto-refresh
    if DROPBOX_REFRESH_TOKEN and DROPBOX_APP_KEY and DROPBOX_APP_SECRET:
        if time.time() < _dropbox_token_expires - 300:  # 5 min buffer
            return _dropbox_access_token
        
        try:
            print("[Dropbox Auth] Refreshing access token...", flush=True)
            resp = http_requests.post('https://api.dropbox.com/oauth2/token', data={
                'grant_type': 'refresh_token',
                'refresh_token': DROPBOX_REFRESH_TOKEN,
                'client_id': DROPBOX_APP_KEY,
                'client_secret': DROPBOX_APP_SECRET,
            }, timeout=30)
            
            if resp.status_code == 200:
                data = resp.json()
                _dropbox_access_token = data['access_token']
                _dropbox_token_expires = time.time() + data.get('expires_in', 14400)
                print(f"[Dropbox Auth] ✓ Token refreshed, expires in {data.get('expires_in', '?')}s", flush=True)
                return _dropbox_access_token
            else:
                print(f"[Dropbox Auth] ✗ Refresh failed ({resp.status_code}): {resp.text[:200]}", flush=True)
        except Exception as e:
            print(f"[Dropbox Auth] ✗ Refresh error: {e}", flush=True)
    
    # Fallback to legacy static token
    return DROPBOX_PHOTOS_TOKEN

TARGET_W = 150
TARGET_H = 150
# Inward padding (px) so image bottom-right anchor stays inside the row.
# Without this, the twoCellAnchor 'to' cell bleeds into the next row and
# images don't hide when Excel filters — causing the "jumbled images" bug.
_IMG_CELL_PAD = 4
COL_WIDTH_UNITS = 22


def _padded_image_opts(img, pad=_IMG_CELL_PAD):
    """Shrink scale & shift offsets so the image stays strictly inside the cell.

    This ensures xlsxwriter's twoCellAnchor 'to' reference never bleeds into
    the next row, which is what causes images to jumble when filtering in Excel.
    The visual difference is ~2 px extra whitespace per side — imperceptible.
    """
    ratio = (TARGET_W - pad) / TARGET_W  # same for W and H (both 150)
    return {
        'image_data': img['image_data'],
        'x_scale':  img['x_scale']  * ratio,
        'y_scale':  img['y_scale']  * ratio,
        'x_offset': img['x_offset'] * ratio + pad / 2,
        'y_offset': img['y_offset'] * ratio + pad / 2,
        'object_position': 1,
        'url': img.get('url', ''),
    }

BRAND_IMAGE_PREFIX = {
    'NAUTICA': 'NA', 'DKNY': 'DK', 'EB': 'EB', 'REEBOK': 'RB', 'VINCE': 'VC',
    'BEN': 'BE', 'USPA': 'US', 'CHAPS': 'CH', 'LUCKY': 'LB', 'JNY': 'JN',
    'BEENE': 'GB', 'NICOLE': 'NM', 'SHAQ': 'SH', 'TAYION': 'TA', 'STRAHAN': 'MS',
    'VD': 'VD', 'VERSA': 'VR', 'CHEROKEE': 'CK', 'AMERICA': 'AC', 'BLO': 'BL', 'BLACK': 'BL', 'DN': 'D9',
    'KL': 'KL', 'RG': 'RG', 'NE': 'NE'
}

BRAND_FULL_NAMES = {
    'NAUTICA': 'Nautica', 'DKNY': 'DKNY', 'EB': 'Eddie Bauer', 'REEBOK': 'Reebok',
    'VINCE': 'Vince Camuto', 'BEN': 'Ben Sherman', 'USPA': 'U.S. Polo Assn.',
    'CHAPS': 'Chaps', 'LUCKY': 'Lucky Brand', 'JNY': 'Jones New York',
    'BEENE': 'Geoffrey Beene', 'NICOLE': 'Nicole Miller', 'SHAQ': "Shaquille O'Neal",
    'TAYION': 'Tayion', 'STRAHAN': 'Michael Strahan', 'VD': 'Von Dutch',
    'VERSA': 'Versa', 'CHEROKEE': 'Cherokee', 'AMERICA': 'American Crew', 'BLO': 'Bloomingdales', 'BLACK': 'Black Label', 'DN': 'Divine 9',
    'KL': 'Karl Lagerfeld Paris', 'RG': 'Robert Graham', 'NE': 'Neiman Marcus'
}

FOLDER_MAPPING = {
    'EB': 'EDDIE+BAUER', 'USPA': 'US+POLO', 'VINCE': 'VINCE+CAMUTO',
    'LUCKY': 'LUCKY+BRAND', 'BEN': 'BEN+SHERMAN', 'BEENE': 'GEOFFREY+BEENE',
    'NICOLE': 'NICOLE+MILLER', 'AMERICA': 'AMERICAN+CREW',
    'TAYION': 'TAYON', 'VD': 'Von+Dutch',
    'KL': 'KARL+LAGERFELD',
    'BLACK': 'Black'
}

# Brand-name normalizer (mirrors the frontend _normalizeBrand). Maps any known
# variant of a brand string to its canonical full name so grouping/filtering
# isn't fragmented across e.g. "BEENE", "GEOFFREY BEENE", "GB", "Geoffrey Beene".
# Built once at import from BRAND_FULL_NAMES + BRAND_IMAGE_PREFIX so adding a
# new brand to those maps is automatically resolvable everywhere.
_BRAND_NAME_ALIASES = {}
def _build_brand_aliases():
    """Populate _BRAND_NAME_ALIASES with every known way to refer to each brand."""
    for abbr, full in BRAND_FULL_NAMES.items():
        # Abbreviation key (e.g. 'BEENE'), full name (e.g. 'Geoffrey Beene'),
        # and a stripped-punctuation form of the full name. All upper-cased.
        _BRAND_NAME_ALIASES[abbr.upper()] = full
        _BRAND_NAME_ALIASES[full.upper()] = full
        stripped = re.sub(r"[.'\-]", '', full).upper()
        _BRAND_NAME_ALIASES[stripped] = full
    for abbr, prefix in BRAND_IMAGE_PREFIX.items():
        if abbr in BRAND_FULL_NAMES:
            _BRAND_NAME_ALIASES[prefix.upper()] = BRAND_FULL_NAMES[abbr]
_build_brand_aliases()
# Alternate SKU brand codes (not in BRAND_IMAGE_PREFIX): NT = Nautica overflow
# serials past 999 — brand strings like "NT" must normalize to Nautica.
_BRAND_NAME_ALIASES['NT'] = 'Nautica'

def _normalize_brand(raw):
    """Map any brand string variant to canonical full name. Falls back to the
    raw input (with whitespace trimmed) if no alias matches — so unknown
    brands still pass through rather than getting silently relabeled."""
    if not raw:
        return ''
    s = str(raw).strip()
    if not s:
        return ''
    return _BRAND_NAME_ALIASES.get(s.upper(), s)


STYLE_CONFIG = {
    'header_bg': '#ADD8E6', 'header_text': '#000000',
    'row_bg_odd': '#FFFFFF', 'row_bg_even': '#F0F4F8',
    'border_color': '#000000', 'font_name': 'Calibri'
}

_s3_client = None

def get_s3():
    global _s3_client
    if _s3_client is None:
        _s3_client = boto3.client('s3',
            aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID', ''),
            aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY', ''),
            region_name=AWS_REGION
        )
    return _s3_client

_inv_lock = threading.Lock()
_inventory = {
    'items': [],
    'brands': {},
    'etag': None,
    'last_sync': None,
    'item_count': 0,
    'source': None,  # 'dropbox' or 's3'
    # Fingerprint of the last accepted sync — used by _sync_passes_sanity_check
    # to reject partial-write reads. Tracks the count of rows with non-zero
    # committed and the total absolute committed sum; both should stay roughly
    # stable across syncs (David's ATS file represents real warehouse state
    # that changes incrementally, not in 90% jumps).
    'committed_nonzero_count': 0,
    'committed_abs_sum': 0,
}

def _sync_passes_sanity_check(new_items):
    """Reject a proposed inventory sync that looks like a partial-write read.

    The "Hourly ATS" Dropbox file is being regenerated on a schedule. If the
    backend fetches it WHILE the upstream process is mid-write, the file we
    download may have its later columns (like Committed) only partially
    populated. openpyxl reads those uncommitted cells as None → parsed as 0,
    silently zeroing out the column for some rows.

    Defense: each accepted sync stamps a fingerprint (count of non-zero
    committed rows, total |committed|). Subsequent syncs are checked against
    the last known good fingerprint. A sync that loses more than 30% of its
    non-zero-committed rows or more than 40% of its total committed magnitude
    compared to the previous good sync is rejected — we keep the prior data
    rather than serve corrupted numbers.

    Returns (accepted: bool, reason: str, fingerprint: dict).
    """
    new_nonzero = sum(1 for i in new_items if i.get('committed', 0) != 0)
    new_abs_sum = sum(abs(i.get('committed', 0)) for i in new_items)
    fingerprint = {
        'committed_nonzero_count': new_nonzero,
        'committed_abs_sum': new_abs_sum,
        'row_count': len(new_items),
    }

    # First-ever sync — always accept, nothing to compare against
    prev_nonzero = _inventory.get('committed_nonzero_count', 0)
    prev_abs_sum = _inventory.get('committed_abs_sum', 0)
    if prev_nonzero == 0 and prev_abs_sum == 0:
        return True, 'first-sync (no baseline to compare)', fingerprint

    # Row count must be in the same ballpark (10% tolerance) — catches the case
    # where parsing got truncated and we only got half the rows
    prev_rows = _inventory.get('item_count', 0)
    if prev_rows > 0 and len(new_items) < prev_rows * 0.7:
        return False, (
            f'row count collapsed ({prev_rows} → {len(new_items)}) — '
            f'likely a partial download or parse failure'
        ), fingerprint

    # Non-zero committed count shouldn't drop by more than 30%
    if prev_nonzero > 0 and new_nonzero < prev_nonzero * 0.70:
        return False, (
            f'non-zero committed rows dropped sharply '
            f'({prev_nonzero} → {new_nonzero}, '
            f'{((prev_nonzero - new_nonzero) / prev_nonzero * 100):.1f}% loss) — '
            f'likely a partial-write read from Dropbox'
        ), fingerprint

    # Absolute committed magnitude shouldn't drop by more than 40%
    if prev_abs_sum > 0 and new_abs_sum < prev_abs_sum * 0.60:
        return False, (
            f'|committed| total dropped sharply '
            f'({prev_abs_sum:,} → {new_abs_sum:,}, '
            f'{((prev_abs_sum - new_abs_sum) / prev_abs_sum * 100):.1f}% loss) — '
            f'likely a partial-write read from Dropbox'
        ), fingerprint

    return True, 'passed', fingerprint

_img_lock = threading.Lock()
_img_cache = {}  # base_style → image result (shared across all size variants)

# --- Dropbox Photos Cache ---
_dropbox_photo_index = {}   # image_code (uppercase) → file_path on disk
_dropbox_thumb_cache = {}   # image_code → thumbnail bytes (for Excel exports)
_dropbox_photo_lock = threading.Lock()
_dropbox_photos_last_sync = 0

# Sportswear photos sub-index — populated alongside _dropbox_photo_index during sync.
# Files inside any "/SPORTSWEAR/" subfolder of the PHOTOS INVENTORY tree get indexed
# here by their original filename (NOT the prefix_number scheme used for shirts).
# This is what powers find_sportswear_image_match() — a longest-prefix lookup against
# SKU style codes, which lets a SKU like "XXGBPJ012SLZ" resolve to image "GBPJ012SL"
# (long-sleeve variant) while "XXGBPJ012RFZ" resolves to "GBPJ012" (short-sleeve).
_sportswear_photo_index = {}    # filename_upper (no ext, no separator normalization) → True
_sportswear_match_cache = {}    # base_style_upper → matched_filename or None
_sportswear_match_lock = threading.Lock()

_export_lock = threading.Lock()
_exports = {
    'brands': {},
    'all_brands': None,
    'generating': False,
    'progress': '',
    'last_generated': None,
}

_overrides_lock = threading.Lock()
_style_overrides = {}
_overrides_loaded = False       # True once S3 load completes — prevents race on fresh workers
_overrides_last_saved = 0.0    # epoch timestamp of last POST save — used for version polling
_prepack_last_saved = 0.0      # epoch timestamp of last prepack-rules save — exports staleness check
_regen_queued = False          # a regen request arrived while one was running — rerun at end
_s3_overrides_etag = None      # ETag from S3 — used to detect cross-worker staleness
_manual_alloc_lock = threading.Lock()
_manual_allocations = []  # list of dicts: {id, sku, customer, po, qty, type, date, notes}
_deduction_assign_lock = threading.Lock()
_deduction_assignments = {}  # dict: { sku: 'warehouse' | 'overseas' }
_prepack_defaults_lock = threading.Lock()
_prepack_defaults = []  # list of dicts: {id, category, fit, label, master_qty, inner_qty, sizes}
_suppression_overrides_lock = threading.Lock()
_suppression_overrides = []  # list of SKU strings that should NOT be suppressed
_banner_rules_lock = threading.Lock()
_banner_rules = []  # list of banner rule dicts

def load_overrides_from_s3():
    """Load style overrides from S3 on startup or when version polling detects
    that another worker saved a change.

    On reload, diff against the current in-memory copy and drop _img_cache
    entries for any style whose override changed. This is the cross-worker
    propagation path: worker A handles POST /overrides and updates its own
    state; worker B picks up the new ETag via /overrides/version and ends up
    here, where we invalidate its local image cache so its exports stop
    serving stale fallback images.
    """
    global _style_overrides, _overrides_loaded, _s3_overrides_etag
    try:
        s3 = get_s3()
        resp = s3.get_object(Bucket=S3_BUCKET, Key=S3_OVERRIDES_KEY)
        data = json.loads(resp['Body'].read().decode('utf-8'))
        etag = resp.get('ETag', '').strip('"')

        # Diff against current to find which styles had their override change.
        # Compare the `image` field specifically — that's what _img_cache mirrors.
        with _overrides_lock:
            old = dict(_style_overrides)

        def _img_of(d):
            return d.get('image') if isinstance(d, dict) else None

        changed_styles = set()
        for style, new_val in data.items():
            if _img_of(new_val) != _img_of(old.get(style)):
                changed_styles.add(style)
        for style in old:
            if style not in data and _img_of(old.get(style)) is not None:
                # Override was deleted — fallback chain takes over, drop cache
                changed_styles.add(style)

        with _overrides_lock:
            _style_overrides = data
            _s3_overrides_etag = etag
        _overrides_loaded = True

        if changed_styles:
            with _img_lock:
                for s in changed_styles:
                    _img_cache.pop(s, None)
                    stale_override_keys = [
                        k for k in list(_img_cache.keys())
                        if isinstance(k, str) and k.startswith(f"__override__:{s}:")
                    ]
                    for k in stale_override_keys:
                        _img_cache.pop(k, None)
            print(f"  ✓ Cross-worker sync: invalidated _img_cache for {len(changed_styles)} changed override styles")

        print(f"  ✓ Loaded {len(_style_overrides)} style overrides from S3 (ETag: {etag[:8]})")
    except ClientError as e:
        if e.response['Error']['Code'] == 'NoSuchKey':
            print("  No existing style overrides in S3 (will create on first save)")
        else:
            print(f"  ⚠ Could not load overrides from S3: {e}")
        _overrides_loaded = True  # mark loaded even on miss
    except Exception as e:
        print(f"  ⚠ Override load error: {e}")
        _overrides_loaded = True  # mark loaded on error

def save_overrides_to_s3():
    """Save style overrides to S3"""
    global _s3_overrides_etag
    try:
        s3 = get_s3()
        with _overrides_lock:
            data = json.dumps(_style_overrides)
        resp = s3.put_object(
            Bucket=S3_BUCKET,
            Key=S3_OVERRIDES_KEY,
            Body=data.encode('utf-8'),
            ContentType='application/json'
        )
        etag = resp.get('ETag', '').strip('"')
        with _overrides_lock:
            _s3_overrides_etag = etag
        print(f"  ✓ Saved {len(_style_overrides)} style overrides to S3 (ETag: {etag[:8]})")
        return True
    except Exception as e:
        print(f"  ✗ Failed to save overrides to S3: {e}")
        return False


def load_manual_allocations_from_s3():
    """Load manual allocation entries from S3 JSON on startup"""
    global _manual_allocations
    try:
        s3 = get_s3()
        resp = s3.get_object(Bucket=S3_BUCKET, Key=S3_MANUAL_ALLOC_KEY)
        data = json.loads(resp['Body'].read().decode('utf-8'))
        with _manual_alloc_lock:
            _manual_allocations = data if isinstance(data, list) else []
        print(f"  ✓ Loaded {len(_manual_allocations)} manual allocation entries from S3")
    except ClientError as e:
        if e.response['Error']['Code'] == 'NoSuchKey':
            print("  No manual allocations in S3 yet (will create on first save)")
        else:
            print(f"  ⚠ Could not load manual allocations from S3: {e}")
    except Exception as e:
        print(f"  ⚠ Manual allocation load error: {e}")


def save_manual_allocations_to_s3():
    """Persist manual allocation entries to S3"""
    try:
        s3 = get_s3()
        with _manual_alloc_lock:
            data = json.dumps(_manual_allocations)
        s3.put_object(
            Bucket=S3_BUCKET,
            Key=S3_MANUAL_ALLOC_KEY,
            Body=data.encode('utf-8'),
            ContentType='application/json'
        )
        print(f"  ✓ Saved {len(_manual_allocations)} manual allocations to S3")
        return True
    except Exception as e:
        print(f"  ✗ Failed to save manual allocations to S3: {e}")
        return False


def load_deduction_assignments_from_s3():
    """Load deduction assignments (sku -> warehouse|overseas) from S3"""
    global _deduction_assignments
    try:
        s3 = get_s3()
        resp = s3.get_object(Bucket=S3_BUCKET, Key=S3_DEDUCTION_ASSIGN_KEY)
        data = json.loads(resp['Body'].read().decode('utf-8'))
        with _deduction_assign_lock:
            _deduction_assignments = data if isinstance(data, dict) else {}
        print(f"  ✓ Loaded {len(_deduction_assignments)} deduction assignments from S3")
    except ClientError as e:
        if e.response['Error']['Code'] == 'NoSuchKey':
            print("  No deduction assignments in S3 yet")
        else:
            print(f"  ⚠ Could not load deduction assignments: {e}")
    except Exception as e:
        print(f"  ⚠ Deduction assignment load error: {e}")


def save_deduction_assignments_to_s3():
    """Save deduction assignments to S3"""
    try:
        s3 = get_s3()
        with _deduction_assign_lock:
            data = json.dumps(_deduction_assignments)
        s3.put_object(
            Bucket=S3_BUCKET,
            Key=S3_DEDUCTION_ASSIGN_KEY,
            Body=data.encode('utf-8'),
            ContentType='application/json'
        )
        print(f"  ✓ Saved {len(_deduction_assignments)} deduction assignments to S3")
        return True
    except Exception as e:
        print(f"  ✗ Failed to save deduction assignments: {e}")
        return False


def load_prepack_defaults_from_s3():
    """Load prepack default rules from S3"""
    global _prepack_defaults
    try:
        s3 = get_s3()
        resp = s3.get_object(Bucket=S3_BUCKET, Key=S3_PREPACK_DEFAULTS_KEY)
        data = json.loads(resp['Body'].read().decode('utf-8'))
        with _prepack_defaults_lock:
            _prepack_defaults = data if isinstance(data, list) else []
        print(f"  ✓ Loaded {len(_prepack_defaults)} prepack defaults from S3")
    except ClientError as e:
        if e.response['Error']['Code'] == 'NoSuchKey':
            print("  No prepack defaults in S3 yet (will create on first save)")
        else:
            print(f"  ⚠ Could not load prepack defaults from S3: {e}")
    except Exception as e:
        print(f"  ⚠ Prepack defaults load error: {e}")


def save_prepack_defaults_to_s3():
    """Persist prepack default rules to S3"""
    try:
        s3 = get_s3()
        with _prepack_defaults_lock:
            data = json.dumps(_prepack_defaults)
        s3.put_object(
            Bucket=S3_BUCKET,
            Key=S3_PREPACK_DEFAULTS_KEY,
            Body=data.encode('utf-8'),
            ContentType='application/json'
        )
        print(f"  ✓ Saved {len(_prepack_defaults)} prepack defaults to S3")
        return True
    except Exception as e:
        print(f"  ✗ Failed to save prepack defaults: {e}")
        return False


def load_suppression_overrides_from_s3():
    global _suppression_overrides
    try:
        s3 = get_s3()
        resp = s3.get_object(Bucket=S3_BUCKET, Key=S3_SUPPRESSION_OVERRIDES_KEY)
        data = json.loads(resp['Body'].read().decode('utf-8'))
        with _suppression_overrides_lock:
            _suppression_overrides = data if isinstance(data, list) else []
        print(f"  ✓ Loaded {len(_suppression_overrides)} suppression overrides from S3")
    except ClientError as e:
        if e.response['Error']['Code'] == 'NoSuchKey':
            print("  No suppression overrides in S3 yet")
        else:
            print(f"  ⚠ Could not load suppression overrides: {e}")
    except Exception as e:
        print(f"  ⚠ Suppression overrides load error: {e}")


def save_suppression_overrides_to_s3():
    try:
        s3 = get_s3()
        with _suppression_overrides_lock:
            data = json.dumps(_suppression_overrides)
        s3.put_object(Bucket=S3_BUCKET, Key=S3_SUPPRESSION_OVERRIDES_KEY, Body=data.encode('utf-8'), ContentType='application/json')
        print(f"  ✓ Saved {len(_suppression_overrides)} suppression overrides to S3")
        return True
    except Exception as e:
        print(f"  ✗ Failed to save suppression overrides: {e}")
        return False


def load_banner_rules_from_s3():
    global _banner_rules
    try:
        s3 = get_s3()
        resp = s3.get_object(Bucket=S3_BUCKET, Key=S3_BANNER_RULES_KEY)
        data = json.loads(resp['Body'].read().decode('utf-8'))
        with _banner_rules_lock:
            _banner_rules = data if isinstance(data, list) else []
        print(f"  ✓ Loaded {len(_banner_rules)} banner rules from S3")
    except ClientError as e:
        if e.response['Error']['Code'] == 'NoSuchKey':
            print("  No banner rules in S3 yet")
        else:
            print(f"  ⚠ Could not load banner rules: {e}")
    except Exception as e:
        print(f"  ⚠ Banner rules load error: {e}")


def save_banner_rules_to_s3():
    try:
        s3 = get_s3()
        with _banner_rules_lock:
            data = json.dumps(_banner_rules)
        s3.put_object(Bucket=S3_BUCKET, Key=S3_BANNER_RULES_KEY, Body=data.encode('utf-8'), ContentType='application/json')
        print(f"  ✓ Saved {len(_banner_rules)} banner rules to S3")
        return True
    except Exception as e:
        print(f"  ✗ Failed to save banner rules: {e}")
        return False


S3_ALLOCATION_KEY = os.environ.get('S3_ALLOCATION_KEY', 'inventory/VIRTUAL WAREHOUSE ALLOCATION.csv')
# Ledger column-I warehouse codes -> platform codes, and the landing
# warehouses whose productions are ADMIN-ONLY (per David, Sep 2026: NJ/AE/AW
# landings must NEVER reach a customer catalog, view, or export — same rule
# as NJ warehouse stock).
_LEDGER_WH_MAP = {'JT': 'JTW', 'JTW': 'JTW', 'TR': 'TR', 'DW': 'DCW', 'DCW': 'DCW',
                  'DC': 'DCW', 'QA': 'QA', 'NJ': 'NJ', 'AE': 'AE', 'AW': 'AW'}
_HIDDEN_LANDING_WH = {'NJ', 'AE', 'AW'}

# Production/Style Ledger — now Dropbox-only (S3 removed)
DROPBOX_PRODUCTION_FOLDER = os.environ.get('DROPBOX_PRODUCTION_FOLDER',
    '/Versa Share Files/David - Dropbox/Style Ledger')

# APO allocation file — Dropbox path
DROPBOX_APO_PATH = os.environ.get('DROPBOX_APO_PATH', '/EDI Team/Nuri/Python Macros/Inventory RAW/APO.csv')

# In-memory APO cache
_apo_data = []
_apo_lock = threading.Lock()
_apo_last_sync = 0

# ── NJ warehouse (3PL BZ FM, Edison NJ) daily on-hand feed ──────────────────
# The 3PL drops INVENTORY_OH.CSV into Dropbox every morning ~08:45 ET
# (ITEM_ID, PIECES_ONHAND, DESCRIPTION, ALT_ITEM_ID1=UPC, ITEM_CLASS,
# DEFAULT_CFG). ITEM_IDs are a mix: Versa SKUs with dash sizes
# (AMBEAW360SLP-14-32), Versa SKUs with space sizes (BONASU543RFP 16-16.5 32),
# bare base styles, '-FBA' earmarked rows, and legacy non-Versa IDs
# (NAU-008-WHITE XL). Aggregation per base style happens at load; results
# merge onto inventory items as the 'nj' field during every sync.
DROPBOX_NJ_PATH = os.environ.get('DROPBOX_NJ_PATH', '/backup/Inventory/INVENTORY_OH.CSV')

_nj_lock = threading.Lock()
_nj_by_base = {}        # base style (upper) -> pieces on hand (shippable)
_nj_sizes = {}          # base style (upper) -> {full ITEM_ID: pieces} breakdown
_nj_fba_units = 0       # '-FBA' earmarked pieces (excluded from shippable)
_nj_unmatched = []      # [{'item_id','qty','description','reason'}] admin report
_nj_last_sync = 0

# Single size token after split('-'): whole or half sizes (15, 15.5, 34).
# Audit fix: the old pattern's combined 'NN-NN.N' branch was dead code post
# split, so half sizes like 16.5 never stripped and those rows were dropped.
_NJ_SIZE_TOKEN = re.compile(r'^\d{1,2}(\.\d)?$')
_NJ_ALPHA_SIZES = {'XS', 'S', 'M', 'L', 'XL', 'XXL', '2XL', '3XL',
                   'ST', 'MT', 'LT', 'XLT'}
# 'V' is NOT a size: -V variants are distinct sellable styles on this
# platform, so an NJ row for BASE-V keys to BASE-V, never folded into BASE.
_NJ_BASE_RE = re.compile(r'^[A-Z]{6}\d{3}[A-Z]{2,3}[A-Z0-9]?(-V)?$')

def _nj_parse_item_id(item_id):
    """Split an NJ ITEM_ID into (style_key, is_fba, looks_versa).
    Handles 'BASE 15-15.5 32', 'BASE-14-32', 'BASE-16.5-34', 'BASE-L-FBA',
    'BASE-V', bare 'BASE'. The style_key keeps a '-V' suffix when present."""
    iid = str(item_id or '').strip().upper()
    if not iid:
        return '', False, False
    token = iid.split(' ')[0]                       # drop space-separated size
    is_fba = token.endswith('-FBA')
    if is_fba:
        token = token[:-4]
    parts = token.split('-')
    while len(parts) > 1 and (_NJ_SIZE_TOKEN.match(parts[-1])
                              or parts[-1] in _NJ_ALPHA_SIZES):
        parts.pop()
    base = '-'.join(parts)
    looks_versa = bool(_NJ_BASE_RE.match(base))
    return base, is_fba, looks_versa

def load_nj_from_dropbox():
    """Fetch + aggregate the NJ warehouse CSV. Keeps prior data on any failure
    so warehouse totals never oscillate on a bad read."""
    global _nj_by_base, _nj_sizes, _nj_fba_units, _nj_unmatched, _nj_last_sync
    token = get_dropbox_token()
    if not token:
        print("[NJ] no Dropbox token — keeping previous NJ data", flush=True)
        return False
    try:
        headers = {
            'Authorization': f'Bearer {token}',
            'Dropbox-API-Arg': json.dumps({'path': DROPBOX_NJ_PATH}),
        }
        ns = os.environ.get('DROPBOX_NJ_NAMESPACE_ID', os.environ.get('DROPBOX_APO_NAMESPACE_ID', ''))
        if ns:
            headers['Dropbox-API-Path-Root'] = json.dumps({'.tag': 'namespace_id', 'namespace_id': ns})
        resp = http_requests.post('https://content.dropboxapi.com/2/files/download',
                                  headers=headers, timeout=30)
        if resp.status_code == 401:
            global _dropbox_token_expires
            _dropbox_token_expires = 0
            token = get_dropbox_token()
            headers['Authorization'] = f'Bearer {token}'
            resp = http_requests.post('https://content.dropboxapi.com/2/files/download',
                                      headers=headers, timeout=30)
        if resp.status_code != 200:
            print(f"[NJ] download failed HTTP {resp.status_code} — keeping previous NJ data", flush=True)
            return False
        text = resp.content.decode('utf-8-sig', errors='replace')
        import csv as _csv
        from io import StringIO as _SIO
        rows = list(_csv.DictReader(_SIO(text)))
        if len(rows) < 500:      # file is ~2,500 rows; a tiny read is a partial upload
            print(f"[NJ] only {len(rows)} rows — likely partial file, keeping previous NJ data", flush=True)
            return False
        by_base, by_sizes, fba_units, unmatched = {}, {}, 0, []
        neg_rows = 0
        for r in rows:
            try:
                qty = int(float(str(r.get('PIECES_ONHAND') or '0').replace(',', '')))
            except Exception:
                qty = 0
            if qty == 0:
                continue
            iid = str(r.get('ITEM_ID') or '').strip()
            if qty < 0:
                neg_rows += 1
                unmatched.append({'item_id': iid, 'qty': qty,
                                  'description': str(r.get('DESCRIPTION') or '')[:60],
                                  'reason': 'negative on-hand'})
                continue
            base, is_fba, looks_versa = _nj_parse_item_id(iid)
            if is_fba:
                fba_units += qty
                unmatched.append({'item_id': iid, 'qty': qty,
                                  'description': str(r.get('DESCRIPTION') or '')[:60],
                                  'reason': 'FBA earmarked (not shippable)'})
                continue
            if not looks_versa:
                unmatched.append({'item_id': iid, 'qty': qty,
                                  'description': str(r.get('DESCRIPTION') or '')[:60],
                                  'reason': 'legacy id (no platform style)'})
                continue
            by_base[base] = by_base.get(base, 0) + qty
            sz = by_sizes.setdefault(base, {})
            sz[iid] = sz.get(iid, 0) + qty
        # Unknown brand codes are decided HERE, once per load (never during the
        # hourly re-merge — appending there duplicated report rows all day).
        for base in list(by_base.keys()):
            if not _nj_brand_for_base(base):
                by_sizes.pop(base, None)
                unmatched.append({'item_id': base, 'qty': by_base.pop(base),
                                  'description': '', 'reason': 'unknown brand code'})
        total = sum(by_base.values())
        # Truncation guard: a partial upload/read must not wipe most of the
        # warehouse. Reject drops below 60% of the previous good load.
        with _nj_lock:
            prev_total = sum(_nj_by_base.values())
        if prev_total > 0 and total < prev_total * 0.6:
            print(f"[NJ] REJECTED load: {total} pcs vs previous {prev_total} "
                  f"(>40% drop — likely partial file). Keeping previous data.", flush=True)
            return False
        with _nj_lock:
            _nj_by_base = by_base
            _nj_sizes = by_sizes
            _nj_fba_units = fba_units
            _nj_unmatched = unmatched
            _nj_last_sync = time.time()
        print(f"[NJ] loaded: {len(by_base)} styles / {total} pcs shippable; "
              f"{fba_units} pcs FBA-earmarked; {len(unmatched)} unmatched rows "
              f"({neg_rows} negative)", flush=True)
        # Persist the good load so a restarting worker can hydrate even if
        # Dropbox is unreachable at boot (best-effort).
        try:
            get_s3().put_object(Bucket=S3_BUCKET, Key='inventory/nj_onhand_cache.json',
                                Body=json.dumps({'by_base': by_base, 'sizes': by_sizes,
                                                 'fba_units': fba_units,
                                                 'unmatched': unmatched, 'ts': time.time()}),
                                ContentType='application/json')
        except Exception as e:
            print(f"[NJ] S3 cache write failed (non-fatal): {e}", flush=True)
        return True
    except Exception as e:
        print(f"[NJ] load failed ({e}) — keeping previous NJ data", flush=True)
        return False

def _nj_brand_for_base(base):
    """Brand for an NJ style key, with the platform's correction rules the
    plain letter map misses. Returns '' when underivable."""
    b = str(base or '').upper()
    if b.endswith('-V'):
        b = b[:-2]
    if b.startswith('LUCK'):
        return 'LUCKY'
    if b.startswith('VP'):
        return 'VERSA'
    code = b[2:4] if len(b) >= 4 else ''
    return SKU_BRAND_CODE_MAP.get(code, '')

def _nj_hydrate_from_s3():
    """Boot fallback: load the last good NJ snapshot from S3 when the Dropbox
    pull fails and this worker has no data yet."""
    global _nj_by_base, _nj_sizes, _nj_fba_units, _nj_unmatched, _nj_last_sync
    try:
        resp = get_s3().get_object(Bucket=S3_BUCKET, Key='inventory/nj_onhand_cache.json')
        data = json.loads(resp['Body'].read().decode('utf-8'))
        by_base = {str(k): int(v) for k, v in (data.get('by_base') or {}).items()}
        if not by_base:
            return False
        with _nj_lock:
            if _nj_by_base:      # a real load beat us — keep it
                return True
            _nj_by_base = by_base
            _nj_sizes = {str(k): {str(s): int(q) for s, q in (v or {}).items()}
                         for k, v in (data.get('sizes') or {}).items()}
            _nj_fba_units = int(data.get('fba_units') or 0)
            _nj_unmatched = list(data.get('unmatched') or [])
            _nj_last_sync = float(data.get('ts') or 0)
        print(f"[NJ] hydrated {len(by_base)} styles / {sum(by_base.values())} pcs "
              f"from S3 cache (Dropbox unavailable at boot)", flush=True)
        return True
    except Exception as e:
        print(f"[NJ] S3 hydrate failed: {e}", flush=True)
        return False

def _apply_nj_to_items(items):
    """Merge NJ per-base quantities onto freshly parsed inventory items.
    The full base quantity lands on exactly ONE row per base (the bare-base
    row when present, else the first row seen) so every base-level sum stays
    correct; siblings get nj=0. Bases with no feed row at all get a synthesized
    row so NJ-only styles still exist on the platform. total_warehouse and
    total_ats on the carrying row grow by nj — NJ is sellable stock."""
    with _nj_lock:
        remaining = dict(_nj_by_base)
        sizes_snap = {k: dict(v) for k, v in _nj_sizes.items()}
    # Idempotent re-apply on COPIES: never mutate dicts that in-flight requests
    # may be serializing (torn-read fix), and drop prior synthesized rows.
    items = [dict(it) for it in items if not it.get('_nj_synth')]
    for it in items:
        prev = int(it.get('nj') or 0)
        if prev:
            it['total_warehouse'] = int(it.get('total_warehouse') or 0) - prev
            it['total_ats'] = int(it.get('total_ats') or 0) - prev
        it['nj'] = 0
        it.pop('nj_sizes', None)
    if not remaining:
        return items
    # Carrier rule (audit fix): NJ always rides the row whose sku EQUALS the
    # NJ style key (bare base, or BASE-V for variant styles). If the feed has
    # no such row, we synthesize one — never park base-level NJ on an
    # arbitrary sized row, which broke exact-SKU consumers (factory view,
    # early-ship pools, per-size AI answers).
    rows_by_sku = {}
    rows_by_base = {}
    for it in items:
        sku = str(it.get('sku') or '').upper()
        rows_by_sku.setdefault(sku, it)
        rows_by_base.setdefault(get_base_style(sku), it)
    for key, qty in remaining.items():
        carrier = rows_by_sku.get(key)
        if carrier is not None:
            carrier['nj'] = qty
            carrier['total_warehouse'] = int(carrier.get('total_warehouse') or 0) + qty
            carrier['total_ats'] = int(carrier.get('total_ats') or 0) + qty
            if sizes_snap.get(key):
                carrier['nj_sizes'] = sizes_snap[key]
            continue
        # Synthesize the carrier. Brand: prefer a real sibling row's brand,
        # else the correction-aware letter map.
        sibling = rows_by_base.get(get_base_style(key))
        brand = str(sibling.get('brand') or '').upper() if sibling else _nj_brand_for_base(key)
        if not brand:
            continue   # loader already routed unknown-brand bases to the report
        name = str(sibling.get('name') or '') if sibling else ''
        row = {
            'sku': key, 'brand': brand, 'brand_abbr': brand,
            'brand_full': BRAND_FULL_NAMES.get(brand, brand), 'name': name,
            'jtw': 0, 'tr': 0, 'dcw': 0, 'qa': 0, 'nj': qty,
            'incoming': 0, 'committed': 0, 'allocated': 0,
            'total_ats': qty, 'total_warehouse': qty,
            'container': '', 'receive_date': '', 'lot_number': '', 'image': '',
            '_nj_synth': True,
        }
        if sizes_snap.get(key):
            row['nj_sizes'] = sizes_snap[key]
        items.append(row)
    return items

def _nj_reapply_to_inventory():
    """Re-merge the NJ cache onto the CURRENT in-memory items (used after the
    10am refresh so new numbers show without waiting for the next ATS sync).
    Rebuilds the brands index too — prebuilt brand workbooks regenerate from
    it, and leaving it stale served yesterday's synthesized rows (audit fix)."""
    with _inv_lock:
        items = _apply_nj_to_items(list(_inventory['items']))
        _inventory['items'] = items
        _inventory['item_count'] = len(items)
        try:
            _inventory['brands'] = _group_by_brand(items)
        except Exception as e:
            print(f"[NJ] brands rebuild failed: {e}", flush=True)

def nj_daily_resync_loop():
    """Refresh the NJ feed ONCE per day at 10:00 AM Eastern (per David: never
    hourly). The 3PL's export lands ~08:45 ET, so 10:00 gives safe margin."""
    while True:
        try:
            try:
                from zoneinfo import ZoneInfo
                from datetime import datetime as _dt
                now = _dt.now(ZoneInfo('America/New_York'))
                target = now.replace(hour=10, minute=0, second=0, microsecond=0)
                if target <= now:
                    target = target + timedelta(days=1)
                wait = (target - now).total_seconds()
            except Exception:
                # tz database unavailable: fall back to 14:00 UTC (10am EDT)
                from datetime import datetime as _dt
                now = _dt.utcnow()
                target = now.replace(hour=14, minute=0, second=0, microsecond=0)
                if target <= now:
                    target = target + timedelta(days=1)
                wait = (target - now).total_seconds()
            time.sleep(max(60, wait))
            if load_nj_from_dropbox():
                _nj_reapply_to_inventory()
        except Exception as e:
            print(f"[NJ] daily loop error: {e}", flush=True)
            time.sleep(3600)


def load_allocation_from_s3():
    """Load allocation CSV from S3 and return as list of dicts"""
    try:
        s3 = get_s3()
        resp = s3.get_object(Bucket=S3_BUCKET, Key=S3_ALLOCATION_KEY)
        text = resp['Body'].read().decode('utf-8-sig')
        lines = text.strip().split('\n')
        if len(lines) < 2:
            return []
        headers = [h.strip() for h in lines[0].split(',')]
        po_idx = next((i for i, h in enumerate(headers) if 'po' in h.lower()), 0)
        cust_idx = next((i for i, h in enumerate(headers) if 'customer' in h.lower()), 1)
        sku_idx = next((i for i, h in enumerate(headers) if 'sku' in h.lower()), 2)
        qty_idx = next((i for i, h in enumerate(headers) if 'qty' in h.lower()), 3)

        results = []
        for line in lines[1:]:
            cols = [c.strip() for c in line.split(',')]
            if len(cols) <= sku_idx or not cols[sku_idx]:
                continue
            try:
                qty = int(cols[qty_idx]) if len(cols) > qty_idx else 0
            except ValueError:
                qty = 0
            results.append({
                'po': cols[po_idx] if len(cols) > po_idx else '',
                'customer': cols[cust_idx] if len(cols) > cust_idx else '',
                'sku': cols[sku_idx].upper(),
                'qty': qty
            })
        print(f"  ✓ Loaded {len(results)} allocation rows from S3")
        return results
    except ClientError as e:
        if e.response['Error']['Code'] == 'NoSuchKey':
            print("  No allocation file found in S3")
        else:
            print(f"  ⚠ Could not load allocation from S3: {e}")
        return []
    except Exception as e:
        print(f"  ⚠ Allocation load error: {e}")
        return []


# Production data cache — loaded from Dropbox, refreshed every 10 min
# (overridable via PRODUCTION_RESYNC_MINUTES env var; see production_resync_loop below)
_production_data = []
_production_last_sync = 0
_production_lock = threading.Lock()
_PRODUCTION_TTL = int(float(os.environ.get('PRODUCTION_RESYNC_MINUTES', 10)) * 60)  # default 10 min


def load_production_from_dropbox():
    """Load Style Ledger xlsx from Dropbox — picks the first .xlsx in the folder.
    Caches result in memory; returns cached data if under 1 hour old."""
    global _production_data, _production_last_sync

    # Return cache if fresh
    with _production_lock:
        if _production_data and (time.time() - _production_last_sync < _PRODUCTION_TTL):
            return _production_data

    token = get_dropbox_token()
    if not token:
        print("  ⚠ Production sync: no Dropbox token available")
        with _production_lock:
            return list(_production_data)  # return stale if any

    try:
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }

        # List the Style Ledger folder — pick first .xlsx file found
        list_resp = http_requests.post(
            'https://api.dropboxapi.com/2/files/list_folder',
            headers=headers,
            json={'path': DROPBOX_PRODUCTION_FOLDER, 'recursive': False, 'limit': 50},
            timeout=20
        )
        if list_resp.status_code != 200:
            print(f"  ⚠ Production: could not list Dropbox folder ({list_resp.status_code}): {list_resp.text[:200]}")
            with _production_lock:
                return list(_production_data)

        entries = list_resp.json().get('entries', [])
        xlsx_files = [e for e in entries if e['.tag'] == 'file' and e['name'].lower().endswith('.xlsx')]
        if not xlsx_files:
            print(f"  ⚠ Production: no .xlsx file found in {DROPBOX_PRODUCTION_FOLDER}")
            with _production_lock:
                return list(_production_data)

        # Pick first xlsx (folder should only have one, but prefer any with "ledger" in name)
        xlsx_files.sort(key=lambda e: (0 if 'ledger' in e['name'].lower() else 1, e['name']))
        chosen = xlsx_files[0]
        print(f"  📋 Loading production data from Dropbox: {chosen['name']}", flush=True)

        # Download file bytes
        dl_resp = http_requests.post(
            'https://content.dropboxapi.com/2/files/download',
            headers={
                'Authorization': f'Bearer {token}',
                'Dropbox-API-Arg': json.dumps({'path': chosen['path_display']})
            },
            timeout=30
        )
        if dl_resp.status_code != 200:
            print(f"  ⚠ Production: download failed ({dl_resp.status_code})")
            with _production_lock:
                return list(_production_data)

        wb = openpyxl.load_workbook(BytesIO(dl_resp.content), read_only=False, data_only=True)
        ws = wb[wb.sheetnames[0]]
        results = []
        row_count = 0
        # Read through column I. Columns of interest:
        #   A=Production#, B=PO Name, C=Style, D=Units, E=Brand,
        #   F=ETD, G=Estimated Arrival to Port, H=Shipment #, I=Warehouse
        # Column I (added Sep 2026) = the warehouse the goods LAND in
        # (JT/TR/NJ/DW/QA/AE/AW on the ledger; normalized to platform codes).
        # Productions landing in a HIDDEN warehouse (NJ/AE/AW) are admin-only:
        # the anonymous catalog scope strips those rows and their quantities.
        # When G is present, it OVERRIDES column F's ETD entirely:
        #   - arrival = G + 10 days  (port-to-warehouse leg)
        #   - etd     = G - 27 days  (factory ship-to-port lead time)
        # When G is empty, fall back to column F ETD + the frontend's transit math
        # (45 days for shirts since Jul 2026 — was 37 — 55 for pants; applied frontend-side).
        # Column H (Shipment #) is admin-only metadata for grouping physical shipments
        # within the same Production#. Never appears in exports.
        from datetime import timedelta as _td
        for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
            row_count += 1
            style = str(row[2] or '').strip().upper()
            if not style:
                continue

            # ── Pull and normalize column G (Estimated Arrival to Port) ──
            port_arrival_dt = None
            if len(row) > 6 and row[6]:
                if isinstance(row[6], datetime):
                    port_arrival_dt = row[6]
                else:
                    try:
                        port_arrival_dt = datetime.strptime(str(row[6])[:10], '%Y-%m-%d')
                    except Exception:
                        port_arrival_dt = None

            # ── Compute etd + arrival ──
            etd = None
            arrival = None
            # etd_raw: column F exactly as the ledger states it, kept even when
            # column G overrides the working dates. ONLY /factory-view reads
            # it — the Open Order to PO factory page shows the true ledger
            # ex-factory date, never the G-27 backdate (David, Jul 21 2026).
            etd_raw = None
            # Track FOB-flagged production rows. When David puts "be ready",
            # "FOB ...", or "FOB-AS READY" in the ETD column, it signals an FOB
            # production batch with no firm ex-factory date. These should NOT
            # participate in date-based feasibility — frontend treats them as
            # date-less but keeps them in the supply pool for FIFO fallback.
            fob_flag = False
            fob_note = ''

            if port_arrival_dt is not None:
                # Column G wins for the WORKING dates: compute both from it.
                etd     = (port_arrival_dt - _td(days=27)).strftime('%Y-%m-%d')
                arrival = (port_arrival_dt + _td(days=10)).strftime('%Y-%m-%d')
                # Still capture the real column-F ETD when it parses as a date.
                if row[5] is not None:
                    if isinstance(row[5], datetime):
                        etd_raw = row[5].strftime('%Y-%m-%d')
                    else:
                        _f = str(row[5]).strip()
                        if re.match(r'^\d{1,4}[-/]\d{1,2}[-/]\d{1,4}', _f):
                            etd_raw = _f
            elif row[5]:
                if isinstance(row[5], datetime):
                    etd = row[5].strftime('%Y-%m-%d')
                else:
                    raw_etd = str(row[5]).strip()
                    raw_upper = raw_etd.upper()
                    # FOB markers: ETD is a non-date string. Common variants:
                    #   "be ready", "FOB", "FOB 6/10", "FOB-AS READY", "as ready"
                    is_fob_marker = (
                        'FOB' in raw_upper
                        or 'BE READY' in raw_upper
                        or raw_upper == 'AS READY'
                        or raw_upper == 'READY'
                        or raw_upper == 'TBD'
                    )
                    # Try to detect if it's at least a parseable date-ish string
                    looks_like_date = bool(re.match(r'^\d{1,4}[-/]\d{1,2}[-/]\d{1,4}', raw_etd))
                    if is_fob_marker or not looks_like_date:
                        # No usable ETD — flag as FOB and leave dates blank
                        fob_flag = True
                        fob_note = raw_etd
                        etd = None
                    else:
                        try:
                            etd = raw_etd
                        except:
                            etd = None
                # Without a column-G override, the working ETD IS the raw one.
                etd_raw = etd

            try:
                units = int(row[3] or 0)
            except (ValueError, TypeError):
                units = 0

            # ── Column H: Shipment # ──
            shipment_no = ''
            if len(row) > 7 and row[7] is not None:
                if isinstance(row[7], (int, float)):
                    shipment_no = str(int(row[7])) if float(row[7]).is_integer() else str(row[7])
                else:
                    shipment_no = str(row[7]).strip()

            # ── Column I: landing Warehouse ──
            warehouse = ''
            if len(row) > 8 and row[8] is not None:
                wr = str(row[8]).strip().upper()
                warehouse = _LEDGER_WH_MAP.get(wr, wr)

            results.append({
                'production': str(row[0] or '').strip(),
                'poName': str(row[1] or '').strip(),
                'style': style,
                'units': units,
                'brand': str(row[4] or '').strip(),
                'etd': etd,
                # The ledger's own column-F date, untouched by the column-G
                # backdate. Only /factory-view consumes it (see etd_raw above).
                'etd_raw': etd_raw,
                # arrival is only set when David provided column G. None means
                # "frontend, please apply your transit rule based on category".
                'arrival': arrival,
                'port_dated': port_arrival_dt is not None,
                # FOB flag: ETD column had "be ready" / "FOB ..." instead of a
                # real date. Frontend treats these as date-less (no feasibility
                # filter) but keeps them in the supply pool. fob_note carries
                # the raw text so the UI can display it.
                'fob_flag': fob_flag,
                'fob_note': fob_note,
                # Shipment # — admin-only column H. Used to break a single Production#
                # PO into physical shipments. Frontend surfaces this in a dedicated
                # admin-only modal; deliberately excluded from all exports.
                'shipmentNo': shipment_no,
                # Landing warehouse — ledger column I, normalized (JT->JTW, DW->DCW).
                # NJ/AE/AW landings are hard-hidden from every customer surface.
                'warehouse': warehouse
            })
        wb.close()

        with _production_lock:
            _production_data = results
            _production_last_sync = time.time()

        print(f"  ✓ Loaded {len(results)} production rows from {row_count} Excel rows ({chosen['name']})")
        return results

    except Exception as e:
        import traceback
        print(f"  ⚠ Production load error: {e}")
        traceback.print_exc()
        with _production_lock:
            return list(_production_data)  # return stale on error


def extract_image_code(sku, brand_abbr):
    """Extract image code from SKU — strips size suffix first.

    Prefix is normally the brand's default (NAUTICA → NA), but when the SKU's
    own brand code (chars 3-4) is an ALTERNATE code for the same brand
    (NT = Nautica overflow, serials restarted past 999), the SKU's code wins:
    NT_201 and NA_201 are different styles and must never share an image code.
    SKU_BRAND_CODE_MAP is defined later in the module — resolved at call time,
    and the first call happens long after import completes.
    """
    s = str(sku or '').upper()
    ab = str(brand_abbr or '').strip().upper()
    # Canonicalize alternate brand spellings to the brand key — ledger/export
    # callers pass whatever the source cell contained ('Nautica', 'NA', 'GB'),
    # and the comparison below must be form-insensitive. For every attested
    # value this is output-equivalent to the old code; it only matters for NT.
    ab = SKU_BRAND_CODE_MAP.get(ab, ab)
    prefix = None
    # Standard layout: brand code at chars 3-4 (after the 2-char customer code);
    # customer-less override keys (e.g. 'NTSU201SLP') carry it at chars 1-2.
    for code in (s[2:4], s[0:2]):
        if len(code) == 2 and ab and SKU_BRAND_CODE_MAP.get(code) == ab:
            prefix = code
            break
    if prefix is None:
        prefix = BRAND_IMAGE_PREFIX.get(ab, ab[:2])
    # Strip size suffix (everything after first dash)
    base_sku = sku.split('-')[0]
    numbers = re.findall(r'\d+', str(base_sku))
    if numbers:
        main_number = max(numbers, key=len)
        return f"{prefix}_{main_number}"
    return f"{prefix}_{base_sku}"


# Bottoms fabric codes (chars 5-6 of the base style). Bottoms must never
# borrow imagery through the brand+serial fallback keys above — those keys
# are fabric-blind, so pants VDBA008 was serving the VDSL008 overshirt photo
# (extract_image_code → 'VD_008' for both). The frontend already restricts
# non-shirt items to precise sources; these helpers give the backend chains
# the same rule. Precise sources (uploaded overrides, STYLE+OVERRIDES,
# sportswear filename match — which includes the fabric code in the match
# string) still apply to bottoms; only the brand_serial rungs are skipped.
# NOTE: US Polo / Geoffrey Beene pants use P##-pattern fabric slots, not
# these codes — their photo conventions are established and left untouched.
BOTTOMS_FABRIC_CODES = {'BA', 'BC', 'BR', 'BH'}


def is_bottoms_style(sku):
    base = str(sku or '').split('-')[0].upper()
    return len(base) >= 6 and base[4:6] in BOTTOMS_FABRIC_CODES


def find_sportswear_image_match(base_style, brand_abbr=None):
    """Find the best-matching sportswear image filename for a SKU.

    Sportswear images live in /SPORTSWEAR/ subfolders of PHOTOS INVENTORY and are
    named with most of the style code minus the customer code prefix. Long-sleeve
    variants include 'SL' in the filename; short-sleeve variants typically don't.

    Algorithm: strip the 2-char customer code prefix from the base style, then
    find the longest filename in the sportswear index that is a prefix of the
    remaining code. Longest-prefix wins, so a long-sleeve variant matches its
    SL-suffixed image and a short-sleeve variant falls through to the bare
    serial-number image.

    Examples (Geoffrey Beene, brand prefix GB):
        SKU 'XXGBPJ012SLZ' → strip XX → 'GBPJ012SLZ' → match 'GBPJ012SL'
        SKU 'XXGBPJ012RFZ' → strip XX → 'GBPJ012RFZ' → match 'GBPJ012'
        SKU 'XXGBPH002Z'   → strip XX → 'GBPH002Z'   → match 'GBPH002'

    Returns the matched filename (uppercase, no extension) or None.
    """
    if not base_style:
        return None
    base = base_style.upper()

    # Cache hit?
    with _sportswear_match_lock:
        if base in _sportswear_match_cache:
            return _sportswear_match_cache[base]
        if not _sportswear_photo_index:
            return None
        index_snap = _sportswear_photo_index  # dict reads are atomic; reading without copy is fine

    # Strip 2-char customer prefix → core style code we'll match against
    core = base[2:] if len(base) > 2 else base

    # Walk progressively shorter prefixes (longest first) until one hits.
    # Stop at length 4 to avoid spurious matches on tiny prefixes like 'GB'.
    matched = None
    min_match_len = 4
    for n in range(len(core), min_match_len - 1, -1):
        candidate = core[:n]
        if candidate in index_snap:
            matched = candidate
            break

    with _sportswear_match_lock:
        if len(_sportswear_match_cache) > 5000:
            _sportswear_match_cache.clear()
        _sportswear_match_cache[base] = matched
    return matched


def get_base_style(sku):
    """Get base style from SKU by stripping size suffix — matches frontend logic"""
    return sku.split('-')[0].upper()


# ── Python-side category + fit derivation (mirrors frontend logic) ────────
# Used by background/server-side exports to annotate items so _add_size_charts
# can match prepack default rules without needing the frontend to send _export_* fields.

_PY_SPORTSWEAR_COLLARS = set('ZUMNOR')
_PY_SPORTSWEAR_FABRICS = {'PH','PJ','PL','PO','PW','TH','HE'}
_PY_BT_FIT_CODES       = {'BT','BB','TT','SB','ST','WB'}  # mirrors frontend BT_FIT_CODES ('WB' was missing)
# Young Men / Sportswear fabric codes — mirrors frontend YOUNG_MEN_FABRIC_CODES.
# Per Style Rules spreadsheet ("YOUNG MEN / SPORTSWEAR" section), these 18 codes belong
# to BOTH the Young Men category AND the Sportswear category.
# PREVIOUS VALUE WAS INCORRECT — had placeholder codes like 'YM','1Y','10'. Fixed.
_PY_YM_FABRIC_CODES    = {
    'KN','WT','SD','SF','SB','SL','BC','BR','BH','BA',
    'CO','TH','PO','PW','PJ','PH','PL','HE',
}
# Sportswear Bottoms — subset of YM/Sportswear that ALSO belongs to the Dress Pants filter
# (per "(BOTTOMS)" marker in spreadsheet: Carpenters, Ripstops, Heavy Weight, Pinstripe).
_PY_SPORTSWEAR_BOTTOM_CODES = {'BC','BR','BH','BA'}
# Brands whose dress pants use the P##X serial convention (position 7 = 'P',
# positions 8-9 digits, position 10 a letter — e.g. "CUUSPPP01SLS", "PMGBDPP01SRS").
# US Polo pioneered it; Geoffrey Beene followed. Mirrors frontend PANTS_SERIAL_BRANDS.
_PY_PANTS_SERIAL_BRANDS = {'US', 'GB'}
# Long-sleeve fit codes — mirrors frontend LONG_SLEEVE_FIT_CODES.
# Used by _py_is_long_sleeve_shirt() so that YM-fabric items (VD shackets etc.) can match
# 'long_sleeve' prepack rules on the basis of their fit code.
_PY_LONG_SLEEVE_FIT_CODES = {'SL','RF','TF','MF','BT','BB','TT','WB','BR','DB'}

# Full fit-code list — mirrors the frontend FIT_CODES table keys EXACTLY.
# Do not trim: any code missing here used to fall through to the old 'RF'
# default below, which made those items match "Regular Fit" (or VD S–XL)
# prepack rules in export bottom grids that the product tiles never showed —
# the "export shows a different prepack than the product" bug.
_PY_ALL_FIT_CODES = {'BB','BR','BT','CE','CH','CR','DB','MF','RF','RR','SB',
                     'SC','SE','SH','SF','SL','SR','SS','ST','TF','TT','WB'}

def _py_extract_fit_code(sku):
    """Extract 2-char fit code — mirrors extractFitCode() in the frontend EXACTLY.
    Parity rules (index.html extractFitCode):
      1. VD B&T special case: WB/BT at base positions 4-6 win first.
      2. Base: 2nd & 3rd chars from the end (last char = collar). NO V-stripping —
         stripping a trailing V shifts the window and misreads e.g. ...01DBV
         (fit DB, collar V) as '1D'.
      3. Dash parts: check EVERY dash-separated part (not just the last one)
         against the full code list, e.g. LUCK-22-170-SL-V → SL.
      4. NO 'RF' default: return the raw candidate. Unknown fit codes must match
         no fit-restricted rule — same as the frontend — instead of silently
         matching Regular Fit rules server-side."""
    parts = sku.upper().split('-')
    base = parts[0]
    if len(base) >= 6 and base[2:4] == 'VD':
        vd_fit = base[4:6]
        if vd_fit in ('WB', 'BT'):
            return vd_fit
    if len(base) >= 3:
        candidate = base[-3:-1]   # 2nd & 3rd from end (last char = collar)
        if candidate in _PY_ALL_FIT_CODES:
            return candidate
    for part in parts[1:]:
        p = part.strip()
        if p in _PY_ALL_FIT_CODES:
            return p
    return base[-3:-1] if len(base) >= 3 else ''

def _py_is_bottom(sku):
    """Structural bottoms test: the P##X dress-pants serial (US Polo / Geoffrey
    Beene) or a sportswear BOTTOMS fabric code.

    Deliberately does NOT go through _py_get_item_category(): that function ends
    by calling _py_is_short_sleeve() for shirts, so calling it from the sleeve
    checks would recurse forever on every shirt."""
    if not sku:
        return False
    base = sku.split('-')[0].upper()
    sku_brand = base[2:4] if len(base) >= 4 else ''
    if (sku_brand in _PY_PANTS_SERIAL_BRANDS and len(base) >= 10
            and base[6] == 'P' and base[7].isdigit() and base[8].isdigit()
            and base[9].isalpha()):
        return True
    return len(base) >= 6 and base[4:6] in _PY_SPORTSWEAR_BOTTOM_CODES

def _py_is_short_sleeve(sku):
    """Short-sleeve check. An explicit fit code wins; the sportswear-collar fallback
    only kicks in when the fit code itself is ambiguous (not recognized as either
    short or long sleeve). This way a long-sleeve polo (e.g. fit RF + Z collar) is
    correctly identified as long sleeve, not force-tagged short by the collar code."""
    # BOTTOMS ARE NEVER SHORT SLEEVE — mirrors the frontend's isShortSleeve()
    # guard. "SR" means Regular Fit SHORT SLEEVE on a shirt but "Slim Fit / Reg
    # Button" on pants, so without this a dress pant (PMGBDPP01SRS) matched the
    # catch-all Short Sleeve prepack and exports printed a shirt size scale under
    # a pants line sheet. The tiles were always right; only exports were wrong.
    if _py_is_bottom(sku):
        return False
    fit = _py_extract_fit_code(sku)
    if fit in {'SS','SR','SB','ST'}:
        return True
    # PARITY: the frontend has NO collar-based short-sleeve fallback for
    # unrecognized fit codes — an unknown fit is simply not short sleeve.
    # (The old fallback here was unreachable while _py_extract_fit_code
    # defaulted to 'RF'; with raw codes returned it would have created a
    # server-only 'short_sleeve' match the tiles never show.)
    return False

def _py_is_young_men(sku):
    base = sku.split('-')[0].upper()
    if len(base) >= 6:
        return base[4:6] in _PY_YM_FABRIC_CODES
    return False

def _py_is_blazer(sku):
    """Blazers — identified SOLELY by the B01-B99 serial at base positions 6-8,
    mirroring frontend isBlazer() exactly (fabric/fit codes deliberately NOT
    used; see the frontend comment about Bloomingdale private-label SKUs).
    Without this, a 'blazers' prepack rule matched on product tiles but
    silently matched NOTHING in any export."""
    if not sku:
        return False
    base = sku.split('-')[0].upper()
    serial = base[6:9]
    return (len(serial) == 3 and serial[0] == 'B'
            and serial[1].isdigit() and serial[2].isdigit())

def _py_is_big_tall(sku):
    base = sku.split('-')[0].upper()
    # Von Dutch B&T: WB/BT at positions 4-5 (the fit slot on VD's shorter SKU
    # format) with ANY collar letter after — mirrors frontend isBigAndTall()
    # exactly (the old prefix whitelist WBJ/BTC/BTS/WBK missed other collars).
    if len(base) >= 6 and base[2:4] == 'VD':
        vd_fit = base[4:6]
        if vd_fit in ('WB', 'BT'):
            return True
    if len(base) >= 11:
        return base[9:11] in _PY_BT_FIT_CODES
    return False

def _py_get_item_category(sku, brand_abbr):
    """Returns category string matching frontend getDetailedCategory() values."""
    base = sku.split('-')[0].upper()
    # Brand code lives at positions 2-3 of the base SKU.
    sku_brand = base[2:4] if len(base) >= 4 else ''
    # Pants: P##X serial pattern, _PY_PANTS_SERIAL_BRANDS only (US Polo, Geoffrey
    # Beene — e.g. "CUUSPPP01SLS", "PMGBDPP01SRS"). Per business rule: other brands
    # don't use the P## convention; their pants are identified by fabric code — only
    # "(BOTTOMS)" codes from Style Rules spreadsheet count as pants (BC/BR/BH/BA),
    # handled by _py_is_pants() via _PY_SPORTSWEAR_BOTTOM_CODES.
    if (sku_brand in _PY_PANTS_SERIAL_BRANDS
            and len(base) >= 10 and base[6] == 'P'
            and base[7].isdigit() and base[8].isdigit() and base[9].isalpha()):
        return 'pants'
    # Sportswear by collar code
    if len(base) >= 11 and base[-1] in _PY_SPORTSWEAR_COLLARS:
        return 'sportswear'
    # Sportswear by fabric code: polo/tee/henley fabrics
    if len(base) >= 6 and base[4:6] in _PY_SPORTSWEAR_FABRICS:
        return 'sportswear'
    # Accessories (Chaps ties, Shaq ties)
    brand_up = (brand_abbr or '').upper()
    if brand_up == 'CHAPS' and base.startswith('CTH'):
        return 'accessories'
    if brand_up == 'SHAQ' and len(base) >= 3 and 'T' in base[:3]:
        return 'accessories'
    # Young Men (fabric-code based)
    if _py_is_young_men(sku):
        return 'young_men'
    # Big & Tall
    if _py_is_big_tall(sku):
        return 'big_tall'
    # Shirts: split by sleeve
    return 'short_sleeve' if _py_is_short_sleeve(sku) else 'long_sleeve'


# Inclusive category helpers — mirror frontend isSportswear/isPants/matchesCategory.
# Items with the 18 YM/Sportswear fabric codes belong to BOTH 'sportswear' AND 'young_men'.
# Items with BC/BR/BH/BA ("bottoms" subset) ALSO belong to 'pants'.
def _py_is_sportswear(sku, brand_abbr):
    if not sku:
        return False
    if _py_get_item_category(sku, brand_abbr) == 'sportswear':
        return True
    return _py_is_young_men(sku)

def _py_is_pants(sku, brand_abbr=''):
    # Same rule as before (P##X serial OR a bottoms fabric code) but routed
    # through the structural helper so nothing in the category chain can recurse.
    return _py_is_bottom(sku)

def _py_is_long_sleeve_shirt(sku):
    """Inclusive long-sleeve check — true for any non-bottom garment whose fit code
    is in _PY_LONG_SLEEVE_FIT_CODES. Mirrors frontend isLongSleeveShirt(). Lets a VD
    shacket / YM-fabric long-sleeve shirt match a 'long_sleeve' prepack rule even
    though _py_get_item_category() classifies it as 'young_men'."""
    if not sku or _py_is_pants(sku, ''):
        return False
    # short sleeve always wins when both could apply
    if _py_is_short_sleeve(sku):
        return False
    return _py_extract_fit_code(sku) in _PY_LONG_SLEEVE_FIT_CODES

def _py_matches_category(sku, brand_abbr, category):
    """Inclusive category matcher. One SKU can match multiple categories
    (e.g. a BC Carpenter matches 'pants', 'sportswear', AND 'young_men';
    a VD shacket with fit SS matches 'young_men', 'sportswear', AND 'short_sleeve')."""
    if not category or category in ('all', 'any'):
        return True
    if category == 'sportswear':
        return _py_is_sportswear(sku, brand_abbr)
    if category == 'pants':
        return _py_is_pants(sku, brand_abbr)
    if category == 'blazers':
        return _py_is_blazer(sku)
    if category == 'young_men':
        return _py_is_young_men(sku)
    if category == 'short_sleeve':
        return _py_is_short_sleeve(sku)
    if category == 'long_sleeve':
        return _py_is_long_sleeve_shirt(sku)
    # Non-overlapping categories fall through to the primary-category equality check
    return _py_get_item_category(sku, brand_abbr) == category

def _fresh_prepack_defaults():
    """Reload prepack rules from S3 and return a snapshot list.

    Export paths must NOT trust this worker's in-memory copy: Render runs
    multiple gunicorn workers with independent memory, so a rule save that
    landed on another worker is invisible here until we re-read S3 (same
    multi-worker pattern as GET /prepack-defaults). Falls back to whatever is
    in memory if the S3 read fails — never blocks an export."""
    try:
        load_prepack_defaults_from_s3()
    except Exception as e:
        print(f"  ⚠ _fresh_prepack_defaults: S3 reload failed, using in-memory copy: {e}")
    with _prepack_defaults_lock:
        return list(_prepack_defaults)


def _annotate_items_for_prepack(items):
    """Add _export_category, _export_fit, and _override_size_pack to raw inventory items (returns new list)."""
    result = []
    with _overrides_lock:
        overrides_snap = dict(_style_overrides)
    for item in items:
        sku        = item.get('sku', '')
        brand_abbr = item.get('brand_abbr', item.get('brand', ''))
        annotated  = dict(item)
        annotated.setdefault('_export_category', _py_get_item_category(sku, brand_abbr))
        annotated.setdefault('_export_fit',      _py_extract_fit_code(sku))
        annotated.setdefault('_export_customer',  sku[:2].upper() if len(sku) >= 2 else '')
        # Check for per-item size pack override (from Override Tool)
        if '_override_size_pack' not in annotated:
            sku_up = sku.upper()
            ov = overrides_snap.get(sku_up)
            if not ov:
                # Prefix match — keys ending with '-'
                for key in overrides_snap:
                    if key.endswith('-') and sku_up.startswith(key[:-1]):
                        ov = overrides_snap[key]
                        break
            if ov and ov.get('sizePack') and ov['sizePack'].get('sizes'):
                annotated['_override_size_pack'] = ov['sizePack']
            else:
                annotated['_override_size_pack'] = None
        result.append(annotated)
    return result


def get_image_url(item, s3_base_url):
    """Get brand-folder fallback URL"""
    brand_abbr = item.get('brand_abbr', item.get('brand', ''))
    folder_name = FOLDER_MAPPING.get(brand_abbr, brand_abbr)
    image_code = extract_image_code(item['sku'], brand_abbr)
    return f"{s3_base_url}/{folder_name}/{image_code}.jpg"


def get_style_override_url(sku):
    """Get STYLE+OVERRIDES URL — primary image source (matches frontend)"""
    base_style = get_base_style(sku)
    return f"{S3_OVERRIDES_IMG_URL}/{base_style}.jpg"


def _process_image_from_url(url, tw=TARGET_W, th=TARGET_H):
    """Download and resize an image from URL, trying .jpg/.png/.jpeg extensions"""
    if not (isinstance(url, str) and url.startswith('http')):
        return None

    headers = {'User-Agent': 'Mozilla/5.0'}
    base_url = url.rsplit('.', 1)[0]

    for ext in ['.jpg', '.png', '.jpeg']:
        try_url = base_url + ext
        try:
            resp = http_requests.get(try_url, headers=headers, timeout=5)
            if resp.status_code != 200:
                continue
            ct = resp.headers.get('Content-Type', '').lower()
            if 'image' not in ct:
                continue

            with PilImage.open(BytesIO(resp.content)) as im:
                im = ImageOps.exif_transpose(im)
                im.thumbnail((tw * 2, th * 2), PilImage.Resampling.LANCZOS)

                fmt = "PNG"
                if im.mode in ("RGBA", "LA") or (im.mode == "P" and "transparency" in im.info):
                    fmt = "PNG"
                else:
                    if im.mode != "RGB":
                        im = im.convert("RGB")
                    fmt = "JPEG"

                buf = BytesIO()
                im.save(buf, format=fmt, quality=85, optimize=True)
                raw = buf.getvalue()
                ow, oh = im.size

            wr = tw / ow
            hr = th / oh
            sf = min(wr, hr)
            return {
                'raw_bytes': raw,
                'x_scale': sf, 'y_scale': sf,
                'x_offset': (tw - ow * sf) / 2,
                'y_offset': (th - oh * sf) / 2,
                'url': try_url
            }
        except Exception:
            continue
    return None



def sync_dropbox_photos():
    """List files in Dropbox PHOTOS INVENTORY folder via API — just metadata, no downloading."""
    global _dropbox_photo_index, _dropbox_photos_last_sync
    token = get_dropbox_token()
    if not token:
        print("[Dropbox Photos] No Dropbox token configured, skipping", flush=True)
        return

    try:
        print(f"[Dropbox Photos] Listing files via API in: {DROPBOX_PHOTOS_PATH}", flush=True)
        headers = {
            'Authorization': f'Bearer {token}',
            'Content-Type': 'application/json'
        }

        new_index = {}  # image_code (uppercase) → dropbox_path
        new_sportswear_index = {}  # sportswear filename (uppercase) → True
        cursor = None
        total_files = 0

        # Initial list_folder call
        payload = {
            'path': DROPBOX_PHOTOS_PATH,
            'recursive': True,
            'limit': 2000
        }
        resp = http_requests.post(
            'https://api.dropboxapi.com/2/files/list_folder',
            headers=headers, json=payload, timeout=60
        )

        if resp.status_code == 401:
            print(f"[Dropbox Photos] Auth failed (401) — forcing token refresh...", flush=True)
            # Force refresh by resetting expiry
            global _dropbox_token_expires
            _dropbox_token_expires = 0
            token = get_dropbox_token()
            if not token:
                print(f"[Dropbox Photos] Could not refresh token, giving up", flush=True)
                return
            headers['Authorization'] = f'Bearer {token}'
            resp = http_requests.post(
                'https://api.dropboxapi.com/2/files/list_folder',
                headers=headers, json=payload, timeout=60
            )
            if resp.status_code == 401:
                print(f"[Dropbox Photos] Still 401 after refresh — check credentials", flush=True)
                return

        # If path not found, try to discover it
        if resp.status_code == 409:
            print(f"[Dropbox Photos] Path '{DROPBOX_PHOTOS_PATH}' not found, searching root...", flush=True)
            # List root to find the right folder
            root_resp = http_requests.post(
                'https://api.dropboxapi.com/2/files/list_folder',
                headers=headers,
                json={'path': '', 'recursive': False, 'limit': 500},
                timeout=30
            )
            if root_resp.status_code == 200:
                root_data = root_resp.json()
                folders = [e['path_display'] for e in root_data.get('entries', []) if e['.tag'] == 'folder']
                print(f"[Dropbox Photos] Root folders: {folders}", flush=True)
                # Try to find PHOTOS INVENTORY anywhere
                for folder in folders:
                    sub_resp = http_requests.post(
                        'https://api.dropboxapi.com/2/files/list_folder',
                        headers=headers,
                        json={'path': folder, 'recursive': False, 'limit': 500},
                        timeout=30
                    )
                    if sub_resp.status_code == 200:
                        sub_data = sub_resp.json()
                        sub_folders = [e['path_display'] for e in sub_data.get('entries', []) if e['.tag'] == 'folder']
                        for sf in sub_folders:
                            if 'photo' in sf.lower() and 'inventory' in sf.lower():
                                print(f"[Dropbox Photos] Found photos folder: {sf}", flush=True)
                                # Re-do the listing with discovered path
                                payload['path'] = sf
                                resp = http_requests.post(
                                    'https://api.dropboxapi.com/2/files/list_folder',
                                    headers=headers, json=payload, timeout=60
                                )
                                break
                        if resp.status_code == 200:
                            break
            if resp.status_code != 200:
                print(f"[Dropbox Photos] Could not find photos folder. API response: {resp.text[:300]}", flush=True)
                return

        if resp.status_code != 200:
            print(f"[Dropbox Photos] API error: {resp.status_code} {resp.text[:200]}", flush=True)
            return

        data = resp.json()

        while True:
            for entry in data.get('entries', []):
                if entry['.tag'] != 'file':
                    continue
                name = entry['name']
                path_lower = entry['path_lower']
                lower = name.lower()

                # Only image files, skip 1x folder and macOS metadata
                if not lower.endswith(('.jpg', '.jpeg', '.png')):
                    continue
                if '/1x/' in path_lower or '__macosx' in path_lower:
                    continue

                name_no_ext = os.path.splitext(name)[0]
                # Clean: remove " copy", " Copy 2", etc.
                clean = re.sub(r'\s*copy\s*\d*$', '', name_no_ext, flags=re.IGNORECASE).strip()
                # Normalize separators: both - and _ → _
                clean = clean.replace('-', '_')
                key = clean.upper()

                # Store the Dropbox path (for on-demand download)
                # Prefer .jpg over .png if duplicates exist
                if key not in new_index or lower.endswith('.jpg'):
                    new_index[key] = entry['path_display']
                total_files += 1

                # ── Sportswear sub-index ──────────────────────────────────
                # Files under any /SPORTSWEAR/ subfolder of PHOTOS INVENTORY
                # are registered for prefix-match lookup against SKU style codes.
                # Use the literal filename (no separator normalization) because
                # sportswear filenames don't contain dashes — they're solid like
                # GBPJ012, GBPJ012SL, GBPH002.
                if '/sportswear/' in path_lower:
                    sw_key = re.sub(r'\s*copy\s*\d*$', '', name_no_ext,
                                    flags=re.IGNORECASE).strip().upper()
                    new_sportswear_index[sw_key] = True
                    # Also register the literal filename in the main index so
                    # get_dropbox_thumbnail(matched_filename) can find the bytes.
                    if sw_key not in new_index or lower.endswith('.jpg'):
                        new_index[sw_key] = entry['path_display']

            if not data.get('has_more'):
                break

            # Continue listing
            resp = http_requests.post(
                'https://api.dropboxapi.com/2/files/list_folder/continue',
                headers=headers,
                json={'cursor': data['cursor']},
                timeout=60
            )
            if resp.status_code != 200:
                print(f"[Dropbox Photos] Continue error: {resp.status_code}", flush=True)
                break
            data = resp.json()

        with _dropbox_photo_lock:
            _dropbox_photo_index = new_index
            _dropbox_photos_last_sync = time.time()

        # Swap in the fresh sportswear sub-index and invalidate the match cache.
        global _sportswear_photo_index
        with _sportswear_match_lock:
            _sportswear_photo_index = new_sportswear_index
            _sportswear_match_cache.clear()

        # Clear image caches so they rebuild with Dropbox awareness
        _dropbox_img_cache.clear()
        _web_img_cache.clear()
        _img_cache.clear()

        print(f"[Dropbox Photos] ✓ Indexed {len(new_index)} unique images ({total_files} total files), {len(new_sportswear_index)} sportswear", flush=True)

    except Exception as e:
        print(f"[Dropbox Photos] Error: {e}", flush=True)
        import traceback
        traceback.print_exc()


# Cache for downloaded Dropbox images — disk-based for persistence across requests
_dropbox_img_cache = {}  # image_code → (bytes, content_type) — small in-memory LRU for hot images
_dropbox_img_cache_lock = threading.Lock()
DROPBOX_DISK_CACHE = os.environ.get('DROPBOX_DISK_CACHE', '/var/data/dropbox_cache')
os.makedirs(DROPBOX_DISK_CACHE, exist_ok=True)



def _download_dropbox_file(dropbox_path):
    """Download a single file from Dropbox via API."""
    token = get_dropbox_token()
    if not token:
        return None, None
    try:
        headers = {
            'Authorization': f'Bearer {token}',
            'Dropbox-API-Arg': json.dumps({'path': dropbox_path})
        }
        resp = http_requests.post(
            'https://content.dropboxapi.com/2/files/download',
            headers=headers, timeout=30
        )
        if resp.status_code == 200:
            ct = resp.headers.get('Content-Type', 'application/octet-stream').lower()
            if 'image' not in ct:
                ext = dropbox_path.lower().rsplit('.', 1)[-1]
                ct = 'image/jpeg' if ext in ('jpg', 'jpeg') else 'image/png'
            return resp.content, ct
    except Exception as e:
        print(f"[Dropbox Photos] Download error for {dropbox_path}: {e}")
    return None, None


def _get_disk_cache_path(image_code):
    """Get the disk cache file path for an image code."""
    return os.path.join(DROPBOX_DISK_CACHE, image_code)


def get_dropbox_image_bytes(image_code):
    """Get raw image bytes from Dropbox (disk-cached). Returns (bytes, content_type) or (None, None)."""
    key = image_code.upper().replace('-', '_')

    # 1. Check in-memory hot cache
    with _dropbox_img_cache_lock:
        cached = _dropbox_img_cache.get(key, 'MISS')
    if cached is None:
        return None, None  # Previously failed
    if cached != 'MISS':
        return cached

    # 2. Check disk cache
    disk_path = _get_disk_cache_path(key)
    if os.path.exists(disk_path + '.jpg'):
        try:
            with open(disk_path + '.jpg', 'rb') as f:
                data = f.read()
            result = (data, 'image/jpeg')
            with _dropbox_img_cache_lock:
                if len(_dropbox_img_cache) > 200:
                    _dropbox_img_cache.clear()
                _dropbox_img_cache[key] = result
            return result
        except Exception:
            pass
    elif os.path.exists(disk_path + '.png'):
        try:
            with open(disk_path + '.png', 'rb') as f:
                data = f.read()
            result = (data, 'image/png')
            with _dropbox_img_cache_lock:
                if len(_dropbox_img_cache) > 200:
                    _dropbox_img_cache.clear()
                _dropbox_img_cache[key] = result
            return result
        except Exception:
            pass

    # 3. Check if this image exists in the index
    dropbox_path = _dropbox_photo_index.get(key)
    if not dropbox_path:
        return None, None

    # 4. Download on demand and save to disk
    data, ct = _download_dropbox_file(dropbox_path)
    if data:
        ext = '.png' if 'png' in ct else '.jpg'
        try:
            with open(disk_path + ext, 'wb') as f:
                f.write(data)
        except Exception:
            pass
        with _dropbox_img_cache_lock:
            if len(_dropbox_img_cache) > 200:
                _dropbox_img_cache.clear()
            _dropbox_img_cache[key] = (data, ct)
        return data, ct

    return None, None


def _upload_to_s3_sync(image_code, data, content_type):
    """Upload an image to S3 DROPBOX_SYNC folder for CDN delivery."""
    try:
        s3 = get_s3()
        ext = '.png' if 'png' in content_type else '.jpg'
        key = f"{S3_DROPBOX_SYNC_PREFIX}/{image_code}{ext}"
        s3.put_object(
            Bucket=S3_BUCKET,
            Key=key,
            Body=data,
            ContentType=content_type,
            CacheControl='public, max-age=86400',
        )
        # Invalidate CloudFront so new/updated image is instantly live
        cf_path = f"/ALL+INVENTORY+Photos/DROPBOX_SYNC/{image_code}{ext}"
        threading.Thread(target=_invalidate_cloudfront, args=([cf_path],), daemon=True).start()
        return True
    except Exception as e:
        print(f"[S3 Sync] Upload failed for {image_code}: {e}")
        return False


def prewarm_dropbox_cache():
    """Background job: download ALL Dropbox images to disk cache. Only one worker runs this."""
    # Use a file lock so only one worker pre-warms
    lock_file = '/tmp/dropbox_prewarm.lock'
    try:
        # Check for stale lock (older than 30 minutes)
        if os.path.exists(lock_file):
            lock_age = time.time() - os.path.getmtime(lock_file)
            if lock_age > 1800:  # 30 min stale threshold
                print(f"[Dropbox Pre-warm] Removing stale lock ({lock_age:.0f}s old)", flush=True)
                os.unlink(lock_file)
            else:
                print(f"[Dropbox Pre-warm] Another worker is already pre-warming, skipping", flush=True)
                return
        fd = os.open(lock_file, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
        os.write(fd, str(os.getpid()).encode())
        os.close(fd)
    except FileExistsError:
        print(f"[Dropbox Pre-warm] Another worker is already pre-warming, skipping", flush=True)
        return

    try:
        if not _dropbox_photo_index:
            return

        # Count how many are already cached
        already_cached = 0
        to_download = []
        for key in _dropbox_photo_index:
            disk_path = _get_disk_cache_path(key)
            if os.path.exists(disk_path + '.jpg') or os.path.exists(disk_path + '.png'):
                already_cached += 1
            else:
                to_download.append(key)

        total = len(_dropbox_photo_index)
        print(f"[Dropbox Pre-warm] {already_cached}/{total} already cached on disk, {len(to_download)} to download...", flush=True)

        # Check which images already exist in S3 DROPBOX_SYNC
        s3_synced = set()
        try:
            s3 = get_s3()
            paginator = s3.get_paginator('list_objects_v2')
            for page in paginator.paginate(Bucket=S3_BUCKET, Prefix=S3_DROPBOX_SYNC_PREFIX + '/'):
                for obj in page.get('Contents', []):
                    fname = obj['Key'].rsplit('/', 1)[-1]
                    code = os.path.splitext(fname)[0].upper()
                    s3_synced.add(code)
            print(f"[Dropbox Pre-warm] {len(s3_synced)} images already in S3 DROPBOX_SYNC", flush=True)
        except Exception as e:
            print(f"[Dropbox Pre-warm] Could not list S3 sync folder: {e}", flush=True)

        # If everything is on disk AND S3, nothing to do
        if not to_download and s3_synced.issuperset(_dropbox_photo_index.keys()):
            print(f"[Dropbox Pre-warm] ✓ All {total} images on disk and S3", flush=True)
            return

        downloaded = [0]
        failed = [0]
        s3_uploaded = [0]

        def _download_one(key):
            dropbox_path = _dropbox_photo_index.get(key)
            if not dropbox_path:
                return

            # Check if already on disk
            disk_path = _get_disk_cache_path(key)
            already_on_disk = os.path.exists(disk_path + '.jpg') or os.path.exists(disk_path + '.png')
            already_on_s3 = key in s3_synced

            # Skip if both disk and S3 are done
            if already_on_disk and already_on_s3:
                return

            # Read from disk if available, otherwise download from Dropbox
            data = None
            ct = None
            if already_on_disk:
                try:
                    ext = '.jpg' if os.path.exists(disk_path + '.jpg') else '.png'
                    with open(disk_path + ext, 'rb') as f:
                        data = f.read()
                    ct = 'image/jpeg' if ext == '.jpg' else 'image/png'
                except Exception:
                    pass

            if not data:
                data, ct = _download_dropbox_file(dropbox_path)
                if data:
                    ext = '.png' if 'png' in ct else '.jpg'
                    try:
                        with open(disk_path + ext, 'wb') as f:
                            f.write(data)
                        downloaded[0] += 1
                    except Exception:
                        failed[0] += 1
                        return
                else:
                    failed[0] += 1
                    return

            # Upload to S3 if not already there
            if not already_on_s3 and data:
                if _upload_to_s3_sync(key, data, ct):
                    s3_uploaded[0] += 1

            # Progress update every 200 images
            done = downloaded[0] + s3_uploaded[0]
            if done % 200 == 0 and done > 0:
                print(f"[Dropbox Pre-warm] Progress: {downloaded[0]} downloaded, {s3_uploaded[0]} synced to S3, {failed[0]} failed", flush=True)

        # Process ALL keys — some may be on disk but not S3 yet
        all_keys = list(_dropbox_photo_index.keys())

        # Use gevent pool — only 3 concurrent to limit memory
        try:
            from gevent.pool import Pool
            pool = Pool(size=3)
            pool.map(_download_one, all_keys)
        except ImportError:
            for key in all_keys:
                _download_one(key)

        print(f"[Dropbox Pre-warm] ✓ Done! {downloaded[0]} new downloads, {s3_uploaded[0]} synced to S3, {failed[0]} failed", flush=True)
    finally:
        try:
            os.unlink(lock_file)
        except Exception:
            pass


DROPBOX_THUMB_CACHE_DIR = os.path.join(os.path.dirname(DROPBOX_DISK_CACHE), 'dropbox_thumbs')
os.makedirs(DROPBOX_THUMB_CACHE_DIR, exist_ok=True)

def _get_thumb_disk_path(key):
    return os.path.join(DROPBOX_THUMB_CACHE_DIR, key + '.thumb')

def get_dropbox_thumbnail(image_code, tw=TARGET_W, th=TARGET_H):
    """Get resized thumbnail from Dropbox for Excel exports.
    Priority: in-memory → disk thumb cache → raw disk cache → Dropbox download.
    Thumbnails are persisted to disk so PIL processing only ever happens once."""
    import struct
    key = image_code.upper().replace('-', '_')

    # 1. In-memory hot cache (fastest)
    if key in _dropbox_thumb_cache:
        return _dropbox_thumb_cache[key]

    # Check index first — skip entirely for unknown images
    if key not in _dropbox_photo_index:
        return None

    def _cache_in_memory(result):
        """Add to in-memory cache with soft cap — disk is the real cache."""
        if len(_dropbox_thumb_cache) > 500:
            for old_key in list(_dropbox_thumb_cache.keys())[:100]:
                del _dropbox_thumb_cache[old_key]
        _dropbox_thumb_cache[key] = result
        return result

    # 2. Disk thumb cache — pre-processed, no PIL needed
    thumb_path = _get_thumb_disk_path(key)
    if os.path.exists(thumb_path):
        try:
            with open(thumb_path, 'rb') as f:
                meta = f.read(32)
                raw = f.read()
            x_scale, y_scale, x_offset, y_offset = struct.unpack('4d', meta)
            result = {
                'raw_bytes': raw,
                'x_scale': x_scale, 'y_scale': y_scale,
                'x_offset': x_offset, 'y_offset': y_offset,
                'url': f'dropbox://{key}'
            }
            return _cache_in_memory(result)
        except Exception:
            pass  # Corrupted — fall through to regenerate

    # 3. Get raw image bytes (disk or Dropbox API)
    raw_bytes, ct = get_dropbox_image_bytes(image_code)
    if not raw_bytes:
        return None

    # 4. PIL resize — happens at most once per image ever
    try:
        with PilImage.open(BytesIO(raw_bytes)) as im:
            im = ImageOps.exif_transpose(im)
            im.thumbnail((tw * 2, th * 2), PilImage.Resampling.LANCZOS)
            if im.mode in ("RGBA", "LA") or (im.mode == "P" and "transparency" in im.info):
                fmt = "PNG"
            else:
                if im.mode != "RGB":
                    im = im.convert("RGB")
                fmt = "JPEG"
            buf = BytesIO()
            im.save(buf, format=fmt, quality=85, optimize=True)
            processed = buf.getvalue()
            ow, oh = im.size
        wr = tw / ow
        hr = th / oh
        sf = min(wr, hr)
        x_off = (tw - ow * sf) / 2
        y_off = (th - oh * sf) / 2
        result = {
            'raw_bytes': processed,
            'x_scale': sf, 'y_scale': sf,
            'x_offset': x_off, 'y_offset': y_off,
            'url': f'dropbox://{key}'
        }
        # Persist to disk — next server restart skips PIL entirely
        try:
            with open(thumb_path, 'wb') as f:
                f.write(struct.pack('4d', sf, sf, x_off, y_off))
                f.write(processed)
        except Exception:
            pass
        return _cache_in_memory(result)
    except Exception:
        return None
def get_image_cached(item, s3_base_url):
    """
    Get image for an item, using cache keyed by base_style.
    Priority: base64 override → STYLE+OVERRIDES → Dropbox → brand folder fallback.
    All size variants of the same style share one cache entry.

    IMPORTANT: Platform base64 overrides are checked BEFORE the cache lookup so
    that a newly-uploaded override always wins, even if a stale fallback image
    is sitting in _img_cache from a prior export run. Override-derived bytes are
    cached under a content-versioned key (`__override__:{style}:{hash}`), so a
    NEW override naturally produces a new cache key and instantly takes effect
    without manual invalidation. The plain `base_style` cache key is reserved
    for non-override fallback images (CloudFront / Dropbox / brand folder).
    """
    sku = item.get('sku', '')
    base_style = get_base_style(sku)

    # ── PRIORITY 1: Platform base64 override (checked BEFORE cache lookup) ──
    # Snapshot under lock so we don't race with /overrides POST.
    with _overrides_lock:
        override_data = _style_overrides.get(base_style)

    if override_data and isinstance(override_data, dict) and override_data.get('image'):
        import base64, hashlib
        img_str_raw = override_data['image']
        # Version the cache key on the override content so a new upload
        # produces a different key. Hash a prefix to keep this cheap.
        try:
            sample = (img_str_raw[:512] + img_str_raw[-128:]) if isinstance(img_str_raw, str) else str(img_str_raw)[:512]
            img_hash = hashlib.md5(sample.encode('utf-8', errors='ignore')).hexdigest()[:12]
        except Exception:
            img_hash = 'nohash'
        override_cache_key = f"__override__:{base_style}:{img_hash}"

        # Check override-versioned cache first — fast path for repeat exports
        with _img_lock:
            cached = _img_cache.get(override_cache_key)
        if cached is not None:
            return {
                'image_data': BytesIO(cached['raw_bytes']),
                'x_scale': cached['x_scale'], 'y_scale': cached['y_scale'],
                'x_offset': cached['x_offset'], 'y_offset': cached['y_offset'],
                'object_position': 1, 'url': cached['url']
            }

        # Decode + render the override image
        try:
            img_str = img_str_raw
            if ',' in img_str:
                img_str = img_str.split(',', 1)[1]
            raw = base64.b64decode(img_str)
            with PilImage.open(BytesIO(raw)) as im:
                im = ImageOps.exif_transpose(im)
                tw, th = TARGET_W, TARGET_H
                im.thumbnail((tw * 2, th * 2), PilImage.Resampling.LANCZOS)
                if im.mode in ("RGBA", "LA") or (im.mode == "P" and "transparency" in im.info):
                    fmt = "PNG"
                else:
                    if im.mode != "RGB":
                        im = im.convert("RGB")
                    fmt = "JPEG"
                buf = BytesIO()
                im.save(buf, format=fmt, quality=85, optimize=True)
                ow, oh = im.size
            wr = tw / ow
            hr = th / oh
            sf = min(wr, hr)
            result = {
                'raw_bytes': buf.getvalue(),
                'x_scale': sf, 'y_scale': sf,
                'x_offset': (tw - ow * sf) / 2,
                'y_offset': (th - oh * sf) / 2,
                'url': f'override://{base_style}'
            }
            with _img_lock:
                _img_cache[override_cache_key] = result
            return {
                'image_data': BytesIO(result['raw_bytes']),
                'x_scale': result['x_scale'], 'y_scale': result['y_scale'],
                'x_offset': result['x_offset'], 'y_offset': result['y_offset'],
                'object_position': 1, 'url': result['url']
            }
        except Exception:
            # Fall through to fallback chain if override processing fails
            pass

    # ── No override (or override failed) — use cache + fallback chain ──
    with _img_lock:
        if base_style in _img_cache:
            c = _img_cache[base_style]
            if c is None:
                return None  # Previously failed — skip
            return {
                'image_data': BytesIO(c['raw_bytes']),
                'x_scale': c['x_scale'], 'y_scale': c['y_scale'],
                'x_offset': c['x_offset'], 'y_offset': c['y_offset'],
                'object_position': 1, 'url': c['url']
            }

    result = None

    # 2. Try STYLE+OVERRIDES via CloudFront.
    #    An override uploaded against a SKU that carries a suffix (ROBLCH001-WS,
    #    the "-WS"/"-V" variants) is stored under that FULL sku, not the base —
    #    so try the full sku first and only then the base. Looking up the base
    #    alone made those styles export as "No Image" even though the catalog
    #    showed them fine (Black Label line sheet, Aug 2026).
    base_style_for_url = get_base_style(sku)
    sku_u = str(sku).upper()
    override_candidates = [sku_u] if sku_u != base_style_for_url else []
    override_candidates.append(base_style_for_url)
    for _cand in override_candidates:
        # S3 direct, not CloudFront — see the matching note in _fetch_raw_image:
        # server-side reads must always see the current file (edge POPs held a
        # replaced image for up to 24h; querystring busting is a no-op here).
        override_url = f"{S3_OVERRIDES_IMG_URL}/{_cand}.jpg"
        result = _process_image_from_url(override_url)
        if result:
            break

    # 2.5. Try sportswear-folder match (longest-prefix on filename — see
    #      find_sportswear_image_match docstring). This is what catches
    #      Geoffrey Beene SPORTSWEAR/ photos and any other brand that
    #      eventually adds a SPORTSWEAR/ subfolder. Runs before the
    #      generic Dropbox call because that one uses extract_image_code's
    #      prefix_number scheme, which doesn't match sportswear filenames.
    if not result:
        sw_match = find_sportswear_image_match(get_base_style(sku),
                                               item.get('brand_abbr', item.get('brand', '')))
        if sw_match:
            result = get_dropbox_thumbnail(sw_match)

    # Bottoms never use the brand+serial rungs below (fabric-blind keys —
    # VD pants were exporting the same-serial shirt photo). See is_bottoms_style.
    _no_serial_fallback = is_bottoms_style(sku)

    # 3. Try Dropbox photos
    if not result and not _no_serial_fallback:
        brand_abbr = item.get('brand_abbr', item.get('brand', ''))
        image_code = extract_image_code(sku, brand_abbr)
        result = get_dropbox_thumbnail(image_code)

    # 4. Fallback to CloudFront brand folder
    if not result and not _no_serial_fallback:
        brand_abbr_cf = item.get('brand_abbr', item.get('brand', ''))
        folder_name_cf = FOLDER_MAPPING.get(brand_abbr_cf, brand_abbr_cf)
        image_code_cf = extract_image_code(sku, brand_abbr_cf)
        brand_url = f"{CLOUDFRONT_PHOTOS_URL}/{folder_name_cf}/{image_code_cf}.jpg"
        result = _process_image_from_url(brand_url)

    # 5. VERY LAST resort: fallback swatch extracted from the brand
    #    style-number files (bottom priority — mirrors the frontend rung so
    #    Excel export thumbnails match what the catalog shows)
    if not result and not _no_serial_fallback:
        swatch_url = f"{CLOUDFRONT_SWATCH_FALLBACK_URL}/{image_code_cf}.jpg"
        result = _process_image_from_url(swatch_url)

    # Cache result (even None to avoid re-fetching failures)
    with _img_lock:
        _img_cache[base_style] = result

    if result:
        return {
            'image_data': BytesIO(result['raw_bytes']),
            'x_scale': result['x_scale'], 'y_scale': result['y_scale'],
            'x_offset': result['x_offset'], 'y_offset': result['y_offset'],
            'object_position': 1, 'url': result['url']
        }
    return None


def download_images_for_items(items, s3_base_url, use_cache=True):
    """Download images for all items using thread pool. Deduplicates by base_style."""
    results = {}

    # Deduplicate: only fetch once per unique base_style
    style_to_indices = {}  # base_style → list of item indices
    unique_items = {}      # base_style → first item with that style
    for i, item in enumerate(items):
        base_style = get_base_style(item.get('sku', ''))
        if base_style not in style_to_indices:
            style_to_indices[base_style] = []
            unique_items[base_style] = item
        style_to_indices[base_style].append(i)

    def _fetch(base_style_item):
        base_style, item = base_style_item
        try:
            img = get_image_cached(item, s3_base_url)
            return base_style, img
        except Exception:
            return base_style, None

    # Clear previously-failed cache entries so they get retried
    with _img_lock:
        failed_keys = [k for k, v in _img_cache.items() if v is None]
        for k in failed_keys:
            del _img_cache[k]

    unique_pairs = list(unique_items.items())
    print(f"    Fetching images: {len(unique_pairs)} unique styles for {len(items)} items")

    def _store_results(base_style, img):
        """Store fetched image data for all item indices sharing this base_style."""
        if not img:
            return
        for idx in style_to_indices.get(base_style, []):
            # Use the returned image data directly — don't re-read from cache
            try:
                raw = img.get('image_data')
                if raw:
                    raw.seek(0)
                    raw_bytes = raw.read()
                    results[idx] = {
                        'image_data': BytesIO(raw_bytes),
                        'x_scale': img.get('x_scale', 1), 'y_scale': img.get('y_scale', 1),
                        'x_offset': img.get('x_offset', 0), 'y_offset': img.get('y_offset', 0),
                        'object_position': 1, 'url': img.get('url', '')
                    }
            except Exception:
                pass

    # Use gevent pool if available (gevent monkey-patches break ThreadPoolExecutor)
    try:
        import gevent.pool
        pool = gevent.pool.Pool(size=30)
        fetch_results = pool.map(_fetch, unique_pairs)
        for result in fetch_results:
            try:
                base_style, img = result
                _store_results(base_style, img)
            except Exception:
                pass
    except ImportError:
        # Fallback to ThreadPoolExecutor if gevent not installed
        with concurrent.futures.ThreadPoolExecutor(max_workers=30) as pool:
            futures = {pool.submit(_fetch, pair): pair[0] for pair in unique_pairs}
            for f in concurrent.futures.as_completed(futures):
                try:
                    base_style, img = f.result()
                    _store_results(base_style, img)
                except Exception:
                    pass

    found = len(results)
    missed = len(items) - found
    print(f"    Image results: {found}/{len(items)} found, {missed} missing")
    return results


# Factory lookup — the first 2 letters of a Production # identify the factory
# producing the goods. ALL exports (admin and customer alike) show only the
# 2-letter prefix; the full-name map below is kept as an internal reference
# for what each prefix means.
FACTORY_NAMES = {
    'TF': 'Top Find',
    'NB': 'Yuxiu',
    'PC': 'Pinnacle',
    'DP': 'David Peng',
    'FR': 'Frank',
    'NK': 'Najmul',
}


def _factory_label(production_ref, full_name=False):
    """Derive the factory from a production reference number.

    full_name=False → 2-letter prefix only (used by ALL exports since Jul 2026).
    full_name=True  → full factory name from FACTORY_NAMES (internal use only).
    """
    ref = str(production_ref or '').strip()
    if len(ref) < 2:
        return ''
    prefix = ref[:2].upper()
    if not prefix.isalpha():
        return ''
    if full_name:
        return FACTORY_NAMES.get(prefix, prefix)
    return prefix


def _setup_worksheet(workbook, worksheet, has_color=False, view_mode='all',
                     is_order=False, incoming_only=False, catalog_mode=False,
                     flow_mode=False, headers_override=None):
    fmt_header = workbook.add_format({
        'bold': True, 'font_name': STYLE_CONFIG['font_name'], 'font_size': 11,
        'bg_color': STYLE_CONFIG['header_bg'], 'font_color': STYLE_CONFIG['header_text'],
        'border': 1, 'border_color': STYLE_CONFIG['border_color'],
        'align': 'center', 'valign': 'vcenter'
    })
    base = {
        'font_name': STYLE_CONFIG['font_name'], 'font_size': 10,
        'text_wrap': True, 'valign': 'vcenter', 'align': 'center',
        'border': 1, 'border_color': STYLE_CONFIG['border_color']
    }
    fmts = {
        'odd':  workbook.add_format({**base, 'bg_color': STYLE_CONFIG['row_bg_odd']}),
        'even': workbook.add_format({**base, 'bg_color': STYLE_CONFIG['row_bg_even']}),
        'num_odd':  workbook.add_format({**base, 'bg_color': STYLE_CONFIG['row_bg_odd'],  'num_format': '#,##0'}),
        'num_even': workbook.add_format({**base, 'bg_color': STYLE_CONFIG['row_bg_even'], 'num_format': '#,##0'}),
        # Highlight rows (item['_yellow']) — Can Ship Today substitution rows
        'yellow':     workbook.add_format({**base, 'bg_color': '#FFF3B0'}),
        'num_yellow': workbook.add_format({**base, 'bg_color': '#FFF3B0', 'num_format': '#,##0'}),
        # Currency columns (Allocation Dollar Value Estimated report)
        'price_odd':  workbook.add_format({**base, 'bg_color': STYLE_CONFIG['row_bg_odd'],  'num_format': '$#,##0.00'}),
        'price_even': workbook.add_format({**base, 'bg_color': STYLE_CONFIG['row_bg_even'], 'num_format': '$#,##0.00'}),
        'val_odd':    workbook.add_format({**base, 'bg_color': STYLE_CONFIG['row_bg_odd'],  'num_format': '$#,##0'}),
        'val_even':   workbook.add_format({**base, 'bg_color': STYLE_CONFIG['row_bg_even'], 'num_format': '$#,##0'}),
    }

    worksheet.hide_gridlines(2)
    worksheet.freeze_panes(1, 0)

    if headers_override is not None:
        # Caller-defined column set (ship-plan line sheets) — same visual
        # treatment, custom columns. Widths/formats still resolve by name.
        headers = list(headers_override)
    elif catalog_mode:
        # ── Catalog exports: no committed/allocated, simplified layout ──
        if view_mode == 'incoming':
            headers = ['IMAGE', 'SKU', 'Brand']
            if has_color:
                headers.append('Color')
            headers.extend(['Fit', 'Fabrication'])
            # 📋 Customer-facing overseas: PO Ref # column (from style ledger column A — the
            # production reference number) when at least one cart row maps to a specific delivery.
            # Factory = 2-letter prefix of the production ref (customers never see full names).
            if flow_mode:
                headers.extend(['PO Ref #', 'Factory'])
            if is_order:
                headers.append('Qty Selected')
            # Customer catalog overseas view: NO Incoming column — per-delivery rows
            # already show Overseas ATS for each arrival date via flow_mode expansion.
            headers.extend(['Overseas ATS', 'Ex-Factory', 'Arrival'])
        else:
            # Warehouse / All view: show warehouse names column instead of per-WH quantities
            headers = ['IMAGE', 'SKU', 'Brand']
            if has_color:
                headers.append('Color')
            headers.extend(['Fit', 'Fabrication'])
            if is_order:
                headers.append('Qty Selected')
            headers.extend(['Warehouse', 'Total ATS'])
            # All Inventory catalog view: include date columns for items with incoming stock
            if view_mode == 'all':
                headers.extend(['Ex-Factory', 'Arrival'])
                # 📋 Mixed-cart support: when customer has overseas items in cart with specific
                # arrivals (flow_mode=true), include PO Ref # + Factory prefix.
                # Warehouse rows get blank cells.
                if flow_mode:
                    headers.extend(['PO Ref #', 'Factory'])
    elif view_mode == 'incoming':
        # Admin overseas view: no warehouse columns, add dates
        headers = ['IMAGE', 'SKU', 'Brand']
        if has_color:
            headers.append('Color')
        headers.extend(['Fit', 'Fabrication'])
        # Flow mode (admin only): add Production #, PO Name and Factory columns.
        # Factory shows the 2-letter prefix, same as customer exports.
        if flow_mode and not catalog_mode:
            headers.extend(['Production #', 'PO Name', 'Factory'])
        if is_order:
            headers.append('Qty Selected')
        headers.extend(['Incoming', 'Committed', 'Allocated', 'Overseas ATS',
                        'Ex-Factory', 'Arrival'])
    else:
        # Admin standard / ATS / All view — full columns
        headers = ['IMAGE', 'SKU', 'Brand']
        if has_color:
            headers.append('Color')
        headers.extend(['Fit', 'Fabrication', 'Delivery'])
        if is_order:
            headers.append('Qty Selected')
        if not incoming_only:
            headers.extend(['JTW', 'TR', 'DCW', 'QA', 'NJ'])
        headers.append('Incoming')
        if not incoming_only:
            headers.append('Total Warehouse')
        headers.append('Total ATS')

    worksheet.set_row(0, 25)
    for c, h in enumerate(headers):
        worksheet.write(0, c, h, fmt_header)

    # Set column widths based on header names
    col_widths = {
        'IMAGE': COL_WIDTH_UNITS, 'SKU': 20, 'Brand': 20, 'Color': 18,
        'Fit': 12, 'Fabrication': 35, 'Delivery': 14, 'Qty Selected': 14,
        'Production #': 16, 'PO Name': 30, 'PO Ref #': 22, 'Factory': 14,
        'JTW': 12, 'TR': 12, 'DCW': 12, 'QA': 12, 'NJ': 12, 'Incoming': 12,
        'Total Warehouse': 14, 'Total ATS': 12, 'Overseas ATS': 14,
        'Committed': 12, 'Allocated': 12, 'Ex-Factory': 14, 'Arrival': 14,
        'Warehouse': 18,
        # Ship-plan line-sheet columns (Confirm Pre-PO exports)
        'Units to Ship': 13, 'Shortfall': 12, 'Can Ship': 15,
        'Held By / Source': 34,
        # Customer Selling line sheets
        'Trajectory': 34, 'Seen': 24,
        # APO "POs Ready to Ship" tab
        'PO #': 16, 'Units': 12, 'Ship Window': 24,
        'Available': 32, 'Original Arrival': 20,
        # Allocation Dollar Value Estimated report
        'Est. Price': 12, 'Est. Value': 14,
    }
    for c, h in enumerate(headers):
        worksheet.set_column(c, c, col_widths.get(h, 12))

    worksheet.set_default_row(112.5)
    return fmts, headers


def _write_rows(workbook, worksheet, data, images, fmts, has_color=False,
                view_mode='all', headers=None, catalog_mode=False):
    """Write data rows using headers list to determine column layout."""
    if not headers:
        headers = []

    # Map header names to data field getters
    FIELD_MAP = {
        'IMAGE': lambda item: '',
        'SKU': lambda item: item.get('sku', ''),
        'Brand': lambda item: item.get('brand_full', ''),
        'Color': lambda item: item.get('color', ''),
        'Fit': lambda item: item.get('fit', 'N/A'),
        'Fabrication': lambda item: item.get('fabrication', 'Standard Fabric'),
        'Delivery': lambda item: item.get('delivery', 'ATS'),
        'Production #': lambda item: item.get('production', ''),
        'PO Name': lambda item: item.get('po_name', ''),
        'PO Ref #': lambda item: item.get('po_ref', ''),
        # Factory derived from production ref prefix — 2-letter abbreviation
        # only, for admin and customer (catalog_mode) exports alike.
        'Factory': lambda item: _factory_label(
            item.get('production') or item.get('po_ref')),
        'Qty Selected': lambda item: item.get('quantity_ordered', 0),
        'JTW': lambda item: item.get('jtw', 0),
        'TR': lambda item: item.get('tr', 0),
        'DCW': lambda item: item.get('dcw', 0),
        'QA': lambda item: item.get('qa', 0),
        'NJ': lambda item: item.get('nj', 0),
        'Incoming': lambda item: item.get('incoming', 0),
        'Total Warehouse': lambda item: item.get('total_warehouse', 0),
        'Total ATS': lambda item: item.get('total_ats', 0),
        'Overseas ATS': lambda item: item.get('total_ats', 0),
        'Committed': lambda item: item.get('committed', 0),
        'Allocated': lambda item: item.get('allocated', 0),
        'Ex-Factory': lambda item: item.get('ex_factory', ''),
        'Arrival': lambda item: item.get('arrival', ''),
        'Warehouse': lambda item: item.get('warehouse', ''),
        # Ship-plan line-sheet columns (Confirm Pre-PO exports)
        'Units to Ship': lambda item: item.get('units_ship', 0),
        'Shortfall': lambda item: item.get('shortfall', 0),
        'Can Ship': lambda item: item.get('can_ship', ''),
        'Held By / Source': lambda item: item.get('source', ''),
        # Customer Selling line sheets (best/okay/worst tabs)
        'Trajectory': lambda item: item.get('trajectory', ''),
        'Seen': lambda item: item.get('seen', ''),
        # APO "POs Ready to Ship" tab
        'PO #': lambda item: item.get('po', ''),
        'Units': lambda item: item.get('units_alloc', 0),
        'Ship Window': lambda item: item.get('ship_window', ''),
        'Available': lambda item: item.get('available', ''),
        'Original Arrival': lambda item: item.get('orig_arrival', ''),
        # Allocation Dollar Value Estimated report
        'Est. Price': lambda item: item.get('est_price') if item.get('est_price') is not None else '—',
        'Est. Value': lambda item: item.get('est_value') if item.get('est_value') is not None else '—',
    }

    # Determine which columns are numeric for formatting
    NUMERIC_HEADERS = {
        'Qty Selected', 'JTW', 'TR', 'DCW', 'QA', 'NJ', 'Incoming',
        'Total Warehouse', 'Total ATS', 'Overseas ATS',
        'Committed', 'Allocated', 'Units to Ship', 'Shortfall',
        'Units'
    }

    for r, item in enumerate(data):
        row = r + 1
        even = r % 2 == 1
        if item.get('_yellow'):
            cf, nf = fmts['yellow'], fmts['num_yellow']
        else:
            cf = fmts['even'] if even else fmts['odd']
            nf = fmts['num_even'] if even else fmts['num_odd']
        pf = fmts.get('price_even' if even else 'price_odd', nf)
        vf = fmts.get('val_even' if even else 'val_odd', nf)

        for c, h in enumerate(headers):
            getter = FIELD_MAP.get(h)
            val = getter(item) if getter else ''
            if h == 'Est. Price':
                fmt = pf
            elif h == 'Est. Value':
                fmt = vf
            else:
                fmt = nf if h in NUMERIC_HEADERS else cf
            worksheet.write(row, c, val, fmt)

        img = images.get(r)
        if img:
            try:
                worksheet.insert_image(row, 0, "img.png", _padded_image_opts(img))
            except Exception:
                worksheet.write(row, 0, "Error", cf)
        else:
            worksheet.write(row, 0, "No Image", cf)
    return len(data)


_SIZED_SKU_BASE_RE = re.compile(r'^[A-Z]{6}\d{3}[A-Z]{2,3}$')
_SIZED_ALPHA_RE = re.compile(r'^(XS|S|M|L|XL|XXL|XXXL|[2-4]XL|ST|MT|LT|XLT|[2-3]XLT)$')

def _sized_sku_label(sku):
    """AMNASU576SLP-1515.53233 -> '15-15.5 / 32-33'. A specific-size SKU is one
    size, so its case pack is a SOLID pack of that size, never the style's
    mixed run. Modern bases only (legacy excluded); bare -V is a style variant,
    not a size. Mirrors the frontend parseSizedSkuLabel."""
    s = str(sku or '').upper().strip()
    di = s.find('-')
    if di < 0:
        return None
    base = s[:di]
    if not _SIZED_SKU_BASE_RE.match(base):
        return None
    suf = s[di + 1:]
    if suf.endswith('-FBA'):
        suf = suf[:-4]
    if suf.endswith('-V'):
        suf = suf[:-2]
    if not suf or suf in ('V', 'FBA'):
        return None
    def neck_ok(n):
        try: return 13 <= float(n) <= 23
        except Exception: return False
    def sl_ok(n):
        try: return 28 <= float(n) <= 40
        except Exception: return False
    m = re.match(r'^(\d{2})(\d{2}(?:\.\d)?)(\d{2})(\d{2})$', suf)
    if m:
        return (f'{m.group(1)}-{m.group(2)} / {m.group(3)}-{m.group(4)}'
                if neck_ok(m.group(1)) and neck_ok(m.group(2)) and sl_ok(m.group(3)) and sl_ok(m.group(4)) else None)
    m = re.match(r'^(\d{2}(?:\.\d)?)-(\d{2}(?:\.\d)?)(\d{2})/(\d{2})$', suf)
    if m:
        return (f'{m.group(1)}-{m.group(2)} / {m.group(3)}-{m.group(4)}'
                if neck_ok(m.group(1)) and neck_ok(m.group(2)) and sl_ok(m.group(3)) and sl_ok(m.group(4)) else None)
    m = re.match(r'^(\d{2}(?:\.\d)?)-(\d{2})(?:/(\d{2}))?$', suf)
    if m:
        if neck_ok(m.group(1)) and sl_ok(m.group(2)) and (not m.group(3) or sl_ok(m.group(3))):
            return f'{m.group(1)} / {m.group(2)}' + (f'-{m.group(3)}' if m.group(3) else '')
        return None
    if _SIZED_ALPHA_RE.match(suf):
        return suf
    return None

def _add_size_charts(workbook, worksheet, start, prepack_defaults=None, items=None):
    """
    Render prepack size scale grids vertically at the bottom of the worksheet.
    One grid per block of rows, stacked top-to-bottom.
    Only shows rules that match items on THIS tab — so each brand tab is independent.
    Falls back to hardcoded Slim/Regular if no rules matched.
    """
    t  = workbook.add_format({'bold':True, 'font_name':'Calibri', 'font_size':11,
                               'bg_color':'#FFFFFF', 'border':0, 'align':'left', 'valign':'vcenter'})
    s  = workbook.add_format({'bold':True, 'font_name':'Calibri', 'font_size':10,
                               'bg_color':'#FFFFFF', 'font_color':'#CC0000', 'border':0,
                               'align':'left', 'valign':'vcenter'})
    gh = workbook.add_format({'bold':True, 'font_name':'Calibri', 'font_size':10,
                               'border':1, 'align':'center', 'valign':'vcenter', 'bg_color':'#F3F4F6'})
    gd = workbook.add_format({'font_name':'Calibri', 'font_size':10,
                               'border':1, 'align':'center', 'valign':'vcenter', 'bg_color':'#FFFFFF'})

    # ── Determine which rules to show ─────────────────────────────────────
    packs_to_render = []

    solid_sizes = {}   # sized SKUs: size label -> master qty (solid pack)

    if prepack_defaults and items:
        seen_keys = {}
        for item in items:
            cat = item.get('_export_category', '')
            fit = item.get('_export_fit', '')
            cust = item.get('_export_customer', '')
            sku = item.get('sku', '').upper()
            base = sku.split('-')[0]
            # A specific-size SKU never contributes the style's mixed grid: it
            # rolls into one consolidated "solid size packs" grid instead.
            sized_label = _sized_sku_label(sku)

            # PRIORITY A: Per-item Override Tool size pack
            override_pack = item.get('_override_size_pack')
            if override_pack and isinstance(override_pack, dict) and override_pack.get('sizes'):
                if sized_label:
                    try:
                        solid_sizes[sized_label] = int(override_pack.get('master_qty') or 36)
                    except Exception:
                        solid_sizes[sized_label] = 36
                    continue
                key = ('__override__', sku)
                if key not in seen_keys:
                    seen_keys[key] = {
                        'label': f"Override: {base}",
                        'master_qty': override_pack.get('master_qty', '?'),
                        'inner_qty':  override_pack.get('inner_qty', '?'),
                        'sizes':      override_pack.get('sizes', []),
                        '_is_override': True
                    }
                continue  # This item is handled — don't also add a prepack rule for it

            # PRIORITY B: SKU-specific assignment on any prepack rule
            sku_matched = None
            for rule in prepack_defaults:
                # 'sk', not 's' — 's' is the red subtitle cell format defined above;
                # shadowing it made every "SIZE SCALE TO USE" row silently disappear.
                for sk in (rule.get('skus') or []):
                    su = sk.upper().strip()
                    if su and (su == sku or su == base or sku.startswith(su)):
                        sku_matched = rule
                        break
                if sku_matched:
                    break

            matched = sku_matched
            if not matched:
                # PRIORITY C: Dimension matching — mirrors frontend matchPrepackDefault().
                # Two-tier scoring:
                #   TIER A (🎯 Specific) = rule has a specific category AND specific fabrics
                #   TIER B (☂️ Umbrella) = everything else
                # Within a tier, +1 per specified dimension (category, fit, customer, brand, fabrics)
                # breaks ties. Ties within the same score fall back to list order (first wins).
                brand_abbr = item.get('brand_abbr', item.get('brand', '')).upper()
                best_fab_spec = -1  # 0 or 1: the primary tier flag
                best_score    = -1  # secondary — dimension count within tier
                best_rule     = None

                # Extract fabric code from SKU (positions 4:6)
                fab = base[4:6] if len(base) >= 6 else ''

                for r in prepack_defaults:
                    # Skip rules that have a SKU list — they should ONLY match via Priority B (exact SKU),
                    # never act as a catch-all through dimension scoring
                    r_skus = r.get('skus') or []
                    if isinstance(r_skus, list) and len([s for s in r_skus if s and s.strip()]) > 0:
                        continue

                    # Category — use INCLUSIVE match so a BC Carpenter (which is pants+sportswear+young_men)
                    # can match a 'pants' rule, a 'sportswear' rule, OR a 'young_men' rule.
                    r_cat = r.get('category', 'any')
                    if r_cat and r_cat != 'any' and not _py_matches_category(sku, brand_abbr, r_cat):
                        continue

                    # Fits: if specified, item must match
                    r_fits = r.get('fits')
                    if isinstance(r_fits, list):
                        r_fits = [x for x in r_fits if x and x != 'any']
                    else:
                        legacy = r.get('fit', '')
                        r_fits = [legacy] if legacy and legacy != 'any' else []
                    if r_fits and fit not in r_fits:
                        continue

                    # Customers: if specified, item must match
                    r_custs = r.get('customers')
                    if isinstance(r_custs, list):
                        r_custs = [c.upper().strip() for c in r_custs if c and c.strip()]
                    else:
                        legacy_c = (r.get('customer') or '').strip()
                        r_custs = [legacy_c.upper()] if legacy_c else []
                    if r_custs and (not cust or cust.upper() not in r_custs):
                        continue

                    # Brands: if specified, item must match
                    r_brands = r.get('brands')
                    if isinstance(r_brands, list):
                        r_brands = [b.upper().strip() for b in r_brands if b and b.strip()]
                    else:
                        r_brands = []
                    if r_brands and (not brand_abbr or brand_abbr not in r_brands):
                        continue

                    # Fabrics: if specified, item's fabric code must match.
                    # (This match was missing on the backend — fabric-narrowed rules used
                    # to be silently ignored by server-side exports.)
                    r_fabs = r.get('fabrics')
                    if isinstance(r_fabs, list):
                        r_fabs = [f.upper().strip() for f in r_fabs if f and f.strip()]
                    else:
                        r_fabs = []
                    if r_fabs and (not fab or fab not in r_fabs):
                        continue

                    # Tier A (Specific) requires BOTH a specific category AND specific fabrics
                    is_specific = 1 if (r_cat and r_cat != 'any' and r_fabs) else 0

                    # Secondary score: weighted dimensions.
                    # Brand and customer count as +2 (harder constraints — "who the item is FOR"),
                    # everything else as +1 (softer constraints — "what the item IS").
                    # This means a brand-specific or customer-specific rule beats a rule that
                    # specifies only category + fit. Mirrors frontend matchPrepackDefault().
                    score = 0
                    if r_cat and r_cat != 'any': score += 1
                    if r_fits:                    score += 1
                    if r_custs:                   score += 2
                    if r_brands:                  score += 2
                    if r_fabs:                    score += 1

                    # Compare: tier first (primary), then dimension count
                    if (is_specific > best_fab_spec
                            or (is_specific == best_fab_spec and score > best_score)):
                        best_fab_spec = is_specific
                        best_score    = score
                        best_rule     = r

                matched = best_rule

            if sized_label:
                try:
                    solid_sizes[sized_label] = int((matched or {}).get('master_qty') or 36)
                except Exception:
                    solid_sizes[sized_label] = 36
            elif matched:
                key = matched.get('id', id(matched))
                if key not in seen_keys:
                    seen_keys[key] = matched

        # Render override grids first (sorted by SKU), then rule-based grids
        override_packs = [(k, v) for k, v in seen_keys.items() if isinstance(k, tuple) and k[0] == '__override__']
        rule_packs     = [(k, v) for k, v in seen_keys.items() if not (isinstance(k, tuple) and k[0] == '__override__')]
        packs_to_render = [v for _, v in sorted(override_packs, key=lambda x: x[0][1])] + [v for _, v in rule_packs]

    elif items:
        # No prepack_defaults supplied — still check for item-level overrides
        seen_overrides = {}
        for item in items:
            override_pack = item.get('_override_size_pack')
            sku = item.get('sku', '').upper()
            base = sku.split('-')[0]
            sized_label = _sized_sku_label(sku)
            if sized_label:
                mq = 36
                if override_pack and isinstance(override_pack, dict):
                    try:
                        mq = int(override_pack.get('master_qty') or 36)
                    except Exception:
                        mq = 36
                solid_sizes[sized_label] = mq
                continue
            if override_pack and isinstance(override_pack, dict) and override_pack.get('sizes'):
                if base not in seen_overrides:
                    seen_overrides[base] = {
                        'label': f"Override: {base}",
                        'master_qty': override_pack.get('master_qty', '?'),
                        'inner_qty':  override_pack.get('inner_qty', '?'),
                        'sizes':      override_pack.get('sizes', []),
                    }
        packs_to_render = list(seen_overrides.values())

    # Consolidated solid-pack grid: one row per specific size on this tab,
    # each shipping as its own full case of that single size.
    if solid_sizes:
        mqs = set(solid_sizes.values())
        solid_master = next(iter(mqs)) if len(mqs) == 1 else '?'
        packs_to_render.append({
            'label': 'SOLID SIZE PACKS (each size ships as its own full case)',
            'master_qty': solid_master, 'inner_qty': '',
            'sizes': sorted([[k, v] for k, v in solid_sizes.items()]),
        })

    # ── Fallback: hardcoded Slim + Regular ────────────────────────────────
    if not packs_to_render:
        packs_to_render = [
            {'label': 'Slim Fit', 'master_qty': 36, 'inner_qty': '4 - 9pc Inners',
             'sizes': [['14-14.5 / 32-33', 4], ['15-15.5 / 32-33', 8], ['15-15.5 / 34-35', 4],
                       ['16-16.5 / 32-33', 4], ['16-16.5 / 34-35', 8], ['17-17.5 / 34-35', 8]]},
            {'label': 'Regular Fit', 'master_qty': 36, 'inner_qty': '4 - 9pc Inners',
             'sizes': [['15-15.5 / 32-33', 8], ['15-15.5 / 34-35', 8], ['16-16.5 / 32-33', 4],
                       ['16-16.5 / 34-35', 4], ['17-17.5 / 34-35', 4], ['17-17.5 / 36-37', 4],
                       ['18-18.5 / 36-37', 4]]},
        ]

    # ── Compute ratios helper ─────────────────────────────────────────────
    from math import gcd
    from functools import reduce
    def _compute_ratios(sizes):
        if not sizes:
            return []
        qtys = [q for _, q in sizes if q > 0]
        if not qtys:
            return [(n, 0) for n, _ in sizes]
        g = reduce(gcd, qtys)
        return [(n, q // g) for n, q in sizes] if g > 0 else sizes

    # ── Render each rule as a vertical block ──────────────────────────────
    r = start

    # "Waist x Length" size-name detector (e.g. dress pants "32 x 30") — such
    # packs pivot into the same compact two-dimensional grid as neck/sleeve
    # shirts (waist across the top, length down the side). Rendered flat (one
    # column per size) they get too wide to read or print past a few sizes.
    x_re = re.compile(r'^\s*([\d.\-]+)\s*[xX×]\s*([\d.\-]+)\s*$')

    for rule in packs_to_render:
        sizes  = rule.get('sizes') or []
        # Ensure sizes is a list of [name, qty] pairs with numeric qty
        safe_sizes = []
        for sz in sizes:
            if isinstance(sz, (list, tuple)) and len(sz) >= 2:
                safe_sizes.append([str(sz[0]), int(sz[1]) if sz[1] else 0])
        sizes = safe_sizes
        if not sizes:
            continue  # Skip rules with no sizes defined
        master = rule.get('master_qty', '?')
        inner  = rule.get('inner_qty', '—')  # free-text string
        if isinstance(inner, (int, float)):
            inner = str(inner)  # legacy numeric → string
        label  = rule.get('label') or rule.get('category') or '?'

        worksheet.set_row(r, 20)
        inner_part = '{}, '.format(inner) if inner and inner != '—' else ''
        worksheet.write(r, 0, '{} | {}{} pcs / box'.format(label, inner_part, master), t)
        r += 1

        worksheet.set_row(r, 16)
        worksheet.write(r, 0, '{} SIZE SCALE TO USE'.format(label.upper()), s)
        r += 1

        ratios = _compute_ratios(sizes)
        is_neck_sleeve = any('/' in str(sz[0]) for sz in sizes if sz)
        is_waist_length = (not is_neck_sleeve and
                           all(x_re.match(str(sz[0])) for sz in sizes))

        if is_neck_sleeve or is_waist_length:
            # Parse two-part sizes — "15-15.5 / 32-33" (neck/sleeve) or
            # "32 x 30" (waist/length) — into columns (first part) and rows (second)
            col_dim, row_dim = ('NECK', 'SLEEVE') if is_neck_sleeve else ('WAIST', 'LENGTH')
            col_map   = {}   # first dim (neck / waist) -> {second dim: ratio}
            col_order = []
            row_order = []
            for (sz, _), (_, ratio) in zip(sizes, ratios):
                if is_neck_sleeve:
                    parts = [p.strip() for p in sz.split('/')]
                    cval  = parts[0]
                    rval  = parts[1] if len(parts) > 1 else ''
                else:
                    m = x_re.match(str(sz))
                    cval, rval = m.group(1), m.group(2)
                if cval not in col_map:
                    col_map[cval] = {}
                    col_order.append(cval)
                col_map[cval][rval] = ratio
                if rval not in row_order:
                    row_order.append(rval)

            # Dimension banner row: merged label over the size columns so the
            # reader knows which number is which (e.g. WAIST across the top)
            worksheet.set_row(r, 18)
            worksheet.write(r, 0, '', gh)
            if len(col_order) > 1:
                worksheet.merge_range(r, 1, r, len(col_order), col_dim, gh)
            else:
                worksheet.write(r, 1, col_dim, gh)
            r += 1

            # Column header row: row-dimension label in the corner (e.g. LENGTH,
            # sitting directly above the row values) + first-dim sizes as columns
            worksheet.set_row(r, 22)
            worksheet.write(r, 0, row_dim, gh)
            for ci, cval in enumerate(col_order):
                worksheet.write(r, 1 + ci, cval, gh)
            r += 1

            # One data row per second-dimension size (sleeve / length)
            for rval in row_order:
                worksheet.set_row(r, 22)
                worksheet.write(r, 0, rval, gh)
                for ci, cval in enumerate(col_order):
                    val = col_map[cval].get(rval, '')
                    worksheet.write(r, 1 + ci, val if val != '' else '', gd)
                r += 1

        else:
            # Flat S/M/L — header row then ratio row
            worksheet.set_row(r, 22)
            worksheet.write(r, 0, 'Size', gh)
            for ci, (sz, _) in enumerate(ratios):
                worksheet.write(r, 1 + ci, sz, gh)
            r += 1

            worksheet.set_row(r, 22)
            worksheet.write(r, 0, 'Ratio', gh)
            for ci, (_, ratio) in enumerate(ratios):
                worksheet.write(r, 1 + ci, ratio, gd)
            r += 1

        # Blank gap row between grids
        worksheet.set_row(r, 10)
        r += 1


def build_brand_excel(brand_name, items, s3_base_url, view_mode='all', is_order=False,
                      catalog_mode=False, prepack_defaults=None, flow_mode=False):
    has_color = any(item.get('color') for item in items)

    # Auto-detect incoming_only: all items have zero warehouse stock
    # (only applies to non-incoming view modes — incoming view already omits warehouse)
    incoming_only = False
    if view_mode != 'incoming' and items and not catalog_mode:
        incoming_only = all(
            item.get('total_warehouse', 0) == 0
            for item in items
        )

    buf = BytesIO()
    wb = xlsxwriter.Workbook(buf, {'in_memory': True, 'strings_to_formulas': False})
    wb.set_properties({'title': f'Versa - {brand_name}', 'author': 'Versa Inventory System'})
    ws = wb.add_worksheet(brand_name[:31])
    # Monkey-patch worksheet.write to catch string-as-format bugs
    _orig_ws_write = ws.write
    def _safe_ws_write(r, c, val=None, fmt=None, *args):
        if fmt is not None and isinstance(fmt, str):
            import traceback as _tb
            print(f"  🚨 STRING-AS-FORMAT at row={r} col={c} val={repr(val)[:80]} fmt={repr(fmt)[:80]}")
            _tb.print_stack()
            return  # skip the bad write
        if fmt is not None:
            _orig_ws_write(r, c, val, fmt, *args)
        elif val is not None:
            _orig_ws_write(r, c, val)
        else:
            _orig_ws_write(r, c)
    ws.write = _safe_ws_write
    print(f"  [build_brand_excel] Step 1: setup worksheet")
    fmts, headers = _setup_worksheet(wb, ws, has_color=has_color, view_mode=view_mode,
                                     is_order=is_order, incoming_only=incoming_only,
                                     catalog_mode=catalog_mode, flow_mode=flow_mode)
    print(f"  [build_brand_excel] Step 2: download images")
    imgs = download_images_for_items(items, s3_base_url, use_cache=True)
    print(f"  [build_brand_excel] Step 3: write {len(items)} rows, headers={headers}")
    n = _write_rows(wb, ws, items, imgs, fmts, has_color=has_color,
                    view_mode=view_mode, headers=headers, catalog_mode=catalog_mode)
    print(f"  [build_brand_excel] Step 4: add size charts (prepack_defaults={type(prepack_defaults).__name__}, len={len(prepack_defaults) if prepack_defaults else 0})")
    try:
        _add_size_charts(wb, ws, n + 2, prepack_defaults=prepack_defaults, items=items)
        print(f"  [build_brand_excel] Step 4: size charts OK")
    except Exception as e:
        import traceback as _tb
        _tb.print_exc()
        print(f"  ⚠ Size charts failed (non-fatal): {e}")
    print(f"  [build_brand_excel] Step 5: wb.close()")
    try:
        wb.close()
    except Exception as e:
        import traceback as _tb
        _tb.print_exc()
        print(f"  ⚠ wb.close() failed: {e}")
        # Try again without size charts — rebuild from scratch
        print(f"  [build_brand_excel] Retrying WITHOUT size charts...")
        buf = BytesIO()
        wb = xlsxwriter.Workbook(buf, {'in_memory': True, 'strings_to_formulas': False})
        wb.set_properties({'title': f'Versa - {brand_name}', 'author': 'Versa Inventory System'})
        ws = wb.add_worksheet(brand_name[:31])
        fmts, headers = _setup_worksheet(wb, ws, has_color=has_color, view_mode=view_mode,
                                         is_order=is_order, incoming_only=incoming_only,
                                         catalog_mode=catalog_mode, flow_mode=flow_mode)
        imgs = download_images_for_items(items, s3_base_url, use_cache=True)
        _write_rows(wb, ws, items, imgs, fmts, has_color=has_color,
                    view_mode=view_mode, headers=headers, catalog_mode=catalog_mode)
        wb.close()
        print(f"  [build_brand_excel] Retry succeeded (no size charts)")
    return buf.getvalue()


def _safe_multi_sheet_name(raw, seen, idx, fallback_prefix='Brand'):
    # xlsxwriter RAISES on duplicate sheet names and rejects names that start
    # or end with an apostrophe. Custom per-tab names (line-sheet cart) make
    # collisions likely — two views of the same brand — so dedupe with a
    # _2.._999 suffix. Excel treats sheet names as CASE-INSENSITIVELY unique
    # ('Nautica Sale' vs 'NAUTICA SALE' still raises), so the seen set holds
    # lowercased names.
    name = re.sub(r'[\\/*?\[\]:]', '', str(raw or ''))[:31].strip().strip("'") or f"{fallback_prefix}_{idx + 1}"
    if name.lower() in seen:
        for i in range(2, 1000):
            cand = f"{name[:31 - len(str(i)) - 1]}_{i}"
            if cand.lower() not in seen:
                name = cand
                break
    seen.add(name.lower())
    return name


def build_multi_brand_excel(brands_list, s3_base_url, catalog_mode=False, view_mode='all', flow_mode=False, prepack_defaults=None):
    # Per-tab overrides (line-sheet cart): an entry may carry its own tab_name,
    # view_mode, flow_mode and keep_order. Anything missing falls back to the
    # workbook-global arguments, so pre-existing callers behave byte-identically.
    # catalog_mode stays workbook-global on purpose — mixing admin and customer
    # tabs in one file is how committed/allocated data leaks to a customer.
    _VALID_VIEWS = ('all', 'ats', 'incoming')

    def _tab_view(b):
        v = str(b.get('view_mode') or view_mode).lower()
        # An unknown value would silently fall into _setup_worksheet's admin
        # else-branch and render a different column set — fall back instead.
        return v if v in _VALID_VIEWS else view_mode

    def _tab_flow(b):
        return bool(b.get('flow_mode', flow_mode))

    for b in brands_list:
        b['items'] = b.get('items') or []
        if b.get('keep_order'):
            continue  # curated tab: the caller's row order IS the layout
        sort_key = 'total_ats' if catalog_mode else 'total_warehouse'
        b['items'] = sorted(b['items'], key=lambda x: x.get(sort_key, 0), reverse=True)

    has_color = any(item.get('color') for b in brands_list for item in b['items'])

    all_items = []
    offsets = []
    off = 0
    for b in brands_list:
        offsets.append((off, len(b['items'])))
        all_items.extend(b['items'])
        off += len(b['items'])

    all_imgs = download_images_for_items(all_items, s3_base_url, use_cache=True)

    buf = BytesIO()
    wb = xlsxwriter.Workbook(buf, {'in_memory': True, 'strings_to_formulas': False})
    wb.set_properties({'title': 'Versa Multi-Brand Export', 'author': 'Versa Inventory System'})

    _seen_names = set()
    for bi, brand in enumerate(brands_list):
        safe = _safe_multi_sheet_name(brand.get('tab_name') or brand.get('brand_name'), _seen_names, bi)
        ws = wb.add_worksheet(safe)
        # Monkey-patch worksheet.write to catch string-as-format bugs
        _orig_ws_write = ws.write
        def _safe_ws_write(r, c, val=None, fmt=None, *args, _orig=_orig_ws_write):
            if fmt is not None and isinstance(fmt, str):
                print(f"  🚨 STRING-AS-FORMAT at row={r} col={c} val={repr(val)[:80]} fmt={repr(fmt)[:80]}")
                return  # skip the bad write
            if fmt is not None:
                _orig(r, c, val, fmt, *args)
            elif val is not None:
                _orig(r, c, val)
            else:
                _orig(r, c)
        ws.write = _safe_ws_write
        fmts, headers = _setup_worksheet(wb, ws, has_color=has_color,
                                         catalog_mode=catalog_mode, view_mode=_tab_view(brand),
                                         flow_mode=_tab_flow(brand))
        start, count = offsets[bi]
        local_imgs = {}
        for li in range(count):
            gi = start + li
            if gi in all_imgs:
                local_imgs[li] = all_imgs[gi]
        n = _write_rows(wb, ws, brand['items'], local_imgs, fmts,
                        has_color=has_color, headers=headers, catalog_mode=catalog_mode)
        # An empty tab would render the hardcoded fallback size grids on an
        # otherwise blank sheet — skip charts entirely when there are no rows.
        if brand['items']:
            try:
                _add_size_charts(wb, ws, n + 2, prepack_defaults=prepack_defaults, items=brand['items'])
            except Exception as e:
                print(f"  ⚠ Size charts failed for {brand.get('brand_name','?')} (non-fatal): {e}")

    try:
        wb.close()
    except Exception as e:
        import traceback as _tb
        _tb.print_exc()
        print(f"  ⚠ wb.close() failed in multi-brand: {e} — retrying WITHOUT size charts...")
        # Retry from scratch without size charts
        buf = BytesIO()
        wb = xlsxwriter.Workbook(buf, {'in_memory': True, 'strings_to_formulas': False})
        wb.set_properties({'title': 'Versa Multi-Brand Export', 'author': 'Versa Inventory System'})
        _retry_seen = set()
        for bi, brand in enumerate(brands_list):
            safe = _safe_multi_sheet_name(brand.get('tab_name') or brand.get('brand_name'), _retry_seen, bi)
            ws = wb.add_worksheet(safe)
            fmts, headers = _setup_worksheet(wb, ws, has_color=has_color,
                                             catalog_mode=catalog_mode, view_mode=_tab_view(brand),
                                             flow_mode=_tab_flow(brand))
            start, count = offsets[bi]
            local_imgs = {}
            for li in range(count):
                gi = start + li
                if gi in all_imgs:
                    local_imgs[li] = all_imgs[gi]
            _write_rows(wb, ws, brand['items'], local_imgs, fmts,
                        has_color=has_color, headers=headers, catalog_mode=catalog_mode)
        wb.close()
        print(f"  ✓ Multi-brand retry succeeded (no size charts)")
    return buf.getvalue()


# ============================================
# PDF TILE EXPORT — 2×2 grid with cached images
# ============================================
def build_brand_pdf(title, items, s3_base_url, subtitle='', show_qty=False):
    """Generate a PDF with 2×2 product tile grid, using server-cached images."""

    page_w, page_h = letter  # 612 × 792 points
    margin = 34  # ~12mm
    cols, rows_per_page = 2, 2
    gap_x, gap_y = 18, 18
    usable_w = page_w - margin * 2
    usable_h = page_h - margin * 2 - 28  # room for header
    tile_w = (usable_w - gap_x * (cols - 1)) / cols
    tile_h = (usable_h - gap_y * (rows_per_page - 1)) / rows_per_page
    img_h = tile_h * 0.65
    img_w = tile_w - 18
    tiles_per_page = cols * rows_per_page

    # Pre-fetch images using existing server cache (fast, parallel, disk-cached)
    imgs = download_images_for_items(items, s3_base_url, use_cache=True)

    buf = BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=letter)
    c.setTitle(title)

    total_pages = max(1, -(-len(items) // tiles_per_page))
    date_str = datetime.now().strftime('%m/%d/%Y')

    for page in range(total_pages):
        if page > 0:
            c.showPage()

        # ── Header ──
        c.setFont('Helvetica-Bold', 14)
        c.setFillColor(HexColor('#1e1e1e'))
        c.drawString(margin, page_h - margin + 4, title)
        c.setFont('Helvetica', 9)
        c.setFillColor(HexColor('#787878'))
        header_right = f"{date_str}  \u2022  {subtitle}  \u2022  Page {page + 1} of {total_pages}"
        c.drawRightString(page_w - margin, page_h - margin + 4, header_right)
        c.setStrokeColor(HexColor('#c8c8c8'))
        c.setLineWidth(0.5)
        c.line(margin, page_h - margin - 6, page_w - margin, page_h - margin - 6)

        start_y = page_h - margin - 10

        page_items = items[page * tiles_per_page:(page + 1) * tiles_per_page]

        for idx, item in enumerate(page_items):
            col_idx = idx % cols
            row_idx = idx // cols
            x = margin + col_idx * (tile_w + gap_x)
            y = start_y - row_idx * (tile_h + gap_y)

            # Tile background + border
            c.setStrokeColor(HexColor('#d2d2d2'))
            c.setLineWidth(0.5)
            c.setFillColor(HexColor('#ffffff'))
            c.roundRect(x, y - tile_h, tile_w, tile_h, 5, fill=1, stroke=1)

            # ── Image ──
            img_x = x + (tile_w - img_w) / 2
            img_y_top = y - 6  # 6pt padding from tile top
            img_y_bottom = img_y_top - img_h
            global_idx = page * tiles_per_page + idx
            img_data = imgs.get(global_idx)

            drew_image = False
            if img_data and img_data.get('image_data'):
                try:
                    img_data['image_data'].seek(0)
                    pil_img = PilImage.open(img_data['image_data'])
                    ow, oh = pil_img.size
                    ratio = ow / oh
                    draw_w, draw_h = img_w, img_h
                    if ratio > img_w / img_h:
                        draw_h = img_w / ratio
                    else:
                        draw_w = img_h * ratio
                    draw_x = img_x + (img_w - draw_w) / 2
                    draw_y = img_y_bottom + (img_h - draw_h) / 2

                    img_buf = BytesIO()
                    if pil_img.mode in ('RGBA', 'LA', 'P'):
                        pil_img = pil_img.convert('RGBA')
                        pil_img.save(img_buf, format='PNG')
                    else:
                        if pil_img.mode != 'RGB':
                            pil_img = pil_img.convert('RGB')
                        pil_img.save(img_buf, format='JPEG', quality=85)
                    img_buf.seek(0)
                    c.drawImage(ImageReader(img_buf), draw_x, draw_y, draw_w, draw_h,
                                preserveAspectRatio=True, mask='auto')
                    drew_image = True
                except Exception:
                    pass

            if not drew_image:
                c.setFillColor(HexColor('#f3f4f6'))
                c.rect(img_x, img_y_bottom, img_w, img_h, fill=1, stroke=0)
                c.setFillColor(HexColor('#a0a0a0'))
                c.setFont('Helvetica', 8)
                c.drawCentredString(x + tile_w / 2, img_y_bottom + img_h / 2, 'No Image')

            # ── Short Sleeve badge ──
            fit_str = item.get('fit', '')
            if 'short sleeve' in fit_str.lower():
                badge_text = 'SHORT SLEEVE'
                c.setFont('Helvetica-Bold', 6)
                bw = c.stringWidth(badge_text, 'Helvetica-Bold', 6) + 8
                bh = 10
                bx = img_x + 3
                by = img_y_bottom + 3
                c.setFillColor(Color(14/255, 165/255, 233/255, 0.9))
                c.roundRect(bx, by, bw, bh, 3, fill=1, stroke=0)
                c.setFillColor(HexColor('#ffffff'))
                c.drawString(bx + 4, by + 3, badge_text)

            # ── Text below image ──
            text_x = x + 9
            ty = img_y_bottom - 12
            line_step = 11

            # SKU
            c.setFont('Helvetica-Bold', 9)
            c.setFillColor(HexColor('#1e1e1e'))
            c.drawString(text_x, ty, str(item.get('sku', '')))
            ty -= line_step

            # Brand
            c.setFont('Helvetica', 7.5)
            c.setFillColor(HexColor('#505050'))
            c.drawString(text_x, ty, str(item.get('brand_full', item.get('brand_abbr', '')))[:35])
            ty -= line_step

            # Color
            color_str = item.get('color', '')
            if color_str:
                c.setFillColor(HexColor('#6d28d9'))
                c.setFont('Helvetica-Bold', 7)
                c.drawString(text_x, ty, color_str[:40])
                ty -= line_step

            # Fit + Fabric
            fabric = item.get('fabrication', '')
            fit = item.get('fit', '')
            fit_fab = f"{fit} \u2022 {fabric[:25] + '...' if len(fabric) > 25 else fabric}"
            c.setFont('Helvetica', 6.5)
            c.setFillColor(HexColor('#787878'))
            c.drawString(text_x, ty, fit_fab[:55])
            ty -= line_step

            # Qty / ATS
            if show_qty:
                qty = item.get('quantity_ordered', item.get('quantity', 0)) or 0
                c.setFont('Helvetica-Bold', 8)
                c.setFillColor(HexColor('#16a34a'))
                c.drawString(text_x, ty, f"Qty: {qty:,}")
                c.setFont('Helvetica', 7)
                c.setFillColor(HexColor('#646464'))
                c.drawRightString(x + tile_w - 9, ty, f"ATS: {item.get('total_ats', 0):,}")
            else:
                c.setFont('Helvetica-Bold', 8)
                c.setFillColor(HexColor('#16a34a'))
                c.drawString(text_x, ty, f"ATS: {item.get('total_ats', 0):,}")

            # Delivery
            delivery = item.get('delivery', '')
            if delivery and delivery not in ('ATS', ''):
                ty -= line_step
                c.setFont('Helvetica', 6.5)
                c.setFillColor(HexColor('#d97706'))
                c.drawString(text_x, ty, f"Delivery: {delivery}")

    c.save()
    return buf.getvalue()


def _col_val(row_dict, name):
    if name in row_dict:
        return row_dict[name]
    lo = name.lower()
    for k, v in row_dict.items():
        if k.lower() == lo:
            return v
    return None


def parse_inventory_excel(file_bytes):
    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows_iter = ws.iter_rows(values_only=False)
    header_row = next(rows_iter)
    headers = [str(cell.value or '').strip() for cell in header_row]

    items = []
    for row in rows_iter:
        rd = {headers[i]: row[i].value for i in range(min(len(headers), len(row)))}

        sku = str(_col_val(rd, 'SKU') or '').strip()
        brand = str(_col_val(rd, 'Brand') or '').strip().upper()
        if brand == 'NT':
            # NT is the Nautica overflow SKU code (serials past 999), not a brand.
            # The frontend force-corrects these rows to NAUTICA; without this the
            # backend would group them as a separate 'NT' brand and the prebuilt
            # Nautica workbooks would silently omit every NT style.
            brand = 'NAUTICA'
        if not sku or sku == 'N/A' or not brand:
            continue

        jtw = int(_col_val(rd, 'JTW') or 0)
        tr  = int(_col_val(rd, 'TR')  or 0)
        dcw = int(_col_val(rd, 'DCW') or 0)
        qa  = int(_col_val(rd, 'QA') or _col_val(rd, 'Q/A') or _col_val(rd, 'Quality') or 0)
        committed = int(_col_val(rd, 'Committed') or 0)
        allocated = int(_col_val(rd, 'Allocated') or 0)
        incoming  = int(_col_val(rd, 'Incoming') or _col_val(rd, 'In Transit') or
                        _col_val(rd, 'InTransit') or _col_val(rd, 'In-Transit') or
                        _col_val(rd, 'On Order') or _col_val(rd, 'PO') or
                        _col_val(rd, 'Incoming Qty') or 0)

        total_ats_raw = _col_val(rd, 'Total ATS') or _col_val(rd, 'Total_ATS') or _col_val(rd, 'TotalATS') or 0
        total_ats = int(total_ats_raw)

        container = str(_col_val(rd, 'Container') or '').strip()
        receive_date = str(_col_val(rd, 'Receive Date') or _col_val(rd, 'ReceiveDate') or '').strip()
        lot_number = str(_col_val(rd, 'Lot Number') or _col_val(rd, 'LotNumber') or '').strip()

        brand_full = BRAND_FULL_NAMES.get(brand, brand)

        items.append({
            'sku': sku,
            'brand': brand,
            'brand_abbr': brand,
            'brand_full': brand_full,
            'name': f"{brand} {sku}",
            'jtw': jtw, 'tr': tr, 'dcw': dcw, 'qa': qa,
            'incoming': incoming,
            'committed': committed, 'allocated': allocated,
            'total_ats': total_ats,
            'total_warehouse': jtw + tr + dcw + qa,
            'container': container,
            'receive_date': receive_date,
            'lot_number': lot_number,
            'image': ''
        })

    wb.close()
    return items


def _group_by_brand(items):
    brands = {}
    for item in items:
        abbr = item['brand']
        if abbr not in brands:
            brands[abbr] = {'name': item['brand_full'], 'items': []}
        brands[abbr]['items'].append(item)
    return brands


# ============================================
# DROPBOX SYNC — primary inventory source
# ============================================
DROPBOX_INVENTORY_PATH = os.environ.get('DROPBOX_INVENTORY_PATH', '/Versa Share Files/Hourly ATS/Inventory_ATS.xlsx')

def sync_from_dropbox():
    """Fetch inventory directly via Dropbox API — uses OAuth, never expires, no shared link needed."""
    token = get_dropbox_token()
    if not token:
        print("  ⚠ No Dropbox token available, skipping Dropbox sync")
        return False

    print(f"  📂 Fetching inventory from Dropbox API: {DROPBOX_INVENTORY_PATH}")
    try:
        headers = {
            'Authorization': f'Bearer {token}',
            'Dropbox-API-Arg': json.dumps({'path': DROPBOX_INVENTORY_PATH})
        }
        resp = http_requests.post(
            'https://content.dropboxapi.com/2/files/download',
            headers=headers, timeout=60
        )

        # Auto-refresh token if expired
        if resp.status_code == 401:
            print("  ⚠ Dropbox auth failed (401) — forcing token refresh...")
            global _dropbox_token_expires
            _dropbox_token_expires = 0
            token = get_dropbox_token()
            if not token:
                print("  ⚠ Could not refresh token, giving up")
                return False
            headers['Authorization'] = f'Bearer {token}'
            resp = http_requests.post(
                'https://content.dropboxapi.com/2/files/download',
                headers=headers, timeout=60
            )

        if resp.status_code != 200:
            print(f"  ⚠ Dropbox API returned HTTP {resp.status_code}: {resp.text[:200]}")
            return False

        data = resp.content
        if len(data) < 1000:
            print(f"  ⚠ Dropbox file too small ({len(data)} bytes), likely error")
            return False

        print(f"  📂 Downloaded {len(data):,} bytes from Dropbox API")

        items = parse_inventory_excel(data)
        if not items:
            print("  ⚠ No valid rows parsed from Dropbox file")
            return False
        items = _apply_nj_to_items(items)

        # ── SANITY CHECK ────────────────────────────────────────────────
        # If the parsed data looks materially worse than the previous accepted
        # sync (probably a partial-write read of the still-being-regenerated
        # Hourly ATS file), reject it and keep the previous good data.
        accepted, reason, fingerprint = _sync_passes_sanity_check(items)
        if not accepted:
            print(f"  ⚠ Dropbox sync REJECTED — {reason}. Keeping previous good data.", flush=True)
            return False
        if reason != 'passed' and reason != 'first-sync (no baseline to compare)':
            print(f"  ℹ Sanity check: {reason}", flush=True)

        brands = _group_by_brand(items)

        with _inv_lock:
            _inventory['items'] = items
            _inventory['brands'] = brands
            _inventory['etag'] = 'dropbox'
            _inventory['last_sync'] = datetime.utcnow().isoformat() + 'Z'
            _inventory['item_count'] = len(items)
            _inventory['source'] = 'dropbox'
            _inventory['committed_nonzero_count'] = fingerprint['committed_nonzero_count']
            _inventory['committed_abs_sum'] = fingerprint['committed_abs_sum']

        print(f"  ✓ Dropbox sync: {len(items)} items across {len(brands)} brands "
              f"(non-zero committed: {fingerprint['committed_nonzero_count']}, "
              f"|committed| sum: {fingerprint['committed_abs_sum']:,})", flush=True)
        return True

    except Exception as e:
        print(f"  ⚠ Dropbox sync failed: {type(e).__name__}: {e}")
        return False


# ============================================
# S3 SYNC — fallback inventory source
# ============================================
def s3_read_inventory():
    try:
        s3 = get_s3()
        resp = s3.get_object(Bucket=S3_BUCKET, Key=S3_INVENTORY_KEY)
        data = resp['Body'].read()
        etag = resp.get('ETag', '')
        print(f"  Downloaded s3://{S3_BUCKET}/{S3_INVENTORY_KEY} ({len(data)} bytes)")
        return data, etag
    except ClientError as e:
        print(f"  S3 read failed: {e.response['Error']['Code']}")
        return None, None
    except NoCredentialsError:
        print("  S3 read failed: no AWS credentials configured")
        return None, None
    except Exception as e:
        print(f"  S3 read failed: {e}")
        return None, None


def s3_check_etag():
    try:
        s3 = get_s3()
        resp = s3.head_object(Bucket=S3_BUCKET, Key=S3_INVENTORY_KEY)
        return resp.get('ETag', '')
    except Exception:
        return None


def s3_upload_export(key, file_bytes):
    try:
        s3 = get_s3()
        s3.put_object(
            Bucket=S3_BUCKET, Key=key, Body=file_bytes,
            ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        print(f"    Uploaded s3://{S3_BUCKET}/{key}")
        return True
    except Exception as e:
        print(f"    Upload failed for {key}: {e}")
        return False


def sync_inventory():
    """Sync inventory: try Dropbox first, then S3 fallback"""
    # Try Dropbox API first
    if DROPBOX_REFRESH_TOKEN or DROPBOX_PHOTOS_TOKEN:
        if sync_from_dropbox():
            return True
        print("  Dropbox failed, falling back to S3...")

    # S3 fallback
    new_etag = s3_check_etag()
    with _inv_lock:
        if new_etag and new_etag == _inventory['etag'] and _inventory['items']:
            print("  Inventory unchanged (same ETag)")
            return False

    data, etag = s3_read_inventory()
    if data is None:
        return False

    try:
        items = _apply_nj_to_items(parse_inventory_excel(data))
    except Exception as e:
        print(f"  Failed to parse inventory: {e}")
        return False

    # Sanity check — same defense applied to the S3 fallback path
    accepted, reason, fingerprint = _sync_passes_sanity_check(items)
    if not accepted:
        print(f"  ⚠ S3 sync REJECTED — {reason}. Keeping previous good data.", flush=True)
        return False

    brands = _group_by_brand(items)

    with _inv_lock:
        _inventory['items'] = items
        _inventory['brands'] = brands
        _inventory['etag'] = etag
        _inventory['last_sync'] = datetime.utcnow().isoformat() + 'Z'
        _inventory['item_count'] = len(items)
        _inventory['source'] = 's3'
        _inventory['committed_nonzero_count'] = fingerprint['committed_nonzero_count']
        _inventory['committed_abs_sum'] = fingerprint['committed_abs_sum']

    print(f"  Parsed {len(items)} items across {len(brands)} brands "
          f"(non-zero committed: {fingerprint['committed_nonzero_count']}, "
          f"|committed| sum: {fingerprint['committed_abs_sum']:,})", flush=True)
    return True


def generate_all_exports():
    global _regen_queued
    with _export_lock:
        if _exports['generating']:
            # A run is already in flight. Queue a follow-up instead of dropping
            # the request — rules/overrides saved MID-RUN would otherwise never
            # be baked into the cached workbooks, yet get stamped "fresh".
            _regen_queued = True
            return
        _exports['generating'] = True
        _exports['progress'] = 'starting...'

    # Everything this run reads (inventory, overrides, prepack rules) is
    # snapshotted near the start, so every export it writes must be stamped
    # with the RUN-START time — not its completion time. Stamping completion
    # time made brands finished AFTER a mid-run save look newer than the save,
    # defeating the _exports_are_stale check.
    run_started_iso = datetime.utcnow().isoformat() + 'Z'

    try:
        with _inv_lock:
            brands = dict(_inventory['brands'])

        if not brands:
            print("  No inventory data for export generation")
            return

        date_str = datetime.utcnow().strftime('%Y-%m-%d')
        total = len(brands)

        # Reload prepack rules from S3 ONCE per generation run — this worker's
        # in-memory copy may be stale (rule saves can land on other gunicorn
        # workers). One read covers every brand tab below.
        pd_snap = _fresh_prepack_defaults()

        print(f"\n{'='*60}")
        print(f"  Generating exports for {total} brands...")
        print(f"  Image strategy: STYLE+OVERRIDES first → brand folder fallback")
        print(f"{'='*60}")

        # Pre-cache images for ALL items (deduplicates by base_style)
        all_items = []
        for abbr, brand in brands.items():
            all_items.extend(brand['items'])

        print(f"  Pre-caching images for {len(all_items)} items...")
        download_images_for_items(all_items, S3_PHOTOS_URL, use_cache=True)
        with _img_lock:
            cached_count = sum(1 for v in _img_cache.values() if v is not None)
            failed_count = sum(1 for v in _img_cache.values() if v is None)
        print(f"  Image cache: {cached_count} found, {failed_count} not found\n")

        brands_list_for_multi = []
        done = 0

        sorted_brands = sorted(brands.items(),
            key=lambda x: sum(i.get('total_warehouse', 0) for i in x[1]['items']),
            reverse=True)

        for abbr, brand in sorted_brands:
            done += 1
            name = brand['name']
            with _export_lock:
                _exports['progress'] = f"{done}/{total}: {name}"

            print(f"  [{done}/{total}] {name} ({len(brand['items'])} items)")

            sorted_items = sorted(brand['items'], key=lambda x: x.get('total_warehouse', 0), reverse=True)

            try:
                annotated_items = _annotate_items_for_prepack(sorted_items)
                # pd_snap loaded fresh from S3 once at generation start
                xl_bytes = build_brand_excel(name, annotated_items, S3_PHOTOS_URL,
                                             prepack_defaults=pd_snap)

                with _export_lock:
                    _exports['brands'][abbr] = {
                        'bytes': xl_bytes,
                        'generated_at': run_started_iso,
                        'name': name,
                        'items_count': len(sorted_items),
                        'size_bytes': len(xl_bytes),
                    }

                s3_key = f"{S3_EXPORT_PREFIX}{name.replace(' ', '_')}_{date_str}.xlsx"
                s3_upload_export(s3_key, xl_bytes)

                brands_list_for_multi.append({
                    'brand_name': name,
                    'items': annotated_items
                })
            except Exception as e:
                print(f"    Failed: {e}")

        if brands_list_for_multi:
            print(f"\n  [ALL] Multi-tab ({len(brands_list_for_multi)} brands)...")
            try:
                # pd_snap loaded fresh from S3 once at generation start
                multi_bytes = build_multi_brand_excel(brands_list_for_multi, S3_PHOTOS_URL,
                                                      prepack_defaults=pd_snap)
                with _export_lock:
                    _exports['all_brands'] = {
                        'bytes': multi_bytes,
                        'generated_at': run_started_iso,
                        'brands_count': len(brands_list_for_multi),
                        'items_count': sum(len(b['items']) for b in brands_list_for_multi),
                        'size_bytes': len(multi_bytes),
                    }
                s3_upload_export(f"{S3_EXPORT_PREFIX}All_Brands_{date_str}.xlsx", multi_bytes)
            except Exception as e:
                print(f"    Failed: {e}")

        with _export_lock:
            _exports['progress'] = 'done'
            _exports['last_generated'] = datetime.utcnow().isoformat() + 'Z'

        print(f"\n{'='*60}")
        print(f"  Export generation complete! {done} brands")
        with _img_lock:
            print(f"  Image cache: {len(_img_cache)} unique styles cached")
        print(f"{'='*60}\n")

    except Exception as e:
        print(f"  Export generation error: {e}")
        with _export_lock:
            _exports['progress'] = f'error: {e}'
    finally:
        # ALWAYS release the flag — the old code leaked generating=True on the
        # empty-inventory early return, permanently blocking every future regen
        # on that worker. Then honor any regen queued while we were running.
        rerun = False
        with _export_lock:
            _exports['generating'] = False
            if _regen_queued:
                _regen_queued = False
                rerun = True
        if rerun:
            print("  Data changed mid-run — regenerating exports with fresh snapshot")
            trigger_background_generation()


def trigger_background_generation():
    t = threading.Thread(target=generate_all_exports, daemon=True)
    t.start()


# ============================================
# ROUTES
# ============================================

@app.route('/', methods=['GET'])
def home():
    return jsonify({
        "service": "Versa Inventory Export API v3",
        "status": "running",
    })


@app.route('/health', methods=['GET'])
def health():
    with _inv_lock:
        inv_count = _inventory['item_count']
        last_sync = _inventory['last_sync']
        source = _inventory['source']
    with _export_lock:
        gen = _exports['generating']
        brands_ready = len(_exports['brands'])
        progress = _exports['progress']
    with _img_lock:
        img_count = len(_img_cache)
        img_found = sum(1 for v in _img_cache.values() if v is not None)

    return jsonify({
        "status": "healthy",
        "inventory_items": inv_count,
        "inventory_source": source,
        "last_sync": last_sync,
        "exports_generating": gen,
        "exports_ready": brands_ready,
        "generation_progress": progress,
        "images_cached": img_count,
        "images_found": img_found,
        "dropbox_photos_indexed": len(_dropbox_photo_index),
        "dropbox_photos_cached": len([f for f in os.listdir(DROPBOX_DISK_CACHE) if f.endswith(('.jpg', '.png'))]) if os.path.exists(DROPBOX_DISK_CACHE) else 0,
        "dropbox_photos_last_sync": _dropbox_photos_last_sync,
        "production_rows": len(_production_data),
        "production_last_sync": _production_last_sync,
        "production_folder": DROPBOX_PRODUCTION_FOLDER,
    })


@app.route('/sync', methods=['GET', 'OPTIONS'])
def sync():
    if request.method == 'OPTIONS':
        return '', 204

    updated = sync_inventory()

    with _inv_lock:
        items = list(_inventory['items'])
        last_sync = _inventory['last_sync']
        brand_count = len(_inventory['brands'])

    with _export_lock:
        has_exports = bool(_exports['brands'])
        is_generating = _exports['generating']

    if (updated or not has_exports) and not is_generating:
        print("  Triggering background export generation...")
        trigger_background_generation()

    return jsonify({
        "status": "ok",
        "updated": updated,
        "last_sync": last_sync,
        "item_count": len(items),
        "brand_count": brand_count,
        "inventory": items,
    })


@app.route('/inventory', methods=['GET', 'OPTIONS'])
def inventory():
    if request.method == 'OPTIONS':
        return '', 204

    # ── STALE-WORKER GUARD ──────────────────────────────────────────────────
    # Render runs multiple gunicorn workers, each with independent in-memory
    # _inventory. The admin's /sync call only updates whichever worker handles
    # it. Customers hitting a different worker would see stale data indefinitely.
    # Fix: if this worker's data is older than INVENTORY_MAX_AGE_SECONDS, force
    # a sync right now before responding — guarantees customers always get fresh data.
    INVENTORY_MAX_AGE_SECONDS = 300  # 5 minutes
    needs_sync = False
    with _inv_lock:
        last_sync_str = _inventory.get('last_sync')
        has_items = bool(_inventory['items'])

    if not has_items:
        needs_sync = True
    elif last_sync_str:
        try:
            last_sync_dt = datetime.fromisoformat(last_sync_str.replace('Z', '+00:00'))
            age = (datetime.now(last_sync_dt.tzinfo) - last_sync_dt).total_seconds()
            if age > INVENTORY_MAX_AGE_SECONDS:
                needs_sync = True
        except Exception:
            needs_sync = True

    if needs_sync:
        try:
            sync_inventory()
        except Exception as e:
            print(f"  [/inventory] Stale-worker sync failed: {e}")

    with _inv_lock:
        return jsonify({
            "status": "ok",
            "last_sync": _inventory['last_sync'],
            "item_count": _inventory['item_count'],
            "inventory": list(_inventory['items']),
        })


# ─────────────────────────────────────────────────────────────────────────────
# DIAGNOSTIC: shows backend state — source, freshness, specific SKU values.
# Useful when frontend numbers disagree with the ATS sheet — proves whether
# the backend is serving fresh/correct data or returning something stale.
# Curl-friendly: curl https://your-backend/inventory/debug?skus=ROGBSA002SLS,BUGBSA001SLS
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/inventory/debug', methods=['GET', 'OPTIONS'])
def inventory_debug():
    if request.method == 'OPTIONS':
        return '', 204

    skus_param = request.args.get('skus', '')
    target_skus = [s.strip().upper() for s in skus_param.split(',') if s.strip()] if skus_param else []

    with _inv_lock:
        items = list(_inventory.get('items', []))
        meta = {
            'source': _inventory.get('source'),
            'etag': _inventory.get('etag'),
            'last_sync': _inventory.get('last_sync'),
            'item_count': _inventory.get('item_count'),
            'dropbox_inventory_path': DROPBOX_INVENTORY_PATH,
            'has_dropbox_token': bool(get_dropbox_token() if DROPBOX_REFRESH_TOKEN or DROPBOX_PHOTOS_TOKEN else None),
        }

    # Sample some specific SKUs the user is checking
    sku_details = {}
    if target_skus:
        for tgt in target_skus:
            matches = [i for i in items if (i.get('sku') or '').strip().upper() == tgt]
            if not matches:
                sku_details[tgt] = {'found': False, 'note': 'Not in backend cache'}
            else:
                sku_details[tgt] = {
                    'found': True,
                    'rows': [{
                        'sku': m.get('sku'),
                        'brand': m.get('brand'),
                        'committed': m.get('committed'),
                        'allocated': m.get('allocated'),
                        'jtw': m.get('jtw'),
                        'tr': m.get('tr'),
                        'dcw': m.get('dcw'),
                        'qa': m.get('qa'),
            'nj': m.get('nj', 0),
                        'incoming': m.get('incoming'),
                        'total_ats': m.get('total_ats'),
                    } for m in matches]
                }

    # Quick stats on the cached data
    nonzero_committed = sum(1 for i in items if i.get('committed', 0) != 0)
    nonzero_allocated = sum(1 for i in items if i.get('allocated', 0) != 0)
    total_committed = sum(i.get('committed', 0) for i in items)
    total_allocated = sum(i.get('allocated', 0) for i in items)

    return jsonify({
        'meta': meta,
        'stats': {
            'total_rows': len(items),
            'rows_with_nonzero_committed': nonzero_committed,
            'rows_with_nonzero_allocated': nonzero_allocated,
            'sum_committed_all_rows': total_committed,
            'sum_allocated_all_rows': total_allocated,
        },
        'sku_lookup': sku_details,
        'hint': 'Pass ?skus=SKU1,SKU2 to inspect specific SKUs',
    })


@app.route('/exports', methods=['GET', 'OPTIONS'])
def exports_manifest():
    if request.method == 'OPTIONS':
        return '', 204

    with _export_lock:
        brands = {}
        for abbr, info in _exports['brands'].items():
            brands[abbr] = {
                'name': info['name'],
                'items_count': info['items_count'],
                'size_bytes': info['size_bytes'],
                'generated_at': info['generated_at'],
            }
        all_b = None
        if _exports['all_brands']:
            a = _exports['all_brands']
            all_b = {
                'brands_count': a['brands_count'],
                'items_count': a['items_count'],
                'size_bytes': a['size_bytes'],
                'generated_at': a['generated_at'],
            }

        return jsonify({
            "generating": _exports['generating'],
            "progress": _exports['progress'],
            "last_generated": _exports['last_generated'],
            "brands": brands,
            "all_brands": all_b,
        })


def _exports_are_stale(generated_at_iso):
    """True if overrides were saved after the pre-built export was generated.
    Used by /download/* to decide whether to rebuild on-demand vs serve cached bytes."""
    if not generated_at_iso:
        return True
    try:
        gen_dt = datetime.fromisoformat(generated_at_iso.replace('Z', '+00:00'))
        gen_epoch = gen_dt.timestamp()
    except Exception:
        return True
    # _overrides_last_saved / _prepack_last_saved are plain epoch floats; small
    # grace to avoid races. A prepack-rule save must invalidate cached exports
    # too — their bottom size-scale grids are built from the rules.
    return max(_overrides_last_saved, _prepack_last_saved) > (gen_epoch + 1)


@app.route('/download/brand/<abbr>', methods=['GET'])
def download_brand(abbr):
    abbr = abbr.upper()
    with _export_lock:
        info = _exports['brands'].get(abbr)
    if not info:
        return jsonify({"error": f"No pre-generated export for '{abbr}'"}), 404

    date_str = datetime.utcnow().strftime('%Y-%m-%d')
    filename = f"{info['name'].replace(' ', '_')}_{date_str}.xlsx"

    # ── Stale-export safety net ───────────────────────────────────────────
    # If the user uploaded an override after this export was generated, the
    # cached bytes have outdated images. Rebuild on demand from current data.
    if _exports_are_stale(info.get('generated_at')):
        print(f"  [/download/brand/{abbr}] Pre-built export is stale vs overrides — rebuilding on demand")
        try:
            with _inv_lock:
                brand = _inventory['brands'].get(abbr)
            if brand:
                sorted_items = sorted(brand['items'], key=lambda x: x.get('total_warehouse', 0), reverse=True)
                annotated_items = _annotate_items_for_prepack(sorted_items)
                pd_snap = _fresh_prepack_defaults()
                fresh_bytes = build_brand_excel(brand['name'], annotated_items, S3_PHOTOS_URL,
                                                prepack_defaults=pd_snap)
                # Update cache so next request is fast
                with _export_lock:
                    _exports['brands'][abbr] = {
                        'bytes': fresh_bytes,
                        'generated_at': datetime.utcnow().isoformat() + 'Z',
                        'name': brand['name'],
                        'items_count': len(sorted_items),
                        'size_bytes': len(fresh_bytes),
                    }
                # Trigger full background regen so other brands (and All_Brands)
                # catch up. Unconditional: if a run is in flight, the generator
                # queues a follow-up instead of dropping the request.
                trigger_background_generation()
                return send_file(
                    BytesIO(fresh_bytes),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=filename
                )
        except Exception as e:
            print(f"  [/download/brand/{abbr}] On-demand rebuild failed, serving cached: {e}")
            # Fall through to serving cached bytes

    return send_file(
        BytesIO(info['bytes']),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=filename
    )


@app.route('/download/all', methods=['GET'])
def download_all():
    with _export_lock:
        info = _exports['all_brands']
    if not info:
        return jsonify({"error": "No pre-generated all-brands export"}), 404

    date_str = datetime.utcnow().strftime('%Y-%m-%d')

    # ── Stale-export safety net (same logic as /download/brand) ──
    if _exports_are_stale(info.get('generated_at')):
        print("  [/download/all] Pre-built all-brands export is stale vs overrides — rebuilding on demand")
        try:
            with _inv_lock:
                brands = dict(_inventory['brands'])
            sorted_brands = sorted(brands.items(),
                key=lambda x: sum(i.get('total_warehouse', 0) for i in x[1]['items']),
                reverse=True)
            brands_list_for_multi = []
            for abbr, brand in sorted_brands:
                sorted_items = sorted(brand['items'], key=lambda x: x.get('total_warehouse', 0), reverse=True)
                annotated_items = _annotate_items_for_prepack(sorted_items)
                brands_list_for_multi.append({'brand_name': brand['name'], 'items': annotated_items})
            if brands_list_for_multi:
                pd_snap = _fresh_prepack_defaults()
                fresh_bytes = build_multi_brand_excel(brands_list_for_multi, S3_PHOTOS_URL,
                                                     prepack_defaults=pd_snap)
                with _export_lock:
                    _exports['all_brands'] = {
                        'bytes': fresh_bytes,
                        'generated_at': datetime.utcnow().isoformat() + 'Z',
                        'brands_count': len(brands_list_for_multi),
                        'items_count': sum(len(b['items']) for b in brands_list_for_multi),
                        'size_bytes': len(fresh_bytes),
                    }
                # Unconditional — the generator queues a follow-up if busy
                trigger_background_generation()
                return send_file(
                    BytesIO(fresh_bytes),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    as_attachment=True, download_name=f"All_Brands_{date_str}.xlsx"
                )
        except Exception as e:
            print(f"  [/download/all] On-demand rebuild failed, serving cached: {e}")

    return send_file(
        BytesIO(info['bytes']),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=f"All_Brands_{date_str}.xlsx"
    )


@app.route('/download/multi', methods=['GET'])
def download_multi_selected():
    brands_param = request.args.get('brands', '')
    if not brands_param:
        return jsonify({"error": "Missing 'brands' param"}), 400

    abbrs = [b.strip().upper() for b in brands_param.split(',') if b.strip()]
    if not abbrs:
        return jsonify({"error": "No valid brands"}), 400

    with _inv_lock:
        all_brands = dict(_inventory['brands'])

    brands_list = []
    for abbr in abbrs:
        if abbr in all_brands:
            sorted_items = sorted(all_brands[abbr]['items'],
                key=lambda x: x.get('total_warehouse', 0), reverse=True)
            brands_list.append({
                'brand_name': all_brands[abbr]['name'],
                'items': sorted_items
            })

    if not brands_list:
        return jsonify({"error": "No matching brands in inventory"}), 404

    brands_list.sort(
        key=lambda b: sum(i.get('total_warehouse', 0) for i in b['items']),
        reverse=True)

    # Annotate items with category/fit for prepack chart matching
    for b in brands_list:
        b['items'] = _annotate_items_for_prepack(b['items'])
    pd_snap = _fresh_prepack_defaults()

    xl_bytes = build_multi_brand_excel(brands_list, S3_PHOTOS_URL, prepack_defaults=pd_snap)
    date_str = datetime.utcnow().strftime('%Y-%m-%d')

    return send_file(
        BytesIO(xl_bytes),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"Versa_{len(brands_list)}_Brands_{date_str}.xlsx"
    )


@app.route('/upload', methods=['POST', 'OPTIONS'])
def upload_inventory():
    if request.method == 'OPTIONS':
        return '', 204

    if 'file' not in request.files:
        return jsonify({"error": "No file. Use multipart form with 'file' field."}), 400

    file = request.files['file']
    data = file.read()
    if not data:
        return jsonify({"error": "Empty file"}), 400

    try:
        s3 = get_s3()
        s3.put_object(Bucket=S3_BUCKET, Key=S3_INVENTORY_KEY, Body=data,
                       ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        return jsonify({"error": f"S3 upload failed: {e}"}), 500

    sync_inventory()
    with _inv_lock:
        count = _inventory['item_count']

    trigger_background_generation()

    return jsonify({"status": "ok", "message": f"Uploaded and synced. {count} items.", "item_count": count})


@app.route('/export-overseas-summary', methods=['POST', 'OPTIONS'])
def export_overseas_summary():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        req = request.get_json()
        if not req or 'data' not in req:
            return jsonify({"error": "Missing 'data'"}), 400
        data = req['data']
        s3_url = req.get('s3_base_url', S3_PHOTOS_URL)
        fname = req.get('filename', 'Overseas_Summary')
        if not data:
            return jsonify({"error": "Empty data"}), 400

        xl_bytes = build_overseas_summary_excel(fname, data, s3_url)
        ts = datetime.now().strftime('%Y-%m-%d')
        return send_file(BytesIO(xl_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=f"{fname}_{ts}.xlsx")
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


def build_overseas_summary_excel(title, items, s3_base_url):
    """Build a formatted overseas summary Excel with images, date banners, brand separators."""
    buf = BytesIO()
    wb = xlsxwriter.Workbook(buf, {'in_memory': True, 'strings_to_formulas': False})
    wb.set_properties({'title': f'Versa - {title}', 'author': 'Versa Inventory System'})
    ws = wb.add_worksheet('Overseas Summary')

    # ── Formats ──
    font = STYLE_CONFIG['font_name']
    fmt_header = wb.add_format({
        'bold': True, 'font_name': font, 'font_size': 11,
        'bg_color': '#1E293B', 'font_color': '#FFFFFF',
        'border': 1, 'border_color': '#1E293B',
        'align': 'center', 'valign': 'vcenter'
    })
    base = {
        'font_name': font, 'font_size': 10,
        'text_wrap': True, 'valign': 'vcenter', 'align': 'center',
        'border': 1, 'border_color': '#D1D5DB'
    }
    fmt_odd = wb.add_format({**base, 'bg_color': '#FFFFFF'})
    fmt_even = wb.add_format({**base, 'bg_color': '#F8FAFC'})
    fmt_num_odd = wb.add_format({**base, 'bg_color': '#FFFFFF', 'num_format': '#,##0'})
    fmt_num_even = wb.add_format({**base, 'bg_color': '#F8FAFC', 'num_format': '#,##0'})
    fmt_neg_odd = wb.add_format({**base, 'bg_color': '#FFFFFF', 'num_format': '#,##0', 'font_color': '#DC2626'})
    fmt_neg_even = wb.add_format({**base, 'bg_color': '#F8FAFC', 'num_format': '#,##0', 'font_color': '#DC2626'})
    fmt_ats_odd = wb.add_format({**base, 'bg_color': '#FFFFFF', 'num_format': '#,##0', 'bold': True, 'font_size': 11})
    fmt_ats_even = wb.add_format({**base, 'bg_color': '#F8FAFC', 'num_format': '#,##0', 'bold': True, 'font_size': 11})
    fmt_date_banner = wb.add_format({
        'bold': True, 'font_name': font, 'font_size': 12,
        'bg_color': '#FEF3C7', 'font_color': '#92400E',
        'border': 1, 'border_color': '#F59E0B',
        'align': 'left', 'valign': 'vcenter'
    })
    fmt_brand_sep = wb.add_format({
        'bold': True, 'font_name': font, 'font_size': 10,
        'bg_color': '#F1F5F9', 'font_color': '#475569',
        'border': 1, 'border_color': '#E2E8F0',
        'align': 'left', 'valign': 'vcenter'
    })
    fmt_sku_odd = wb.add_format({**base, 'bg_color': '#FFFFFF', 'bold': True, 'align': 'left'})
    fmt_sku_even = wb.add_format({**base, 'bg_color': '#F8FAFC', 'bold': True, 'align': 'left'})

    # ── Headers ──
    headers = ['IMAGE', 'SKU', 'Brand', 'Color', 'Fit', 'Fabrication',
               'Production', 'Factory', 'PO', 'Ex-Factory', 'Arrival',
               'Produced', 'Deducted', 'Flow ATS']
    col_widths = [COL_WIDTH_UNITS, 22, 12, 20, 12, 32, 22, 14, 22, 14, 14, 12, 12, 12]

    ws.hide_gridlines(2)
    ws.freeze_panes(1, 0)
    ws.set_row(0, 28)
    for c, h in enumerate(headers):
        ws.write(0, c, h, fmt_header)
        ws.set_column(c, c, col_widths[c])

    # ── Download images ──
    img_items = [{'sku': d.get('sku', ''), 'brand_abbr': d.get('brand_abbr', ''),
                  'brand': d.get('brand_abbr', '')} for d in items]
    imgs = download_images_for_items(img_items, s3_base_url, use_cache=True)

    # ── Write rows with date banners & brand separators ──
    row = 1
    data_row_idx = 0  # for alternating colors
    last_date_key = None
    last_brand = None
    num_cols = len(headers)

    for i, item in enumerate(items):
        date_key = item.get('arrival', '') or item.get('ex_factory', '') or 'No Date'

        # Date group banner
        if date_key != last_date_key:
            etd = item.get('ex_factory', '—') or '—'
            arr = item.get('arrival', '—') or '—'
            ws.set_row(row, 26)
            ws.merge_range(row, 0, row, num_cols - 1,
                           f"  📅  Ex-Factory: {etd}  →  Arrival: {arr}", fmt_date_banner)
            row += 1
            last_date_key = date_key
            last_brand = None
            data_row_idx = 0

        # Brand separator
        brand = item.get('brand_full', '') or item.get('brand_abbr', '')
        brand_key = item.get('brand_abbr', '')
        if brand_key != last_brand:
            ws.set_row(row, 20)
            ws.merge_range(row, 0, row, num_cols - 1,
                           f"  ▸  {brand}", fmt_brand_sep)
            row += 1
            last_brand = brand_key
            data_row_idx = 0

        # Data row
        even = data_row_idx % 2 == 1
        cf = fmt_even if even else fmt_odd
        nf = fmt_num_even if even else fmt_num_odd
        sf = fmt_sku_even if even else fmt_sku_odd
        deducted_fmt = fmt_neg_even if even else fmt_neg_odd
        ats_fmt = fmt_ats_even if even else fmt_ats_odd

        ws.set_row(row, 112.5)

        # Image
        img = imgs.get(i)
        if img:
            try:
                ws.insert_image(row, 0, "img.png", _padded_image_opts(img))
            except Exception:
                ws.write(row, 0, "Error", cf)
        else:
            ws.write(row, 0, "No Image", cf)

        ws.write(row, 1, item.get('sku', ''), sf)
        ws.write(row, 2, item.get('brand_abbr', ''), cf)
        ws.write(row, 3, item.get('color', ''), cf)
        ws.write(row, 4, item.get('fit', ''), cf)
        ws.write(row, 5, item.get('fabrication', ''), cf)
        ws.write(row, 6, item.get('production', ''), cf)
        # Factory abbreviation only (matches every other export)
        ws.write(row, 7, _factory_label(item.get('production')), cf)
        ws.write(row, 8, item.get('po', ''), cf)
        ws.write(row, 9, item.get('ex_factory', ''), cf)
        ws.write(row, 10, item.get('arrival', ''), cf)
        ws.write(row, 11, item.get('produced', 0), nf)
        deducted = item.get('deducted', 0)
        ws.write(row, 12, deducted, deducted_fmt if deducted else nf)
        ws.write(row, 13, item.get('flow_ats', 0), ats_fmt)

        row += 1
        data_row_idx += 1

    wb.close()
    return buf.getvalue()


@app.route('/export', methods=['POST', 'OPTIONS'])
def export_single():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        req = request.get_json()
        if not req or 'data' not in req:
            return jsonify({"error": "Missing 'data'"}), 400
        data = req['data']
        # Anonymous catalog holders: NJ stock and hidden-landing productions are
        # admin-only. Scrub even client-posted rows so a stale cached page can't
        # export them.
        if getattr(g, '_catalog_scope', None) is not None and isinstance(data, list):
            data = _strip_hidden_prod_export_rows(_strip_nj_rows(data))
        s3_url = req.get('s3_base_url', S3_PHOTOS_URL)
        fname = req.get('filename', 'Export')
        view_mode = req.get('view_mode', 'all')
        is_order = req.get('is_order', False)
        catalog_mode = req.get('catalog_mode', False)
        flow_mode = req.get('flow_mode', False)
        # Omitted, null, or EMPTY client rules all mean "server, use your own"
        # (reloaded from S3 so a stale worker copy can't bake old grids in).
        # The frontend can legitimately send [] during its first seconds of
        # load, before /prepack-defaults resolves — substituting server rules
        # there matches the old behavior and avoids silently exporting the
        # hardcoded fallback grids.
        prepack_defaults = req.get('prepack_defaults')
        if not prepack_defaults:
            prepack_defaults = _fresh_prepack_defaults()
        if not data:
            return jsonify({"error": "Empty data"}), 400

        print(f"[Export] {fname}: {len(data)} items, view={view_mode}, flow={flow_mode}, prepack_rules={len(prepack_defaults)}")

        xl_bytes = build_brand_excel(fname, data, s3_url, view_mode=view_mode,
                                     is_order=is_order, catalog_mode=catalog_mode,
                                     prepack_defaults=prepack_defaults,
                                     flow_mode=flow_mode)
        ts = datetime.now().strftime('%Y-%m-%d')
        return send_file(BytesIO(xl_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=f"{fname}_{ts}.xlsx")
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"[Export] FAILED for {fname}: {e}")
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route('/export-pdf', methods=['POST', 'OPTIONS'])
def export_pdf():
    if request.method == 'OPTIONS':
        return '', 204
    if not HAS_REPORTLAB:
        return jsonify({"error": "reportlab not installed on server. Add 'reportlab' to requirements.txt and redeploy."}), 500
    try:
        req = request.get_json()
        if not req or 'data' not in req:
            return jsonify({"error": "Missing 'data'"}), 400
        data = req['data']
        if getattr(g, '_catalog_scope', None) is not None and isinstance(data, list):
            data = _strip_hidden_prod_export_rows(_strip_nj_rows(data))
        if not data:
            return jsonify({"error": "Empty data"}), 400
        s3_url = req.get('s3_base_url', S3_PHOTOS_URL)
        title = req.get('title', 'Export')
        subtitle = req.get('subtitle', f"{len(data)} styles")
        show_qty = req.get('show_qty', False)
        fname = req.get('filename', 'Export')

        pdf_bytes = build_brand_pdf(title, data, s3_url,
                                    subtitle=subtitle, show_qty=show_qty)
        ts = datetime.now().strftime('%Y-%m-%d')
        return send_file(BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True, download_name=f"{fname}_{ts}.pdf")
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route('/export-multi', methods=['POST', 'OPTIONS'])
def export_multi():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        req = request.get_json()
        if not req or 'brands' not in req:
            return jsonify({"error": "Missing 'brands'"}), 400
        brands_data = req['brands']
        if getattr(g, '_catalog_scope', None) is not None and isinstance(brands_data, list):
            for _b in brands_data:
                if isinstance(_b, dict) and isinstance(_b.get('items'), list):
                    _b['items'] = _strip_hidden_prod_export_rows(_strip_nj_rows(_b['items']))
        s3_url = req.get('s3_base_url', S3_PHOTOS_URL)
        fname = req.get('filename', 'Multi_Brand')
        catalog_mode = req.get('catalog_mode', False)
        view_mode = req.get('view_mode', 'all')
        # 📋 flow_mode enables PO Name column in catalog overseas exports
        flow_mode = req.get('flow_mode', False)
        # Omitted, null, or EMPTY client rules all mean "server, use your own"
        # (reloaded from S3 so a stale worker copy can't bake old grids in).
        # The frontend can legitimately send [] during its first seconds of
        # load, before /prepack-defaults resolves — substituting server rules
        # there matches the old behavior and avoids silently exporting the
        # hardcoded fallback grids.
        prepack_defaults = req.get('prepack_defaults')
        if not prepack_defaults:
            prepack_defaults = _fresh_prepack_defaults()
        if not brands_data:
            return jsonify({"error": "Empty brands"}), 400

        xl_bytes = build_multi_brand_excel(brands_data, s3_url,
                                           catalog_mode=catalog_mode, view_mode=view_mode,
                                           flow_mode=flow_mode,
                                           prepack_defaults=prepack_defaults)
        ts = datetime.now().strftime('%Y-%m-%d')
        return send_file(BytesIO(xl_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=f"{fname}_{ts}.xlsx")
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


# ── Ship-plan line-sheet export (Confirm Pre-PO) ─────────────────────────
# Multi-tab workbook in the standard line-sheet format (embedded images,
# styled headers) with ship-plan columns. Tabs are CALLER-defined — the tool
# sends flat / per-brand / per-brand+color groupings; the server just renders.
SHIP_PLAN_HEADERS = ['IMAGE', 'SKU', 'Brand', 'Color', 'Fit', 'Fabrication',
                     'Production #', 'Arrival', 'Warehouse',
                     'Units to Ship', 'Shortfall']


def build_ship_plan_excel(tabs, s3_base_url, headers=None):
    """headers=None → the full SHIP_PLAN_HEADERS set (Confirm Pre-PO exports);
    callers may pass a trimmed list (the weekly APO report drops Shortfall)."""
    sheet_headers = list(headers) if headers else SHIP_PLAN_HEADERS
    output = BytesIO()
    wb = xlsxwriter.Workbook(output, {'in_memory': True, 'strings_to_formulas': False})
    # One image fetch for the whole workbook (dedupes by base style), then
    # slice per tab — same pattern as build_multi_brand_excel.
    all_items = []
    offsets = []
    for t in tabs:
        rows = t.get('items') or []
        offsets.append((len(all_items), len(rows)))
        all_items.extend(rows)
    images = download_images_for_items(all_items, s3_base_url)
    seen = set()
    for ti, t in enumerate(tabs):
        # Same sanitize + case-insensitive dedupe as the multi-brand builder
        # (Excel sheet names are case-insensitively unique; xlsxwriter raises).
        name = _safe_multi_sheet_name(t.get('name'), seen, ti, fallback_prefix='Tab')
        ws = wb.add_worksheet(name)
        fmts, hdrs = _setup_worksheet(wb, ws, headers_override=(t.get('headers') or sheet_headers))
        off, cnt = offsets[ti]
        local = {i - off: img for i, img in images.items() if off <= i < off + cnt}
        # No size charts here — a ship plan is about allocation, not prepacks.
        _write_rows(wb, ws, t.get('items') or [], local, fmts, headers=hdrs)
    wb.close()
    output.seek(0)
    return output.getvalue()


@app.route('/export-ship-plan', methods=['POST', 'OPTIONS'])
def export_ship_plan():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        req = request.get_json() or {}
        tabs = req.get('tabs') or []
        if not tabs or not any(t.get('items') for t in tabs):
            return jsonify({"error": "No rows to export"}), 400
        s3_url = req.get('s3_base_url', S3_PHOTOS_URL)
        fname = req.get('filename', 'Ship_Plan')
        # Optional caller-defined column set (Customer Selling sheets send
        # their own). Unknown header names render blank; IMAGE must stay in
        # column 0 because the renderer anchors images there.
        cols = req.get('columns')
        if cols:
            cols = [str(c) for c in cols][:20]
            if not cols or cols[0] != 'IMAGE':
                cols = ['IMAGE'] + [c for c in cols if c != 'IMAGE']
        xl_bytes = build_ship_plan_excel(tabs, s3_url, headers=cols)
        ts = datetime.now().strftime('%Y-%m-%d')
        return send_file(BytesIO(xl_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=f"{fname}_{ts}.xlsx")
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


# ── APO brand-color line sheets (weekly email report) ────────────────────────
# GET /export-apo-brandcolor?customer=<name>[&exclude_po=TK[,TOK2]]
# ALL open allocations (the /apo Dropbox feed) for one customer, rendered as a
# line-sheet workbook with one tab per "<Brand> Solids"/"<Brand> Fancies"
# bucket — the same visual format and the same color rule as the Confirm
# Pre-PO tool's "By Brand + Color" export. The helpers below are faithful
# Python ports of the frontend functions of the same name (index.html);
# if the frontend color rules change, change them here too.
#
# Supply columns use a documented SIMPLE rule (not the frontend routing
# engine): warehouse free stock first (committed/allocated are NEGATIVE in
# the feed; the line's own qty is added back since 'allocated' already
# contains it), then production FIFO by arrival date. The Confirm Pre-PO
# tool remains the authority on exact engine placement.

_APO_BRAND_FULL = {  # frontend BRAND_MAPPING key → full_name
    'NAUTICA': 'Nautica', 'DKNY': 'DKNY', 'EB': 'Eddie Bauer', 'REEBOK': 'Reebok',
    'VINCE': 'Vince Camuto', 'BEN': 'Ben Sherman', 'USPA': 'U.S. Polo Assn.',
    'CHAPS': 'Chaps', 'LUCKY': 'Lucky Brand', 'JNY': 'Jones New York',
    'BEENE': 'Geoffrey Beene', 'NICOLE': 'Nicole Miller', 'SHAQ': "Shaquille O'Neal",
    'TAYION': 'Tayion', 'STRAHAN': 'Michael Strahan', 'VD': 'Von Dutch',
    'VERSA': 'Versa', 'CHEROKEE': 'Cherokee', 'AMERICA': 'American Crew',
    'BLO': 'Bloomingdales Private Label', 'BLACK': 'Black Label', 'DN': 'Divine 9',
    'KL': 'Karl Lagerfeld Paris', 'NE': 'Neiman Marcus', 'RG': 'Robert Graham',
    'DH': 'Daniel Hechter',
}

_APO_BRAND_PREFIX = {  # frontend BRAND_IMAGE_PREFIX (brand key → 2-char SKU code)
    'NAUTICA': 'NA', 'DKNY': 'DK', 'EB': 'EB', 'REEBOK': 'RB', 'VINCE': 'VC',
    'BEN': 'BE', 'USPA': 'US', 'CHAPS': 'CH', 'LUCKY': 'LB', 'JNY': 'JN',
    'BEENE': 'GB', 'NICOLE': 'NM', 'SHAQ': 'SH', 'TAYION': 'TA', 'STRAHAN': 'MS',
    'VD': 'VD', 'VERSA': 'VR', 'CHEROKEE': 'CK', 'AMERICA': 'AC', 'BLO': 'BL',
    'BLACK': 'BL', 'DN': 'D9', 'KL': 'KL', 'RG': 'RG', 'NE': 'NE', 'DH': 'DH',
}

_apo_color_map_cache = {'map': None, 'time': 0.0}
_APO_COLOR_MAP_TTL = 3600  # frontend refetches hourly; match it

def _apo_color_map():
    """style_color_map.xlsx from S3 as {KEY: raw color description}. Mirrors
    the frontend loadColorMapFromS3: 'Color Map' sheet (else first), key col
    'Key' (fallback 'Style_Number', 'Style_Num'), value 'Color_Description'."""
    now = time.time()
    if _apo_color_map_cache['map'] is not None and now - _apo_color_map_cache['time'] < _APO_COLOR_MAP_TTL:
        return _apo_color_map_cache['map']
    try:
        from swatch_extractor import _download_color_map, _resolve_color_map_sheet
        wb, _etag, _size = _download_color_map(get_s3, S3_BUCKET)
        ws = _resolve_color_map_sheet(wb)
        rows = ws.iter_rows(values_only=True)
        headers = [str(c or '').strip() for c in next(rows)]
        def _ci(name):
            try: return headers.index(name)
            except ValueError: return -1
        i_key, i_sn, i_sn2, i_val = _ci('Key'), _ci('Style_Number'), _ci('Style_Num'), _ci('Color_Description')
        cmap = {}
        for r in rows:
            def _cell(i):
                return str(r[i]).strip() if (0 <= i < len(r) and r[i] is not None) else ''
            key = _cell(i_key) or _cell(i_sn) or _cell(i_sn2)
            val = _cell(i_val)
            if key and val:
                cmap[key.upper()] = val
        wb.close()
        _apo_color_map_cache['map'] = cmap
        _apo_color_map_cache['time'] = now
        print(f'[APO Export] color map loaded: {len(cmap)} keys')
        return cmap
    except Exception as e:
        print(f'[APO Export] color map load failed: {e}')
        return _apo_color_map_cache['map'] or {}

_APO_COLOR_ABBR = {'BLK': 'Black', 'WHT': 'White', 'BLU': 'Blue', 'NVY': 'Navy', 'GRY': 'Grey'}

def _apo_format_color(raw):
    """Port of frontend formatColorName."""
    if not raw:
        return ''
    s = str(raw).strip()
    def _word(m):
        w = m.group(1)
        hit = _APO_COLOR_ABBR.get(w.upper())
        return hit if hit else (w[:1].upper() + w[1:].lower())
    s = re.sub(r'\b([A-Za-z]+)\b', _word, s)
    return re.sub(r'\s{2,}', ' ', s).strip()

def _apo_has_pants_serial(base):
    """Port of frontend hasPantsSerial (US/GB brands, P## serial)."""
    return (len(base) >= 10 and base[2:4] in _PY_PANTS_SERIAL_BRANDS and base[6] == 'P'
            and base[7].isdigit() and base[8].isdigit() and base[9].isalpha())

def _apo_is_sportswear(base):
    """Port of frontend isSportswear (collar/fabric checks skip pants serials)."""
    if not _apo_has_pants_serial(base):
        if len(base) >= 11 and base[-1] in _PY_SPORTSWEAR_COLLARS:
            return True
        if len(base) >= 6 and base[4:6] in _PY_SPORTSWEAR_FABRICS:
            return True
    return len(base) >= 6 and base[4:6] in _PY_YM_FABRIC_CODES

def _apo_sku_prefix(base, brand_abbr):
    """Port of frontend skuImagePrefix."""
    for code in (base[2:4], base[0:2]):
        if len(code) == 2 and brand_abbr and SKU_BRAND_CODE_MAP.get(code) == brand_abbr:
            return code
    return _APO_BRAND_PREFIX.get(brand_abbr, (brand_abbr or '')[:2])

def _apo_color_fallback_keys(base, brand_abbr):
    """Port of frontend styleColorFallbackKeys — category namespace first,
    legacy bare key always last."""
    b = (base or '').upper()
    prefix = _apo_sku_prefix(b, brand_abbr)
    keys = []
    serial = b[6:9] if len(b) >= 9 else ''
    numbers = re.findall(r'\d+', b)
    padded = max(numbers, key=len).rjust(3, '0') if numbers else None
    if re.match(r'^P\d\d$', serial) or _apo_has_pants_serial(b):
        keys.append(f'{prefix}_{serial}')
    elif re.match(r'^B\d\d$', serial):  # isBlazer: serial only, by design
        keys.append(f'{prefix}_{serial}')
    elif padded and _apo_is_sportswear(b):
        keys.append(f'{prefix}_SW_{padded}')
    if padded:
        keys.append(f'{prefix}_{padded}')
    return keys

def _apo_style_color(base, brand_abbr):
    """Port of frontend getStyleColorInfo → the display string (or '')."""
    ov = _style_overrides.get(base) or {}
    if ov.get('color'):
        return str(ov['color'])
    cmap = _apo_color_map()
    raw = cmap.get(base)
    if not raw:
        for k in _apo_color_fallback_keys(base, brand_abbr):
            raw = cmap.get(k)
            if raw:
                break
    if not raw:
        return ''
    if '||' in raw:  # Nicole Miller dual-color "ground||print"
        parts = raw.split('||')
        return (_apo_format_color(parts[0]) + ' ' + _apo_format_color(parts[1])).strip()
    return _apo_format_color(raw)

_APO_BLUE_FAMILY = re.compile(r'\bnavy\b|\bblue\b|\bindigo\b')
_APO_PRINT_RE = re.compile(r'\bprint\b|\bprnt\b|\bgrnd\b|\bstripe\b|\bstripes\b|\bgeo\b|\bcheck\b')

def _apo_classify_color(color_display, brand_abbr):
    """Port of frontend classifyColor. Buckets: white/black/navy/other_solids/
    fancies; the report folds the first four into 'Solids'."""
    if not color_display:
        return 'fancies'
    c = str(color_display).strip().lower()
    # Two-part colors: split ONLY on slash with whitespace on BOTH sides;
    # "W/" means "with" and must never split. Recurse on the first half.
    parts = re.split(r'\s+/\s+', c)
    if len(parts) > 1 and parts[0].strip():
        return _apo_classify_color(parts[0].strip(), brand_abbr)
    has_print = bool(_APO_PRINT_RE.search(c))
    # Dobby is a woven texture, not a print — ANY name containing "dobby" is a
    # solid (David, Aug 24 2026), even when the name carries a stripe/check/
    # print word ("White Jacquard Stripe Dobby"). Mirrors frontend classifyColor.
    if re.search(r'\bdobby\b', c):
        if re.search(r'\bwhite\b|\bivory\b|\bcream\b', c):
            return 'white'
        if re.search(r'\bblack\b', c):
            return 'black'
        if _APO_BLUE_FAMILY.search(c):
            return 'navy'
        return 'other_solids'
    m = re.match(r'^(.*?)\s*\bs(?:olid|ld)\b', c)
    if not has_print and m and _APO_BLUE_FAMILY.search(m.group(1)):
        return 'navy'
    m2 = re.match(r'^(\S+)\s+s(?:olid|ld)$', c)
    if m2 and not has_print:
        b = m2.group(1)
        if b == 'white':
            return 'white'
        if b == 'black':
            return 'black'
        return 'other_solids'
    if brand_abbr == 'USPA' and not has_print:
        m3 = re.match(r'^(\S+)\s+s(?:olid|ld)', c)
        if m3:
            b = m3.group(1)
            if b in ('white', 'ivory', 'cream'):
                return 'white'
            if b == 'black':
                return 'black'
            return 'other_solids'
    if not has_print and re.search(r'\bs(?:olid|ld)\b', c):
        return 'other_solids'
    return 'fancies'

_APO_FABRIC_RULES = {  # frontend FABRIC_RULES, verbatim
    "AW": "4 Way Stretch", "CA": "Catatonic 95% Polyester / 5% Spandex",
    "TD": "CVC Dobby 60% Polyester / 40% Cotton", "CH": "Chambray TC Stretch",
    "CS": "Cooling Stretch", "CV": "Cotton / Poly CVC",
    "DS": "4 Way Stretch Dobby 95% Polyester / 5% Spandex",
    "OX": "PINPOINT Oxford 65%/35% Poly/Cotton", "PP": "100% Polyester- 150D",
    "SA": "150D - Sateen 100% POLYESTER", "LN": "100% Slab Linen",
    "ST": "97% Cotton 3% spandex", "SW": "97% Cotton 3% Stretch Twill",
    "SU": "Stretch Supershirt (95% Polyester, 5% Spandex)", "TR": "Traverler Stretch",
    "TW": "4 way stretch twill", "TS": "TC Stretch (77% POLYESTER /20% COTTON /3%SPANDEX)",
    "WS": "4 Way Stretch (95%,5%) Sateen", "PC": "TC Poplin 65%/35% Poly/Cotton",
    "PT": "97% Poly 3% Stretch - 150D STRETCH", "VS": "Viscose (31%) Stretch",
    "VP": "50% Viscose 50%Polyester", "LP": "Linen Polyester/Spandex",
    "MR": "50% microfiber 50% rayon", "CT": "100% Cotton", "CP": "98% Cotton / 2% Spandex",
    "BP": "50% BAMBOO / 50% POLYESTER", "TC": "TC Stretch (52P,45C,3S %)",
    "SC": "60% Cotton, 38% Poly, 2% Spand",
    "BM": "30% Rayon made from Bamboo / 30% Microfiber / 36% Poly / 4% Spandex Twill",
    "VM": "62% Poly 35% Viscose made from Bamboo 3% Spendex",
    "SP": "52% Poly 45% COTTON 3% Spand CVC YARN DYE",
    "TP": "Solid Twill 21%Rayon/75.5%Poly/3.5%Spandex", "LC": "Linen 51% Cotton / 49% Poly",
    "CX": "97% Cotton / 3% Polyster", "WF": "96% Poly 4% Spandex waffle",
    "FT": "97% POLY/ 3% SPANDEX - FLAX TEXTURE",
    "CE": "88% Polyester/ 7% Cellulose/ 5% spandex - Tech", "PK": "100% Polyester - knit",
    "PD": "60% Cotton/ 40% polyester - Dobby",
    "PY": "50% cotton / 47% polyester/ 3% spandex - CVC OXFORD",
    "UP": "95% poly / 5%spandex ---Perforated", "NY": "78% Nylon / 22% Spandex - 165GSM",
    "CL": "35% Lyocell/35%Cotton/27% Nylon/3%Spandex", "PM": "50% Polyester / 50% Microfiber",
    "PX": "95% Polyester / 5% Spandex - Core", "CN": "71% Cotton / 27% Nylon / 2% Spandex",
    "MP": "74% Modal / 26% Polyester", "LE": "100% Linen",
    "PE": "96% POLYESTER / 4% SPANDEX - END ON END", "OC": "100% Cotton - OXFORD",
    "CD": "65% Polyester / 35% Cotton - Dobby", "CY": "100% Cotton - Yarn Dye",
    "CW": "100% Cotton - Twill", "CJ": "100% Cotton - Jacquard", "LT": "45% Cotton / 55% Linen",
    "DP": "95% POLYESTER / 5% SPANDEX - KNIT PERFORMANCE",
    "PR": "87% Polyamide / 13% Elastic - 149GSM - Rhone",
    "PS": "94% Polyester / 6% Spandex - 210GSM Knit", "CG": "100% Cotton - Poplin 105gsm",
    "PA": "88% Polyester / 12% Spandex 160GSM - Seamless Lux Knit",
    "PN": "88% Polyester / 12% Spandex 160GSM - Non-Seamless", "CF": "100% Cotton 50s 2 ply",
    "CB": "98% Cotton / 2% Spandex (Bloomingdale)", "KN": "KNITS", "WT": "WOVEN TOPS",
    "SD": "SWEATERS", "SF": "Flannel (Shacket)", "SB": "Trucker (Shacket)",
    "CO": "Corduroy (Overshirt)", "SL": "Twill (Shacket)",
    "YD": "65% Polyester / 35% Cotton Yarn Dye", "KS": "Knit Sport Coat",
    "LA": "8% Lyocell / 88% Polyester / 4% Spandex 120GSM",
    "NP": "78% Nylon / 22% Spandex - 180GSM Premium Nylon",
    "PB": "100% Polyester - Imitation Cotton 130GSM",
    "PF": "92% Polyester / 8% Spandex 150GSM - Dobby",
    "PG": "73% Polyester / 5% Spandex / 22% Recycled Fiber 130GSM - Chambray End-on-End",
    "PH": "100% Polyester Polo - Mesh Sweater Knit", "PO": "100% Polyester Polo - Pique",
    "PJ": "100% Polyester Polo - Jersey", "PL": "100% Polyester Polo - Sweater Knit",
    "PU": "92% Polyester / 8% Spandex 150GSM - Lux Twisted Dobby",
    "PV": "94% Polyester / 6% Spandex - 210GSM", "PW": "100% Polyester Polo - Waffle",
    "PZ": "94% Polyester / 6% Spandex - 210GSM", "SE": "88% Polyester / 12% Spandex - 180GSM",
    "TB": "Polyester Rayon Blend", "WB": "Wool Polyester Rayon Blend", "TH": "TSHIRT",
    "HE": "HENLEY", "BC": "CARPENTERS (Bottoms)", "BR": "RIPSTOPS (Bottoms)",
    "BH": "HEAVY WEIGHT (Bottoms)", "BA": "PINSTRIPE (Bottoms)",
    "CK": "100% Cotton 100s 2 Ply (Kirkland)", "CM": "100% Cotton - Dobby for KLP",
    # Style rules Ver 3 (Aug 6, 2026)
    "CQ": "97% Cotton / 3% Spandex - Sateen 135GSM", "CR": "100% Cotton - Sateen 112GSM",
    "CU": "100% Cotton - Knit", "PQ": "65% Polyester / 35% Cotton - TC Twill",
}

def _apo_format_fabric(raw):
    """Port of frontend formatFabricName (typo fixes, spacing, Title Case)."""
    if not raw or len(str(raw)) <= 2:
        return str(raw or '')
    s = str(raw)
    s = re.sub(r'Polyster\b', 'Polyester', s, flags=re.I)
    s = re.sub(r'Spendex\b', 'Spandex', s, flags=re.I)
    s = re.sub(r'Spand\b', 'Spandex', s, flags=re.I)
    s = re.sub(r'Traverler\b', 'Traveler', s, flags=re.I)
    s = re.sub(r'Cataonic\b', 'Cationic', s, flags=re.I)
    s = re.sub(r'-{2,}', '-', s)
    s = s.replace('/', ' / ')
    s = re.sub(r'(\d)\s*%\s*', r'\1% ', s)
    s = re.sub(r'\s*-\s*', ' - ', s)
    s = re.sub(r'\s{2,}', ' ', s).strip()
    def _word(m):
        w = m.group(1)
        if w.lower() in ('from', 'made', 'with') and m.start() > 0:
            return w.lower()
        return w[:1].upper() + w[1:].lower()
    s = re.sub(r'\b([a-zA-Z]+)\b', _word, s)
    s = re.sub(r'\bCvc\b', 'CVC', s)
    s = re.sub(r'\bTc\b', 'TC', s)
    s = re.sub(r'\bGsm\b', 'GSM', s)
    return s

def _apo_fabrication(base):
    """Port of frontend getFabricFromSKU → description string."""
    ov = _style_overrides.get(base) or {}
    # Live overrides carry 'fabrication' (frontend field); 'fabric' is legacy.
    if ov.get('fabrication') or ov.get('fabric'):
        return str(ov.get('fabrication') or ov.get('fabric'))
    brand = base[2:4] if len(base) >= 4 else ''
    fab = base[4:6] if len(base) >= 6 else ''
    if not fab:
        return 'Standard Fabric'
    name = _APO_FABRIC_RULES.get(fab, '')
    # Brand-specific substitutions (frontend parseSkuComponents)
    if brand == 'CH':
        if fab == 'YD':
            name = '50% Microfiber / 50% Polyester Yarn Dye'
        if fab == 'PT':
            name = '97% Polyester / 3% Spandex (150D STRETCH)'
    if brand == 'BE' and fab == 'YD':
        name = '77% Poly / 20% Cotton / 3% Spandex'
    if brand == 'US' and fab == 'CD':
        name = '65% Polyester / 35% Cotton'
    if fab == 'PP' and len(base) >= 9 and base[6] == 'P' and base[7].isdigit():
        name = '100% Polyester Woven Pant'
    if not name:
        return 'Standard Fabric'
    return _apo_format_fabric(name)

_APO_FIT_LABELS = {  # frontend fitCodeToLabel short forms
    'SL': 'Slim Fit', 'RF': 'Regular Fit', 'TF': 'Tailored Fit', 'MF': 'Modern Fit',
    'BT': 'Big & Tall', 'WB': 'Big & Tall (Von Dutch)', 'BB': 'Big Fit', 'TT': 'Tall Fit',
    'SB': 'Short Sleeve Big', 'ST': 'Short Sleeve Tall', 'CF': 'Classic Fit',
    'AF': 'Athletic Fit', 'SS': 'Slim Fit Short Sleeve', 'SR': 'Regular Fit Short Sleeve',
    'SE': 'Slim Fit Extended Button', 'SH': 'Slim Fit Hook & Eye',
    'CE': 'Classic Fit Extended Button', 'CH': 'Classic Fit Hook & Eye',
    'CR': 'Classic Fit Reg Button', 'SF': 'Straight Fit', 'SC': 'Straight Fit Hook & Eye',
    'RR': 'Relaxed Fit', 'BR': 'Single Breaster', 'DB': 'Double Breaster',
}
_APO_PANTS_FIT_FULL = {  # frontend PANTS_FIT_CODES (full labels win on pants)
    'SE': 'Slim Fit / Extended Button', 'SH': 'Slim Fit / Hook & Eye Closure',
    'SR': 'Slim Fit / Reg Button', 'CE': 'Classic Fit / Extended Button',
    'CH': 'Classic Fit / Hook & Eye Closure', 'CR': 'Classic Fit / Reg Button',
    'SF': 'Straight Fit / Reg Button', 'SC': 'Straight Fit / Hook & Eye Closure',
    'RR': 'Relaxed Fit / Reg Button',
}

def _apo_fit_label(base):
    """Port of frontend getFitFromSKU (override → label; pants adjustments)."""
    ov = _style_overrides.get(base) or {}
    if ov.get('fit'):
        return str(ov['fit'])
    try:
        code = _py_extract_fit_code(base) or ''
    except Exception:
        code = ''
    pants = _py_is_bottom(base)
    if pants and code in _APO_PANTS_FIT_FULL:
        return _APO_PANTS_FIT_FULL[code]
    label = _APO_FIT_LABELS.get(code, 'Slim Fit')
    if pants:
        label = re.sub(r'\s*(Long|Short)\s*Sleeve\s*', ' ', label, flags=re.I)
        label = re.sub(r'\s{2,}', ' ', label).strip() or 'Slim Fit'
    return label

def _apo_parse_date(v):
    from datetime import date as _d
    if not v:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, _d):
        return v
    try:
        return datetime.fromisoformat(str(v)[:19]).date()
    except Exception:
        return None

def _apo_fmt_date(d):
    return f'{d.strftime("%b")} {d.day}, {d.year}' if d else ''

def _apo_prod_arrival(p, pants):
    """Effective arrival for a production row, mirroring the frontend's
    loadProductionFromS3 transit rule: the ledger's real arrival (port date
    + 10) when present, else ETD + 45 days (55 for pants). Rows with neither
    date are genuinely TBD."""
    from datetime import timedelta
    ad = _apo_parse_date(p.get('arrival'))
    if ad:
        return ad
    etd = _apo_parse_date(p.get('etd'))
    if etd:
        return etd + timedelta(days=55 if pants else 45)
    return None

# A2000 open orders come from the open-orders platform's public feed; cached
# briefly so the three per-customer workbooks in one report share a fetch.
_oo_orders_cache = {'data': None, 'time': 0.0}

def _fetch_a2000_orders():
    now = time.time()
    if _oo_orders_cache['data'] is not None and now - _oo_orders_cache['time'] < 600:
        return _oo_orders_cache['data']
    try:
        r = http_requests.get('https://open-orders-api.onrender.com/api/orders', timeout=45)
        if r.status_code == 200:
            orders = (r.json() or {}).get('orders') or []
            _oo_orders_cache['data'] = orders
            _oo_orders_cache['time'] = now
            return orders
        print(f"[APO report] open-orders fetch HTTP {r.status_code}", flush=True)
    except Exception as e:
        print(f"[APO report] open-orders fetch failed: {e}", flush=True)
    return _oo_orders_cache['data'] or []


# ── Allocation Dollar Value Estimated (David, Aug 10 2026) ───────────────────
# Average sales price by (customer, brand) from live A2000 open orders,
# units-weighted, recomputed on every report run. Brand is derived from the
# STYLE CODE (chars 3-4) on both the price side and the allocation side so
# the two always join, regardless of feed brand-name quirks (e.g. 'NM').

_APO_DOLLAR_CUST = {   # report customer → A2000 feed customer codes
    'tjx': {'TJMA', 'MARS'},
    'ross': {'ROSS'},
    'burlington': {'BURL'},
}

# Positional brand decode is ONLY valid on modern-format SKUs. A legacy style
# number like LUCK2221SL would read chars 3-4 ('CK') as Cherokee otherwise —
# that exact mislabel shipped in the Allocation Dollar Value email once.
_APO_MODERN_RE = re.compile(r'^[A-Z]{6}\d{3}[A-Z]{2,3}$')
_APO_LEGACY_PREFIX = (('LUCK', 'LUCKY'), ('CHER', 'CHEROKEE'), ('NAU', 'NAUTICA'),
                      ('DKNY', 'DKNY'), ('CHAPS', 'CHAPS'), ('USPA', 'USPA'),
                      ('JNY', 'JNY'), ('BEENE', 'BEENE'))

def _apo_style_brand_abbr(base):
    b = str(base or '').upper().split(' ')[0].split('-')[0]
    if _APO_MODERN_RE.match(b):
        return SKU_BRAND_CODE_MAP.get(b[2:4], '')
    for pre, ab in _APO_LEGACY_PREFIX:
        if b.startswith(pre):
            return ab
    return ''

def _apo_style_brand(base):
    abbr = _apo_style_brand_abbr(base)
    return _APO_BRAND_FULL.get(abbr, abbr or 'Other')

def _apo_avg_price_maps():
    """Returns (cust_avg, brand_avg): units-weighted average price keyed by
    (report-customer, brand_full) and by brand_full across the big 3."""
    per_cust = {}
    per_brand = {}
    for o in _fetch_a2000_orders():
        if o.get('reportType') != 'a2000':
            continue
        code = str(o.get('customer') or '').strip().upper()
        cust_l = next((k for k, codes in _APO_DOLLAR_CUST.items() if code in codes), None)
        if not cust_l:
            continue
        try:
            price = float(o.get('salesPrice') or 0)
        except Exception:
            price = 0.0
        units = int(o.get('openQty') or 0) + int(o.get('pickQty') or 0)
        if price <= 0 or units <= 0:
            continue
        base = get_base_style(str(o.get('style') or ''))
        brand = _apo_style_brand(base)
        if not brand or brand == 'Other':
            continue
        a = per_cust.setdefault((cust_l, brand), [0, 0.0])
        a[0] += units
        a[1] += units * price
        b = per_brand.setdefault(brand, [0, 0.0])
        b[0] += units
        b[1] += units * price
    cust_avg = {k: v[1] / v[0] for k, v in per_cust.items() if v[0] > 0}
    brand_avg = {k: v[1] / v[0] for k, v in per_brand.items() if v[0] > 0}
    return cust_avg, brand_avg

def build_apo_dollar_summary(customer, exclude_tokens=None):
    """Units + estimated $ by brand for one customer's open allocations —
    the numbers behind the Allocation Dollar Value email summary. Price
    fallback order: this customer's avg for the brand → big-3 cross-customer
    avg for the brand → unpriced (value omitted, flagged)."""
    with _apo_lock:
        apo_rows = list(_apo_data)
    if not apo_rows:
        apo_rows = load_apo_from_dropbox() or []
    cust_l = str(customer or '').strip().lower()
    excl = [t.strip().upper() for t in (exclude_tokens or []) if t and t.strip()]
    by_brand = {}
    for a in apo_rows:
        if str(a.get('customer') or '').strip().lower() != cust_l:
            continue
        try:
            qty = int(a.get('qty') or 0)
        except Exception:
            qty = 0
        if qty <= 0:
            continue
        po = str(a.get('po') or '').upper()
        if excl and any(t in po for t in excl):
            continue
        base = get_base_style(str(a.get('style') or ''))
        if not base:
            continue
        by_brand[_apo_style_brand(base)] = by_brand.get(_apo_style_brand(base), 0) + qty
    cust_avg, brand_avg = _apo_avg_price_maps()
    rows = []
    for brand, units in sorted(by_brand.items(), key=lambda kv: -kv[1]):
        price = cust_avg.get((cust_l, brand)) or brand_avg.get(brand)
        # Value uses the ROUNDED (displayed) price so units × Avg Price always
        # reproduces Est. Value exactly when someone checks it in Excel.
        p2 = round(price, 2) if price else None
        rows.append({
            'brand': brand,
            'units': units,
            'price': p2,
            'value': round(units * p2) if p2 else None,
            'price_source': 'customer' if (cust_l, brand) in cust_avg
                            else ('cross' if brand in brand_avg else None),
        })
    return {
        'customer': customer,
        'brands': rows,
        'total_units': sum(r['units'] for r in rows),
        'total_value': sum(r['value'] for r in rows if r['value'] is not None),
        'unpriced_brands': [r['brand'] for r in rows if r['value'] is None],
    }


def build_apo_brandcolor_excel(customer, exclude_tokens=None, dollars=False):
    """All open allocations for one customer → line-sheet xlsx bytes with
    <Brand> Solids / <Brand> Fancies tabs. Returns (xlsx_bytes, n_lines).
    dollars=True (Allocation Dollar Value Estimated report): adds Est. Price /
    Est. Value columns from live A2000 average prices and SKIPS the
    'POs Ready to Ship' tab."""
    with _apo_lock:
        apo_rows = list(_apo_data)
    if not apo_rows:
        apo_rows = load_apo_from_dropbox() or []
    cust_l = str(customer or '').strip().lower()
    excl = [t.strip().upper() for t in (exclude_tokens or []) if t and t.strip()]

    # Aggregate this customer's open allocations to base style (mirrors the
    # CPP picker: qty <= 0 rows dropped; styles compared on base style).
    agg = {}
    for a in apo_rows:
        if str(a.get('customer') or '').strip().lower() != cust_l:
            continue
        try:
            qty = int(a.get('qty') or 0)
        except Exception:
            qty = 0
        if qty <= 0:
            continue
        po = str(a.get('po') or '').upper()
        if excl and any(t in po for t in excl):
            continue
        base = get_base_style(str(a.get('style') or ''))
        if not base:
            continue
        agg[base] = agg.get(base, 0) + qty
    if not agg:
        return None, 0

    # Per-base inventory aggregation (feed rows are per size-variant SKU).
    with _inv_lock:
        inv_items = list(_inventory.get('items') or [])
    inv_by_base = {}
    for it in inv_items:
        b = get_base_style(it.get('sku') or '')
        if not b:
            continue
        d = inv_by_base.setdefault(b, {'jtw': 0, 'tr': 0, 'dcw': 0, 'qa': 0, 'nj': 0,
                                       'committed': 0, 'allocated': 0,
                                       'brand_abbr': it.get('brand_abbr') or '',
                                       'brand_full': it.get('brand_full') or ''})
        for k in ('jtw', 'tr', 'dcw', 'qa', 'nj', 'committed', 'allocated'):
            try:
                d[k] += int(it.get(k) or 0)
            except Exception:
                pass

    # Production per base, dated arrivals first (FIFO), undated last.
    prod_by_base = {}
    for p in (load_production_from_dropbox() or []):
        b = get_base_style(str(p.get('style') or ''))
        try:
            units = int(p.get('units') or 0)
        except Exception:
            units = 0
        if not b or units <= 0:
            continue
        prod_by_base.setdefault(b, []).append(p)
    # (sorted per style below — the effective arrival is transit-rule dependent
    # and the pants/45-vs-55 split is a property of the style)

    d_cust_avg, d_brand_avg = _apo_avg_price_maps() if dollars else ({}, {})

    rows = []
    for base, qty in agg.items():
        inv = inv_by_base.get(base)
        brand_abbr = (inv and inv['brand_abbr']) or _apo_style_brand_abbr(base)
        brand_full = (inv and inv['brand_full']) or _APO_BRAND_FULL.get(brand_abbr, brand_abbr or 'Other')
        color = _apo_style_color(base, brand_abbr)

        # Warehouse first: committed/allocated are NEGATIVE; add the line's own
        # qty back since 'allocated' already includes this allocation.
        wh_names = []
        wh_free = 0
        if inv:
            wh_total = inv['jtw'] + inv['tr'] + inv['dcw'] + inv['qa'] + inv.get('nj', 0)
            wh_free = max(0, min(wh_total, wh_total + inv['committed'] + inv['allocated'] + qty))
            for k, nm in (('jtw', 'JTW'), ('tr', 'TR'), ('dcw', 'DCW'), ('qa', 'QA'), ('nj', 'NJ')):
                if inv[k] > 0:
                    wh_names.append(nm)
        take_wh = min(qty, wh_free)
        remaining = qty - take_wh

        pants = _py_is_bottom(base)
        plist = sorted(prod_by_base.get(base, []),
                       key=lambda p: (_apo_prod_arrival(p, pants) is None,
                                      _apo_prod_arrival(p, pants) or datetime.max.date()))
        pulls, gate, tbd = [], None, False
        for p in plist:
            if remaining <= 0:
                break
            take = min(remaining, int(p.get('units') or 0))
            if take <= 0:
                continue
            remaining -= take
            pulls.append(str(p.get('production') or '—'))
            ad = _apo_prod_arrival(p, pants)
            if ad is None:
                tbd = True
            elif gate is None or ad > gate:
                gate = ad

        row = {
            'sku': base, 'brand_abbr': brand_abbr, 'brand_full': brand_full,
            'color': color,
            'fit': _apo_fit_label(base),
            'fabrication': _apo_fabrication(base),
            'production': ', '.join(dict.fromkeys(pulls)),
            'arrival': ('TBD' if tbd else _apo_fmt_date(gate)) if pulls else '',
            'warehouse': (', '.join(wh_names) or '—') if take_wh > 0 else '',
            'units_ship': qty,
            'shortfall': remaining,
            '_bucket': _apo_classify_color(color, brand_abbr),
            '_arr': gate,   # sort key: gating arrival date (None = now or TBD)
        }
        if dollars:
            sb = _apo_style_brand(base)
            p = d_cust_avg.get((cust_l, sb)) or d_brand_avg.get(sb)
            # Rounded price for BOTH columns so units × price = value in Excel.
            p2 = round(p, 2) if p else None
            row['est_price'] = p2
            row['est_value'] = round(qty * p2) if p2 else None
        rows.append(row)

    tabs_by = {}
    for r in rows:
        bucket = 'Fancies' if r.pop('_bucket') == 'fancies' else 'Solids'
        tabs_by.setdefault(f"{r['brand_full']} {bucket}", []).append(r)

    # Chronological within each tab (David, Aug 4 2026): in-warehouse rows
    # first (available now), then production-covered rows by soonest arrival,
    # then date-TBD production, no-supply lines last; units desc breaks ties.
    def _row_order(r):
        has_prod = bool(r['production'])
        has_wh = bool(r['warehouse'])
        if has_wh and not has_prod:
            grp = 0
        elif has_prod and r['arrival'] != 'TBD':
            grp = 1
        elif has_prod:
            grp = 2
        else:
            grp = 3
        return (grp, r['_arr'] or datetime.max.date(), -r['units_ship'])
    tabs = []
    for name, items in sorted(tabs_by.items()):
        items.sort(key=_row_order)
        for it in items:
            it.pop('_arr', None)
        tabs.append({'name': name, 'items': items})

    # ── "POs Ready to Ship" tab (David, Aug 10 2026 v2) ──────────────────────
    # REAL A2000 POs only (open-orders feed — allocations stay on the brand
    # tabs). A PO makes this tab ONLY if every open line can ship COMPLETE
    # today: own free warehouse stock, or — for a style that isn't available —
    # the identical design under another customer prefix with enough free
    # stock ("available now", yellow row, original style's arrival kept).
    # All-or-nothing per PO: a PO with any uncoverable line is left off.
    _A2000_CUST = {
        'tjx': {'TJMA', 'MARS'},   # TJ Maxx + Marshalls (US — never TK Maxx/UK/CA/AU)
        'ross': {'ROSS'},
        'burlington': {'BURL'},
    }

    # Free-now stock per inventory base (no add-back — a sub carries its own
    # commitments), grouped by design (base minus 2-char customer prefix).
    sub_free = {}
    design_ix = {}
    for b, d in inv_by_base.items():
        wt = d['jtw'] + d['tr'] + d['dcw'] + d['qa'] + d.get('nj', 0)
        free = max(0, min(wt, wt + d['committed'] + d['allocated']))
        if free > 0 and len(b) >= 8:
            sub_free[b] = free
            design_ix.setdefault(b[2:], []).append(b)
    for k in design_ix:
        design_ix[k].sort(key=lambda b: -sub_free[b])

    def _earliest_arrival_text(base):
        pants = _py_is_bottom(base)
        best, tbd = None, False
        for p in prod_by_base.get(base, []):
            ad = _apo_prod_arrival(p, pants)
            if ad is None:
                tbd = True
            elif best is None or ad < best:
                best = ad
        if best:
            return f"arrives {_apo_fmt_date(best)}"
        return 'production date TBD' if tbd else 'no supply'

    po_rows = []
    # Dollar-value report: same brand tabs, NO "POs Ready to Ship" tab.
    a2k_codes = None if dollars else _A2000_CUST.get(cust_l)
    my_lines = []
    if a2k_codes:
        for o in _fetch_a2000_orders():
            if (o.get('reportType') == 'a2000'
                    and str(o.get('customer') or '').strip().upper() in a2k_codes
                    and int(o.get('openQty') or 0) > 0):
                my_lines.append(o)
    if my_lines:
        # 'committed' in the ATS feed carries A2000 open orders, so add back
        # this customer's own open qty per base (same add-back idea as the
        # allocation tabs use for 'allocated').
        own_qty = {}
        for o in my_lines:
            b = get_base_style(str(o.get('style') or ''))
            own_qty[b] = own_qty.get(b, 0) + int(o.get('openQty') or 0)
        own_pool = {}
        for b, q in own_qty.items():
            d = inv_by_base.get(b)
            if not d:
                own_pool[b] = 0
                continue
            wt = d['jtw'] + d['tr'] + d['dcw'] + d['qa'] + d.get('nj', 0)
            own_pool[b] = max(0, min(wt, wt + d['committed'] + d['allocated'] + q))

        def _oiso(s):
            try:
                return datetime.fromisoformat(str(s)[:19]).date()
            except Exception:
                return None

        by_po = {}
        for o in my_lines:
            by_po.setdefault(str(o.get('orderNo') or '—').strip(), []).append(o)

        def _po_sort(kv):
            po, ls = kv
            ds = [d for d in (_oiso(l.get('startDate')) for l in ls) if d]
            return (min(ds) if ds else datetime.max.date(), po)

        for po, ls in sorted(by_po.items(), key=_po_sort):
            # Tentative pass — commit pool consumption only if the WHOLE PO fits.
            plan, ok = [], True
            tent_own, tent_sub = {}, {}
            for o in ls:
                b = get_base_style(str(o.get('style') or ''))
                need = int(o.get('openQty') or 0)
                if own_pool.get(b, 0) - tent_own.get(b, 0) >= need:
                    tent_own[b] = tent_own.get(b, 0) + need
                    plan.append((o, b, need, None))
                    continue
                sub = None
                if len(b) >= 8:
                    for cand in design_ix.get(b[2:], []):
                        if cand != b and sub_free.get(cand, 0) - tent_sub.get(cand, 0) >= need:
                            sub = cand
                            break
                if sub:
                    tent_sub[sub] = tent_sub.get(sub, 0) + need
                    plan.append((o, b, need, sub))
                else:
                    ok = False
                    break
            if not ok:
                continue
            for b, q in tent_own.items():
                own_pool[b] = own_pool.get(b, 0) - q
            for cand, q in tent_sub.items():
                sub_free[cand] = sub_free.get(cand, 0) - q
            for o, b, need, sub in plan:
                inv = inv_by_base.get(b) or (inv_by_base.get(sub) if sub else None)
                s_abbr = (inv and inv['brand_abbr']) or SKU_BRAND_CODE_MAP.get(b[2:4] if len(b) >= 4 else '', '')
                s_full = (inv and inv['brand_full']) or _APO_BRAND_FULL.get(s_abbr, s_abbr or 'Other')
                sd, cd = _oiso(o.get('startDate')), _oiso(o.get('cancelDate'))
                window = ' – '.join(x for x in (
                    _apo_fmt_date(sd) if sd else '', _apo_fmt_date(cd) if cd else '') if x)
                po_rows.append({
                    'sku': b, 'brand_abbr': s_abbr, 'brand_full': s_full,
                    'color': _apo_style_color(b, s_abbr),
                    'fit': _apo_fit_label(b),
                    'fabrication': _apo_fabrication(b),
                    'po': po,
                    'units_alloc': need,
                    'ship_window': window,
                    'available': f"{sub} — available now" if sub else 'In warehouse',
                    'orig_arrival': _earliest_arrival_text(b) if sub else '',
                    '_yellow': bool(sub),
                })
    if po_rows:
        tabs.insert(0, {
            'name': 'POs Ready to Ship',
            'items': po_rows,
            'headers': ['IMAGE', 'SKU', 'Brand', 'Color', 'Fit', 'Fabrication',
                        'PO #', 'Units', 'Ship Window',
                        'Available', 'Original Arrival'],
        })

    # No Shortfall column on the weekly APO report (David, Aug 4 2026) —
    # the CPP tool's own exports keep it.
    apo_headers = [h for h in SHIP_PLAN_HEADERS if h != 'Shortfall']
    if dollars:
        apo_headers = apo_headers + ['Est. Price', 'Est. Value']
    return build_ship_plan_excel(tabs, S3_PHOTOS_URL, headers=apo_headers), len(rows)


@app.route('/export-apo-brandcolor', methods=['GET', 'OPTIONS'])
def export_apo_brandcolor():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        customer = (request.args.get('customer') or '').strip()
        if not customer:
            return jsonify({'error': 'customer parameter required'}), 400
        excl = (request.args.get('exclude_po') or '').split(',')
        dollars = str(request.args.get('dollars') or '').strip() in ('1', 'true', 'yes')
        xl_bytes, n = build_apo_brandcolor_excel(customer, excl, dollars=dollars)
        if not xl_bytes:
            return jsonify({'error': f'No open allocations for {customer}'}), 404
        ts = datetime.now().strftime('%Y-%m-%d')
        label = 'Dollar Value Estimated' if dollars else 'By Color-Brand'
        fname = re.sub(r'[\\/:*?"<>|]+', '', f'{customer} Allocations - {label}')
        return send_file(BytesIO(xl_bytes),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=f'{fname}_{ts}.xlsx')
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/apo-dollar-summary', methods=['GET', 'OPTIONS'])
def apo_dollar_summary_route():
    """Units + estimated $ by brand for one customer's open allocations —
    powers the Allocation Dollar Value Estimated email summary. Prices are
    the same average-price maps the dollars=1 workbook uses."""
    if request.method == 'OPTIONS':
        return '', 204
    try:
        customer = (request.args.get('customer') or '').strip()
        if not customer:
            return jsonify({'error': 'customer parameter required'}), 400
        excl = (request.args.get('exclude_po') or '').split(',')
        return jsonify(build_apo_dollar_summary(customer, excl))
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


# ── Style Overrides ────────────────────────────────
@app.route('/overrides', methods=['GET', 'OPTIONS'])
def get_overrides():
    if request.method == 'OPTIONS':
        return '', 204
    # ALWAYS reload from S3 to prevent stale-worker data loss.
    # This endpoint is only called on page load and when version polling detects a change,
    # so the extra S3 read is infrequent and worth the consistency guarantee.
    load_overrides_from_s3()
    with _overrides_lock:
        return jsonify({"overrides": _style_overrides})

@app.route('/overrides/version', methods=['GET', 'OPTIONS'])
def get_overrides_version():
    """Lightweight version endpoint for frontend polling — uses S3 ETag as cross-worker source of truth."""
    if request.method == 'OPTIONS':
        return '', 204
    try:
        s3 = get_s3()
        head = s3.head_object(Bucket=S3_BUCKET, Key=S3_OVERRIDES_KEY)
        s3_etag = head['ETag'].strip('"')

        # If S3 ETag differs from our local copy, another worker saved — reload
        with _overrides_lock:
            local_etag = _s3_overrides_etag
        if local_etag != s3_etag:
            print(f"  [overrides/version] ETag mismatch (local={local_etag}, s3={s3_etag[:8]}) — reloading from S3", flush=True)
            load_overrides_from_s3()

        with _overrides_lock:
            count = len(_style_overrides)
        return jsonify({"version": s3_etag, "count": count, "last_saved": _overrides_last_saved})
    except ClientError as e:
        if e.response['Error']['Code'] == 'NoSuchKey':
            return jsonify({"version": "none", "count": 0, "last_saved": 0})
        # Fallback to local hash on S3 errors
        import hashlib
        with _overrides_lock:
            count = len(_style_overrides)
            keys_str = ','.join(sorted(_style_overrides.keys()))
        fingerprint = hashlib.md5(f"{keys_str}:{_overrides_last_saved}".encode()).hexdigest()[:12]
        return jsonify({"version": fingerprint, "count": count, "last_saved": _overrides_last_saved})
    except Exception:
        import hashlib
        with _overrides_lock:
            count = len(_style_overrides)
            keys_str = ','.join(sorted(_style_overrides.keys()))
        fingerprint = hashlib.md5(f"{keys_str}:{_overrides_last_saved}".encode()).hexdigest()[:12]
        return jsonify({"version": fingerprint, "count": count, "last_saved": _overrides_last_saved})

def _backup_overrides_to_s3(overrides_dict):
    """Write a timestamped backup copy of overrides to S3. Fire-and-forget."""
    try:
        s3 = get_s3()
        ts = datetime.utcnow().strftime('%Y-%m-%d_%H-%M-%S')
        backup_key = f"inventory/overrides_backups/style_overrides_{ts}.json"
        s3.put_object(
            Bucket=S3_BUCKET,
            Key=backup_key,
            Body=json.dumps(overrides_dict).encode('utf-8'),
            ContentType='application/json'
        )
        print(f"  ✓ Backup saved: {backup_key}")
    except Exception as e:
        print(f"  ⚠ Backup failed (non-fatal): {e}")


@app.route('/overrides', methods=['POST'])
def save_overrides():
    try:
        req = request.get_json()
        if not req or 'overrides' not in req:
            return jsonify({"error": "Missing 'overrides' in request body"}), 400

        overrides = req['overrides']
        if not isinstance(overrides, dict):
            return jsonify({"error": "'overrides' must be an object"}), 400

        global _style_overrides
        # ALWAYS reload from S3 before merging — same rule as GET /overrides.
        # A cold worker's empty in-memory map must never be the merge base: on
        # 2026-08-21 a POST hit a just-deployed instance, merged against {},
        # and rewrote the canonical file with only the payload — wiping every
        # override (restored from a local backup; the async overrides_backups
        # writer had silently stopped in March 2026). S3 is the source of truth.
        load_overrides_from_s3()
        with _overrides_lock:
            current = dict(_style_overrides)
            loaded_etag = _s3_overrides_etag
        # If the reload failed AND memory holds nothing, refuse to merge-save —
        # the result would overwrite the canonical file with just this payload.
        # replace_all is exempt: its payload IS the whole intended state.
        if not current and not loaded_etag and not req.get('replace_all', False):
            return jsonify({"error": "Overrides not loaded from S3 — save refused to avoid data loss. Retry shortly."}), 503

        # SAFETY: merge incoming into current — never allow a smaller payload to wipe existing data
        # Incoming keys add/update; keys not present in incoming are preserved from S3 state.
        # replace_all=True is the explicit opt-out: the payload becomes the entire dict,
        # which is the only way to delete an override key through this endpoint.
        incoming_count = len(overrides)
        current_count = len(current)
        if req.get('replace_all', False):
            merged = dict(overrides)
            if incoming_count < current_count:
                print(f"  ⚠ replace_all: shrinking overrides {current_count} → {incoming_count}")
        else:
            merged = dict(current)
            merged.update(overrides)
            if incoming_count < current_count:
                # Merge only — do not shrink
                print(f"  ⚠ Incoming overrides ({incoming_count}) < current ({current_count}), merging (not replacing)")

        # Per-key tombstones: the frontend's delete flow removes a key from its map,
        # but merge mode preserves missing keys — so UI deletions never persisted
        # (the editor's "deletion may not persist" toast). 'deleted' lists exact keys
        # to remove; a key also present in 'overrides' wins (kept), so delete+recreate
        # in one save can never lose the new value. Ignored under replace_all (the
        # payload is already the entire dict there).
        deleted_applied = 0
        deleted_keys = req.get('deleted') or []
        if deleted_keys and not req.get('replace_all', False):
            for k in deleted_keys:
                if isinstance(k, str) and k not in overrides and merged.pop(k, None) is not None:
                    deleted_applied += 1
            if deleted_applied:
                print(f"  🗑 Tombstones applied: {deleted_applied} of {len(deleted_keys)} keys removed")

        # Find which styles have new/changed images (for CloudFront invalidation)
        changed_styles = [
            style for style, data in overrides.items()
            if isinstance(data, dict) and data.get('image') and
               data.get('image') != (current.get(style) or {}).get('image')
        ]

        with _overrides_lock:
            _style_overrides = merged
        global _overrides_last_saved
        _overrides_last_saved = time.time()

        # Async backup before saving canonical file
        threading.Thread(target=_backup_overrides_to_s3, args=(merged,), daemon=True).start()

        success = save_overrides_to_s3()

        if success:
            # Invalidate CloudFront cache for changed override images — instant update
            if changed_styles:
                paths = [f"/ALL+INVENTORY+Photos/STYLE+OVERRIDES/{s}.jpg" for s in changed_styles]
                paths += [f"/ALL+INVENTORY+Photos/STYLE+OVERRIDES/{s}.png" for s in changed_styles]
                threading.Thread(target=_invalidate_cloudfront, args=(paths,), daemon=True).start()

            # ── Invalidate in-memory _img_cache for changed styles ──
            # The plain `base_style` cache key holds whatever fallback image
            # (Dropbox / brand folder) was fetched before the override existed.
            # Drop it so subsequent exports use the new override path. (The
            # versioned `__override__:{style}:{hash}` keys are self-invalidating,
            # but old ones for previous override versions can be cleared too.)
            if changed_styles:
                with _img_lock:
                    for s in changed_styles:
                        _img_cache.pop(s, None)
                        stale_override_keys = [
                            k for k in list(_img_cache.keys())
                            if isinstance(k, str) and k.startswith(f"__override__:{s}:")
                        ]
                        for k in stale_override_keys:
                            _img_cache.pop(k, None)
                # The /image proxy keeps its own per-worker cache with no TTL —
                # without this eviction an updated override image kept serving
                # stale bytes until an instance restart (VD pants fix, Aug 2026).
                with _web_img_lock:
                    for s in changed_styles:
                        _web_img_cache.pop(str(s).upper(), None)
                        _web_img_cache.pop(get_base_style(s), None)
                print(f"  ✓ Cleared _img_cache for {len(changed_styles)} changed override styles")

            # ── Trigger background regen of pre-built exports ──
            # /download/brand and /download/all serve bytes from _exports['brands'],
            # which were built with old images. Without this, downloads stay stale
            # until the next /sync or /regenerate.
            if changed_styles:
                print(f"  Override change → triggering background export regeneration ({len(changed_styles)} styles)")
                # Unconditional — the generator queues a follow-up if a run is in flight
                trigger_background_generation()

            return jsonify({"success": True, "count": len(merged), "invalidated": len(changed_styles),
                            "deleted_applied": deleted_applied})
        else:
            return jsonify({"error": "Failed to save to S3"}), 500
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


# ── OVERRIDE IMAGE EXTRACTION (background job) ─────────────────────────
# Moves embedded base64 override images out of the overrides JSON into the
# S3 STYLE OVERRIDES folder (one JPEG per BASE style #), then strips them
# from the JSON and saves. Automates the recurring 50MB-overrides fix.
# Runs in a daemon thread: 300+ S3 uploads far exceed gunicorn's request
# timeout, so POST returns immediately and progress is polled via
# GET /overrides/extract-images/status.
# Safety: full timestamped backup (images included) is written to S3 BEFORE
# anything is stripped; an entry's image is only stripped if its upload
# succeeded; the stripped JSON is saved once at the end. Idempotent.
_extract_lock = threading.Lock()
_extract_state = {'running': False, 'done': False, 'error': None,
                  'total': 0, 'uploaded': 0, 'failed': [], 'stripped': 0,
                  'backup': None, 'json_size_mb': None, 'started_at': None}


def _run_override_image_extraction():
    global _style_overrides, _overrides_last_saved
    import base64 as _b64
    try:
        with _overrides_lock:
            snapshot = json.loads(json.dumps(_style_overrides))  # deep copy
        img_keys = [k for k, v in snapshot.items()
                    if isinstance(v, dict) and isinstance(v.get('image'), str)
                    and v['image'].startswith('data:image/')]
        if not img_keys:
            with _extract_lock:
                _extract_state.update(running=False, done=True, error=None, total=0)
            print("  Override image extraction: no embedded images found")
            return

        # Full backup FIRST — must succeed before anything is touched
        s3 = get_s3()
        ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
        backup_key = f"inventory/overrides_backups/style_overrides_pre_image_extract_{ts}.json"
        s3.put_object(Bucket=S3_BUCKET, Key=backup_key,
                      Body=json.dumps(snapshot), ContentType='application/json')

        # Group by BASE style — the S3 lookup is per base. Prefer the exact
        # base key's image, else the largest payload (observed dupes across
        # size-suffixed keys are byte-identical anyway).
        by_base = {}
        for k in img_keys:
            by_base.setdefault(get_base_style(k), []).append(k)
        with _extract_lock:
            _extract_state.update(total=len(by_base), backup=backup_key)

        uploaded, failed = [], []
        for base, keys in by_base.items():
            pick = sorted(keys, key=lambda kk: (0 if get_base_style(kk) == kk.upper().strip() else 1,
                                                -len(snapshot[kk]['image'])))[0]
            try:
                raw = _b64.b64decode(snapshot[pick]['image'].split(',', 1)[1])
                with PilImage.open(BytesIO(raw)) as im:
                    im = ImageOps.exif_transpose(im)
                    if im.mode in ('RGBA', 'LA', 'P'):
                        rgba = im.convert('RGBA')
                        bg = PilImage.new('RGB', rgba.size, (255, 255, 255))
                        bg.paste(rgba, mask=rgba.split()[-1])
                        im = bg
                    elif im.mode != 'RGB':
                        im = im.convert('RGB')
                    out = BytesIO()
                    im.save(out, format='JPEG', quality=90)
                s3.put_object(Bucket=S3_BUCKET,
                              Key=f"ALL INVENTORY Photos/STYLE OVERRIDES/{base}.jpg",
                              Body=out.getvalue(), ContentType='image/jpeg')
                uploaded.append(base)
            except Exception as e:
                failed.append({"base": base, "error": str(e)})
            with _extract_lock:
                _extract_state.update(uploaded=len(uploaded), failed=list(failed))

        # Strip image fields ONLY where the S3 upload succeeded
        ok_bases = set(uploaded)
        stripped = 0
        for k in img_keys:
            if get_base_style(k) in ok_bases:
                snapshot[k].pop('image', None)
                stripped += 1
                if not snapshot[k]:
                    snapshot.pop(k, None)

        with _overrides_lock:
            _style_overrides = snapshot
        _overrides_last_saved = time.time()
        if not save_overrides_to_s3():
            with _extract_lock:
                _extract_state.update(running=False, done=True, stripped=0,
                                      error="Images uploaded but saving the stripped JSON failed — re-run. Backup: " + backup_key)
            return

        # New S3 images must show immediately: bust CloudFront + image cache,
        # regenerate cached exports (generator queues itself if busy).
        paths = [f"/ALL+INVENTORY+Photos/STYLE+OVERRIDES/{b}.jpg" for b in uploaded]
        threading.Thread(target=_invalidate_cloudfront, args=(paths,), daemon=True).start()
        with _img_lock:
            for b in uploaded:
                _img_cache.pop(b, None)
        # Also evict the /image proxy's per-worker cache (no TTL — see the
        # matching eviction in save_overrides).
        with _web_img_lock:
            for b in uploaded:
                _web_img_cache.pop(str(b).upper(), None)
        trigger_background_generation()

        size_after = len(json.dumps(snapshot))
        with _extract_lock:
            _extract_state.update(running=False, done=True, error=None, stripped=stripped,
                                  json_size_mb=round(size_after / 1024 / 1024, 2))
        print(f"  ✓ Override image extraction: {len(uploaded)} images → S3, {stripped} fields stripped, "
              f"JSON now {size_after/1024/1024:.1f}MB, backup {backup_key}")
    except Exception as e:
        import traceback
        traceback.print_exc()
        with _extract_lock:
            _extract_state.update(running=False, done=True, error=str(e))


@app.route('/overrides/extract-images', methods=['POST', 'OPTIONS'])
def extract_override_images():
    if request.method == 'OPTIONS':
        return '', 204
    with _extract_lock:
        if _extract_state['running']:
            return jsonify({"ok": True, "already_running": True, **{k: v for k, v in _extract_state.items()}})
        _extract_state.update(running=True, done=False, error=None, total=0, uploaded=0,
                              failed=[], stripped=0, backup=None, json_size_mb=None,
                              started_at=datetime.utcnow().isoformat() + 'Z')
    threading.Thread(target=_run_override_image_extraction, daemon=True).start()
    return jsonify({"ok": True, "started": True,
                    "note": "Running in background — poll GET /overrides/extract-images/status"})


@app.route('/overrides/extract-images/status', methods=['GET', 'OPTIONS'])
def extract_override_images_status():
    if request.method == 'OPTIONS':
        return '', 204
    with _extract_lock:
        return jsonify({k: v for k, v in _extract_state.items()})


# ── CONFIRM PRE-PO: saved check history (S3-backed snapshots) ───────────
# A saved check is a SNAPSHOT of a Confirm Pre-PO run (verdict, per-line
# pulls, ship plan) plus the inputs needed to re-run it live. List GET
# strips the heavy payloads; the item GET returns the full record.
S3_CPP_CHECKS_KEY = os.environ.get('S3_CPP_CHECKS_KEY', 'inventory/cpp_check_history.json')
_cpp_checks_lock = threading.Lock()


def _load_cpp_checks():
    """Always read from S3 — saves can land on any worker."""
    try:
        s3 = get_s3()
        resp = s3.get_object(Bucket=S3_BUCKET, Key=S3_CPP_CHECKS_KEY)
        data = json.loads(resp['Body'].read())
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _save_cpp_checks(checks):
    s3 = get_s3()
    s3.put_object(Bucket=S3_BUCKET, Key=S3_CPP_CHECKS_KEY,
                  Body=json.dumps(checks), ContentType='application/json')


@app.route('/cpp-checks', methods=['GET', 'POST', 'OPTIONS'])
def cpp_checks():
    if request.method == 'OPTIONS':
        return '', 204
    if request.method == 'GET':
        checks = _load_cpp_checks()
        slim = [{k: v for k, v in c.items() if k not in ('snapshot', 'inputs')} for c in checks]
        return jsonify({"checks": slim})
    try:
        req = request.get_json()
        check = (req or {}).get('check')
        if not isinstance(check, dict) or not check.get('id'):
            return jsonify({"error": "Missing 'check' object with an 'id'"}), 400
        if len(json.dumps(check)) > 3_000_000:
            return jsonify({"error": "Check too large to save (over 3MB) — too many lines"}), 400
        with _cpp_checks_lock:
            checks = _load_cpp_checks()
            checks = [c for c in checks if c.get('id') != check['id']]
            checks.insert(0, check)
            checks = checks[:50]  # keep the most recent 50
            _save_cpp_checks(checks)
        return jsonify({"ok": True, "count": len(checks)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/cpp-checks/<check_id>', methods=['GET', 'DELETE', 'OPTIONS'])
def cpp_check_item(check_id):
    if request.method == 'OPTIONS':
        return '', 204
    if request.method == 'GET':
        for c in _load_cpp_checks():
            if c.get('id') == check_id:
                return jsonify({"check": c})
        return jsonify({"error": "Not found"}), 404
    try:
        with _cpp_checks_lock:
            checks = _load_cpp_checks()
            kept = [c for c in checks if c.get('id') != check_id]
            if len(kept) == len(checks):
                return jsonify({"error": "Not found"}), 404
            _save_cpp_checks(kept)
        return jsonify({"ok": True, "count": len(kept)})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/overrides/backups', methods=['GET', 'OPTIONS'])
def list_overrides_backups():
    """List available override backup files in S3, newest first."""
    if request.method == 'OPTIONS':
        return '', 204
    try:
        s3 = get_s3()
        resp = s3.list_objects_v2(Bucket=S3_BUCKET, Prefix='inventory/overrides_backups/')
        items = resp.get('Contents', [])
        backups = sorted(
            [{'key': o['Key'], 'size': o['Size'], 'last_modified': o['LastModified'].isoformat()} for o in items],
            key=lambda x: x['last_modified'], reverse=True
        )
        return jsonify({'backups': backups, 'count': len(backups)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/overrides/backups/<path:backup_key>', methods=['GET', 'OPTIONS'])
def restore_overrides_backup(backup_key):
    """Download a specific backup file contents."""
    if request.method == 'OPTIONS':
        return '', 204
    try:
        s3 = get_s3()
        full_key = f"inventory/overrides_backups/{backup_key}"
        obj = s3.get_object(Bucket=S3_BUCKET, Key=full_key)
        data = json.loads(obj['Body'].read().decode('utf-8'))
        return jsonify({'overrides': data, 'count': len(data), 'key': full_key})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/allocations', methods=['GET', 'OPTIONS'])
def get_allocations():
    if request.method == 'OPTIONS':
        return '', 204
    # Merge S3 sheet entries with manual entries
    s3_data = load_allocation_from_s3()
    for row in s3_data:
        row['source'] = 's3'
    with _manual_alloc_lock:
        manual = list(_manual_allocations)
    for row in manual:
        row_copy = dict(row)
        row_copy['source'] = 'manual'
    merged = s3_data + [dict(r, source='manual') for r in manual]
    return jsonify({"allocations": merged})


@app.route('/allocations/version', methods=['GET', 'OPTIONS'])
def get_allocations_version():
    """Returns a lightweight version hash of manual allocations — for live sync polling"""
    if request.method == 'OPTIONS':
        return '', 204
    import hashlib
    # Read from S3 for multi-worker consistency
    try:
        s3 = get_s3()
        resp = s3.get_object(Bucket=S3_BUCKET, Key=S3_MANUAL_ALLOC_KEY)
        data = json.loads(resp['Body'].read().decode('utf-8'))
        allocs = data if isinstance(data, list) else []
        # Also update in-memory
        global _manual_allocations
        with _manual_alloc_lock:
            _manual_allocations = allocs
    except Exception:
        with _manual_alloc_lock:
            allocs = list(_manual_allocations)
    sorted_allocs = sorted(allocs, key=lambda x: x.get('id', ''))
    version = hashlib.md5(json.dumps(sorted_allocs, sort_keys=True).encode()).hexdigest()[:12]
    return jsonify({"version": version, "count": len(allocs)})


@app.route('/deduction-assignments', methods=['GET', 'OPTIONS'])
def get_deduction_assignments():
    if request.method == 'OPTIONS':
        return '', 204
    # Always reload from S3 for multi-worker consistency
    load_deduction_assignments_from_s3()
    with _deduction_assign_lock:
        data = dict(_deduction_assignments)
    return jsonify({"assignments": data})


@app.route('/deduction-assignments', methods=['POST'])
def save_deduction_assignments():
    try:
        req = request.get_json()
        if not req or 'assignments' not in req:
            return jsonify({"error": "Missing 'assignments'"}), 400
        assignments = req['assignments']
        if not isinstance(assignments, dict):
            return jsonify({"error": "'assignments' must be an object"}), 400
        global _deduction_assignments
        with _deduction_assign_lock:
            _deduction_assignments = assignments
        success = save_deduction_assignments_to_s3()
        if success:
            return jsonify({"success": True, "count": len(assignments)})
        return jsonify({"error": "Failed to save to S3"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/prepack-defaults', methods=['GET', 'OPTIONS'])
def get_prepack_defaults():
    if request.method == 'OPTIONS':
        return '', 204
    # Always reload from S3 for multi-worker consistency
    load_prepack_defaults_from_s3()
    with _prepack_defaults_lock:
        return jsonify({"defaults": list(_prepack_defaults)})


@app.route('/prepack-defaults', methods=['POST'])
def save_prepack_defaults_route():
    try:
        global _prepack_defaults
        req = request.get_json()
        if not req or 'defaults' not in req:
            return jsonify({"error": "Missing 'defaults' in request body"}), 400
        defaults = req['defaults']
        if not isinstance(defaults, list):
            return jsonify({"error": "'defaults' must be an array"}), 400
        # SAFETY: refuse to save empty array if we already have rules
        # (prevents accidental wipe from failed client load). Reload from S3
        # first — a worker whose startup load failed would otherwise report
        # count 0 and let the wipe through.
        try:
            load_prepack_defaults_from_s3()
        except Exception:
            pass
        with _prepack_defaults_lock:
            current_count = len(_prepack_defaults)
        if len(defaults) == 0 and current_count > 0:
            print(f"  ⚠ BLOCKED: attempted to save 0 prepack defaults (server has {current_count})")
            return jsonify({"error": f"Refusing to delete all {current_count} rules. Use the UI to delete rules individually."}), 400
        with _prepack_defaults_lock:
            _prepack_defaults = defaults
        success = save_prepack_defaults_to_s3()
        if success:
            # Rule edits change the export bottom grids — invalidate cached
            # workbooks and kick off a background regeneration so instant
            # downloads pick up the new rules without waiting for the next
            # override save / inventory sync.
            global _prepack_last_saved
            _prepack_last_saved = time.time()
            try:
                trigger_background_generation()
            except Exception as e:
                print(f"  ⚠ prepack save: could not trigger export regen: {e}")
            return jsonify({"ok": True, "count": len(defaults)})
        return jsonify({"error": "Failed to save to S3"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/suppression-overrides', methods=['GET', 'OPTIONS'])
def get_suppression_overrides():
    if request.method == 'OPTIONS':
        return '', 204
    load_suppression_overrides_from_s3()
    with _suppression_overrides_lock:
        return jsonify({"overrides": list(_suppression_overrides)})


@app.route('/suppression-overrides', methods=['POST'])
def save_suppression_overrides_route():
    try:
        req = request.get_json()
        if not req or 'overrides' not in req:
            return jsonify({"error": "Missing 'overrides'"}), 400
        overrides = req['overrides']
        if not isinstance(overrides, list):
            return jsonify({"error": "'overrides' must be an array"}), 400
        global _suppression_overrides
        with _suppression_overrides_lock:
            _suppression_overrides = overrides
        success = save_suppression_overrides_to_s3()
        if success:
            return jsonify({"ok": True, "count": len(overrides)})
        return jsonify({"error": "Failed to save to S3"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/banner-rules', methods=['GET', 'OPTIONS'])
def get_banner_rules():
    if request.method == 'OPTIONS':
        return '', 204
    load_banner_rules_from_s3()
    with _banner_rules_lock:
        return jsonify({"rules": list(_banner_rules)})


@app.route('/banner-rules', methods=['POST'])
def save_banner_rules_route():
    try:
        global _banner_rules
        req = request.get_json()
        if not req or 'rules' not in req:
            return jsonify({"error": "Missing 'rules'"}), 400
        rules = req['rules']
        if not isinstance(rules, list):
            return jsonify({"error": "'rules' must be an array"}), 400
        with _banner_rules_lock:
            _banner_rules = rules
        success = save_banner_rules_to_s3()
        if success:
            return jsonify({"ok": True, "count": len(rules)})
        return jsonify({"error": "Failed to save to S3"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/manual-allocations', methods=['GET', 'OPTIONS'])
def get_manual_allocations():
    if request.method == 'OPTIONS':
        return '', 204
    # Always read from S3 for multi-worker consistency
    # (POST saves to S3, but other workers keep stale in-memory copies)
    try:
        s3 = get_s3()
        resp = s3.get_object(Bucket=S3_BUCKET, Key=S3_MANUAL_ALLOC_KEY)
        data = json.loads(resp['Body'].read().decode('utf-8'))
        fresh = data if isinstance(data, list) else []
        # Update in-memory copy so this worker stays current
        global _manual_allocations
        with _manual_alloc_lock:
            _manual_allocations = fresh
        return jsonify({"allocations": fresh})
    except ClientError as e:
        if e.response['Error']['Code'] == 'NoSuchKey':
            return jsonify({"allocations": []})
        # Fallback to in-memory if S3 read fails
        with _manual_alloc_lock:
            data = list(_manual_allocations)
        return jsonify({"allocations": data})
    except Exception:
        with _manual_alloc_lock:
            data = list(_manual_allocations)
        return jsonify({"allocations": data})


@app.route('/manual-allocations', methods=['POST'])
def save_manual_allocations():
    try:
        req = request.get_json()
        if not req or 'allocations' not in req:
            return jsonify({"error": "Missing 'allocations' in request body"}), 400
        entries = req['allocations']
        if not isinstance(entries, list):
            return jsonify({"error": "'allocations' must be a list"}), 400
        global _manual_allocations
        with _manual_alloc_lock:
            _manual_allocations = entries
        success = save_manual_allocations_to_s3()
        if success:
            return jsonify({"success": True, "count": len(entries)})
        else:
            return jsonify({"error": "Failed to save to S3"}), 500
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route('/allocations/delete-entries', methods=['POST', 'OPTIONS'])
def delete_allocation_entries():
    """Delete specific entries from S3 CSV and/or manual allocations.
    Accepts: { entries: [{sku, po, customer, qty, source}], delete_po: "PO#123" }
    - If delete_po is set, removes ALL entries (S3+manual) matching that PO.
    - If entries is set, removes each matching entry individually.
    """
    if request.method == 'OPTIONS':
        return '', 204
    try:
        req = request.get_json()
        if not req:
            return jsonify({"error": "Missing request body"}), 400

        delete_po = req.get('delete_po', '').strip()
        entries_to_delete = req.get('entries', [])
        deleted_s3 = 0
        deleted_manual = 0

        # ── Load current CSV from S3 ──
        s3 = get_s3()
        try:
            resp = s3.get_object(Bucket=S3_BUCKET, Key=S3_ALLOCATION_KEY)
            csv_text = resp['Body'].read().decode('utf-8-sig')
            csv_lines = csv_text.strip().split('\n')
        except ClientError:
            csv_lines = []

        if len(csv_lines) >= 2:
            header_line = csv_lines[0]
            headers = [h.strip() for h in header_line.split(',')]
            po_idx = next((i for i, h in enumerate(headers) if 'po' in h.lower()), 0)
            cust_idx = next((i for i, h in enumerate(headers) if 'customer' in h.lower()), 1)
            sku_idx = next((i for i, h in enumerate(headers) if 'sku' in h.lower()), 2)
            qty_idx = next((i for i, h in enumerate(headers) if 'qty' in h.lower()), 3)

            original_count = len(csv_lines) - 1  # minus header

            if delete_po:
                # Remove ALL rows matching this PO
                kept_lines = [header_line]
                for line in csv_lines[1:]:
                    cols = [c.strip() for c in line.split(',')]
                    row_po = cols[po_idx] if len(cols) > po_idx else ''
                    if row_po.strip().upper() == delete_po.upper():
                        deleted_s3 += 1
                    else:
                        kept_lines.append(line)
                csv_lines = kept_lines
            elif entries_to_delete:
                # Build set of entries to delete (sku+po+qty)
                delete_set = set()
                for e in entries_to_delete:
                    if e.get('source') == 's3':
                        key = (e.get('sku', '').upper(), e.get('po', '').strip(), str(e.get('qty', 0)))
                        delete_set.add(key)
                if delete_set:
                    kept_lines = [header_line]
                    for line in csv_lines[1:]:
                        cols = [c.strip() for c in line.split(',')]
                        sku = cols[sku_idx].upper() if len(cols) > sku_idx else ''
                        po = cols[po_idx] if len(cols) > po_idx else ''
                        qty = cols[qty_idx] if len(cols) > qty_idx else '0'
                        key = (sku, po, qty)
                        if key in delete_set:
                            deleted_s3 += 1
                            delete_set.discard(key)  # only remove first match
                        else:
                            kept_lines.append(line)
                    csv_lines = kept_lines

            # Write updated CSV back to S3
            if deleted_s3 > 0:
                new_csv = '\n'.join(csv_lines)
                s3.put_object(
                    Bucket=S3_BUCKET,
                    Key=S3_ALLOCATION_KEY,
                    Body=new_csv.encode('utf-8'),
                    ContentType='text/csv'
                )
                print(f"  ✓ Deleted {deleted_s3} S3 CSV entries, {len(csv_lines) - 1} remaining")

        # ── Also handle manual allocation deletions ──
        global _manual_allocations
        if delete_po:
            with _manual_alloc_lock:
                before = len(_manual_allocations)
                _manual_allocations = [r for r in _manual_allocations
                                       if (r.get('po', '').strip().upper() != delete_po.upper())]
                deleted_manual = before - len(_manual_allocations)
            if deleted_manual > 0:
                save_manual_allocations_to_s3()
        elif entries_to_delete:
            manual_ids = [e.get('id') for e in entries_to_delete if e.get('source') == 'manual' and e.get('id')]
            if manual_ids:
                id_set = set(manual_ids)
                with _manual_alloc_lock:
                    before = len(_manual_allocations)
                    _manual_allocations = [r for r in _manual_allocations if r.get('id') not in id_set]
                    deleted_manual = before - len(_manual_allocations)
                if deleted_manual > 0:
                    save_manual_allocations_to_s3()

        return jsonify({
            "success": True,
            "deleted_s3": deleted_s3,
            "deleted_manual": deleted_manual,
            "total_deleted": deleted_s3 + deleted_manual
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "trace": traceback.format_exc()}), 500


@app.route('/production', methods=['GET', 'OPTIONS'])
def get_production():
    if request.method == 'OPTIONS':
        return '', 204
    data = load_production_from_dropbox()
    return jsonify({"production": data})


# ============================================
# APO ALLOCATION — Dropbox CSV (hourly sync)
# Path: /EDI Team/Nuri/Python Macros/APO.csv
# Columns: B=Customer, E=Style#
# ============================================

def load_apo_from_dropbox():
    """Download APO.csv from Dropbox, parse customer (col B) and style (col E),
    cache in memory. Returns list of {customer, style, qty} dicts."""
    global _apo_data, _apo_last_sync

    token = get_dropbox_token()
    if not token:
        print("  ⚠ APO sync: no Dropbox token available")
        return False

    ns_id = os.environ.get('DROPBOX_APO_NAMESPACE_ID', '')
    print(f"  📋 Fetching APO allocation file from Dropbox: {DROPBOX_APO_PATH} (namespace: {ns_id or 'none'})", flush=True)
    try:
        headers = {
            'Authorization': f'Bearer {token}',
            'Dropbox-API-Arg': json.dumps({'path': DROPBOX_APO_PATH})
        }
        if ns_id:
            headers['Dropbox-API-Path-Root'] = json.dumps({'.tag': 'namespace_id', 'namespace_id': ns_id})
        resp = http_requests.post(
            'https://content.dropboxapi.com/2/files/download',
            headers=headers, timeout=30
        )

        if resp.status_code == 401:
            print("  ⚠ APO sync: token expired, refreshing...")
            global _dropbox_token_expires
            _dropbox_token_expires = 0
            token = get_dropbox_token()
            if not token:
                return False
            headers['Authorization'] = f'Bearer {token}'
            resp = http_requests.post(
                'https://content.dropboxapi.com/2/files/download',
                headers=headers, timeout=30
            )

        if resp.status_code != 200:
            print(f"  ⚠ APO sync: HTTP {resp.status_code}: {resp.text[:200]}")
            return False

        text = resp.content.decode('utf-8-sig', errors='replace')
        import csv as _csv, io as _io
        reader = list(_csv.reader(_io.StringIO(text)))
        rows = [r for r in reader if any(c.strip() for c in r)]
        if len(rows) < 2:
            print(f"  ⚠ APO sync: file has fewer than 2 lines (got {len(rows)})")
            return False

        # Log raw headers and first 3 data rows so we can verify structure
        raw_headers = [h.strip() for h in rows[0]]
        print(f"  📋 APO headers ({len(raw_headers)} cols): {raw_headers[:10]}")
        for i, row in enumerate(rows[1:4]):
            print(f"  📋 APO row {i+1}: {row[:11]}")

        headers_lower = [h.lower() for h in raw_headers]

        # Find columns by header name, fall back to fixed positions (B=1, E=4, 0-indexed)
        def find_col(names, fallback):
            for n in names:
                for i, h in enumerate(headers_lower):
                    if n in h:
                        return i
            return fallback

        cust_idx  = find_col(['customer', 'cust', 'account'], 1)   # col B
        style_idx = find_col(['style', 'sku', 'item #', 'item#', 'style #', 'style#', 'part'], 4)  # col E
        qty_idx   = find_col(['qty', 'quantity', 'units', 'allocated', 'open qty', 'openqty'], -1)
        po_idx    = find_col(['po', 'order', 'po #', 'po#', 'order #'], -1)

        print(f"  📋 APO cols → cust:{cust_idx} style:{style_idx} qty:{qty_idx} po:{po_idx}")

        results = []
        for cols in rows[1:]:
            if len(cols) <= max(cust_idx, style_idx):
                continue
            customer = cols[cust_idx] if len(cols) > cust_idx else ''
            style_raw = cols[style_idx] if len(cols) > style_idx else ''
            # PRESERVE FULL SKU (no suffix strip).
            # Smart routing matches APO demand to inventory by exact SKU, so
            # "TMDKPK001SLS-M" must stay as-is to land on the -M inventory pool.
            # If the ledger has base-style APO rows like "ASU201SLS", they pass
            # through unchanged and continue to match base-style inventory.
            # Only whitespace is trimmed.
            style = style_raw.upper().strip()
            if not style:
                continue
            qty = 0
            if qty_idx >= 0 and len(cols) > qty_idx:
                try:
                    qty = int(float(cols[qty_idx].replace(',', '') or 0))
                except (ValueError, TypeError):
                    qty = 0
            po = cols[po_idx].strip() if po_idx >= 0 and len(cols) > po_idx else ''
            results.append({
                'customer': customer,
                'style': style,
                'qty': qty,
                'po': po
            })

        print(f"  📋 APO sample styles: {list(set(r['style'] for r in results[:10]))}")

        with _apo_lock:
            _apo_data = results
            _apo_last_sync = time.time()

        print(f"  ✓ APO sync: {len(results)} allocation rows loaded", flush=True)
        return True

    except Exception as e:
        print(f"  ⚠ APO sync failed: {type(e).__name__}: {e}")
        return False


@app.route('/apo/explore', methods=['GET', 'OPTIONS'])
def get_apo_explore():
    """Drill into Dropbox folder structure to find APO.csv exact path."""
    if request.method == 'OPTIONS':
        return '', 204
    token = get_dropbox_token()
    if not token:
        return jsonify({'error': 'No Dropbox token'}), 500

    def list_folder(path):
        try:
            r = http_requests.post(
                'https://api.dropboxapi.com/2/files/list_folder',
                headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'},
                json={'path': path, 'recursive': False},
                timeout=15
            )
            if r.status_code == 200:
                return [{'name': e['name'], 'type': e['.tag']} for e in r.json().get('entries', [])]
            return {'error': r.status_code, 'body': r.text[:200]}
        except Exception as e:
            return {'error': str(e)}

    results = {
        'configured_path': DROPBOX_APO_PATH,
        'edi_team_root': list_folder('/EDI Team'),
    }

    # Drill into Nuri if it exists
    edi_entries = results['edi_team_root']
    if isinstance(edi_entries, list):
        nuri_match = next((e for e in edi_entries if 'nuri' in e['name'].lower()), None)
        if nuri_match:
            nuri_path = f"/EDI Team/{nuri_match['name']}"
            results['nuri_folder_name'] = nuri_match['name']
            results['nuri_contents'] = list_folder(nuri_path)

            # Drill into Python Macros if it exists
            nuri_entries = results['nuri_contents']
            if isinstance(nuri_entries, list):
                macro_match = next((e for e in nuri_entries if 'macro' in e['name'].lower() or 'python' in e['name'].lower()), None)
                if macro_match:
                    macro_path = f"{nuri_path}/{macro_match['name']}"
                    results['macros_folder_name'] = macro_match['name']
                    results['macros_contents'] = list_folder(macro_path)

    return jsonify(results)


@app.route('/apo/debug', methods=['GET', 'OPTIONS'])
def get_apo_debug():
    """Returns raw first rows of APO file for debugging column structure."""
    if request.method == 'OPTIONS':
        return '', 204
    token = get_dropbox_token()
    if not token:
        return jsonify({'error': 'No Dropbox token'}), 500
    try:
        # Try with namespace header if APO_NAMESPACE_ID env var is set
        ns_id = os.environ.get('DROPBOX_APO_NAMESPACE_ID', '')
        api_arg = {'path': DROPBOX_APO_PATH}
        req_headers = {
            'Authorization': f'Bearer {token}',
            'Dropbox-API-Arg': json.dumps(api_arg)
        }
        if ns_id:
            req_headers['Dropbox-API-Path-Root'] = json.dumps({'.tag': 'namespace_id', 'namespace_id': ns_id})

        resp = http_requests.post(
            'https://content.dropboxapi.com/2/files/download',
            headers=req_headers, timeout=30
        )
        if resp.status_code != 200:
            return jsonify({'error': f'HTTP {resp.status_code}', 'body': resp.text[:300], 'namespace_used': ns_id or 'none'}), 500
        text = resp.content.decode('utf-8-sig', errors='replace')
        lines = [l for l in text.strip().splitlines() if l.strip()]
        return jsonify({
            'path': DROPBOX_APO_PATH,
            'namespace_used': ns_id or 'none',
            'total_lines': len(lines),
            'headers': lines[0] if lines else '',
            'headers_split': [h.strip() for h in lines[0].split(',')] if lines else [],
            'sample_rows': [l for l in lines[1:6]]
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/apo', methods=['GET', 'OPTIONS'])
def get_apo():
    if request.method == 'OPTIONS':
        return '', 204
    with _apo_lock:
        data = list(_apo_data)
    last_sync = _apo_last_sync
    age_minutes = int((time.time() - last_sync) / 60) if last_sync else None
    return jsonify({
        'apo': data,
        'count': len(data),
        'last_sync': datetime.utcfromtimestamp(last_sync).isoformat() + 'Z' if last_sync else None,
        'age_minutes': age_minutes
    })


@app.route('/regenerate', methods=['POST', 'OPTIONS'])
def regenerate_exports():
    """Force re-sync inventory and regenerate all exports"""
    if request.method == 'OPTIONS':
        return '', 204

    with _export_lock:
        if _exports['generating']:
            return jsonify({"status": "already_generating", "progress": _exports['progress']})

    # Clear image cache to force re-download
    with _img_lock:
        _img_cache.clear()
    with _web_img_lock:
        _web_img_cache.clear()
    _dropbox_thumb_cache.clear()
    _dropbox_img_cache.clear()

    updated = sync_inventory()
    trigger_background_generation()

    with _inv_lock:
        count = _inventory['item_count']
        source = _inventory['source']

    return jsonify({
        "status": "regenerating",
        "inventory_source": source,
        "item_count": count,
        "message": f"Re-synced {count} items from {source}, regenerating exports..."
    })


###############################################################################
# IMAGE PROXY — Serve product images through the API
# Solves S3 browser-access issues (CORS, bucket policies, etc.)
# Serves original-resolution images (not resized like Excel thumbnails)
###############################################################################

_web_img_cache = {}   # base_style → (content_bytes, content_type)
_web_img_lock = threading.Lock()


def _fetch_raw_image(base_style, brand_abbr):
    """Download raw image bytes: base64 override → S3 override → Dropbox → S3 brand folder"""
    headers = {'User-Agent': 'Mozilla/5.0'}
    image_code = extract_image_code(base_style, brand_abbr)
    # Bottoms never use the brand+serial rungs (fabric-blind keys — VD pants
    # were serving the same-serial shirt photo). See is_bottoms_style.
    _no_serial_fallback = is_bottoms_style(base_style)

    # 0. Platform base64 override (highest priority)
    override_data = _style_overrides.get(base_style)
    if override_data and isinstance(override_data, dict) and override_data.get('image'):
        try:
            import base64
            img_str = override_data['image']
            if ',' in img_str:
                img_str = img_str.split(',', 1)[1]
            raw = base64.b64decode(img_str)
            # Detect content type
            ct = 'image/png' if raw[:4] == b'\x89PNG' else 'image/jpeg'
            return raw, ct
        except Exception:
            pass

    # 1. Try STYLE+OVERRIDES — S3 DIRECT, deliberately not CloudFront.
    #    Server-side fetches must always see the CURRENT file: the distribution
    #    ignores query strings (no cache-busting possible) and the invalidation
    #    call silently skips when CLOUDFRONT_DISTRIBUTION_ID is unset, so a
    #    replaced override image kept serving stale from Render's edge POP for
    #    up to the 24h TTL (VD pants fix, Aug 2026). Results land in
    #    _web_img_cache anyway, so the extra origin hit is once per style.
    override_base = f"{S3_OVERRIDES_IMG_URL}/{base_style}"
    for ext in ['.jpg', '.png', '.jpeg']:
        try:
            url = override_base + ext
            resp = http_requests.get(url, headers=headers, timeout=10)
            if resp.status_code == 200:
                ct = resp.headers.get('Content-Type', '').lower()
                if 'image' in ct:
                    return resp.content, ct
        except Exception:
            continue

    # 2. Try CloudFront DROPBOX_SYNC (fastest — edge cached)
    if not _no_serial_fallback and image_code.upper() in _dropbox_photo_index:
        sync_base = f"{CLOUDFRONT_DROPBOX_SYNC_URL}/{image_code}"
        for ext in ['.jpg', '.png']:
            try:
                url = sync_base + ext
                resp = http_requests.get(url, headers=headers, timeout=10)
                if resp.status_code == 200:
                    ct = resp.headers.get('Content-Type', '').lower()
                    if 'image' in ct:
                        return resp.content, ct
            except Exception:
                continue

    # 2.5. Sportswear-folder match (longest-prefix lookup) — see
    #      find_sportswear_image_match docstring. Catches Geoffrey Beene
    #      SPORTSWEAR/ photos and any other brand with a SPORTSWEAR/ subfolder.
    sw_match = find_sportswear_image_match(base_style, brand_abbr)
    if sw_match:
        # Try CloudFront DROPBOX_SYNC first (edge-cached) for the matched filename
        if sw_match in _dropbox_photo_index:
            sw_sync_base = f"{CLOUDFRONT_DROPBOX_SYNC_URL}/{sw_match}"
            for ext in ['.jpg', '.png']:
                try:
                    url = sw_sync_base + ext
                    resp = http_requests.get(url, headers=headers, timeout=10)
                    if resp.status_code == 200:
                        ct = resp.headers.get('Content-Type', '').lower()
                        if 'image' in ct:
                            return resp.content, ct
                except Exception:
                    continue
        # Fallback to direct Dropbox bytes
        sw_bytes, sw_ct = get_dropbox_image_bytes(sw_match)
        if sw_bytes:
            return sw_bytes, sw_ct

    # 3. Try Dropbox photos (disk cache → API download)
    if not _no_serial_fallback:
        dbx_bytes, dbx_ct = get_dropbox_image_bytes(image_code)
        if dbx_bytes:
            return dbx_bytes, dbx_ct

    # 4. Fallback to S3 brand folder
    if not _no_serial_fallback:
        folder_name = FOLDER_MAPPING.get(brand_abbr, brand_abbr)
        brand_base = f"{CLOUDFRONT_PHOTOS_URL}/{folder_name}/{image_code}"
        for ext in ['.jpg', '.png', '.jpeg']:
            try:
                url = brand_base + ext
                resp = http_requests.get(url, headers=headers, timeout=10)
                if resp.status_code == 200:
                    ct = resp.headers.get('Content-Type', '').lower()
                    if 'image' in ct:
                        return resp.content, ct
            except Exception:
                continue

    # 5. VERY LAST resort: fallback swatch from the brand style-number files
    #    (bottom priority — mirrors the frontend's swatch-fallback rung)
    if not _no_serial_fallback:
        try:
            url = f"{CLOUDFRONT_SWATCH_FALLBACK_URL}/{image_code}.jpg"
            resp = http_requests.get(url, headers=headers, timeout=10)
            if resp.status_code == 200 and 'image' in resp.headers.get('Content-Type', '').lower():
                return resp.content, resp.headers.get('Content-Type')
        except Exception:
            pass

    return None, None


@app.route('/image/<base_style>', methods=['GET', 'OPTIONS'])
def proxy_image(base_style):
    """Serve a product image by base style code, with server-side S3 caching."""
    if request.method == 'OPTIONS':
        return '', 204

    base_style = base_style.upper().split('.')[0]  # strip extension if present

    # Check web image cache first
    with _web_img_lock:
        cached = _web_img_cache.get(base_style, 'MISS')
    if cached is None:
        return '', 404  # Previously failed — skip
    if cached != 'MISS':
        resp = make_response(cached[0])
        resp.headers['Content-Type'] = cached[1]
        resp.headers['Cache-Control'] = 'public, max-age=86400'
        resp.headers['Access-Control-Allow-Origin'] = '*'
        return resp

    # Get brand from query param or look up from inventory
    brand_abbr = request.args.get('brand', '').upper()
    if not brand_abbr:
        with _inv_lock:
            for item in (_inventory.get('items') or []):
                if item.get('sku', '').split('-')[0].upper() == base_style:
                    brand_abbr = item.get('brand_abbr', item.get('brand', ''))
                    break

    # Fetch raw image from S3
    raw_bytes, content_type = _fetch_raw_image(base_style, brand_abbr)

    if raw_bytes:
        # Cache for future requests (limit cache to ~500 images to control memory)
        with _web_img_lock:
            if len(_web_img_cache) > 200:
                # Evict oldest ~50 entries
                keys = list(_web_img_cache.keys())[:50]
                for k in keys:
                    del _web_img_cache[k]
            _web_img_cache[base_style] = (raw_bytes, content_type)

        resp = make_response(raw_bytes)
        resp.headers['Content-Type'] = content_type
        resp.headers['Cache-Control'] = 'public, max-age=86400'
        resp.headers['Access-Control-Allow-Origin'] = '*'
        return resp

    # Cache the miss too (avoid re-fetching failures)
    with _web_img_lock:
        _web_img_cache[base_style] = None

    return '', 404


@app.route('/dropbox-photos', methods=['GET', 'OPTIONS'])
def dropbox_photo_list():
    """Return list of available Dropbox image codes for frontend."""
    if request.method == 'OPTIONS':
        return '', 204
    return jsonify({
        'codes': list(_dropbox_photo_index.keys()),
        'count': len(_dropbox_photo_index),
        'last_sync': _dropbox_photos_last_sync,
        's3_sync_url': S3_DROPBOX_SYNC_URL
    })


@app.route('/sportswear-photos', methods=['GET', 'OPTIONS'])
def sportswear_photo_list():
    """List sportswear-folder image filenames currently indexed.
    Useful for verifying the SPORTSWEAR/ subfolder sync after deploy."""
    if request.method == 'OPTIONS':
        return '', 204
    with _sportswear_match_lock:
        keys = sorted(_sportswear_photo_index.keys())
    return jsonify({
        'count': len(keys),
        'filenames': keys,
        'last_sync': _dropbox_photos_last_sync
    })


@app.route('/sportswear-match/<base_style>', methods=['GET', 'OPTIONS'])
def sportswear_match_test(base_style):
    """Return which sportswear image filename a given SKU base style maps to.
    Diagnostic only — used to verify the longest-prefix matcher is doing
    the right thing before/after a Dropbox sync."""
    if request.method == 'OPTIONS':
        return '', 204
    brand_abbr = request.args.get('brand', '').upper() or None
    base_style_up = base_style.upper().split('-')[0]
    matched = find_sportswear_image_match(base_style_up, brand_abbr)
    return jsonify({
        'base_style': base_style_up,
        'stripped_after_customer_code': base_style_up[2:] if len(base_style_up) > 2 else base_style_up,
        'matched_filename': matched,
        'in_sportswear_index': matched is not None,
        'in_dropbox_index': matched in _dropbox_photo_index if matched else False
    })


@app.route('/dropbox-photos/sync', methods=['POST', 'OPTIONS'])
def trigger_dropbox_photo_sync():
    """Manually trigger Dropbox photo sync + pre-warm."""
    if request.method == 'OPTIONS':
        return '', 204
    def _sync_and_warm():
        sync_dropbox_photos()
        prewarm_dropbox_cache()
    threading.Thread(target=_sync_and_warm, daemon=True).start()
    return jsonify({'status': 'sync_started', 'current_count': len(_dropbox_photo_index)})


# ============================================
# ADMIN — Force-refresh sync endpoints
# ============================================
# Each /admin/refresh/* endpoint clears the TTL marker for the corresponding data
# source and triggers an immediate re-pull. Used by the Admin Tools panel.

# ─────────────────────────────────────────────────────────────────────────────
# DROPBOX → S3 COLOR MAP SYNC
#
# The platform reads color descriptions from ONE file:
# s3://<bucket>/"Inventory Colors Data/style_color_map.xlsx" (frontend
# loadColorMapFromS3, refetched hourly). The team maintains colors in
# per-brand Dropbox workbooks — this sync pulls each configured (file, tab)
# pair and merges the rows into the S3 master so it stays the single
# centralized source of truth.
#
# SAFETY MODEL (same spirit as the swatch extractor):
#   • ADD + UPDATE only — a key missing from a Dropbox tab is NEVER deleted
#     from the master (manual/swatch-committed rows survive).
#   • A source that fails to download/parse is skipped and reported; it can
#     never wipe or shrink the master.
#   • The current master is backed up in S3 before every write.
#   • No write happens when nothing changed.
#
# Each source maps ONE tab of ONE Dropbox workbook:
#   {
#     'label':      'Nautica — NT tab',      # for logs + refresh report
#     'path':       '/Versa Share Files/.../Nautica Colors.xlsx',
#     'sheet':      'NT',                    # tab name; ''/absent = first tab
#     'key_col':    'Style',                 # header text (case-insensitive)
#                                            # or 0-based column index
#     'color_col':  'Color Description',     # same
#     'key_prefix': 'NT_',                   # bare-digit cells ('201') become
#                                            # 'NT_201' (frontend XX_NNN keys);
#                                            # full style #s pass through as-is;
#                                            # ''/absent = skip bare digits
#     'header_row': 1,                       # optional, 1-based, default 1
#   }
# Populated with David brand by brand — an empty list makes the sync a no-op.
# ONLY files David has explicitly mapped go in this list.
COLOR_SYNC_SOURCES = [
    {
        'label': 'Ben Sherman',
        'path': '/Versa Share Files/New Style Numbers/BEN SHERMAN NEW STYLES.xlsx',
        'sheet': 'MASTER 1',          # row 1 is a title banner; headers live on row 2
        'key_col': 'STYLE NUMBER',    # BE_NNN keys (XX_NNN form, matches frontend lookup)
        'color_col': 'DESCRIPTION',
        'header_row': 2,
        'approved': True,             # David approved the full merge Jul 29 2026
    },
    {
        'label': 'Chaps',
        'path': '/Versa Share Files/New Style Numbers/CHAPS NEW STYLE NUMBERS.xlsx',
        'sheet': 'DRESS SHIRTS',      # ONLY this tab — the TIES tab REUSES the same
                                      # CH_NNN numbers for different products, and
                                      # David excluded ties/hanky/shirt-and-tie sets
        'key_col': 'STYLE NUMBER',
        'color_col': 'DESCRIPTION',
        # No key_prefix: this sheet's bare-digit rows (420, 430, ...) are section
        # banners, not styles — leave them skipped.
        'header_row': 2,
        # Section-banner rows whose DESCRIPTION cell holds a heading, not a color
        # ("NEW WALMART CHAPS STYLES", "COSTCO USA SWATCHES"). Remove a key from
        # this list once the file carries a real color for it.
        'skip_keys': ['CH_258', 'CH_279', 'CH_328', 'CH_361'],
        'approved': True,            # pending David's review of the dry run
    },
    {
        'label': 'DKNY',
        'path': '/Versa Share Files/New Style Numbers/DKNY STYLES NUMBERS.xlsx',
        'sheet': 'MASTER',            # single-tab file, same banner+header layout
        'key_col': 'STYLE NUMBER',
        'color_col': 'DESCRIPTION',
        'header_row': 2,
        'approved': True,            # pending David's review of the dry run
    },
    {
        'label': 'Eddie Bauer',
        'path': '/Versa Share Files/New Style Numbers/Eddie Bauer New Style #.xlsx',
        'sheet': 'MASTER',
        'key_col': 'STYLE #',         # this file's header differs from the others
        'color_col': 'DESCRIPTION',   # actual header has a trailing space (stripped)
        'header_row': 2,
        # CAUTION for the final review: unlike Ben/Chaps/DKNY, this file's 32
        # overwrites are ENTIRELY DIFFERENT colors vs the master (not respelled
        # text) — one side is misnumbered/stale. David decides which wins.
        'approved': True,
    },
    # Geoffrey Beene reuses the SAME GB_NNN numbers per category — each tab
    # gets its own key namespace (see _apply_key_namespace). The bare GB_NNN
    # rows in the master are the dress-shirt set.
    {
        'label': 'Geoffrey Beene — Dress Shirts',
        'path': '/Versa Share Files/New Style Numbers/GEOFFREY BEENE NEW STYLE NUMBERS 3.24.2023.xlsx',
        'sheet': 'DRESS SHIRTS',
        'key_col': 'STYLE #',
        'color_col': 'DESCRIPTION',
        'header_row': 2,
        'approved': True,
    },
    {
        'label': 'Geoffrey Beene — Sportswear',
        'path': '/Versa Share Files/New Style Numbers/GEOFFREY BEENE NEW STYLE NUMBERS 3.24.2023.xlsx',
        'sheet': 'SPORTSWEAR',
        'key_col': 'STYLE #',
        'color_col': 'DESCRIPTION',
        'header_row': 2,
        'key_namespace': 'sportswear',   # GB_036 → GB_SW_036
        'approved': True,
    },
    {
        'label': 'Geoffrey Beene — Blazers',
        'path': '/Versa Share Files/New Style Numbers/GEOFFREY BEENE NEW STYLE NUMBERS 3.24.2023.xlsx',
        'sheet': 'BLAZERS',
        'key_col': 'STYLE #',
        'color_col': 'DESCRIPTION',
        'header_row': 2,
        'key_namespace': 'blazer',       # GB_001 → GB_B01 (B## serial range)
        'approved': True,
    },
    # Von Dutch: SHIRTS and SHACKET tabs REUSE the same VD_NNN numbers for
    # different products, and the master's bare VD rows are currently a MIX
    # (77 shirt + 32 shacket values). Shackets (CO/SL/SF young-men fabrics)
    # already resolve VD_SW_NNN-first, so: SHACKET → sportswear namespace,
    # SHIRTS → takes over the bare namespace (fixes live shirt items that
    # today display shacket colors). PANTS tab is natively VD_P## — literal
    # P-serial SKUs (TMVDBAP07SCP) resolve it via the frontend's extended
    # pants color-fallback; plain-serial BA pants keep their base-SKU rows.
    {
        'label': 'Von Dutch — Shirts',
        'path': '/Versa Share Files/New Style Numbers/VON DUTCH STYLES NUMBERS.xlsx',
        'sheet': 'SHIRTS',
        'key_col': 'STYLE NUMBER',
        'color_col': 'DESCRIPTION',
        'header_row': 2,
        'skip_keys': ['SHIR'],        # stray junk key in the tab
        'approved': True,
    },
    {
        'label': 'Von Dutch — Shackets',
        'path': '/Versa Share Files/New Style Numbers/VON DUTCH STYLES NUMBERS.xlsx',
        'sheet': 'SHACKET',
        'key_col': 'STYLE NUMBER',
        'color_col': 'DESCRIPTION',
        'header_row': 2,
        'key_namespace': 'sportswear',   # VD_NNN → VD_SW_NNN
        'approved': True,
    },
    {
        'label': 'Von Dutch — Pants',
        'path': '/Versa Share Files/New Style Numbers/VON DUTCH STYLES NUMBERS.xlsx',
        'sheet': 'PINSTRIPE PANTS',
        'key_col': 'STYLE NUMBER',
        'color_col': 'DESCRIPTION',
        'header_row': 2,
        'key_namespace': 'pants',        # tab already uses VD_P## natively
        'approved': True,
    },
    {
        'label': 'Vince Camuto',
        'path': '/Versa Share Files/New Style Numbers/VINCE CAMUTO NEW STYLE NUMBERS.xlsx',
        'sheet': 'MASTER',
        'key_col': 'STYLE NUMBER',
        'color_col': 'DESCRIPTION',
        'header_row': 2,
        'approved': True,            # pending David's review of the dry run
    },
    {
        'label': 'U.S. Polo',
        'path': '/Versa Share Files/New Style Numbers/USPA NEW STYLES 4.18.2023.xlsx',
        'sheet': 'MASTER LIST.',      # period in the tab name is real; OLD MASTER tab ignored
        'key_col': 'Style Number',    # column C
        'color_col': 3,               # column D: authored full 'SHIRT / HORSE LOGO' text
                                      # (cols E/F hold the split components — D is richer)
        'header_row': 2,              # row 1 is blank in this file
        'approved': True,            # pending David's review of the dry run
    },
    # Shaq: three tabs with naturally distinct key namespaces (SH_NNN shirts /
    # TSH_NN ties / JT_NNN shirt-tie sets) — no collisions. Shirts + ties split
    # their description across TWO cells (col A + headerless col B) joined with
    # a space; the master's 147 truncated "WHT GRND W/" Shaq rows came from a
    # past sync that only read col A. Set SKUs (JTSHST002...) number by the
    # CONTAINED SHIRT's serial, not the JT_ tab (verified mismatch) — so JT_
    # rows are reference data, deliberately not auto-wired.
    {
        'label': 'Shaq — Shirts',
        'path': '/Versa Share Files/New Style Numbers/SHAQ NEW STYLE NUMBERS.xlsx',
        'sheet': 'SHIRTS ',           # trailing space in the tab name is real
        'key_col': 'STYLE #',         # column C
        'color_col': 'DESCRIPTION',   # column A ("WHT GRND W/")
        'color_col2': 1,              # headerless column B ("BLUE MICRO CHECK")
        'color_join': ' ',            # one description split across two cells
        'header_row': 2,
        'approved': True,
    },
    {
        'label': 'Shaq — Ties',
        'path': '/Versa Share Files/New Style Numbers/SHAQ NEW STYLE NUMBERS.xlsx',
        'sheet': 'TIES',
        'key_col': 'STYLE #',         # column C, native TSH_NN keys
        'color_col': 'DESCRIPTION',
        'color_col2': 1,
        'color_join': ' ',
        'header_row': 2,
        'approved': True,
    },
    {
        'label': 'Shaq — Shirt-Tie Sets',
        'path': '/Versa Share Files/New Style Numbers/SHAQ NEW STYLE NUMBERS.xlsx',
        'sheet': 'SHIRT-TIE SETS',
        'key_col': 'STYLE #',         # column A, native JT_NNN keys
        'color_col': 'DESCRIPTION',   # column B, full text incl. the tie
        'header_row': 1,              # no banner on this tab
        'approved': True,
    },
    {
        'label': 'Robert Graham',
        'path': '/Versa Share Files/New Style Numbers/ROBERT GRAHAM NEW STYLE NUMBERS.xlsx',
        'sheet': 'MASTER',
        'key_col': 'STYLE #',
        'color_col': 'DESCRIPTION',
        'header_row': 2,
        'approved': True,            # dry run showed zero diffs — master already in sync
    },
    # Reebok REMOVED entirely per David (Jul 30 2026): the file's 21 overwrites
    # contradicted live-correct master data (actively-stocked BURBAW001-005
    # styles display the master's print colors; the file claimed solids).
    {
        'label': 'Nicole Miller',
        'path': '/Versa Share Files/New Style Numbers/NICOLE MILLER STYLE NUMBERS (version 1).xlsb.xlsx',
        'sheet': 'Sheet1',            # Sheet2 is a one-row scratch tab — ignored
        'key_col': 'STYLE #',         # column C in this file
        'color_col': 'GROUND COLOR',  # base shirt color
        'color_col2': 'PRINT',        # contrast cuff/collar design
        'color_join': '||',           # NM dual-color convention: the frontend
                                      # splits GROUND||PRINT into ground + print
        'header_row': 1,              # headers on row 1, no banner
        'approved': True,            # pending David's review of the dry run
    },
    {
        'label': 'Neiman Marcus',
        'path': '/Versa Share Files/New Style Numbers/NEIMAN MARCUS - NEW STYLES.xlsx',
        'sheet': 'Sheet1',
        'key_col': 'STYLE NUMBER',
        'color_col': 'DESCRIPTION',
        'header_row': 2,
        'approved': True,            # pending David's review of the dry run
    },
    # Nautica: MASTER - NA = classic NA_NNN namespace; MASTER - NT = the NT
    # overflow namespace (serials past 999, NT 201 ≠ NA 201); BLAZER tab is
    # natively keyed NA_B## (same convention as Geoffrey Beene). OLD MASTER /
    # SUPERSHIRT / TRAVELER / STRETCH tabs ignored per David.
    {
        'label': 'Nautica — NA',
        'path': '/Versa Share Files/New Style Numbers/NAUTICA NEW STYLES 4.10.xlsx',
        'sheet': 'MASTER - NA',
        'key_col': 'Style Number',
        'color_col': 'Label Description',
        'header_row': 1,              # this file has no banner row
        'key_prefix_aliases': {'NAU_': 'NA_'},   # a few rows typed the long code
        'approved': True,
    },
    {
        'label': 'Nautica — NT',
        'path': '/Versa Share Files/New Style Numbers/NAUTICA NEW STYLES 4.10.xlsx',
        'sheet': 'MASTER - NT',
        'key_col': 'Style Number',
        'color_col': 'Label Description',
        'header_row': 1,
        'approved': True,
    },
    {
        'label': 'Nautica — Blazers',
        'path': '/Versa Share Files/New Style Numbers/NAUTICA NEW STYLES 4.10.xlsx',
        'sheet': 'BLAZER',
        'key_col': 'Style Number',
        'color_col': 'Label Description',
        'header_row': 1,
        'key_namespace': 'blazer',    # tab already uses NA_B## natively
        'approved': True,
    },
    {
        'label': 'Lucky Brand',
        'path': '/Versa Share Files/New Style Numbers/LUCKY BRAND NEW STYLES.xlsx',
        'sheet': 'MASTER',
        'key_col': 'STYLE NUMBER',
        'color_col': 'DESCRIPTION',
        'header_row': 2,
        'approved': True,            # pending David's review of the dry run
    },
    {
        'label': 'Karl Lagerfeld Paris',
        'path': '/Versa Share Files/New Style Numbers/KLP NEW STYLE NUMBERS.xlsx',
        'sheet': 'MASTER',
        'key_col': 'STYLE NUMBER',
        'color_col': 'DESCRIPTION',
        'header_row': 2,
        'approved': True,            # pending David's review of the dry run
    },
    {
        'label': 'Jones New York',
        'path': '/Versa Share Files/New Style Numbers/JONES NEW YORK NEW STYLES 3.24.2023.xlsx',
        'sheet': 'MASTER NEW',        # David: this tab ONLY (MASTER + CVC DOBBY ignored)
        'key_col': 'STYLE #',
        'color_col': 'DESCRIPTION',
        'key_prefix': 'JN_',          # a few rows list the bare serial ('129')
        'key_prefix_aliases': {'JNY_': 'JN_'},   # some rows typed the long brand code
        'header_row': 2,
        'approved': True,            # pending David's review of the dry run
    },
    # NOTE: the GB workbook's own PANTS tab was dropped as a source — the
    # dedicated ALL PANTS file below is the pants authority per David, and the
    # two contradict on GB_P03/GB_P04 (swapped; ALL PANTS carries Pantone
    # specs). David to fix/confirm the GB workbook's PANTS tab.
    {
        'label': 'All Pants — Nicole Miller',
        'path': '/Versa Share Files/New Style Numbers/ALL PANTS - NEW STYLES.xlsx',
        'sheet': 'NICOLE MILLER - PANTS',   # real tab name has a trailing space (handled)
        'key_col': 'Style Number',
        'color_col': 'DESCRIPTION',
        'header_row': 2,
        'key_namespace': 'pants',           # native NM_P## keys
        'approved': True,
    },
    {
        'label': 'All Pants — U.S. Polo',
        'path': '/Versa Share Files/New Style Numbers/ALL PANTS - NEW STYLES.xlsx',
        'sheet': 'USPA - PANTS',
        'key_col': 'Style Number',
        'color_col': 'Label Description',   # full 'PANTS / HORSE' text (col C is pants-only)
        'header_row': 2,
        'key_namespace': 'pants',           # native US_P## keys
        'approved': True,
    },
    {
        'label': 'All Pants — Geoffrey Beene',
        'path': '/Versa Share Files/New Style Numbers/ALL PANTS - NEW STYLES.xlsx',
        'sheet': 'GEOFFREY BEENE - PANTS',
        'key_col': 'Style Number',
        'color_col': 'DESCRIPTION',
        'header_row': 2,
        'key_namespace': 'pants',           # native GB_P## keys, incl. Pantone specs
        'approved': True,
    },
]

# Hourly automation gate. When True, the hourly loop merges APPROVED sources
# only ('approved': True) — sources still under review with David never run
# unattended; they are exercised via the manual endpoint
# (POST /admin/refresh/colors?dry_run=1&source=<label>) until approved.
COLOR_SYNC_AUTORUN = True

_color_sync_lock = threading.Lock()
_color_sync_status = {'last_run': None, 'last_result': None}
S3_COLOR_MAP_BACKUP_PREFIX = 'Inventory Colors Data/backups/'


def _dropbox_download_path(path, timeout=60):
    """Generic Dropbox file download with one 401 token-refresh retry.
    Returns bytes or None."""
    global _dropbox_token_expires
    token = get_dropbox_token()
    if not token:
        return None
    headers = {
        'Authorization': f'Bearer {token}',
        'Dropbox-API-Arg': json.dumps({'path': path})
    }
    resp = http_requests.post('https://content.dropboxapi.com/2/files/download',
                              headers=headers, timeout=timeout)
    if resp.status_code == 401:
        _dropbox_token_expires = 0
        token = get_dropbox_token()
        if not token:
            return None
        headers['Authorization'] = f'Bearer {token}'
        resp = http_requests.post('https://content.dropboxapi.com/2/files/download',
                                  headers=headers, timeout=timeout)
    if resp.status_code != 200:
        print(f"  [ColorSync] ⚠ HTTP {resp.status_code} downloading {path}: {resp.text[:150]}")
        return None
    return resp.content


def _color_sync_normalize_key(raw, prefix):
    """Map a source cell to a color-map key. Full style numbers pass through
    uppercased; bare digits become '<PREFIX><zero-padded-3>' to match the
    frontend's XX_NNN fallback keys; XX_NNN-shaped keys get stray whitespace
    collapsed ('BE _260' → 'BE_260') and digits zero-padded to 3, because the
    frontend only ever looks up the padded form. Returns '' for unusable cells."""
    k = str(raw or '').strip().upper()
    if not k or len(k) > 40 or k.endswith('_'):
        return ''    # trailing-underscore keys ('NM_') have no serial — garbage
    if k.isdigit():
        return f"{prefix}{k.zfill(3)}" if prefix else ''
    m = re.match(r'^([A-Z0-9]{2})[\s_\-]+(\d{1,3})$', k)
    if m:
        # Repairs separator typos: 'BE _260' / 'NM-372' / 'US 274' → XX_NNN form
        return f"{m.group(1)}_{int(m.group(2)):03d}"
    if re.search(r'\s', k):
        # Whitespace remaining after the XX_NNN repair means the cell is not a
        # single key (e.g. two style numbers pasted into one cell) — skip it;
        # the parser reports these as malformed for manual fixing.
        return ''
    return k


def _apply_key_namespace(key, namespace):
    """Rewrite an XX_NNN key into a category namespace, mirroring how the
    frontend categorizes SKUs. Brands like Geoffrey Beene reuse the same
    style numbers per category (dress-shirt GB_036 ≠ sportswear GB_036 ≠
    blazer GB_036), so each tab's keys must land in a distinct namespace:
      sportswear: GB_036  → GB_SW_036   (frontend: sportswear fabric/collar)
      blazer:     GB_001  → GB_B01      (frontend: B## serial, B01–B99)
      pants:      GB_P01  passes through; GB_001 → GB_P01 (P## serial, US/GB)
    Returns '' when the key can't be namespaced (reported as malformed)."""
    if not namespace:
        return key
    m = re.match(r'^([A-Z0-9]{2})_(\d{3})$', key)
    if namespace == 'sportswear':
        if m:
            return f"{m.group(1)}_SW_{m.group(2)}"
        return key if re.match(r'^[A-Z0-9]{2}_SW_\d{3}$', key) else ''
    if namespace == 'blazer':
        if m and int(m.group(2)) <= 99:
            return f"{m.group(1)}_B{int(m.group(2)):02d}"
        return key if re.match(r'^[A-Z0-9]{2}_B\d\d$', key) else ''
    if namespace == 'pants':
        if m and int(m.group(2)) <= 99:
            return f"{m.group(1)}_P{int(m.group(2)):02d}"
        return key if re.match(r'^[A-Z0-9]{2}_P\d\d$', key) else ''
    return ''


def _parse_color_source(src, data):
    """Parse one source config against its workbook bytes.
    Returns (rows_dict key→color, error_string_or_None)."""
    try:
        wb = openpyxl.load_workbook(BytesIO(data), data_only=True)
    except Exception as e:
        return {}, f"workbook load failed: {e}"
    try:
        sheet_name = str(src.get('sheet') or '').strip()
        if sheet_name:
            # Tolerate leading/trailing whitespace in the workbook's tab names
            # (the Shaq file's shirts tab is literally named 'SHIRTS ').
            actual = next((n for n in wb.sheetnames if n.strip() == sheet_name), None)
            if actual is None:
                return {}, f"tab '{sheet_name}' not found (tabs: {', '.join(wb.sheetnames[:10])})"
            ws = wb[actual]
        else:
            ws = wb[wb.sheetnames[0]]

        header_row = int(src.get('header_row') or 1)

        def resolve_col(spec):
            if isinstance(spec, int):
                return spec
            want = str(spec or '').strip().lower()
            if not want:
                return None
            for idx, cell in enumerate(ws[header_row]):
                if str(cell.value or '').strip().lower() == want:
                    return idx
            return None

        key_col = resolve_col(src.get('key_col'))
        color_col = resolve_col(src.get('color_col'))
        # Optional second color component (Nicole Miller: GROUND + PRINT columns
        # joined with '||' — the platform's dual-color convention, split by the
        # frontend into ground/print display parts).
        color_col2 = resolve_col(src.get('color_col2')) if src.get('color_col2') is not None else None
        if key_col is None or color_col is None or (src.get('color_col2') is not None and color_col2 is None):
            headers = [str(c.value or '').strip() for c in ws[header_row]]
            return {}, (f"columns not found (key_col={src.get('key_col')!r}, "
                        f"color_col={src.get('color_col')!r}, "
                        f"color_col2={src.get('color_col2')!r}; headers: {headers[:12]})")
        color_join = str(src.get('color_join') or '||')

        prefix = str(src.get('key_prefix') or '')
        # Optional prefix aliases: some files mix key spellings for one brand
        # (Jones New York lists both JN_408 and JNY_408). Map alias prefixes to
        # the canonical form BEFORE normalization so every row lands on the key
        # the frontend actually looks up.
        aliases = {str(a).strip().upper(): str(b).strip().upper()
                   for a, b in (src.get('key_prefix_aliases') or {}).items()}
        # skip_keys are written as they appear IN THE FILE — run each through
        # the same normalize+namespace pipeline as data keys so they still
        # match on namespaced sources (a raw 'GB_042' entry must catch the
        # post-transform 'GB_SW_042').
        skip = set()
        for sk in (src.get('skip_keys') or []):
            sk_raw = str(sk).strip().upper()
            skip.add(sk_raw)
            sk_norm = _color_sync_normalize_key(sk_raw, prefix)
            if sk_norm:
                skip.add(sk_norm)
                sk_ns = _apply_key_namespace(sk_norm, src.get('key_namespace'))
                if sk_ns:
                    skip.add(sk_ns)
        rows = {}
        dup_keys = set()
        blank_color = 0
        malformed = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if len(rows) > 20000:
                return {}, "more than 20,000 rows — refusing (wrong tab/columns?)"
            key_raw = row[key_col] if key_col < len(row) else None
            color_raw = row[color_col] if color_col < len(row) else None
            key_in = str(key_raw or '')
            if aliases:
                ku = key_in.strip().upper()
                for a, b in aliases.items():
                    if ku.startswith(a):
                        key_in = b + ku[len(a):]
                        break
            key = _color_sync_normalize_key(key_in, prefix)
            if key:
                if src.get('key_namespace'):
                    key = _apply_key_namespace(key, src.get('key_namespace'))
                elif re.match(r'^[A-Z0-9]{2}_(SW_\d{3}|[BP]\d\d)$', key):
                    # A namespaced-form key on a PLAIN tab is a stray row from a
                    # category tab — writing it would poison the frontend's
                    # first-priority lookup slot. Reject and report instead.
                    key = ''
            # Collapse ALL internal whitespace (double spaces, Excel in-cell
            # line breaks) — otherwise cosmetic cell formatting shows up as a
            # phantom overwrite of an identical color. Excel error tokens from
            # broken formulas (#VALUE! in Chaps CH_022) are NOT colors — treat
            # as blank so the master keeps whatever it has.
            def _clean_color(v):
                s = ' '.join(str(v or '').split())
                return '' if re.fullmatch(r'#(VALUE!|REF!|N/A|NAME\??|DIV/0!?|NULL!?)', s.upper()) else s
            color = _clean_color(color_raw)
            if color_col2 is not None:
                c2 = _clean_color(row[color_col2] if color_col2 < len(row) else None)
                if color and c2:
                    color = f"{color}{color_join}{c2}"
                elif c2 and not color:
                    color = ''   # print with no ground — unusable, counts as blank
            if not key:
                raw_txt = ' '.join(str(key_raw or '').split())
                if raw_txt and color:
                    malformed.append(raw_txt[:40])   # unusable key next to a color
                continue
            if not color:
                blank_color += 1     # style number reserved, no color entered yet
                continue
            if key in skip or len(color) > 200:
                continue
            if key in rows and rows[key] != color:
                dup_keys.add(key)    # same key listed twice with DIFFERENT colors
            rows[key] = color        # last one wins — flagged in the report
        src['_stats'] = {'dup_conflicting_keys': sorted(dup_keys)[:20],
                         'blank_color_rows': blank_color,
                         'malformed_keys': malformed[:15]}
        return rows, None
    finally:
        wb.close()


def _merge_color_rows(ws, desired, apply=True):
    """Merge desired {key: color} into the master sheet. Updates existing rows
    in place (ALL duplicate occurrences of a key — the frontend reader is
    last-duplicate-wins) and returns (updated, adds): updated is a list of
    (key, old_color, new_color) triples, adds the (key, color) pairs not yet
    present. apply=False computes both WITHOUT touching the sheet (dry run).
    Returns (None, None) when the master schema is unrecognizable — callers
    must abort, never write."""
    headers = [str(c.value or '').strip().lower() for c in ws[1]]

    def find_col(*cands):
        for cand in cands:
            if cand.lower() in headers:
                return headers.index(cand.lower())
        return None

    key_col = find_col('Key', 'Style_Number', 'Style_Num')
    color_col = find_col('Color_Description')
    if key_col is None or color_col is None:
        return None, None

    existing = {}
    for r in range(2, ws.max_row + 1):
        k = str(ws.cell(row=r, column=key_col + 1).value or '').strip().upper()
        if k:
            existing.setdefault(k, []).append(r)

    updated = []
    adds = []
    for key, color in desired.items():
        if key in existing:
            changed = False
            for r in existing[key]:
                cur = str(ws.cell(row=r, column=color_col + 1).value or '').strip()
                if cur != color:
                    if apply:
                        ws.cell(row=r, column=color_col + 1).value = color
                    if not changed:
                        updated.append((key, cur, color))
                        changed = True
        else:
            adds.append((key, color))
    return updated, adds


def sync_color_map_from_dropbox(dry_run=False, only_label=None, approved_only=False):
    """Pull configured Dropbox color sources and merge into the S3 master
    color map. Add/update only — never deletes. dry_run computes and reports
    the full merge without writing anything; only_label restricts the run to
    one source; approved_only (the hourly path) skips sources still under
    review. Returns a result dict."""
    result = {'status': 'ok', 'sources': [], 'adds': 0, 'updates': 0,
              'wrote': False, 'dry_run': bool(dry_run)}
    if not COLOR_SYNC_SOURCES:
        result['status'] = 'no_sources'
        _color_sync_status.update(last_run=time.time(), last_result=result)
        return result

    if not _color_sync_lock.acquire(blocking=False):
        return {'status': 'busy'}
    try:
        # 1. Download each distinct file once, parse each configured tab
        downloads = {}
        desired = {}
        ok_sources = 0
        considered = 0
        for src in COLOR_SYNC_SOURCES:
            label = src.get('label') or src.get('path')
            if only_label and label != only_label:
                continue
            if approved_only and not src.get('approved'):
                continue
            considered += 1
            path = src.get('path')
            if path not in downloads:
                downloads[path] = _dropbox_download_path(path)
            data = downloads[path]
            if not data:
                result['sources'].append({'label': label, 'status': 'download_failed'})
                continue
            rows, err = _parse_color_source(src, data)
            stats = src.pop('_stats', {})
            if err:
                result['sources'].append({'label': label, 'status': f'parse_error: {err}'})
                continue
            entry = {'label': label, 'status': 'ok', 'rows': len(rows)}
            entry.update(stats)
            # Cross-source collisions (two sources claiming the same key with
            # different colors) are almost always a stray row — surface them
            # instead of silently letting the later source win.
            clashes = [k for k in rows if k in desired and desired[k] != rows[k]]
            if clashes:
                entry['cross_source_conflicts'] = [
                    {'key': k, 'discarded': desired[k], 'kept': rows[k]}
                    for k in sorted(clashes)[:10]
                ]
            desired.update(rows)
            ok_sources += 1
            result['sources'].append(entry)

        if considered == 0:
            result['status'] = 'no_matching_sources'
            _color_sync_status.update(last_run=time.time(), last_result=result)
            return result
        if ok_sources == 0:
            result['status'] = 'all_sources_failed'
            _color_sync_status.update(last_run=time.time(), last_result=result)
            return result

        # 2. Load the S3 master and compute the merge (reusing the swatch
        #    extractor's schema-aware helpers so both writers stay consistent)
        from swatch_extractor import (_download_color_map, _resolve_color_map_sheet,
                                      _append_rows, _upload_color_map, S3_COLOR_MAP_KEY)
        wb, _etag, _size = _download_color_map(get_s3, S3_BUCKET)
        ws = _resolve_color_map_sheet(wb)
        updated, adds = _merge_color_rows(ws, desired, apply=not dry_run)
        if updated is None:
            result['status'] = 'master_schema_unrecognized'
            _color_sync_status.update(last_run=time.time(), last_result=result)
            return result

        result['updates'] = len(updated)
        result['adds'] = len(adds)
        result['sample_adds'] = adds[:15]
        result['sample_updates'] = [
            {'key': k, 'old': old, 'new': new} for k, old, new in updated[:15]
        ]
        if not updated and not adds:
            print(f"  [ColorSync] ✓ {len(desired)} keys from {ok_sources} source(s) — master already current")
            _color_sync_status.update(last_run=time.time(), last_result=result)
            return result

        if dry_run:
            # Dry runs are the review surface — include the FULL change lists
            # (bounded), not just the 15-row samples.
            result['all_adds'] = adds[:500]
            result['all_updates'] = [
                {'key': k, 'old': old, 'new': new} for k, old, new in updated[:500]
            ]
            print(f"  [ColorSync] (dry run) {len(adds)} adds, {len(updated)} updates pending")
            _color_sync_status.update(last_run=time.time(), last_result=result)
            return result

        # 3. Backup the master in S3, then write
        try:
            ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
            get_s3().copy_object(
                Bucket=S3_BUCKET,
                CopySource={'Bucket': S3_BUCKET, 'Key': S3_COLOR_MAP_KEY},
                Key=f"{S3_COLOR_MAP_BACKUP_PREFIX}style_color_map_{ts}.xlsx",
            )
        except Exception as e:
            print(f"  [ColorSync] ⚠ Backup copy failed (continuing): {e}")

        if adds:
            _append_rows(ws, adds)
        _upload_color_map(get_s3, S3_BUCKET, wb)
        result['wrote'] = True
        print(f"  [ColorSync] ✓ Merged {ok_sources} source(s): {len(adds)} added, {len(updated)} updated key(s)")
        _color_sync_status.update(last_run=time.time(), last_result=result)
        return result
    except Exception as e:
        print(f"  [ColorSync] ⚠ Sync failed: {e}")
        result['status'] = f'error: {e}'
        _color_sync_status.update(last_run=time.time(), last_result=result)
        return result
    finally:
        _color_sync_lock.release()


# ── Claude file delivery (standing, David-requested Jul 30 2026) ─────────────
# Chat attachments don't reach David when he works remotely, so files Claude
# produces for him are delivered as public S3 URLs under the 'claude uploaded/'
# folder. The key prefix is forced server-side — this route can only write
# into that folder, never anywhere else in the bucket.
@app.route('/admin/claude-uploads', methods=['POST', 'OPTIONS'])
def admin_claude_uploads():
    if request.method == 'OPTIONS':
        return '', 204
    import base64
    body = request.get_json(silent=True) or {}
    files = body.get('files') or []
    if not files:
        return jsonify({'error': 'No files provided'}), 400
    s3 = get_s3()
    written, errors = [], []
    for f in files:
        try:
            name = str(f.get('name') or '').strip().lstrip('/')
            if not name or '..' in name:
                raise ValueError('bad name')
            key = f"claude uploaded/{name}"   # forced folder — never anywhere else
            data = base64.b64decode(f.get('b64') or '')
            s3.put_object(Bucket=S3_BUCKET, Key=key, Body=data,
                          ContentType=str(f.get('content_type') or 'application/octet-stream'))
            written.append({'key': key, 'bytes': len(data)})
        except Exception as e:
            errors.append({'name': f.get('name'), 'error': str(e)[:120]})
    return jsonify({'written': written, 'errors': errors})


@app.route('/admin/refresh/colors', methods=['POST', 'OPTIONS'])
def admin_refresh_colors():
    """Force a Dropbox → S3 color-map sync and report what changed.
    ?dry_run=1 (or JSON {"dry_run": true}) reports the merge without writing;
    ?source=<label> (or JSON {"source": ...}) restricts the run to one source.

    WRITE GATE: a real (writing) run only touches APPROVED sources — a bare
    POST (or a forgotten dry_run flag) can never merge sources still under
    review. Real-running an unapproved source requires BOTH naming it via
    ?source= AND passing ?confirm=1."""
    if request.method == 'OPTIONS':
        return '', 204
    body = request.get_json(silent=True) or {}
    dry = request.args.get('dry_run') in ('1', 'true', 'yes') or bool(body.get('dry_run'))
    only = request.args.get('source') or body.get('source') or None
    confirm = request.args.get('confirm') in ('1', 'true', 'yes') or bool(body.get('confirm'))
    approved_only = False
    if not dry:
        if only:
            src = next((s for s in COLOR_SYNC_SOURCES
                        if (s.get('label') or s.get('path')) == only), None)
            if src is not None and not src.get('approved') and not confirm:
                return jsonify({
                    'status': 'confirm_required',
                    'message': (f"Source '{only}' is not approved yet. Re-run with "
                                f"?confirm=1 to write it anyway, or ?dry_run=1 to preview."),
                }), 409
        else:
            approved_only = True
    return jsonify(sync_color_map_from_dropbox(dry_run=dry, only_label=only,
                                               approved_only=approved_only))


@app.route('/admin/refresh/production', methods=['POST', 'OPTIONS'])
def admin_refresh_production():
    """Force a fresh pull of the Style Ledger from Dropbox."""
    if request.method == 'OPTIONS':
        return '', 204
    global _production_last_sync
    _production_last_sync = 0
    try:
        load_production_from_dropbox()
        return jsonify({
            'status': 'ok',
            'last_sync': _production_last_sync,
            'row_count': len(_production_data)
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/admin/refresh/apo', methods=['POST', 'OPTIONS'])
def admin_refresh_apo():
    """Force a fresh pull of the APO allocations from Dropbox."""
    if request.method == 'OPTIONS':
        return '', 204
    global _apo_last_sync
    _apo_last_sync = 0
    try:
        load_apo_from_dropbox()
        return jsonify({
            'status': 'ok',
            'last_sync': _apo_last_sync,
            'row_count': len(_apo_data)
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/admin/refresh/nj', methods=['POST', 'OPTIONS'])
def admin_refresh_nj():
    """Force a fresh pull of the NJ warehouse on-hand CSV from Dropbox.
    (Automatic refresh is once daily at 10am ET — this is the manual lever.)"""
    if request.method == 'OPTIONS':
        return '', 204
    try:
        ok = load_nj_from_dropbox()
        if ok:
            _nj_reapply_to_inventory()
        with _nj_lock:
            return jsonify({
                'status': 'ok' if ok else 'kept-previous',
                'last_sync': _nj_last_sync,
                'base_styles': len(_nj_by_base),
                'shippable_units': sum(_nj_by_base.values()),
                'fba_earmarked_units': _nj_fba_units,
                'unmatched_rows': len(_nj_unmatched),
            })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/admin/nj-unmatched', methods=['GET', 'OPTIONS'])
def admin_nj_unmatched():
    """NJ rows that could not be counted as shippable platform stock:
    legacy non-Versa IDs, FBA-earmarked pieces, negatives, unknown brand codes."""
    if request.method == 'OPTIONS':
        return '', 204
    with _nj_lock:
        return jsonify({
            'last_sync': _nj_last_sync,
            'shippable_base_styles': len(_nj_by_base),
            'shippable_units': sum(_nj_by_base.values()),
            'fba_earmarked_units': _nj_fba_units,
            'unmatched': sorted(_nj_unmatched, key=lambda r: -abs(r.get('qty') or 0)),
        })


@app.route('/admin/refresh/inventory', methods=['POST', 'OPTIONS'])
def admin_refresh_inventory():
    """Force a fresh pull of inventory (ATS) from S3."""
    if request.method == 'OPTIONS':
        return '', 204
    try:
        sync_inventory()
        with _inv_lock:
            last_sync = _inventory['last_sync']
            count = _inventory['item_count']
        return jsonify({'status': 'ok', 'last_sync': last_sync, 'item_count': count})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/admin/sync-status', methods=['GET', 'OPTIONS'])
def admin_sync_status():
    """Single endpoint returning the last-sync timestamp for every data source.
    Used by the Admin Tools panel to show one unified view of system freshness."""
    if request.method == 'OPTIONS':
        return '', 204
    import time as _time
    now = _time.time()

    def _epoch_to_iso(epoch_seconds):
        if not epoch_seconds:
            return None
        return datetime.utcfromtimestamp(epoch_seconds).isoformat() + 'Z'

    def _iso_age(iso_str):
        if not iso_str:
            return None
        try:
            dt = datetime.fromisoformat(iso_str.replace('Z', '+00:00'))
            return int((datetime.now(dt.tzinfo) - dt).total_seconds())
        except Exception:
            return None

    with _inv_lock:
        inv_sync = _inventory.get('last_sync')
        inv_count = _inventory.get('item_count', 0)

    return jsonify({
        'inventory': {
            'last_sync_iso': inv_sync,
            'age_seconds': _iso_age(inv_sync),
            'row_count': inv_count,
            'auto_interval_minutes': 60,
        },
        'production': {
            'last_sync_iso': _epoch_to_iso(_production_last_sync) if _production_last_sync else None,
            'age_seconds': int(now - _production_last_sync) if _production_last_sync else None,
            'row_count': len(_production_data),
            'auto_interval_minutes': PRODUCTION_RESYNC_INTERVAL // 60,
        },
        'apo': {
            'last_sync_iso': _epoch_to_iso(_apo_last_sync) if _apo_last_sync else None,
            'age_seconds': int(now - _apo_last_sync) if _apo_last_sync else None,
            'row_count': len(_apo_data),
            'auto_interval_minutes': 60,
        },
        'nj_warehouse': {
            'last_sync_iso': _epoch_to_iso(_nj_last_sync) if _nj_last_sync else None,
            'age_seconds': int(now - _nj_last_sync) if _nj_last_sync else None,
            'row_count': len(_nj_by_base),
            'shippable_units': sum(_nj_by_base.values()),
            'auto_interval_minutes': 1440,
        },
        'dropbox_photos': {
            'last_sync_iso': _epoch_to_iso(_dropbox_photos_last_sync) if _dropbox_photos_last_sync else None,
            'age_seconds': int(now - _dropbox_photos_last_sync) if _dropbox_photos_last_sync else None,
            'row_count': len(_dropbox_photo_index),
            'auto_interval_minutes': int(DROPBOX_RESYNC_INTERVAL / 60),
        }
    })




# ============================================
# SAVED CATALOGS — S3 persistence
# ============================================

def _load_saved_catalogs():
    """Load saved catalogs from S3."""
    try:
        s3 = get_s3()
        resp = s3.get_object(Bucket=S3_BUCKET, Key=S3_SAVED_CATALOGS_KEY)
        return json.loads(resp['Body'].read().decode('utf-8'))
    except Exception as e:
        err_str = str(e)
        if 'NoSuchKey' in err_str or '404' in err_str or 'does not exist' in err_str.lower():
            return []
        print(f"[Saved Catalogs] Load error: {e}", flush=True)
        return []

def _save_saved_catalogs(catalogs):
    """Save catalogs list to S3 with timestamped backup."""
    try:
        s3 = get_s3()
        body = json.dumps(catalogs, indent=2)
        s3.put_object(
            Bucket=S3_BUCKET,
            Key=S3_SAVED_CATALOGS_KEY,
            Body=body,
            ContentType='application/json'
        )
        try:
            ts = datetime.utcnow().strftime('%Y-%m-%d_%H-%M-%S')
            s3.put_object(
                Bucket=S3_BUCKET,
                Key=f"inventory/catalogs_backups/saved_catalogs_{ts}.json",
                Body=body,
                ContentType='application/json'
            )
        except Exception as be:
            print(f"[Saved Catalogs] Backup failed (non-fatal): {be}", flush=True)
        try:
            # Invalidate the authz gate's short-lived catalog cache so an edited
            # catalog's slug resolves to the new filters immediately.
            with _catalog_cache_lock:
                _catalog_cache['list'] = None
                _catalog_cache['ts'] = 0.0
        except Exception:
            pass
        return True
    except Exception as e:
        print(f"[Saved Catalogs] Save to S3 error: {e}", flush=True)
        return False

def _cors_json(data, status=200):
    """Return JSON response with explicit CORS headers."""
    resp = make_response(jsonify(data), status)
    resp.headers['Access-Control-Allow-Origin'] = '*'
    resp.headers['Access-Control-Allow-Methods'] = 'GET, POST, DELETE, OPTIONS'
    resp.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Api-Key'
    return resp

def _invalidate_cloudfront(paths):
    """Invalidate specific CloudFront paths so updates are instant.
    paths: list of strings like ['/ALL+INVENTORY+Photos/STYLE+OVERRIDES/NA5001.jpg']
    First 1000 invalidation paths/month are free."""
    try:
        distribution_id = os.environ.get('CLOUDFRONT_DISTRIBUTION_ID', '')
        if not distribution_id:
            return  # Silently skip if not configured
        cf = boto3.client('cloudfront',
            aws_access_key_id=os.environ.get('AWS_ACCESS_KEY_ID', ''),
            aws_secret_access_key=os.environ.get('AWS_SECRET_ACCESS_KEY', ''),
            region_name='us-east-1'  # CloudFront is always us-east-1
        )
        cf.create_invalidation(
            DistributionId=distribution_id,
            InvalidationBatch={
                'Paths': {'Quantity': len(paths), 'Items': paths},
                'CallerReference': str(int(time.time()))
            }
        )
        print(f"[CloudFront] Invalidated {len(paths)} path(s): {paths}", flush=True)
    except Exception as e:
        print(f"[CloudFront] Invalidation failed (non-fatal): {e}", flush=True)

@app.route('/saved-catalogs', methods=['GET', 'POST', 'OPTIONS'])
def handle_saved_catalogs():
    if request.method == 'OPTIONS':
        resp = make_response('', 204)
        resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Access-Control-Allow-Methods'] = 'GET, POST, DELETE, OPTIONS'
        resp.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Api-Key'
        return resp

    try:
        if request.method == 'GET':
            return _cors_json(_load_saved_catalogs())
        
        # POST
        data = request.get_json(force=True, silent=True)
        if not data or not data.get('name') or not data.get('url'):
            return _cors_json({'error': 'name and url required'}, 400)

        catalogs = _load_saved_catalogs()

        existing_idx = next((i for i, c in enumerate(catalogs) if c['name'].lower() == data['name'].lower()), -1)

        # Preserve existing slug if updating, otherwise use the one sent from frontend
        existing_slug = catalogs[existing_idx].get('slug', '') if existing_idx >= 0 else ''
        slug = existing_slug or data.get('slug', '')
        
        # Preserve existing folder if not explicitly provided
        existing_folder = catalogs[existing_idx].get('folder', '') if existing_idx >= 0 else ''
        folder = data.get('folder') if 'folder' in data else existing_folder

        entry = {
            'name': data['name'],
            'url': data['url'],
            'brands': data.get('brands', []),
            'slug': slug,
            'folder': folder,
            'savedAt': int(time.time() * 1000)
        }

        if existing_idx >= 0:
            catalogs[existing_idx] = entry
            action = 'updated'
        else:
            catalogs.insert(0, entry)
            action = 'created'

        if _save_saved_catalogs(catalogs):
            return _cors_json({'status': action, 'catalogs': catalogs})
        else:
            return _cors_json({'error': 'Failed to save to S3'}, 500)
    except Exception as e:
        print(f"[Saved Catalogs] Endpoint error: {e}", flush=True)
        import traceback; traceback.print_exc()
        return _cors_json({'error': str(e)}, 500)

@app.route('/saved-catalogs/reorder', methods=['POST', 'OPTIONS'])
def reorder_saved_catalogs():
    if request.method == 'OPTIONS':
        resp = make_response('', 204)
        resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Access-Control-Allow-Methods'] = 'POST, OPTIONS'
        resp.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Api-Key'
        return resp
    try:
        data = request.get_json()
        order = data.get('order', [])  # list of original indices in new order
        catalogs = _load_saved_catalogs()
        if len(order) != len(catalogs):
            return _cors_json({'error': 'order length mismatch'}, 400)
        reordered = [catalogs[i] for i in order if 0 <= i < len(catalogs)]
        if len(reordered) != len(catalogs):
            return _cors_json({'error': 'invalid indices'}, 400)
        _save_saved_catalogs(reordered)
        return _cors_json({'status': 'reordered', 'catalogs': reordered})
    except Exception as e:
        print(f"[Saved Catalogs] Reorder error: {e}", flush=True)
        return _cors_json({'error': str(e)}, 500)


@app.route('/saved-catalogs/<int:idx>', methods=['DELETE', 'OPTIONS'])
def delete_saved_catalog(idx):
    if request.method == 'OPTIONS':
        resp = make_response('', 204)
        resp.headers['Access-Control-Allow-Origin'] = '*'
        resp.headers['Access-Control-Allow-Methods'] = 'GET, POST, DELETE, OPTIONS'
        resp.headers['Access-Control-Allow-Headers'] = 'Content-Type, Authorization, X-Api-Key'
        return resp
    try:
        catalogs = _load_saved_catalogs()
        if idx < 0 or idx >= len(catalogs):
            return _cors_json({'error': 'invalid index'}, 404)
        removed = catalogs.pop(idx)
        _save_saved_catalogs(catalogs)
        return _cors_json({'status': 'deleted', 'name': removed['name'], 'catalogs': catalogs})
    except Exception as e:
        print(f"[Saved Catalogs] Delete error: {e}", flush=True)
        return _cors_json({'error': str(e)}, 500)


DROPBOX_RESYNC_INTERVAL = int(float(os.environ.get('DROPBOX_RESYNC_HOURS', 1)) * 3600)  # Default: 1 hour; supports fractions (e.g. 0.25 = 15 min, 0.5 = 30 min)

_worker_initialized = False

@app.before_request
def ensure_worker_initialized():
    """Trigger startup sync on the first request to this worker process.
    This is the only reliable way to initialize with gunicorn's fork model."""
    global _worker_initialized
    if not _worker_initialized:
        _worker_initialized = True
        print("\n  [before_request] First request — triggering startup sync...")
        threading.Thread(target=startup_sync, daemon=True).start()


# ═════════════════════════════════════════════════════════════════════════════
# ACCESS CONTROL (authz gate)
#
# Identity tiers:
#   staff    — Versa-Docs Supabase session (profiles.role = 'staff'): full access
#   factory  — Versa-Docs Supabase session (role = 'factory'): /production and
#              /api/ai-proxy only (Versa-Docs ledger + note translation);
#              /factory-view and /factory-report keep their own inline auth
#   machine  — X-Api-Key header == INVENTORY_API_KEY env: read-only data +
#              report endpoints (open-orders-api emails, FBA tools, AM ship tool)
#   oo       — open-orders team bearer token (validated against the open-orders
#              API /api/auth/me): read-only data + POST /suppression-overrides
#   anonymous— catalog links only: GET reads carrying a valid catalog_slug get
#              a response SCOPED to that saved catalog's brands/styles
#
# AUTH_MODE env: 'off' (default) = never blocks, only logs would-deny lines and
# applies catalog scoping when an anonymous request presents a valid slug.
# 'on' = enforce. A single request can preview enforcement with
# ?authz_preview=1 regardless of mode (used to validate before flipping).
#
# Always open (no identity needed): /health, /, /image/*, /dropbox-photos,
# /sportswear-*, /mcp* (own path token), /factory-view + /factory-report*
# (own inline auth), /catalog/match. /image stays public forever: image URLs
# are hard-embedded in already-sent customer presentations and <img> tags
# cannot carry headers.
# ═════════════════════════════════════════════════════════════════════════════

AUTH_MODE = (os.environ.get('AUTH_MODE', 'off') or 'off').strip().lower()   # 'off' | 'on'
INVENTORY_API_KEY = (os.environ.get('INVENTORY_API_KEY', '') or '').strip()

print(f"[authz] AUTH_MODE={AUTH_MODE} | machine key {'set' if INVENTORY_API_KEY else 'NOT SET'} | "
      f"MCP_TOKEN {'set' if os.environ.get('MCP_TOKEN') else 'NOT SET — /mcp is open to anyone!'}",
      flush=True)

_AUTHZ_OPEN_EXACT = {'/', '/health', '/favicon.ico',
                     # image-index reads used by catalog pages and image fallback
                     # chains; exact paths only so /dropbox-photos/sync (a write
                     # trigger) is NOT covered and stays staff-only
                     '/dropbox-photos', '/sportswear-photos'}
_AUTHZ_OPEN_PREFIXES = ('/image/', '/mcp', '/factory-view', '/factory-report',
                        '/sportswear-match/', '/catalog/')
# GET endpoints a customer catalog page needs. Anonymous access requires a valid
# catalog_slug and the response is scoped; staff/machine/oo tiers get them full.
_AUTHZ_CATALOG_READS = {
    '/inventory', '/sync', '/overrides', '/overrides/version',
    '/allocations', '/allocations/version', '/manual-allocations',
    '/deduction-assignments', '/banner-rules', '/prepack-defaults',
    '/suppression-overrides', '/production', '/apo', '/saved-catalogs',
}
# Extra GET endpoints the machine key may use (report/export pulls).
_AUTHZ_MACHINE_EXTRA = {'/export-apo-brandcolor', '/apo-dollar-summary', '/exports'}
# POST endpoints a customer catalog page legitimately uses (Excel/PDF exports
# of the already-scoped data the page holds). Anonymous access still requires
# a valid catalog_slug, like the reads.
_AUTHZ_CATALOG_POSTS = {'/export', '/export-multi', '/export-pdf'}

_AUTHZ_ALL_BRAND_FLAGS = ('all', 'all_brands', 'wh_all_brands', 'os_all_brands')

# ── Saved-catalog cache (60s) so slug lookups don't hit S3 per request ──────
_catalog_cache = {'ts': 0.0, 'list': None}
_catalog_cache_lock = threading.Lock()

def _load_saved_catalogs_cached(max_age=60):
    now = time.time()
    with _catalog_cache_lock:
        if _catalog_cache['list'] is not None and now - _catalog_cache['ts'] < max_age:
            return _catalog_cache['list']
    cats = _load_saved_catalogs()
    with _catalog_cache_lock:
        _catalog_cache['ts'] = time.time()
        _catalog_cache['list'] = cats
    return cats

# ── Open-orders team token introspection (for the open-orders frontends) ────
_oo_token_cache = {}          # token -> (expires_at, bool)
_oo_token_lock = threading.Lock()

def _oo_token_valid(token):
    now = time.time()
    with _oo_token_lock:
        hit = _oo_token_cache.get(token)
        if hit and hit[0] > now:
            return hit[1]
    ok = False
    transient = False
    try:
        r = http_requests.get(f"{OPEN_ORDERS_API_URL}/api/auth/me",
                              headers={'Authorization': f'Bearer {token}'}, timeout=8)
        ok = (r.status_code == 200)
    except Exception as e:
        # Network/timeout, not a real rejection — cache only briefly so a
        # cold-starting open-orders service doesn't 401 a valid session for
        # two minutes (observed during browser testing).
        transient = True
        print(f"[authz] open-orders introspection failed: {e}", flush=True)
    with _oo_token_lock:
        ttl = 600 if ok else (20 if transient else 120)
        _oo_token_cache[token] = (now + ttl, ok)
        if len(_oo_token_cache) > 500:
            for k in [k for k, v in _oo_token_cache.items() if v[0] <= now]:
                _oo_token_cache.pop(k, None)
    return ok

def _request_identity():
    """Resolve the current request's credential to a tier dict, or None."""
    if hasattr(g, '_authz_ident'):
        return g._authz_ident
    ident = None
    try:
        key = (request.headers.get('X-Api-Key') or '').strip()
        if key and INVENTORY_API_KEY and hmac.compare_digest(key, INVENTORY_API_KEY):
            ident = {'tier': 'machine'}
        else:
            auth = request.headers.get('Authorization', '')
            token = auth[7:].strip() if auth.lower().startswith('bearer ') else ''
            if token.startswith('eyJ') and token.count('.') == 2:
                # Supabase JWT (Versa-Docs accounts) — shape-checked before any
                # network call so garbage tokens can't spam Supabase.
                prof = _caller_identity(token)
                if prof:
                    role = (prof.get('role') or '').lower()
                    ident = {'tier': 'staff' if role == 'staff' else 'factory',
                             'prefix': prof.get('factory_prefix') or ''}
            elif token and token.count('.') == 1:
                # open-orders HMAC token shape (b64payload.signature)
                if _oo_token_valid(token):
                    ident = {'tier': 'oo'}
    except Exception as e:
        print(f"[authz] identity resolution error: {e}", flush=True)
    g._authz_ident = ident
    return ident

# ── Catalog scope: what an anonymous holder of a slug may see ───────────────

def _catalog_scope_for_slug(slug):
    """Build a scope dict for a valid saved-catalog slug, else None."""
    if not slug:
        return None
    entry = next((c for c in _load_saved_catalogs_cached()
                  if str(c.get('slug') or '') == slug), None)
    if entry is None:
        return None
    pub_entry = {k: entry.get(k) for k in ('name', 'url', 'brands', 'slug')}
    scope = {'slug': slug, 'entry': pub_entry, 'all': False,
             'brands': set(), 'brands_norm': set(), 'styles': set()}
    from urllib.parse import urlparse, parse_qsl
    try:
        q = dict(parse_qsl(urlparse(str(entry.get('url') or '')).query,
                           keep_blank_values=True))
    except Exception:
        q = {}
    if any(q.get(f) for f in _AUTHZ_ALL_BRAND_FLAGS):
        scope['all'] = True
        return scope
    for key in ('brands', 'wh_brands', 'os_brands'):
        for b in str(q.get(key) or '').split(','):
            b = b.strip().upper()
            # __SEG__ is the frontend's placeholder token for segment links
            # whose brand list was stripped — it is not a brand.
            if b and b != '__SEG__':
                scope['brands'].add(b)
    for b in (entry.get('brands') or []):
        b = str(b).strip().upper()
        if b and b != '__SEG__':
            scope['brands'].add(b)
    for s in str(q.get('skus') or '').split(','):
        s = s.strip().upper()
        if s:
            scope['styles'].add(get_base_style(s))
    # pos= catalogs (and segment wh_pos/os_pos variants) are PO-scoped: pull
    # the matching production rows' styles in.
    po_filter = set()
    for key in ('pos', 'wh_pos', 'os_pos'):
        po_filter.update(p.strip().upper() for p in str(q.get(key) or '').split(',') if p.strip())
    if po_filter:
        try:
            for row in (_production_data or []):
                if (str(row.get('production') or '').strip().upper() in po_filter or
                        str(row.get('poName') or '').strip().upper() in po_filter):
                    st = str(row.get('style') or '').strip().upper()
                    if st:
                        scope['styles'].add(get_base_style(st))
        except Exception:
            pass
    for b in list(scope['brands']):
        try:
            scope['brands_norm'].add(_normalize_brand(b).upper())
        except Exception:
            pass
    if not scope['brands'] and not scope['styles']:
        # Degenerate entry (no url params, no brands recorded). Serving nothing
        # would break the link; serve unscoped and say so loudly.
        print(f"[authz] WARNING catalog slug {slug} has no derivable scope — serving unscoped",
              flush=True)
        scope['all'] = True
    return scope

_authz_brand_idx = {'etag': None, 'map': {}}

def _authz_base_brand(base):
    """base style -> item brand (upper), from the in-memory inventory.
    Cache key = identity of the items list: the Dropbox sync stamps a constant
    etag string, so keying on etag alone froze this map for the worker's whole
    life (audit fix) — every accepted sync/NJ re-apply swaps the list object."""
    with _inv_lock:
        items = _inventory.get('items') or []
        etag = (id(items), len(items))
    idx = _authz_brand_idx
    if idx['etag'] != etag or not idx['map']:
        m = {}
        for it in items:
            b = str(it.get('brand') or '').upper()
            sku = str(it.get('sku') or '')
            if b and sku:
                m.setdefault(get_base_style(sku), b)
        idx['map'] = m
        idx['etag'] = etag
    return idx['map'].get(base, '')

def _scope_brand_match(scope, b):
    """Brand string vs scope, tolerant of the frontend's alias spellings."""
    if not b:
        return False
    if b in scope['brands']:
        return True
    try:
        return _normalize_brand(b).upper() in scope['brands_norm']
    except Exception:
        return False

def _scope_has_style(scope, sku):
    if scope.get('all'):
        return True
    base = get_base_style(str(sku or ''))
    if not base:
        return False
    if base in scope['styles']:
        return True
    if _scope_brand_match(scope, _authz_base_brand(base)):
        return True
    # Fallbacks for styles absent from inventory: brand letters in the SKU,
    # then the customer-prefix brand groupings the frontend also uses.
    code = base[2:4] if len(base) >= 4 else ''
    ab = str(SKU_BRAND_CODE_MAP.get(code, '') or '').upper()
    if ab and _scope_brand_match(scope, ab):
        return True
    return base[:2] in scope['brands']

def _scope_has_brandstr(scope, raw):
    if scope.get('all'):
        return True
    s = str(raw or '').strip().upper()
    if not s:
        return False
    if s in scope['brands']:
        return True
    try:
        return _normalize_brand(s).upper() in scope['brands_norm']
    except Exception:
        return False

def _scope_has_item(scope, it):
    if scope.get('all'):
        return True
    if _scope_brand_match(scope, str(it.get('brand') or '').upper()):
        return True
    return _scope_has_style(scope, it.get('sku'))

def _strip_nj_rows(rows):
    """NJ warehouse data is ADMIN-ONLY (per David): customer catalog views and
    exports must never see it. Drops NJ-synthesized rows entirely and removes
    NJ quantities (and their share of the totals) from every other row.
    Returns new row dicts — never mutates the caller's list."""
    out = []
    for it in rows:
        if not isinstance(it, dict):
            out.append(it)
            continue
        if it.get('_nj_synth'):
            continue
        nj = int(it.get('nj') or 0)
        if nj or 'nj' in it or 'nj_sizes' in it:
            it = dict(it)
            if nj:
                it['total_warehouse'] = max(0, int(it.get('total_warehouse') or 0) - nj)
                it['total_ats'] = int(it.get('total_ats') or 0) - nj
            it.pop('nj', None)
            it.pop('nj_sizes', None)
        # Client-computed warehouse-names strings ride export payloads.
        wh = it.get('warehouse')
        if isinstance(wh, str) and 'NJ' in wh.upper():
            it = dict(it)
            parts = [p for p in re.split(r'[,/]\s*', wh) if p.strip().upper() != 'NJ']
            it['warehouse'] = ', '.join(p.strip() for p in parts if p.strip())
        out.append(it)
    return out

def _flt_inventory(data, scope):
    items = data.get('inventory')
    if isinstance(items, list):
        kept = _strip_hidden_landing_rows(_strip_nj_rows([it for it in items if _scope_has_item(scope, it)]))
        data['inventory'] = kept
        data['item_count'] = len(kept)
    return data

def _flt_overrides(data, scope):
    ov = data.get('overrides')
    if isinstance(ov, dict):
        data['overrides'] = {k: v for k, v in ov.items() if _scope_has_style(scope, k)}
    return data

def _flt_alloc(data, scope):
    rows = data.get('allocations')
    if isinstance(rows, list):
        data['allocations'] = [r for r in rows if _scope_has_style(scope, r.get('sku'))]
    return data

def _flt_deductions(data, scope):
    d = data.get('assignments')
    if isinstance(d, dict):
        data['assignments'] = {k: v for k, v in d.items() if _scope_has_style(scope, k)}
    return data

def _flt_suppression(data, scope):
    ov = data.get('overrides')
    if isinstance(ov, list):
        data['overrides'] = [s for s in ov if _scope_has_style(scope, s)]
    return data

def _hidden_landing_maps():
    """({style: hidden units}, {style: True if any VISIBLE production}) from the
    in-memory ledger. Ledger styles match inventory SKUs exactly (1:1)."""
    with _production_lock:
        rows = list(_production_data)
    hidden, visible = {}, {}
    for p in rows:
        st = str(p.get('style') or '').upper()
        if not st:
            continue
        if str(p.get('warehouse') or '').upper() in _HIDDEN_LANDING_WH:
            try:
                hidden[st] = hidden.get(st, 0) + int(p.get('units') or 0)
            except Exception:
                pass
        else:
            visible[st] = True
    return hidden, visible

def _strip_hidden_landing_rows(rows):
    """Customer scope: remove the incoming quantity attributable to hidden-landing
    productions. A style whose ONLY supply is a hidden-landing production drops
    entirely (per David: TMNASU201SLS landing NJ must never appear to customers)."""
    hidden, visible = _hidden_landing_maps()
    if not hidden:
        return rows
    out = []
    for it in rows:
        if not isinstance(it, dict):
            out.append(it)
            continue
        sku = str(it.get('sku') or '').upper()
        h = hidden.get(sku, 0)
        if h <= 0:
            out.append(it)
            continue
        it = dict(it)
        inc = int(it.get('incoming') or 0)
        # every production hidden -> ALL overseas supply is invisible; mixed ->
        # subtract the hidden productions' units, clamped
        cut = inc if sku not in visible else min(inc, h)
        if cut > 0:
            it['incoming'] = inc - cut
            it['total_ats'] = int(it.get('total_ats') or 0) - cut
            if int(it.get('total_warehouse') or 0) <= 0 and it['incoming'] <= 0:
                continue   # exists only as hidden-landing production
        out.append(it)
    return out

def _hidden_prod_numbers():
    with _production_lock:
        return {str(p.get('production') or '').strip().upper() for p in _production_data
                if p.get('production') and str(p.get('warehouse') or '').upper() in _HIDDEN_LANDING_WH}

def _strip_hidden_prod_export_rows(rows):
    """Customer exports: drop any posted per-delivery row that references a
    hidden-landing production (stale-cache defense; live pages never form them)."""
    hp = _hidden_prod_numbers()
    if not hp:
        return rows
    out = []
    for it in rows:
        if isinstance(it, dict):
            ref = str(it.get('production') or it.get('po_ref') or '').strip().upper()
            if ref and ref in hp:
                continue
        out.append(it)
    return out

def _flt_production(data, scope):
    rows = data.get('production')
    if isinstance(rows, list):
        data['production'] = [r for r in rows
                              if str(r.get('warehouse') or '').upper() not in _HIDDEN_LANDING_WH
                              and (_scope_has_style(scope, r.get('style'))
                                   or _scope_has_brandstr(scope, r.get('brand')))]
    return data

def _flt_apo(data, scope):
    rows = data.get('apo')
    if isinstance(rows, list):
        kept = [r for r in rows if _scope_has_style(scope, r.get('style'))]
        data['apo'] = kept
        data['count'] = len(kept)
    return data

def _flt_saved_catalogs(data, scope):
    # Anonymous slug holders see exactly one entry: their own catalog.
    return [scope['entry']]

_SCOPE_FILTERS = {
    '/inventory': _flt_inventory,
    '/sync': _flt_inventory,
    '/overrides': _flt_overrides,
    '/allocations': _flt_alloc,
    '/manual-allocations': _flt_alloc,
    '/deduction-assignments': _flt_deductions,
    '/suppression-overrides': _flt_suppression,
    '/production': _flt_production,
    '/apo': _flt_apo,
    '/saved-catalogs': _flt_saved_catalogs,
    # /overrides/version, /allocations/version, /banner-rules, /prepack-defaults
    # pass through unfiltered: version stamps and global display config.
}

# ── Gate ────────────────────────────────────────────────────────────────────
_authz_log_seen = {}

def _authz_log(path, method, reason, enforced):
    key = (path, method, reason, enforced)
    now = time.time()
    if now - _authz_log_seen.get(key, 0) < 60:
        return
    if len(_authz_log_seen) > 500:
        _authz_log_seen.clear()
    _authz_log_seen[key] = now
    verb = 'DENY' if enforced else 'would-deny'
    print(f"[authz] {verb} {method} {path} — {reason}", flush=True)

@app.before_request
def authz_gate():
    if request.method == 'OPTIONS':
        return None
    path = request.path.rstrip('/') or '/'
    if path in _AUTHZ_OPEN_EXACT or path.startswith(_AUTHZ_OPEN_PREFIXES):
        return None
    method = 'GET' if request.method == 'HEAD' else request.method
    ident = _request_identity()
    if ident:
        tier = ident.get('tier')
        if tier == 'staff':
            return None
        if tier == 'machine':
            # No /saved-catalogs for the machine key: the catalog list holds
            # every customer's slug (the anonymous scope keys) and no machine
            # consumer needs it. POST /admin/claude-uploads is the standing
            # Claude file-delivery channel.
            if method == 'GET' and ((path in _AUTHZ_CATALOG_READS and path != '/saved-catalogs')
                                    or path in _AUTHZ_MACHINE_EXTRA
                                    or path.startswith('/download/')):
                return None
            if method == 'POST' and path == '/admin/claude-uploads':
                return None
        if tier == 'oo' and ((method == 'GET' and path in _AUTHZ_CATALOG_READS)
                             or (method == 'POST' and path == '/suppression-overrides')):
            return None
        if tier == 'factory' and path in ('/production', '/api/ai-proxy'):
            return None
        reason = f"tier '{tier}' not allowed here"
    else:
        if ((method == 'GET' and path in _AUTHZ_CATALOG_READS)
                or (method == 'POST' and path in _AUTHZ_CATALOG_POSTS)):
            slug = (request.args.get('catalog_slug') or '').strip()
            scope = _catalog_scope_for_slug(slug) if slug else None
            if scope is not None:
                g._catalog_scope = scope
                return None
            had_auth = bool(request.headers.get('Authorization') or request.headers.get('X-Api-Key'))
            reason = ('anonymous, unknown catalog_slug' if slug
                      else ('credential rejected' if had_auth else 'anonymous, no credential'))
        else:
            had_auth = bool(request.headers.get('Authorization') or request.headers.get('X-Api-Key'))
            reason = 'credential rejected' if had_auth else 'anonymous, no credential'
    enforced = (AUTH_MODE == 'on') or (request.args.get('authz_preview') == '1')
    _authz_log(path, request.method, reason, enforced)
    if enforced:
        return jsonify({'error': 'auth required',
                        'detail': reason,
                        'hint': 'sign in on the platform, or open a valid catalog link'}), 401
    return None

@app.after_request
def apply_catalog_scope(resp):
    scope = getattr(g, '_catalog_scope', None)
    if scope is None or resp.status_code != 200:
        return resp
    path = request.path.rstrip('/') or '/'
    fn = _SCOPE_FILTERS.get(path)
    if fn is None:
        return resp
    if 'application/json' not in (resp.content_type or ''):
        return resp
    try:
        data = json.loads(resp.get_data(as_text=True))
        data = fn(data, scope)
        resp.set_data(json.dumps(data, default=str))
        resp.headers['X-Catalog-Scope'] = str(scope.get('slug') or '')
    except Exception as e:
        # Fail open (the link must keep working) but say so loudly.
        print(f"[authz] scope filter error on {path}: {e}", flush=True)
    return resp

# ── Public slug resolver for legacy full-parameter catalog links ────────────

def _authz_canon_qs(qs):
    """Canonicalize a catalog query string for matching: drop volatile params,
    apply the same *_all_brands cleanup the frontend does, sort."""
    from urllib.parse import parse_qsl, urlencode
    try:
        pairs = parse_qsl(str(qs or '').lstrip('?'), keep_blank_values=True)
    except Exception:
        return ''
    keys = {k for k, _v in pairs}
    drop = {'_', 'catalog', 'catalog_slug', 'authz_preview'}
    if 'wh_all_brands' in keys:
        drop.add('wh_brands')
    if 'os_all_brands' in keys:
        drop.add('os_brands')
    kept = sorted((k, v) for k, v in pairs if k not in drop)
    return urlencode(kept)

@app.route('/catalog/match', methods=['GET', 'OPTIONS'])
def catalog_match():
    """Resolve a full-parameter catalog link to its saved slug (if any).
    Public: answers only 'this exact saved link's slug', never lists catalogs."""
    if request.method == 'OPTIONS':
        return '', 204
    canon = _authz_canon_qs(request.args.get('qs') or '')
    if canon:
        from urllib.parse import urlparse
        for c in _load_saved_catalogs_cached():
            try:
                if (c.get('slug') and
                        _authz_canon_qs(urlparse(str(c.get('url') or '')).query) == canon):
                    return _cors_json({'slug': c['slug']})
            except Exception:
                continue
    return _cors_json({'slug': None})

# ═══════════════════════════ end access control ═════════════════════════════


def hourly_resync():
    """Background loop: re-sync from Dropbox and regenerate exports every hour"""
    while True:
        time.sleep(DROPBOX_RESYNC_INTERVAL)
        print(f"\n  ⏰ Hourly re-sync triggered...")

        # Periodic Dropbox photo sync + S3 upload
        if (DROPBOX_PHOTOS_TOKEN or DROPBOX_REFRESH_TOKEN) and (time.time() - _dropbox_photos_last_sync > DROPBOX_PHOTOS_SYNC_HOURS * 3600):
            print(f"  📷 Dropbox photos sync due (every {DROPBOX_PHOTOS_SYNC_HOURS}h)...")
            try:
                sync_dropbox_photos()
                prewarm_dropbox_cache()  # Downloads new images + uploads to S3
            except Exception as e:
                print(f"  ⚠ Dropbox photos sync failed: {e}")

        with _export_lock:
            if _exports['generating']:
                print("  ⏭ Skipping — export generation already in progress")
                continue

        try:
            updated = sync_inventory()
            if updated:
                with _inv_lock:
                    count = _inventory['item_count']
                    source = _inventory['source']
                print(f"  ✓ Re-synced {count} items from {source}, regenerating exports...")
                trigger_background_generation()
            else:
                print("  No changes detected, skipping export regeneration")
        except Exception as e:
            print(f"  ⚠ Hourly re-sync failed: {e}")

        # Refresh APO allocation file every hour
        try:
            load_apo_from_dropbox()
        except Exception as e:
            print(f"  ⚠ APO hourly refresh failed: {e}")

        # Refresh Style Ledger (production data) every hour
        try:
            # Force reload by resetting timestamp
            global _production_last_sync
            _production_last_sync = 0
            load_production_from_dropbox()
        except Exception as e:
            print(f"  ⚠ Production hourly refresh failed: {e}")

        # Refresh the color map from Dropbox sources every hour
        # (gated off until the mapped sources are approved for automation)
        try:
            if COLOR_SYNC_AUTORUN:
                sync_color_map_from_dropbox(approved_only=True)
        except Exception as e:
            print(f"  ⚠ Color map hourly sync failed: {e}")


# ────────────────────────────────────────────────────────────────────────
# STYLE LEDGER FAST-LANE — pulls production data every 10 min (configurable
# via PRODUCTION_RESYNC_MINUTES env var). This runs INDEPENDENTLY of the
# heavy hourly_resync above so we get fresh ledger data without thrashing
# inventory, APO, or export regeneration on the same cadence.
# ────────────────────────────────────────────────────────────────────────
PRODUCTION_RESYNC_INTERVAL = int(float(os.environ.get('PRODUCTION_RESYNC_MINUTES', 10)) * 60)


def production_resync_loop():
    """Background loop: re-pull the Style Ledger from Dropbox every N minutes."""
    print(f"  ⏰ Style Ledger fast-lane enabled (every {PRODUCTION_RESYNC_INTERVAL//60} min)")
    while True:
        try:
            time.sleep(PRODUCTION_RESYNC_INTERVAL)
            global _production_last_sync
            # Reset timestamp to bypass the TTL cache and force a fresh pull
            _production_last_sync = 0
            load_production_from_dropbox()
        except Exception as e:
            print(f"  ⚠ Style Ledger fast-lane refresh failed: {e}")
            # Brief back-off on persistent errors so we don't tight-loop
            time.sleep(60)


# ────────────────────────────────────────────────────────────────────────
# DAILY FULL REGEN — clears ALL image caches and rebuilds every export
# ────────────────────────────────────────────────────────────────────────
# Set DAILY_REGEN_HOUR_UTC in env to change the trigger time (0–23, UTC).
# Defaults to 04:00 UTC = 11pm Eastern (Standard) / 9pm Pacific (Standard).
# East Coast: 11pm EST → 04 UTC, 11pm EDT → 03 UTC
# West Coast: 11pm PST → 07 UTC, 11pm PDT → 06 UTC
# This is what catches new S3 STYLE+OVERRIDES uploads, new Dropbox photos,
# and new brand-folder additions that were dropped in directly (bypassing
# the platform). Platform overrides still propagate instantly via the POST
# /overrides path — this is the safety net for everything else.
DAILY_REGEN_HOUR_UTC = int(os.environ.get('DAILY_REGEN_HOUR_UTC', '4'))
DAILY_REGEN_MINUTE_UTC = int(os.environ.get('DAILY_REGEN_MINUTE_UTC', '0'))


def _do_full_regen(reason='scheduled'):
    """Clear all image caches, re-sync inventory, and trigger export regeneration.
    Same logic as POST /regenerate but callable internally from the scheduler."""
    with _export_lock:
        if _exports['generating']:
            print(f"  [DailyRegen:{reason}] ⏭ Already generating, skipping cache clear")
            return False

    print(f"  [DailyRegen:{reason}] Clearing image caches...")
    with _img_lock:
        cleared_img = len(_img_cache)
        _img_cache.clear()
    with _web_img_lock:
        cleared_web = len(_web_img_cache)
        _web_img_cache.clear()
    cleared_thumb = len(_dropbox_thumb_cache)
    _dropbox_thumb_cache.clear()
    cleared_dbx = len(_dropbox_img_cache)
    _dropbox_img_cache.clear()
    with _sportswear_match_lock:
        cleared_sw = len(_sportswear_match_cache)
        _sportswear_match_cache.clear()
    print(f"  [DailyRegen:{reason}] Cleared: {cleared_img} img, {cleared_web} web, {cleared_thumb} thumb, {cleared_dbx} dbx, {cleared_sw} sw-match")

    try:
        sync_inventory()
    except Exception as e:
        print(f"  [DailyRegen:{reason}] ⚠ Sync failed (continuing anyway): {e}")

    trigger_background_generation()
    return True


def _seconds_until_next_daily_regen():
    """Compute seconds from now until the next scheduled regen (UTC hour/min)."""
    now = datetime.utcnow()
    target = now.replace(hour=DAILY_REGEN_HOUR_UTC, minute=DAILY_REGEN_MINUTE_UTC,
                         second=0, microsecond=0)
    if target <= now:
        target = target + timedelta(days=1)
    return (target - now).total_seconds(), target


def daily_regen_loop():
    """Background daemon thread: triggers a full cache flush + regen once per day.
    Runs independently in each gunicorn worker so every worker's in-memory
    _exports['brands'] stays fresh — small redundant cost (3x image fetches at
    off-peak hours), but it means downloads stay instant from any worker."""
    print(f"  ⏰ Daily regen scheduler armed (target: {DAILY_REGEN_HOUR_UTC:02d}:{DAILY_REGEN_MINUTE_UTC:02d} UTC)")
    while True:
        try:
            sleep_seconds, target = _seconds_until_next_daily_regen()
            hrs = sleep_seconds / 3600
            print(f"  ⏰ [DailyRegen] Next run at {target.isoformat()}Z (in {hrs:.1f}h)")
            time.sleep(sleep_seconds)

            print(f"\n  🌙 [DailyRegen] {DAILY_REGEN_HOUR_UTC:02d}:{DAILY_REGEN_MINUTE_UTC:02d} UTC trigger fired")
            _do_full_regen(reason='daily')
        except Exception as e:
            print(f"  ⚠ [DailyRegen] Loop error: {e}")
            # Back off an hour and retry rather than tight-looping on persistent errors
            time.sleep(3600)


def startup_sync():
    print("\n" + "="*60)
    print("  VERSA INVENTORY EXPORT API v3 — Startup")
    print(f"  Dropbox URL configured: {'YES' if DROPBOX_URL else 'NO'}")
    _dbx_configured = bool(DROPBOX_PHOTOS_TOKEN or DROPBOX_REFRESH_TOKEN)
    print(f"  Dropbox Photos configured: {'YES (refresh token)' if DROPBOX_REFRESH_TOKEN else ('YES (static token)' if DROPBOX_PHOTOS_TOKEN else 'NO')}")
    print("="*60)

    load_overrides_from_s3()
    load_manual_allocations_from_s3()
    load_deduction_assignments_from_s3()
    load_prepack_defaults_from_s3()
    load_suppression_overrides_from_s3()
    load_banner_rules_from_s3()
    try:
        load_apo_from_dropbox()
    except Exception as e:
        print(f"  ⚠ APO startup load failed: {e}")

    # NJ warehouse daily feed: one hydrating load at boot (same day's 08:45
    # snapshot), then ONLY the 10am daily loop — never hourly, per David.
    try:
        if not load_nj_from_dropbox():
            _nj_hydrate_from_s3()
    except Exception as e:
        print(f"  ⚠ NJ startup load failed: {e}")
        try:
            _nj_hydrate_from_s3()
        except Exception:
            pass

    # Load Style Ledger (production data) from Dropbox
    try:
        load_production_from_dropbox()
    except Exception as e:
        print(f"  ⚠ Production startup load failed: {e}")

    # Sync Dropbox photos index (before inventory so images are ready for export generation)
    if _dbx_configured:
        print("  → Syncing Dropbox photos index...", flush=True)
        sync_dropbox_photos()
        # Start background pre-warm of all images to disk
        if _dropbox_photo_index:
            import threading as _th
            _th.Thread(target=prewarm_dropbox_cache, daemon=True).start()

    # Sync inventory: Dropbox first, then S3 fallback
    try:
        updated = sync_inventory()
        with _inv_lock:
            count = _inventory['item_count']
            source = _inventory['source']
        if count > 0:
            print(f"  ✓ Startup: {count} items loaded from {source}")
            print(f"  → Generating exports (images + Excel)...")
            trigger_background_generation()
        else:
            print("  ⚠ Startup: no inventory data")
    except Exception as e:
        print(f"  Startup sync failed: {e}")

    # Start daily full-regen scheduler (clears all caches + rebuilds at 11pm/configurable)
    threading.Thread(target=daily_regen_loop, daemon=True, name='daily-regen').start()

    # Start Style Ledger fast-lane (10-min Dropbox pull, independent of hourly_resync)
    threading.Thread(target=production_resync_loop, daemon=True, name='production-resync').start()

    # NJ warehouse: once-daily 10:00 AM ET refresh (the 3PL exports ~08:45 ET)
    threading.Thread(target=nj_daily_resync_loop, daemon=True, name='nj-daily').start()

    # Start daily selling-data Dropbox sync + warm caches now (non-blocking)
    threading.Thread(target=daily_selling_sync_loop, daemon=True, name='selling-sync').start()
    threading.Thread(target=lambda: refresh_selling_data('Ross'),
                     daemon=True, name='selling-warmup').start()

    # Start hourly re-sync loop
    if DROPBOX_URL:
        print(f"  ⏰ Hourly Dropbox re-sync enabled (every {DROPBOX_RESYNC_INTERVAL//3600}h)")
        hourly_resync()  # This runs forever in the same thread


# ─────────────────────────────────────────────────────────────────────────────
# CUSTOMER SELLING DATA
# Daily Dropbox sync of /Selling Data/<Customer>/*.xlsx files. Each customer's
# workbook has one sheet per recap date or season (e.g. "9.08", "Fall 2025"),
# with columns: BRAND | TYPE | STYLE NO. | DESCRIPTION | LABEL (best/okay/worst).
# Cached in memory + persisted to S3 so cold starts don't have to wait for Dropbox.
# ─────────────────────────────────────────────────────────────────────────────

_selling_cache = {}            # { customer_lower: {sheets, fetched_at, source_filename} }
_selling_last_synced = {}      # { customer_lower: epoch_seconds }
_SELLING_S3_PREFIX = 'selling_data'  # S3 backup path

def _normalize_label(raw):
    """Normalize label cell values to one of: best | okay | worst | None."""
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if not s or s == 'nan':
        return None
    if 'best' in s or 'good' in s:    return 'best'
    if 'worst' in s or 'bad' in s:    return 'worst'
    if 'okay' in s or 'ok' in s or 'medium' in s: return 'okay'
    return None  # unknown labels dropped silently

# Style numbers in selling recaps: 6 brand/customer letters (+ optional 7th for
# pants P-serials like ROUSPPP02SRS), 2-3 digit serial, 2-3 letter suffix.
_SELLING_SKU_RE = re.compile(r'^[A-Z]{6}[A-Z]?\d{2,3}[A-Z]{2,3}(-[A-Z0-9]+)?$')

def _selling_header_map(header_row):
    """Map a header row's column names → field indices, or None if the row
    doesn't look like a recognizable header (caller falls back to the legacy
    positional layout). Handles both layouts seen in the wild:
      legacy: BRAND | TYPE | STYLE NO. | DESCRIPTION | LABEL
      new   : BRAND | STYLE | DESCRIPTION | COLOR | LABEL   (Burlington 8.5.26)
    """
    if not header_row:
        return None
    cells = [(str(c).strip().lower() if c is not None else '') for c in header_row]
    if not cells or cells[0] not in ('brand', 'brands'):
        return None
    idx = {}
    for i, c in enumerate(cells):
        key = c.replace('.', '').replace('#', '').replace('_', ' ').strip()
        if key in ('brand', 'brands'):
            idx.setdefault('brand', i)
        elif key == 'type':
            idx.setdefault('type', i)
        elif key in ('style no', 'style number', 'style', 'styles', 'sku', 'style  no'):
            idx.setdefault('sku', i)
        elif key in ('description', 'desc'):
            idx.setdefault('desc', i)
        elif key in ('color', 'colour', 'colors', 'color name'):
            idx.setdefault('color', i)
        elif key in ('label', 'labels', 'rating'):
            idx.setdefault('label', i)
    # Without a style column the map is useless — fall back to positional.
    if 'sku' not in idx:
        return None
    return idx

def _parse_selling_workbook(xlsx_bytes):
    """Parse the selling-data xlsx into a JSON-friendly structure.
    Returns: { 'sheets': [ {name, rows: [{brand, type, sku, description, color, label}]} ] }
    Sheet order in the output matches the source workbook (typically chronological).
    Tolerant to:
      - Variable column counts beyond the expected 5
      - Per-sheet column layouts (mapped by header names; positional fallback)
      - Extra notes/columns in some sheets (e.g. 12.2 has commentary on the right)
      - Layout drift: if the mapped style cell doesn't look like a style #, the
        row is rescanned for a cell that does (guards against future reorders
        on sheets whose headers we can't map)
    """
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    out = {'sheets': []}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        col_map = None  # per-sheet header mapping (None → legacy positional)
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx == 0:
                col_map = _selling_header_map(row)
                if col_map is not None:
                    continue
                # Header-ish but unmappable (or blank) row — skip it
                first_cell = (str(row[0]) if row and row[0] is not None else '').strip().lower()
                if first_cell in ('brand', 'brands', ''):
                    continue
            if not row or row[0] is None:
                continue

            def _cell(i):
                if i is None or len(row) <= i or row[i] is None:
                    return ''
                return str(row[i]).strip()

            if col_map:
                brand = _cell(col_map.get('brand'))
                type_ = _cell(col_map.get('type'))
                sku   = _cell(col_map.get('sku')).upper()
                desc  = _cell(col_map.get('desc'))
                color = _cell(col_map.get('color'))
                label = _normalize_label(_cell(col_map.get('label')) or None)
            else:
                brand = _cell(0)
                type_ = _cell(1)
                sku   = _cell(2).upper()
                desc  = _cell(3)
                color = ''
                label = _normalize_label(row[4] if len(row) > 4 else None)

            # Safety net: if what we grabbed doesn't look like a style number,
            # scan the row for a cell that does (catches column reorders).
            if sku and not _SELLING_SKU_RE.match(sku):
                for c in row:
                    cu = str(c).strip().upper() if c is not None else ''
                    if cu and _SELLING_SKU_RE.match(cu):
                        sku = cu
                        break
            if not sku:
                continue
            rows.append({
                'brand': brand,
                'type': type_,
                'sku': sku,
                'description': desc,
                'color': color,
                'label': label,
            })
        if rows:
            out['sheets'].append({'name': sheet_name, 'rows': rows})
    wb.close()
    return out

def _fetch_selling_from_dropbox(customer):
    """Download the latest selling xlsx for `customer` from Dropbox and parse it.
    Returns dict with sheets+metadata, or None on failure. The folder may contain
    multiple xlsx files (e.g. dated versions) — we pick the most recently modified.
    """
    token = get_dropbox_token()
    if not token:
        print(f"[Selling Sync] No Dropbox token, can't fetch {customer}", flush=True)
        return None

    folder_path = f"{DROPBOX_SELLING_BASE}/{customer}"
    headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}

    # List files in the customer's folder
    try:
        list_resp = http_requests.post(
            'https://api.dropboxapi.com/2/files/list_folder',
            headers=headers,
            json={'path': folder_path, 'recursive': False, 'limit': 100},
            timeout=30
        )
        if list_resp.status_code != 200:
            print(f"[Selling Sync] list_folder for '{folder_path}' returned {list_resp.status_code}: "
                  f"{list_resp.text[:200]}", flush=True)
            return None
        entries = list_resp.json().get('entries', [])
        # Only .xlsx files, picking the one with the most recent server_modified date
        xlsx_entries = [e for e in entries
                        if e.get('.tag') == 'file'
                        and e.get('name', '').lower().endswith('.xlsx')
                        and not e.get('name', '').startswith('~$')]   # exclude lock files
        if not xlsx_entries:
            print(f"[Selling Sync] No .xlsx files in {folder_path}", flush=True)
            return None
        xlsx_entries.sort(key=lambda e: e.get('server_modified', ''), reverse=True)
        latest = xlsx_entries[0]
        file_path = latest['path_lower']
        file_name = latest['name']
        print(f"[Selling Sync] Fetching {file_name} for customer={customer}", flush=True)
    except Exception as e:
        print(f"[Selling Sync] list_folder failed: {e}", flush=True)
        return None

    # Download the file
    try:
        dl_resp = http_requests.post(
            'https://content.dropboxapi.com/2/files/download',
            headers={'Authorization': f'Bearer {token}',
                     'Dropbox-API-Arg': json.dumps({'path': file_path})},
            timeout=60
        )
        if dl_resp.status_code != 200:
            print(f"[Selling Sync] download for {file_name} returned {dl_resp.status_code}", flush=True)
            return None
        parsed = _parse_selling_workbook(dl_resp.content)
        parsed['source_filename'] = file_name
        parsed['fetched_at'] = int(time.time())
        # Persist a backup to S3 so cold starts don't hit Dropbox
        try:
            s3 = get_s3()
            if s3:
                s3.put_object(
                    Bucket=S3_BUCKET,
                    Key=f"{_SELLING_S3_PREFIX}/{customer.lower()}.json",
                    Body=json.dumps(parsed).encode('utf-8'),
                    ContentType='application/json'
                )
        except Exception as s3e:
            print(f"[Selling Sync] S3 backup write failed (non-fatal): {s3e}", flush=True)
        total_rows = sum(len(s['rows']) for s in parsed['sheets'])
        print(f"[Selling Sync] ✓ {customer}: {len(parsed['sheets'])} sheets, "
              f"{total_rows} rows from {file_name}", flush=True)
        return parsed
    except Exception as e:
        print(f"[Selling Sync] download/parse failed: {e}", flush=True)
        return None

def _load_selling_from_s3(customer):
    """Fast cold-start path: read last known good data from S3."""
    try:
        s3 = get_s3()
        if not s3:
            return None
        obj = s3.get_object(Bucket=S3_BUCKET, Key=f"{_SELLING_S3_PREFIX}/{customer.lower()}.json")
        return json.loads(obj['Body'].read())
    except Exception:
        return None  # No backup yet — first run

def refresh_selling_data(customer='Ross'):
    """Refresh one customer's selling cache (Dropbox first, S3 backup on failure)."""
    key = customer.lower()
    data = _fetch_selling_from_dropbox(customer)
    if data:
        _selling_cache[key] = data
        _selling_last_synced[key] = int(time.time())
        return True
    # Dropbox failed — keep whatever's currently in memory; if empty, try S3
    if key not in _selling_cache:
        backup = _load_selling_from_s3(customer)
        if backup:
            _selling_cache[key] = backup
            print(f"[Selling Sync] Using S3 backup for {customer} (Dropbox unavailable)", flush=True)
            return True
    return False

def daily_selling_sync_loop():
    """Background daemon: refresh selling-data caches once per day."""
    print(f"  ⏰ Daily selling-data sync armed "
          f"(target: {SELLING_REFRESH_HOUR_UTC:02d}:{SELLING_REFRESH_MIN_UTC:02d} UTC)",
          flush=True)
    customers = ['Ross']  # legacy best/okay/worst format
    weekly_customers = ['Costco']  # new quantitative weekly format
    while True:
        try:
            now = datetime.utcnow()
            target = now.replace(hour=SELLING_REFRESH_HOUR_UTC,
                                 minute=SELLING_REFRESH_MIN_UTC,
                                 second=0, microsecond=0)
            if target <= now:
                target = target + timedelta(days=1)
            sleep_seconds = (target - now).total_seconds()
            time.sleep(sleep_seconds)
            print(f"\n  🛒 [SellingSync] {SELLING_REFRESH_HOUR_UTC:02d}:"
                  f"{SELLING_REFRESH_MIN_UTC:02d} UTC trigger fired", flush=True)
            for c in customers:
                refresh_selling_data(c)
            for c in weekly_customers:
                refresh_weekly_selling_data(c)
        except Exception as e:
            print(f"  ⚠ [SellingSync] Loop error: {e}", flush=True)
            time.sleep(3600)  # back off

@app.route('/selling-data/<customer>', methods=['GET', 'OPTIONS'])
def get_selling_data(customer):
    """Return parsed selling data for a customer. Triggers an on-demand refresh
    if cache is empty (e.g. first request after deploy)."""
    if request.method == 'OPTIONS':
        return ('', 204)
    key = (customer or '').lower()
    if key not in _selling_cache:
        # Try S3 backup synchronously so the first user doesn't see an empty page
        backup = _load_selling_from_s3(customer)
        if backup:
            _selling_cache[key] = backup
        else:
            # Last resort: pull from Dropbox right now
            refresh_selling_data(customer)
    data = _selling_cache.get(key)
    if not data:
        return jsonify({'error': 'No selling data available', 'customer': customer}), 404
    return jsonify({
        'customer': customer,
        'fetched_at': data.get('fetched_at'),
        'source_filename': data.get('source_filename'),
        'sheets': data.get('sheets', []),
    })

@app.route('/selling-data/<customer>/refresh', methods=['POST', 'OPTIONS'])
def refresh_selling_data_endpoint(customer):
    """Manual refresh trigger — useful if you upload new data outside the daily cycle."""
    if request.method == 'OPTIONS':
        return ('', 204)
    ok = refresh_selling_data(customer)
    if ok:
        data = _selling_cache.get(customer.lower(), {})
        return jsonify({
            'ok': True,
            'sheets': len(data.get('sheets', [])),
            'source_filename': data.get('source_filename'),
        })
    return jsonify({'ok': False, 'error': 'Refresh failed (see server logs)'}), 502


# ─────────────────────────────────────────────────────────────────────────────
# WEEKLY SELLING DATA (Costco-style)
# ─────────────────────────────────────────────────────────────────────────────
# Parallel system to the legacy Ross "best/okay/worst" labels. Costco files:
#   • Multiple files per customer folder (one per week)
#   • Filename: Costco_EB_20260523_-_WK_3.xlsx (Customer_Brand_YYYYMMDD_-_WK_N.xlsx)
#   • Each file has parent styles broken into item-rows by size/color
#   • Quantitative metrics: $ sold, units, DPW, sell-through, inventory, etc.
#   • Current week + lifetime totals

import re as _re

_WEEKLY_FILENAME_RE = _re.compile(
    r'^(?P<customer>[^_]+)_(?P<brand>.+?)_(?P<date>\d{8})_-_WK_(?P<week>\d+)\.xlsx$',
    _re.IGNORECASE
)

# Alternate pattern matching David's actual naming convention:
#   "Costco Selling EB 5.23.xlsx"  →  customer=Costco, brand=EB, date=2026-05-23
# Tokens: Customer Selling Brand M.D[.YYYY]
# - Date can be M.D (year inferred) or M.D.YYYY
# - "Selling" literal between customer and brand can also be other words; we
#   ignore the middle and take the last non-date whitespace token as brand.
_WEEKLY_FILENAME_RE_ALT = _re.compile(
    r'^(?P<customer>[A-Za-z][A-Za-z0-9]*)\s+'
    r'(?P<middle>.+?)\s+'
    r'(?P<brand>[A-Za-z][A-Za-z0-9]*)\s+'
    r'(?P<m>\d{1,2})\.(?P<d>\d{1,2})(?:\.(?P<y>\d{2,4}))?'
    r'\.xlsx$',
    _re.IGNORECASE
)

def _parse_weekly_filename(filename):
    """Pull brand / date / week# out of a weekly-selling filename.

    Supports two naming conventions:
      1. Strict:   Customer_Brand_YYYYMMDD_-_WK_N.xlsx
      2. Loose:    Customer Selling Brand M.D[.YYYY].xlsx  (Costco's actual format)

    For the loose format the week number is derived from the calendar week of
    the parsed date. Year is inferred from the current year, or the previous
    year if M.D falls in the future (which means it's actually last year's).
    """
    name = filename or ''

    # Try strict format first
    m = _WEEKLY_FILENAME_RE.match(name)
    if m:
        try:
            date_str = m.group('date')
            snapshot_date = datetime.strptime(date_str, '%Y%m%d').date().isoformat()
        except ValueError:
            return None
        return {
            'customer': m.group('customer'),
            'brand': m.group('brand').upper(),
            'snapshot_date': snapshot_date,
            'week': int(m.group('week')),
        }

    # Fall back to loose format
    m = _WEEKLY_FILENAME_RE_ALT.match(name)
    if m:
        try:
            mo = int(m.group('m'))
            day = int(m.group('d'))
            year_raw = m.group('y')
            if year_raw:
                yr = int(year_raw)
                if yr < 100:
                    yr += 2000
            else:
                # Infer year: if the M.D is in the future relative to today,
                # it's last year's data. Otherwise current year.
                today = datetime.utcnow().date()
                yr = today.year
                try:
                    candidate = datetime(yr, mo, day).date()
                    if candidate > today:
                        yr -= 1
                except ValueError:
                    pass
            dt = datetime(yr, mo, day).date()
            iso_week = dt.isocalendar()[1]
            return {
                'customer': m.group('customer'),
                'brand': m.group('brand').upper(),
                'snapshot_date': dt.isoformat(),
                'week': iso_week,
            }
        except (ValueError, AttributeError):
            return None

    return None


def _parse_weekly_workbook(xlsx_bytes, file_meta):
    """Parse a single weekly selling workbook into a JSON-friendly structure."""
    import io as _io
    import openpyxl as _openpyxl
    wb = _openpyxl.load_workbook(_io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    sheet_name = None
    for n in wb.sheetnames:
        if n.lower() != 'data':
            sheet_name = n
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    def _metric_block(row, start_col):
        def _num(idx):
            v = row[idx] if idx < len(row) else None
            try:
                return float(v) if v is not None else 0.0
            except (TypeError, ValueError):
                return 0.0
        return {
            'dollar_sales':    _num(start_col + 0),
            'unit_sales':      _num(start_col + 1),
            'avg_price':       _num(start_col + 2),
            'dpw':             _num(start_col + 3),
            'refund_dollars':  _num(start_col + 4),
            'inventory':       _num(start_col + 5),
            'on_order':        _num(start_col + 6),
            'in_transit':      _num(start_col + 7),
            'qty_received':    _num(start_col + 8),
            'nsi_units':       _num(start_col + 9),
            'retail_value':    _num(start_col + 10),
        }

    styles = []
    current_style = None
    for row in rows[2:]:
        if not row:
            continue
        col_c = (str(row[2]).strip() if len(row) > 2 and row[2] is not None else '')
        col_a_item = (str(row[0]).strip() if len(row) > 0 and row[0] is not None else '')
        col_d_item = (str(row[3]).strip() if len(row) > 3 and row[3] is not None else '')

        if not col_c and not col_a_item:
            continue

        is_total = col_c.upper().startswith('TOTAL')
        is_header = col_c.upper() in ('STYLE', 'BRAND', 'ITEM')
        is_new_parent = bool(col_c) and not is_total and not is_header

        if is_header:
            continue

        if is_new_parent:
            current_style = {
                'style': col_c.upper(),
                'description': str(row[5]).strip() if len(row) > 5 and row[5] else '',
                'brand_label': str(row[4]).strip() if len(row) > 4 and row[4] else '',
                'items': [],
                'totals_current': None,
                'totals_lifetime': None,
            }
            styles.append(current_style)

        if is_total:
            if current_style is not None:
                current_style['totals_current']  = _metric_block(row, 6)
                current_style['totals_lifetime'] = _metric_block(row, 17)
            continue

        if current_style is None or not (col_a_item or col_d_item):
            continue

        current_style['items'].append({
            'item_code': col_a_item or col_d_item,
            'description': str(row[5]).strip() if len(row) > 5 and row[5] else '',
            'current':  _metric_block(row, 6),
            'lifetime': _metric_block(row, 17),
        })

    # Compute derived metrics per style
    for s in styles:
        tc = s.get('totals_current') or {}
        tl = s.get('totals_lifetime') or {}
        if tc:
            sold_lt = tl.get('unit_sales', 0) or 0
            inv_lt  = tl.get('inventory', 0) or 0
            denom = sold_lt + inv_lt
            s['sell_through_pct'] = (sold_lt / denom * 100) if denom > 0 else 0.0
            wk_units = tc.get('unit_sales', 0) or 0
            s['weeks_of_cover'] = (tc.get('inventory', 0) / wk_units) if wk_units > 0 else None

    return {
        'snapshot_date': file_meta['snapshot_date'],
        'brand': file_meta['brand'],
        'week': file_meta['week'],
        'source_filename': file_meta.get('source_filename', ''),
        'styles': styles,
    }


_weekly_selling_cache = {}
_weekly_selling_last_synced = {}


def _fetch_weekly_selling_from_dropbox(customer):
    """List every .xlsx in the customer's folder, parse each one, group by brand."""
    token = get_dropbox_token()
    if not token:
        print(f"[Weekly Selling] No Dropbox token, can't fetch {customer}", flush=True)
        return None

    folder_path = f"{DROPBOX_SELLING_BASE}/{customer}"
    headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}

    try:
        list_resp = http_requests.post(
            'https://api.dropboxapi.com/2/files/list_folder',
            headers=headers,
            json={'path': folder_path, 'recursive': False, 'limit': 500},
            timeout=30
        )
        if list_resp.status_code != 200:
            print(f"[Weekly Selling] list_folder({folder_path}) → {list_resp.status_code}: "
                  f"{list_resp.text[:200]}", flush=True)
            return None
        entries = list_resp.json().get('entries', [])
    except Exception as e:
        print(f"[Weekly Selling] List failed for {customer}: {e}", flush=True)
        return None

    parseable = []
    for e in entries:
        if e.get('.tag') != 'file':
            continue
        name = e.get('name', '')
        if name.startswith('~$') or not name.lower().endswith('.xlsx'):
            continue
        meta = _parse_weekly_filename(name)
        if meta is None:
            continue
        meta['source_filename'] = name
        parseable.append((meta, e))

    if not parseable:
        print(f"[Weekly Selling] No matching files in {folder_path} "
              "(expected Customer_Brand_YYYYMMDD_-_WK_N.xlsx)", flush=True)
        return {'weeks': [], 'by_brand': {}}

    parseable.sort(key=lambda pe: pe[0]['snapshot_date'])
    weeks = []
    for meta, entry in parseable:
        try:
            dl_resp = http_requests.post(
                'https://content.dropboxapi.com/2/files/download',
                headers={
                    'Authorization': f'Bearer {token}',
                    'Dropbox-API-Arg': json.dumps({'path': entry['path_display']})
                },
                timeout=60
            )
            if dl_resp.status_code != 200:
                print(f"[Weekly Selling] Download {entry['name']} → {dl_resp.status_code}",
                      flush=True)
                continue
            parsed = _parse_weekly_workbook(dl_resp.content, meta)
            weeks.append(parsed)
        except Exception as e:
            print(f"[Weekly Selling] Parse error for {entry.get('name')}: {e}", flush=True)
            continue

    by_brand = {}
    for w in weeks:
        brand = w['brand']
        by_brand.setdefault(brand, []).append(w)

    return {
        'weeks': weeks,
        'by_brand': by_brand,
        'fetched_at': datetime.utcnow().isoformat() + 'Z',
    }


def refresh_weekly_selling_data(customer):
    """Refresh one customer's weekly-selling cache."""
    data = _fetch_weekly_selling_from_dropbox(customer)
    if data is not None:
        key = customer.lower()
        _weekly_selling_cache[key] = data
        _weekly_selling_last_synced[key] = int(time.time())
        return True
    return False


@app.route('/selling-data-weekly/<customer>', methods=['GET', 'OPTIONS'])
def get_weekly_selling_data(customer):
    """Return all weekly selling data for a customer (e.g. Costco)."""
    if request.method == 'OPTIONS':
        return ('', 204)
    key = (customer or '').lower()
    if key not in _weekly_selling_cache:
        refresh_weekly_selling_data(customer)
    data = _weekly_selling_cache.get(key)
    if not data:
        return jsonify({'error': 'No weekly selling data available', 'customer': customer}), 404
    return jsonify({
        'customer': customer,
        'fetched_at': data.get('fetched_at'),
        'last_synced': _weekly_selling_last_synced.get(key),
        'by_brand': data.get('by_brand', {}),
        'brand_count': len(data.get('by_brand', {})),
        'week_count': len(data.get('weeks', [])),
    })


@app.route('/selling-data-weekly/<customer>/refresh', methods=['POST', 'OPTIONS'])
def refresh_weekly_selling_endpoint(customer):
    """Force re-pull of all weekly files for this customer from Dropbox."""
    if request.method == 'OPTIONS':
        return ('', 204)
    ok = refresh_weekly_selling_data(customer)
    if ok:
        key = customer.lower()
        d = _weekly_selling_cache.get(key, {})
        return jsonify({
            'ok': True,
            'week_count': len(d.get('weeks', [])),
            'brand_count': len(d.get('by_brand', {})),
        })
    return jsonify({'ok': False, 'error': 'Refresh failed'}), 502


@app.route('/selling-data-customers', methods=['GET', 'OPTIONS'])
def list_selling_customers():
    """List all customer folders under /Selling Data."""
    if request.method == 'OPTIONS':
        return ('', 204)
    token = get_dropbox_token()
    if not token:
        return jsonify({'customers': []}), 200
    try:
        resp = http_requests.post(
            'https://api.dropboxapi.com/2/files/list_folder',
            headers={'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'},
            json={'path': DROPBOX_SELLING_BASE, 'recursive': False, 'limit': 200},
            timeout=30
        )
        if resp.status_code != 200:
            return jsonify({'customers': [], 'error': f'list_folder {resp.status_code}'}), 200
        entries = resp.json().get('entries', [])
        customers = sorted([e['name'] for e in entries if e.get('.tag') == 'folder'])
        return jsonify({'customers': customers})
    except Exception as e:
        return jsonify({'customers': [], 'error': str(e)}), 200


# ─────────────────────────────────────────────────────────────────────────────
# AI BOOKING SUGGESTIONS  (Phase 2)
# Takes selling data for one customer/season + inventory color descriptions
# and asks Claude Opus to recommend styles to add/drop for the upcoming season.
# Returns structured recommendations the frontend can render as image tiles.
# ─────────────────────────────────────────────────────────────────────────────

ANTHROPIC_API_KEY = (os.environ.get('ANTHROPIC_API_KEY')
                     or os.environ.get('CLAUDE_API_KEY', ''))
# Pricing for cost-estimate display. Pulled from anthropic.com/pricing for Opus.
# These are per-million-token rates and may drift — used only for the UI estimate,
# not for billing. Override via env if Anthropic raises prices.
AI_OPUS_INPUT_PER_MTOK  = float(os.environ.get('AI_OPUS_INPUT_PER_MTOK',  15.0))
AI_OPUS_OUTPUT_PER_MTOK = float(os.environ.get('AI_OPUS_OUTPUT_PER_MTOK', 75.0))
AI_OPUS_MODEL = os.environ.get('AI_OPUS_MODEL', 'claude-opus-4-5')

# ─────────────────────────────────────────────────────────────────────────────
# SKU code maps — module level so they're shared across resolvers.
# Brand abbreviation lives at positions 2-3 of the base SKU (after the 2-char
# customer code). Example: RONASU371SLP → customer=RO, brand=NA → NAUTICA.
# These mirror the frontend's BRAND_IMAGE_PREFIX (reversed).
# ─────────────────────────────────────────────────────────────────────────────
SKU_BRAND_CODE_MAP = {
    'NA': 'NAUTICA', 'DK': 'DKNY', 'EB': 'EB', 'RB': 'REEBOK', 'VC': 'VINCE',
    'BE': 'BEN', 'US': 'USPA', 'CH': 'CHAPS', 'LB': 'LUCKY', 'JN': 'JNY',
    'GB': 'BEENE', 'NM': 'NICOLE', 'SH': 'SHAQ', 'TA': 'TAYION', 'MS': 'STRAHAN',
    'VD': 'VD', 'VR': 'VERSA', 'CK': 'CHEROKEE', 'AC': 'AMERICA', 'BL': 'BLO',
    'D9': 'DN', 'KL': 'KL', 'RG': 'RG', 'NE': 'NE',
    # NT = Nautica overflow: serials hit 999 (Jul 2026), new Nautica styles use
    # NT with the serial namespace restarted — NT 201 ≠ NA 201. extract_image_code
    # keys images/colors to NT_NNN for these SKUs.
    'NT': 'NAUTICA',
}

# Friendly fit-code labels (from positions 9-10 of base SKU, e.g. "SL" → Slim/Long)
SKU_FIT_FRIENDLY = {
    'SL': 'Slim/Long Sleeve', 'RF': 'Regular Fit',
    'TF': 'Trim Fit', 'MF': 'Modern Fit',
    'BT': 'Big & Tall', 'BB': 'Big & Tall', 'TT': 'Tall',
    'SS': 'Short Sleeve', 'SR': 'Short Sleeve Regular',
    'SB': 'Short Sleeve B&T', 'ST': 'Short Sleeve Tall',
}

# Friendly fabric-code labels (from positions 4-5 of base SKU)
SKU_FABRIC_FRIENDLY = {
    'PH': 'Pique (Polo)', 'PJ': 'Pique Jersey', 'PL': 'Pique LS',
    'PO': 'Pique Open', 'PW': 'Pique Woven', 'TH': 'Tee Heather',
    'HE': 'Heather', 'KN': 'Knit', 'WT': 'Woven Tee',
    'SD': 'Soft Dobby', 'SF': 'Soft Fabric',
    'YD': 'Yarn Dye', 'PT': 'Print/Pattern', 'TS': 'TC Stretch',
    'PK': 'Polyester Knit', 'FT': 'Flax Stretch',
    'BC': 'Carpenter Bottom', 'BR': 'Ripstop Bottom',
    'BH': 'Heavy Bottom', 'BA': 'Pinstripe Bottom',
    'SU': 'Suiting',
}

def _synthesize_style_from_sku(sku):
    """Build a synthetic style metadata dict from SKU code parsing alone.
    Used as the last-resort fallback in _resolve_style_metadata so we NEVER
    drop a model recommendation — every suggestion gets shown, just with
    whatever fidelity we can derive from the code structure.

    SKU structure: [CUSTOMER 2][BRAND 2][FABRIC 2][SERIAL 3][FIT 2][COLLAR 1]
    Example: RONASU371SLP → RO|NA|SU|371|SL|P → Ross Nautica Suiting #371 Slim Long-sleeve Polo-collar
    """
    sku_up = (sku or '').upper().split('-')[0]
    brand_code = sku_up[2:4] if len(sku_up) >= 4 else ''
    brand_abbr = SKU_BRAND_CODE_MAP.get(brand_code, brand_code)
    fab_code = sku_up[4:6] if len(sku_up) >= 6 else ''
    fab_label = SKU_FABRIC_FRIENDLY.get(fab_code, fab_code)
    fit = ''
    try:
        fit_code = _py_extract_fit_code(sku_up)
        # Display default: unknown/junk codes (raw serial fragments like '37')
        # read as Regular Fit here, preserving this feed's old behavior from
        # when _py_extract_fit_code itself defaulted to 'RF'.
        if fit_code and fit_code not in _PY_ALL_FIT_CODES:
            fit_code = 'RF'
        if fit_code:
            fit = SKU_FIT_FRIENDLY.get(fit_code, fit_code)
    except Exception:
        pass
    return {
        'style': sku_up,
        'brand_abbr': brand_abbr,
        'brand_full': BRAND_FULL_NAMES.get(brand_abbr, brand_abbr),
        'color': '',          # can't derive color from SKU code; UI will show blank
        'fabric': fab_label,
        'fit': fit,
        'total_warehouse': 0,
        'incoming': 0,
        'total_ats': 0,
        'in_current_pipeline': False,
        '_synthesized': True,   # flag the UI uses to tag this as "new style suggestion"
    }

def _build_full_catalog_index():
    """Snapshot the live inventory + overrides DB into a unified catalog dict.
    Also builds a secondary 'design' index — keyed by the style code WITHOUT the
    2-char customer prefix — so cross-customer suggestions match too.

    Example: if overrides DB has 'NASU371SLP' (no customer prefix), and the AI
    suggests 'RONASU371SLP' (Ross), we find it under the design index by
    stripping the customer code: 'RONASU371SLP'[2:] = 'NASU371SLP' → match.

    Returns (full_catalog, design_index). Both are dicts of catalog entry dicts.
    """
    with _inv_lock:
        items_snap = list(_inventory['items'])
    with _overrides_lock:
        overrides_snap = dict(_style_overrides)

    full_catalog = {}
    # Live inventory entries (highest fidelity)
    for it in items_snap:
        base = (it.get('sku', '') or '').split('-')[0].upper()
        if base and base not in full_catalog:
            full_catalog[base] = {
                'style': base,
                'brand_abbr': (it.get('brand_abbr') or it.get('brand') or '').upper(),
                'brand_full': it.get('brand_full', ''),
                'color': it.get('color', ''),
                'fabric': it.get('fabrication', ''),
                'fit': it.get('fit', ''),
                'total_warehouse': (it.get('jtw',0)+it.get('tr',0)+it.get('dcw',0)+it.get('qa',0)+it.get('nj',0)),
                'incoming': it.get('incoming', 0),
                'total_ats': it.get('total_ats', 0),
                'in_current_pipeline': True,
            }
    # Overrides DB entries (color/fabric/fit descriptions, may be revival styles)
    for base, ov in overrides_snap.items():
        base_up = (base or '').upper()
        if not base_up or base_up in full_catalog or not isinstance(ov, dict):
            continue
        brand_abbr = (ov.get('brand') or '').upper() or (base_up[2:4] if len(base_up) >= 4 else '')
        full_catalog[base_up] = {
            'style': base_up,
            'brand_abbr': brand_abbr,
            'brand_full': BRAND_FULL_NAMES.get(brand_abbr, brand_abbr),
            'color': ov.get('color', ''),
            'fabric': ov.get('fabric', ''),
            'fit': ov.get('fit', ''),
            'total_warehouse': 0,
            'incoming': 0,
            'total_ats': 0,
            'in_current_pipeline': False,
        }

    # Build the design index — strip the 2-char customer prefix. This lets us
    # match cross-customer suggestions: an override stored as 'NASU371SLP' will
    # match an AI suggestion of 'RONASU371SLP' (Ross prefix) or 'TJNASU371SLP'
    # (TJ prefix), since they share the same design code.
    design_index = {}
    for style_key, entry in full_catalog.items():
        if len(style_key) >= 4:
            design = style_key[2:]   # strip customer prefix
            if design and design not in design_index:
                design_index[design] = entry
            # Also try the full key as a "design" (in case the override IS keyed
            # without a customer prefix in the first place — e.g. 'NASU371SLP')
            if style_key not in design_index:
                design_index[style_key] = entry
    return full_catalog, design_index

def _resolve_style_metadata(style, eligible_styles, full_catalog, design_index):
    """Progressive lookup chain. NEVER returns None — at minimum, synthesizes
    metadata from the SKU code itself. Returns (metadata_dict, source_label).

    source_label tells you which tier the match came from:
      'exact'         — found in eligible_styles (the AI's intended pool)
      'inventory'     — found in live inventory feed
      'inventory_base'— matched after stripping size suffix
      'design'        — fuzzy-matched by stripping customer prefix
      'synthesized'   — no match anywhere; metadata derived from SKU code only

    The 'synthesized' tier means the AI proposed a new style combo that doesn't
    exist as an exact catalog key. We still show it — the user can decide if it's
    a useful new bet. Synthesized entries are flagged so the UI can label them.
    """
    s = (style or '').upper().strip()
    if not s:
        return (_synthesize_style_from_sku(''), 'synthesized')

    # Tier 1: Exact match in the eligible list (the AI's intended pool)
    if s in eligible_styles:
        return (eligible_styles[s], 'exact')

    # Tier 2: Exact match in full catalog (live inventory + overrides)
    if s in full_catalog:
        return (full_catalog[s], 'inventory')

    # Tier 3: Strip size suffix and retry (e.g. "RONASU371SLP-M" → "RONASU371SLP")
    base = s.split('-')[0]
    if base in eligible_styles:
        return (eligible_styles[base], 'exact')
    if base in full_catalog:
        return (full_catalog[base], 'inventory_base')

    # Tier 4: Design-code fuzzy match. Strip the customer prefix and check the
    # design index. This catches the common case where the overrides DB stores
    # color descriptions under a brand-only key (e.g. 'NASU371SLP') but the AI
    # is suggesting a customer-prefixed code (e.g. 'RONASU371SLP').
    if len(base) >= 4:
        design = base[2:]
        if design in design_index:
            # Found by design — but the AI's suggested style code is what we
            # want to display (it's the customer-prefixed version), so we copy
            # the metadata from the design match and override the style code.
            entry = dict(design_index[design])
            entry['style'] = base
            # Update brand from the AI's SKU code in case it suggests a brand
            # variant — but the brand code is positions 2-4 of the AI's suggestion.
            sku_brand_code = base[2:4]
            entry['brand_abbr'] = SKU_BRAND_CODE_MAP.get(sku_brand_code, sku_brand_code)
            entry['brand_full'] = BRAND_FULL_NAMES.get(entry['brand_abbr'], entry['brand_abbr'])
            return (entry, 'design')

    # Tier 5: Synthesize from SKU code. Last resort — never fails.
    return (_synthesize_style_from_sku(s), 'synthesized')

def _anthropic_post_with_retry(payload, *, timeout=180, max_retries=4):
    """POST to Anthropic's /v1/messages with retry-on-transient-error.

    Anthropic returns HTTP 529 (overloaded) and 503 (service unavailable) when
    their backend is over capacity. These typically clear in a few seconds.
    Without retry, even one transient hiccup makes the entire user-facing
    operation fail — and in per-brand parallel mode, ANY one of 5 parallel
    calls hitting 529 throws the whole run.

    Retry policy: exponential backoff with jitter — 2s, 4s, 8s, 16s. Stop after
    `max_retries` total attempts. 529, 503, and 429 (rate limit) are retried;
    everything else (400 bad request, 401 auth, 5xx other than the above) is
    returned as-is so we don't mask real errors.
    """
    headers = {
        'x-api-key': ANTHROPIC_API_KEY,
        'anthropic-version': '2023-06-01',
        'content-type': 'application/json',
    }
    RETRYABLE_STATUSES = {429, 503, 529}
    last_resp = None
    for attempt in range(max_retries):
        try:
            resp = http_requests.post(
                'https://api.anthropic.com/v1/messages',
                headers=headers, json=payload, timeout=timeout
            )
            if resp.status_code == 200 or resp.status_code not in RETRYABLE_STATUSES:
                # Success, or a non-retryable error — return immediately
                return resp
            last_resp = resp
            # Honor server's Retry-After header if present; otherwise exponential backoff
            retry_after = resp.headers.get('Retry-After')
            if retry_after:
                try:
                    delay = min(float(retry_after), 30.0)
                except (ValueError, TypeError):
                    delay = (2 ** attempt) + random.random()
            else:
                delay = (2 ** attempt) + random.random()
            print(f"  [anthropic] {resp.status_code} on attempt {attempt+1}/{max_retries} — "
                  f"backing off {delay:.1f}s", flush=True)
            time.sleep(delay)
        except http_requests.exceptions.RequestException as e:
            # Network-level error (timeout, connection reset, DNS) — also retry
            print(f"  [anthropic] network error on attempt {attempt+1}/{max_retries}: {e} — retry",
                  flush=True)
            delay = (2 ** attempt) + random.random()
            time.sleep(delay)
    # Exhausted retries — return last response (or raise if we never got one)
    if last_resp is None:
        raise http_requests.exceptions.RequestException(
            f"Anthropic API unreachable after {max_retries} attempts")
    return last_resp

def _ai_season_from_sheet(sheet_name):
    """Same auto-mapping as the frontend's _sellingInferSeason. Kept here so the
    backend can pre-filter sheets by season if the request asks for it."""
    s = (sheet_name or '').lower()
    if 'fall'   in s: return 'fall'
    if 'winter' in s: return 'winter'
    if 'spring' in s: return 'spring'
    if 'summer' in s: return 'summer'
    m = re.match(r'^(\d{1,2})[.\-/](\d{1,2})', s)
    if m:
        month = int(m.group(1))
        if 9 <= month <= 11: return 'fall'
        if month == 12 or month <= 2: return 'winter'
        if 3 <= month <= 5: return 'spring'
        if 6 <= month <= 8: return 'summer'
    return 'unknown'

def _ai_date_from_sheet(sheet_name):
    """Best-effort parse of a sheet name into an actual date.
    Sheet names look like:
      '9.08'      → month/day, assume current selling-cycle year (Sep+ = prior year,
                    Jan-Aug = current year — matches how the recap reports are dated)
      '2.25.26'   → month.day.year (2-digit year)
      'Fall 2025' → no specific date; use mid-season midpoint (Oct 15)
      'Spring 26' → similar; use mid-season midpoint (Apr 15)
    Returns a datetime object on success, None if unparseable. The dates are used
    for ordering selling windows and matching ship dates to seasons — they don't
    have to be perfectly accurate, just consistent.
    """
    if not sheet_name:
        return None
    s = sheet_name.strip().lower()

    # Explicit season-name forms — "Fall 2025", "Spring 26", "Summer", etc.
    season_match = re.match(r'(fall|winter|spring|summer)\s*(\d{2,4})?', s)
    if season_match:
        season = season_match.group(1)
        year_str = season_match.group(2)
        if year_str:
            year = int(year_str)
            if year < 100: year += 2000
        else:
            year = datetime.utcnow().year
        # Midpoint of each season — gives the AI a reasonable date to compare ship dates against
        mid = {'spring': (4, 15), 'summer': (7, 15), 'fall': (10, 15), 'winter': (1, 15)}[season]
        # 'winter X' typically refers to the winter that spans Dec X-1 → Feb X
        if season == 'winter':
            return datetime(year, mid[0], mid[1])
        return datetime(year, mid[0], mid[1])

    # Numeric date forms — "9.08", "2.25.26", "11.10"
    m = re.match(r'^(\d{1,2})[.\-/](\d{1,2})(?:[.\-/](\d{2,4}))?$', s)
    if not m:
        return None
    month = int(m.group(1))
    day = int(m.group(2))
    if not (1 <= month <= 12 and 1 <= day <= 31):
        return None
    if m.group(3):
        year = int(m.group(3))
        if year < 100: year += 2000
    else:
        # No year given. Heuristic: months 9-12 belong to the *prior* selling year,
        # months 1-8 belong to the *current* selling year. This matches how the
        # Ross recap sheets are dated (Sep-Dec 2025 sheets are dated 9.xx-12.xx;
        # Jan-Aug 2026 sheets are dated 1.xx-8.xx — they're all part of one cycle).
        # The selling cycle is anchored on the upcoming year unless we're past Aug.
        now = datetime.utcnow()
        anchor_year = now.year if now.month <= 6 else now.year + 1
        year = anchor_year - 1 if month >= 9 else anchor_year
    try:
        return datetime(year, month, day)
    except ValueError:
        return None

def _ai_season_from_date(dt):
    """Bucket a date into one of fall/winter/spring/summer. Northern-hemisphere retail
    season convention: Sep-Nov = Fall, Dec-Feb = Winter, Mar-May = Spring, Jun-Aug = Summer."""
    if not dt:
        return 'unknown'
    m = dt.month
    if 9 <= m <= 11: return 'fall'
    if m == 12 or m <= 2: return 'winter'
    if 3 <= m <= 5: return 'spring'
    return 'summer'

def _ai_floor_season_from_ship_date(dt):
    """Bucket a SHIP DATE into the season the goods will actually be on the floor.

    Why this isn't just _ai_season_from_date: goods leaving the dock on 8/12
    aren't actually selling on 8/12. Between shipping out, transit, DC receiving,
    QA, and store distribution, real time-on-floor lags the ship date by ~3 weeks.
    So a ship date of 8/12 should be classified as FALL (not late summer) because
    the goods will be selling in early September. We add 21 days then bucket.
    """
    if not dt:
        return 'unknown'
    from datetime import timedelta
    return _ai_season_from_date(dt + timedelta(days=21))

def _ai_parse_order_ship_date(order):
    """Pull the start ship date from an open-order row. Open-orders-api returns
    startDate as an ISO string ('2026-03-15') or similar. Returns datetime or None."""
    raw = order.get('startDate') or order.get('start_date') or order.get('shipDate')
    if not raw:
        return None
    if isinstance(raw, datetime):
        return raw
    s = str(raw).strip()
    if not s:
        return None
    # Try a few common formats
    for fmt in ('%Y-%m-%d', '%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M:%SZ',
                '%m/%d/%Y', '%m/%d/%y', '%Y/%m/%d'):
        try:
            return datetime.strptime(s[:len(fmt)+2], fmt)   # +2 for buffer on T-suffixed forms
        except ValueError:
            continue
    # Fallback: best-effort regex on YYYY-MM-DD pattern
    m = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m:
        try:
            return datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    return None

def _ai_build_sku_windows(sheets_with_dates):
    """Convert per-sheet selling labels into per-SKU time windows.

    Input:  list of {name, date (datetime), season, rows: [...]} in chronological order
    Output: dict mapping each SKU → list of windows, each:
              { start_date, end_date (or None=open-ended), label, season, sheet_name }

    Model (matches what the user described):
      Every appearance of a SKU defines the START of a window with that label.
      The window stays open until the NEXT sheet that mentions the same SKU,
      at which point a new window opens with whatever label that next sheet gave.
      The last window per SKU extends from its sheet date until now (open-ended).
      Before the first appearance, the SKU has no window (unknown — not assumed).
    """
    # Sort sheets by parsed date (chronological, oldest first). Sheets without a
    # parseable date go to the end and are treated as open-ended at the tail.
    dated   = sorted([s for s in sheets_with_dates if s.get('date')], key=lambda s: s['date'])
    undated = [s for s in sheets_with_dates if not s.get('date')]
    ordered = dated + undated

    # First pass: collect every appearance per SKU, preserving order
    by_sku = {}
    for sh in ordered:
        for r in sh['rows']:
            sku = (r.get('sku') or '').upper()
            if not sku or r.get('label') not in ('best', 'okay', 'worst'):
                continue
            by_sku.setdefault(sku, []).append({
                'sheet_name': sh['name'],
                'date': sh.get('date'),
                'season': sh.get('season') or _ai_season_from_sheet(sh['name']),
                'label': r['label'],
            })

    # Second pass: stitch consecutive appearances into windows.
    # Window i runs from appearance[i].date until appearance[i+1].date, with label[i].
    # The last appearance's window is open-ended (end_date=None).
    windows_by_sku = {}
    for sku, appearances in by_sku.items():
        ws = []
        for i, a in enumerate(appearances):
            end = appearances[i + 1]['date'] if i + 1 < len(appearances) else None
            ws.append({
                'start_date': a['date'],
                'end_date': end,
                'label': a['label'],
                'season': a['season'],
                'sheet_name': a['sheet_name'],
            })
        windows_by_sku[sku] = ws
    return windows_by_sku

def _ai_format_window(w):
    """Render one window as a compact string for the AI prompt."""
    def _fmt(d):
        return d.strftime('%Y-%m-%d') if d else 'open'
    start = _fmt(w['start_date'])
    end = _fmt(w['end_date'])
    return f"{start}→{end} [{w['label'].upper()}, {w['season']}]"

def _ai_build_inventory_context(ats_source='all', target_brands=None,
                                ats_only=False, ats_min=0, ats_scope='both'):
    """Build the inventory context for AI prompts.

    ats_source: 'all' (default), 'warehouse_only', or 'all_styles_ever'
    target_brands: optional list of brand identifiers. Accepts either abbreviations
        (BEENE), full names (Geoffrey Beene), or any case variant — each is run
        through _normalize_brand and resolved to a canonical full name, then
        matched against the inventory item's normalized brand_full. This lets the
        frontend send whatever it has (full name from the dropdown, abbr from
        per-brand counts) without separate code paths.
    """
    # Pre-compute the set of acceptable canonical brand names so the per-item check
    # is a single dict lookup. Empty/None means "any brand".
    target_brand_canonicals = None
    if target_brands:
        target_brand_canonicals = set()
        for b in target_brands:
            canon = _normalize_brand(b)
            if canon:
                target_brand_canonicals.add(canon)
                # Also add the uppercase abbreviation form just in case some legacy
                # entries store the abbr in brand_full
                target_brand_canonicals.add(canon.upper())

    with _inv_lock:
        items_snap = list(_inventory['items'])
    with _overrides_lock:
        overrides_snap = dict(_style_overrides)

    # Dedupe by base style (one row per style, regardless of size variants)
    by_style = {}
    _ctx_nj_by_base = {}
    for it in items_snap:
        sku = it.get('sku', '')
        base = sku.split('-')[0].upper()
        if not base:
            continue
        brand_abbr = (it.get('brand_abbr') or it.get('brand') or '').upper()
        brand_full = it.get('brand_full', '') or BRAND_FULL_NAMES.get(brand_abbr, brand_abbr)
        # Brand filter — normalize the item's brand and compare to the target set.
        # The target set already contains canonical full names AND uppercase abbrs,
        # so this catches any variant the frontend might send.
        if target_brand_canonicals is not None:
            item_canon = _normalize_brand(brand_full or brand_abbr)
            if item_canon not in target_brand_canonicals and brand_abbr not in target_brand_canonicals:
                continue

        wh = (it.get('jtw',0)+it.get('tr',0)+it.get('dcw',0)+it.get('qa',0)+it.get('nj',0))
        inc = it.get('incoming', 0)
        if ats_source == 'warehouse_only' and wh <= 0:
            continue
        # 'all' and 'all_styles_ever' both include anything in the live inventory
        # feed, even if wh+inc == 0 (e.g. a fully-committed style still tells the
        # AI that this style exists and was being made). 'all_styles_ever' adds
        # overrides-only styles on top of that — handled below.
        if ats_source == 'all' and (wh + inc) <= 0:
            continue

        # Color/description from style override if present, else from raw item
        ov = overrides_snap.get(base) or {}
        if not isinstance(ov, dict):
            ov = {}
        color  = ov.get('color')   or it.get('color')        or ''
        fabric = ov.get('fabric')  or it.get('fabrication')  or ''
        fit    = ov.get('fit')     or it.get('fit')          or ''

        # Keep the highest-stock variant per base style (most representative).
        # NJ rides ONE carrier row per base, so the pick is made nj-neutral and
        # the base's full NJ total is re-added after the loop — otherwise a
        # non-carrier max row silently drops the whole NJ quantity (audit fix).
        njq = int(it.get('nj') or 0)
        _ctx_nj_by_base[base] = _ctx_nj_by_base.get(base, 0) + njq
        prev = by_style.get(base)
        if prev and (prev['total_warehouse'] + prev['incoming']) > ((wh - njq) + inc):
            continue
        by_style[base] = {
            'style': base,
            'brand_abbr': brand_abbr,
            'brand_full': it.get('brand_full', brand_abbr),
            'color': color,
            'fabric': fabric,
            'fit': fit,
            'total_warehouse': wh - njq,
            'incoming': inc,
            'total_ats': int(it.get('total_ats', 0) or 0) - njq,
            'in_current_pipeline': True,
        }

    # Re-add each base's full NJ quantity onto its kept representative row.
    for _b, _row in by_style.items():
        _njt = _ctx_nj_by_base.get(_b, 0)
        if _njt:
            _row['total_warehouse'] += _njt
            _row['total_ats'] += _njt

    # 'all_styles_ever' — augment with every style that exists in the overrides
    # database but isn't already in the live inventory feed. These represent
    # historical/discontinued styles whose color/fabric/fit metadata is still
    # available — the AI can recommend re-introducing them based on selling patterns.
    if ats_source == 'all_styles_ever':
        for base, ov in overrides_snap.items():
            base_up = (base or '').upper()
            if not base_up or base_up in by_style:
                continue
            if not isinstance(ov, dict):
                continue
            # Brand abbr isn't stored on overrides, so derive it from SKU positions 2-4.
            # If the override has an explicit brand field (some entries do), prefer that.
            brand_abbr = (ov.get('brand') or '').upper() or (base_up[2:4] if len(base_up) >= 4 else '')
            # Apply the normalized brand filter the same way as the live-inventory branch.
            if target_brand_canonicals is not None:
                brand_full_guess = BRAND_FULL_NAMES.get(brand_abbr, brand_abbr)
                item_canon = _normalize_brand(brand_full_guess)
                if item_canon not in target_brand_canonicals and brand_abbr not in target_brand_canonicals:
                    continue
            color  = ov.get('color')  or ''
            fabric = ov.get('fabric') or ''
            fit    = ov.get('fit')    or ''
            # Skip overrides that have nothing useful for the AI to reason from.
            if not (color or fabric or fit):
                continue
            by_style[base_up] = {
                'style': base_up,
                'brand_abbr': brand_abbr,
                'brand_full': brand_abbr,  # we don't have full names for overrides-only
                'color': color,
                'fabric': fabric,
                'fit': fit,
                'total_warehouse': 0,
                'incoming': 0,
                'total_ats': 0,
                'in_current_pipeline': False,
            }

    out = list(by_style.values())

    # ── ATS availability filter (post-process) ──
    # Triggered by the "ATS available only" toggle in the UI. We filter on a
    # per-scope basis so the user can say e.g. "only styles with at least 500
    # units in the warehouse" vs "at least 100 units of either WH or incoming".
    # When ats_only=False, we don't filter at all — the previous behavior.
    if ats_only:
        threshold = max(0, int(ats_min or 0))
        def _passes(s):
            wh  = int(s.get('total_warehouse', 0) or 0)
            inc = int(s.get('incoming', 0) or 0)
            ats = int(s.get('total_ats', 0) or 0)
            if ats_scope == 'warehouse':
                # Use total_ats as the gating signal so over-allocated styles
                # (where committed > stock → ats negative) are correctly excluded.
                # Combined with wh > 0 to make sure we're not picking up overseas-only.
                return wh > 0 and ats >= threshold
            if ats_scope == 'overseas':
                return inc >= threshold
            # 'both' = combined available (WH + incoming, net of deductions = total_ats)
            return ats >= threshold
        out = [s for s in out if _passes(s)]

    return out

def _ai_build_selling_context(customer, season_filter=None, sheet_filter=None):
    """Build selling-data context for the prompt. If season_filter is supplied,
    only includes sheets matching that season. If sheet_filter is supplied, only
    that specific sheet. Otherwise, all sheets.

    Each sheet entry now includes a parsed `date` (datetime or None) so downstream
    consumers can build SKU time-windows for season-aware prediction.

    Returns: list of {name, season, date, rows: [{sku, label, brand, type, description}]}
    """
    key = customer.lower()
    if key not in _selling_cache:
        refresh_selling_data(customer)
    data = _selling_cache.get(key, {})
    sheets_out = []
    for sh in data.get('sheets', []):
        season = _ai_season_from_sheet(sh['name'])
        if sheet_filter and sh['name'] != sheet_filter:
            continue
        if season_filter and season_filter != 'all' and season != season_filter:
            continue
        sheets_out.append({
            'name': sh['name'],
            'season': season,
            'date': _ai_date_from_sheet(sh['name']),   # datetime or None
            'rows': sh['rows'],
        })
    return sheets_out

def _ai_estimate_tokens(selling_sheets, inventory_styles):
    """Cheap token-count heuristic for the cost estimate UI. We use ~4 chars/token
    as a rule of thumb. Doesn't need to be exact — it's a budgeting aid."""
    # Rough char counts of what we'll serialize into the prompt
    selling_chars = sum(
        len(r.get('sku','')) + len(r.get('description','')) + len(r.get('label','')) + 20
        for sh in selling_sheets for r in sh['rows']
    ) + 200
    inv_chars = sum(
        len(s['style']) + len(s.get('color','')) + len(s.get('fabric','')) + len(s.get('fit','')) + 30
        for s in inventory_styles
    ) + 800  # system prompt + instructions
    input_tokens = (selling_chars + inv_chars) // 4
    # Output: assume 200 tokens per recommendation + 300 overhead
    return input_tokens

def _ai_estimate_cost(input_tokens, output_tokens):
    """Convert token counts to USD estimate using the configured Opus rates."""
    return (input_tokens  / 1_000_000) * AI_OPUS_INPUT_PER_MTOK + \
           (output_tokens / 1_000_000) * AI_OPUS_OUTPUT_PER_MTOK

@app.route('/ai-suggestions/estimate', methods=['POST', 'OPTIONS'])
def ai_suggestions_estimate():
    """Pre-call cost estimate. Same inputs as /ai-suggestions but doesn't call
    the model — just returns the projected token + dollar cost so the UI can
    show '~$0.62' before the user commits."""
    if request.method == 'OPTIONS':
        return ('', 204)
    if not ANTHROPIC_API_KEY:
        return jsonify({'error': 'ANTHROPIC_API_KEY not configured on server'}), 500
    try:
        req = request.get_json(silent=True) or {}
        customer       = req.get('customer', 'Ross')
        season         = (req.get('season') or 'all').lower()
        sheet_filter   = req.get('sheet') or None
        target_brands  = req.get('brands') or None
        ats_source     = req.get('ats_source', 'all')
        ats_only       = bool(req.get('ats_only', False))
        ats_min        = int(req.get('ats_min', 0) or 0)
        ats_scope      = req.get('ats_scope', 'both')
        rec_count      = int(req.get('count', 10))
        mode           = req.get('mode', 'add_only')  # add_only | drop_only | both

        sheets = _ai_build_selling_context(customer, season, sheet_filter)
        inv    = _ai_build_inventory_context(ats_source, target_brands,
                                              ats_only=ats_only, ats_min=ats_min, ats_scope=ats_scope)
        # Same sort + cap as the live endpoint so the estimate matches reality.
        inv.sort(key=lambda s: (-(s.get('total_ats',0)), 0 if s.get('in_current_pipeline') else 1))
        inv = inv[:1500]

        in_toks = _ai_estimate_tokens(sheets, inv)
        # Output budget: scales with rec count; longer rationales now → 320 toks per rec.
        out_per_rec = 320
        out_toks = rec_count * out_per_rec + (rec_count * 200 if mode == 'both' else 0) + 500
        cost = _ai_estimate_cost(in_toks, out_toks)
        return jsonify({
            'input_tokens_est':  in_toks,
            'output_tokens_est': out_toks,
            'cost_usd_est': round(cost, 3),
            'selling_rows': sum(len(s['rows']) for s in sheets),
            'inventory_styles': len(inv),
            'sheets_included': [s['name'] for s in sheets],
            'model': AI_OPUS_MODEL,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/ai-suggestions', methods=['POST', 'OPTIONS'])
def ai_suggestions():
    """Generate booking recommendations from selling data using Claude Opus.

    Request body:
      customer:    'Ross' (str)
      season:      'fall'|'winter'|'spring'|'summer'|'all' — which sheets to feed in
      sheet:       optional specific sheet name (overrides season)
      brands:      optional list of brand abbreviations to restrict recs to
      ats_source:  'all' | 'warehouse_only' — which inventory styles are eligible
      count:       number of recommendations to return (per direction)
      mode:        'add_only' | 'drop_only' | 'both'
      target_season: 'fall'|'winter'|'spring'|'summer' — the season being booked FOR
                     (this is the recommendation target, season above is the historical filter)

    Returns: { recommendations: { add: [...], drop: [...] }, usage: {...}, cost_usd: 0.XX }
    Each rec: { style, brand, recommended_color, strength, rationale,
                referenced_sellers: ['SKU1','SKU2'] }
    """
    if request.method == 'OPTIONS':
        return ('', 204)
    if not ANTHROPIC_API_KEY:
        return jsonify({'error': 'ANTHROPIC_API_KEY not configured on server. '
                       'Add it under Render → Environment.'}), 500
    try:
        req = request.get_json(silent=True) or {}
        customer       = req.get('customer', 'Ross')
        season         = (req.get('season') or 'all').lower()
        sheet_filter   = req.get('sheet') or None
        target_brands  = req.get('brands') or None
        ats_source     = req.get('ats_source', 'all')
        ats_only       = bool(req.get('ats_only', False))
        ats_min        = int(req.get('ats_min', 0) or 0)
        ats_scope      = req.get('ats_scope', 'both')
        rec_count      = max(1, min(50, int(req.get('count', 10))))   # clamp 1..50
        mode           = req.get('mode', 'add_only')
        target_season  = (req.get('target_season') or season or 'fall').lower()
        # Per-brand counts: optional dict {brand_abbr: count}. When set, we fire one
        # call per brand in parallel and concatenate the results. The top-level
        # `count` field is IGNORED in this mode. Each brand sub-call uses the same
        # prompt and inventory context, but with `brands=[abbr]` and `count=N`.
        per_brand_counts = req.get('per_brand_counts') or {}
        # Sanitize: positive ints only, cap each at 50
        per_brand_counts = {
            str(k).upper(): max(1, min(50, int(v)))
            for k, v in per_brand_counts.items()
            if v and int(v) > 0
        }

        # ── Gather context ──
        sheets = _ai_build_selling_context(customer, season, sheet_filter)
        if not sheets:
            return jsonify({'error': 'No selling data matched the season/sheet filter'}), 404
        inv = _ai_build_inventory_context(ats_source, target_brands,
                                          ats_only=ats_only, ats_min=ats_min, ats_scope=ats_scope)
        # ── Smart cap with revival quota ──
        # The earlier behavior was "sort by ATS desc, take first 1500". That meant if
        # 'all_styles_ever' was selected and you had 1500+ pipeline styles in current
        # stock, NO revival styles ever reached the model — defeating the point of the
        # source option. Fix: split the cap so revivals (in_current_pipeline=False) get
        # a reserved share. Default split: 70% pipeline / 30% revival. Within each
        # bucket, sort by ATS desc. Both lists are then concatenated.
        TOTAL_CAP = 3000
        if ats_source == 'all_styles_ever':
            REVIVAL_SHARE = 0.30  # 30% of slots reserved for revival styles
        else:
            REVIVAL_SHARE = 0.0
        revival_slots = int(TOTAL_CAP * REVIVAL_SHARE)
        pipeline_slots = TOTAL_CAP - revival_slots
        pipeline_styles = [s for s in inv if s.get('in_current_pipeline')]
        revival_styles  = [s for s in inv if not s.get('in_current_pipeline')]
        pipeline_styles.sort(key=lambda s: -(s.get('total_ats', 0)))
        revival_styles.sort(key=lambda s: (s.get('style') or ''))   # alpha for stability
        # If one bucket has fewer than its share, the other absorbs the slack.
        if len(pipeline_styles) < pipeline_slots:
            revival_slots = min(TOTAL_CAP - len(pipeline_styles), len(revival_styles))
        elif len(revival_styles) < revival_slots:
            pipeline_slots = min(TOTAL_CAP - len(revival_styles), len(pipeline_styles))
        inv = pipeline_styles[:pipeline_slots] + revival_styles[:revival_slots]
        if not inv:
            return jsonify({'error': 'No inventory styles match the filters. '
                            'Try widening the ATS minimum or switching scope.'}), 404

        # ── Serialize for the prompt ──
        # Selling data: compact "SKU | LABEL | brand | description" lines grouped by sheet.
        # Sheets are presented in their original order (which is chronological in the source)
        # so the model can read trajectories: a style going Worst→Okay→Best vs Best→Worst.
        selling_lines = []
        for sh in sheets:
            selling_lines.append(f"\n## Sheet: {sh['name']}  (season: {sh['season']})")
            for r in sh['rows']:
                lbl = (r.get('label') or 'unknown').upper()
                selling_lines.append(
                    f"  [{lbl}] {r.get('sku','')} | {r.get('brand','')} | {r.get('description','')[:120]}"
                )
        selling_block = '\n'.join(selling_lines)

        # Build a recurrence index so the model gets aggregate signal alongside per-sheet rows.
        # "BEST x3, WORST x0" is much sharper than reading the same SKU labeled BEST in three
        # separate sheets and having to count it yourself.
        from collections import defaultdict
        sku_track = defaultdict(lambda: {'best':0, 'okay':0, 'worst':0, 'desc':'', 'brand':''})
        for sh in sheets:
            for r in sh['rows']:
                sku = (r.get('sku') or '').upper()
                if not sku:
                    continue
                lbl = r.get('label')
                if lbl in ('best','okay','worst'):
                    sku_track[sku][lbl] += 1
                if r.get('description') and not sku_track[sku]['desc']:
                    sku_track[sku]['desc'] = r['description'][:100]
                if r.get('brand') and not sku_track[sku]['brand']:
                    sku_track[sku]['brand'] = r['brand']
        # Sort: most-appearances first; tie-break: more Best is higher signal than more Worst
        ranked = sorted(sku_track.items(),
                        key=lambda kv: (-(kv[1]['best']+kv[1]['okay']+kv[1]['worst']),
                                        -(kv[1]['best'] - kv[1]['worst'])))
        recurrence_lines = ['## RECURRENCE TABLE (same SKU across multiple sheets — strong patterns)']
        for sku, t in ranked:
            appearances = t['best']+t['okay']+t['worst']
            if appearances < 2:
                continue  # only show SKUs that appeared in 2+ sheets
            recurrence_lines.append(
                f"  {sku} | {t['brand']} | BEST×{t['best']} OKAY×{t['okay']} WORST×{t['worst']} | {t['desc']}"
            )
        # If everything appeared exactly once, omit the section so we don't waste tokens
        recurrence_block = '\n'.join(recurrence_lines) if len(recurrence_lines) > 1 else ''

        # Inventory: include the pipeline flag so the AI knows which styles are
        # currently being made vs. which are historical/discontinued candidates for revival.
        inv_lines = ['## ELIGIBLE STYLES (style | brand | color | fabric | fit | WH | INC | pipeline)',
                     '## Note: pipeline=YES means currently in stock or on order; pipeline=NO means historical style with no current stock (only present if "All styles ever" was selected).']
        for s in inv:
            pl = 'YES' if s.get('in_current_pipeline') else 'NO'
            inv_lines.append(
                f"  {s['style']} | {s['brand_abbr']} | {s.get('color','')} | "
                f"{s.get('fabric','')} | {s.get('fit','')} | "
                f"WH:{s['total_warehouse']} | INC:{s['incoming']} | pipeline:{pl}"
            )
        inv_block = '\n'.join(inv_lines)

        # ── Build the prompt ──
        strength_labels_add = (
            "Strong = high conviction, justified by 2+ winning patterns; book in large quantity. "
            "Medium = solid signal but with some uncertainty; book in moderate quantity. "
            "Trial = experimental pick, weaker signal but worth testing; book in small quantity."
        )
        strength_labels_drop = (
            "Cut = clear consistent loser, cancel/cut entirely. "
            "Reduce = mixed signals or single-sheet loser; cut quantity meaningfully but don't fully drop. "
            "Watch = borderline; flag for review but don't necessarily cut yet."
        )

        # ── Drop semantics ──
        # Drops are most actionable on INCOMING styles (you can still cancel/swap production).
        # We tell the model this explicitly so it stops recommending dropping warehouse stock
        # (which is already made and can't be unmade).
        if mode == 'add_only':
            mode_instr = (
                f"Recommend exactly {rec_count} styles to ADD/BOOK for {target_season.upper()} {customer}.\n"
                f"Return an empty 'drop' array."
            )
        elif mode == 'drop_only':
            mode_instr = (
                f"Recommend exactly {rec_count} styles to DROP or REDUCE.\n"
                f"IMPORTANT: drops are only actionable on styles still on incoming production order "
                f"(INC > 0) — you cannot 'drop' finished warehouse stock since it's already made. "
                f"Restrict drop recommendations to styles with INC > 0.\n"
                f"Return an empty 'add' array."
            )
        else:
            mode_instr = (
                f"Recommend exactly {rec_count} styles to ADD/BOOK for {target_season.upper()} {customer}, "
                f"AND exactly {rec_count} styles to DROP/REDUCE.\n"
                f"IMPORTANT: drops are only actionable on styles still on incoming production order "
                f"(INC > 0) — you cannot 'drop' finished warehouse stock. Restrict drop recommendations "
                f"to styles with INC > 0."
            )

        system_prompt = (
            "You are an elite senior apparel buyer and merchandiser working at a private-label "
            "manufacturer that sells to major US retailers. You have 20 years of experience reading "
            "selling data and translating it into next-season production bets.\n\n"
            "You are RIGOROUS. You don't reason from one data point. You distinguish coincidence "
            "from pattern. You separate 'color won' from 'fabric won' from 'print style won' from "
            "'fit won' from 'brand won' — five different explanations for the same selling result "
            "that imply five different next bets.\n\n"
            "Your most valuable skill is CROSS-BRAND PATTERN TRANSFER. When you see a pattern win "
            "in one brand, your job is to ask: 'is this pattern brand-specific (works because of "
            "the brand's cachet/customer base) or is it attribute-driven (works because the print, "
            "color, fit, or fabric is genuinely on-trend)?' The latter type of pattern transfers — "
            "a geometric print that won under Nautica likely has legs under Jones New York or "
            "Eddie Bauer because the print itself, not the badge, drove the sale. This is the "
            "main way you generate non-obvious recommendations.\n\n"
            "Hard rules you NEVER break:\n"
            "1. You ONLY recommend styles that appear verbatim in the ELIGIBLE STYLES list. Never "
            "   invent a style code.\n"
            "2. You MUST return the exact count requested. If you can't find that many high-"
            "   conviction picks, fill the remainder with HONEST Trial-strength picks and say so "
            "   in the rationale ('Trial because the pattern signal is thin, but XYZ suggests it's "
            "   worth a small bet'). Never undershoot the count.\n"
            "3. Every recommendation MUST include a `rationale` of 2-4 sentences that names a "
            "   specific pattern AND cites the evidence. Say WHICH selling SKUs and WHY the "
            "   pattern projects onto your pick. If you're doing cross-brand transfer, name BOTH "
            "   the source brand (where the pattern won) and your reasoning for the transfer.\n"
            "4. Every recommendation MUST include `referenced_sellers` — 1 to 4 actual SKUs from "
            "   the selling data that justify it. If you can't cite at least one, drop that pick "
            "   and find another (but still hit the count).\n"
            "5. DIVERSITY IS MANDATORY. No more than 2 picks from the same color family (e.g. at "
            "   most 2 white solids in the whole 'add' list). No more than 3 picks from the same "
            "   brand. Spread your bets across brands, color families, fabrics, and fits.\n"
            "6. Strength levels must be calibrated honestly. 'Strong' = 2+ supporting patterns or "
            "   3+ consistent sellers AND attribute-level (not just brand-level) signal. 'Medium' "
            "   = one solid pattern. 'Trial' = thin signal but worth testing. Don't inflate.\n"
        )

        user_prompt = (
            f"# CUSTOMER: {customer}\n"
            f"# BOOKING FOR: {target_season.upper()} {datetime.utcnow().year + (1 if datetime.utcnow().month >= 7 else 0)}\n"
            f"# MODE: {mode}\n"
            f"# REQUESTED COUNT: {rec_count} {'add' if mode != 'drop_only' else 'drop'}"
            + (f" AND {rec_count} drop" if mode == 'both' else "") + " recommendation(s) — return EXACTLY this count.\n"
            f"# Strength levels (ADD): {strength_labels_add}\n"
            f"# Strength levels (DROP): {strength_labels_drop}\n\n"
            f"## HISTORICAL SELLING DATA (chronological)\n"
            f"{selling_block}\n\n"
            + (f"{recurrence_block}\n\n" if recurrence_block else "")
            + f"{inv_block}\n\n"
            + f"## TASK\n"
            + f"{mode_instr}\n\n"
            + "## REASONING METHOD (apply explicitly, in order):\n"
            + "1. **Identify PATTERNS at the ATTRIBUTE level, not the SKU level.** When a style "
            + "wins, ask: was it the COLOR (e.g. navy ground), the PRINT TYPE (e.g. small geometric "
            + "check, gingham, stripe), the FABRIC (e.g. yarn-dye dobby, tc stretch, jersey knit), "
            + "the FIT (e.g. slim long-sleeve), or the BRAND? Most patterns mix several of these. "
            + "Tag each winning pattern with its dominant attribute(s).\n"
            + "2. **Classify each pattern as TRANSFERABLE or BRAND-SPECIFIC.** Attribute-level "
            + "patterns (a print style, a fabric weight, a fit silhouette) generally TRANSFER "
            + "across brands. Brand-cachet patterns (a specific brand's customer loyalty for solid "
            + "colors, for example) DON'T transfer. Examples:\n"
            + "   • 'Small navy gingham checks win across all brands tested' → TRANSFERABLE\n"
            + "   • 'White solids in DKNY specifically' → BRAND-SPECIFIC (DKNY's brand pulls customers)\n"
            + "   • 'Geometric prints in Nautica' → likely TRANSFERABLE if the pattern is the draw\n"
            + "3. **CROSS-BRAND TRANSFER IS A PRIMARY TOOL.** When you have a transferable winning "
            + "pattern (geometric print, navy ground, slim fit, yarn-dye, etc.) and you find styles "
            + "in OTHER brands with the same attribute profile, those are some of your strongest "
            + "picks. Make at least 30% of your recommendations cross-brand transfers — meaning "
            + "the rationale should explicitly say 'this pattern won in Brand X, transferring to "
            + "this Brand Y style because [reason transfer is justified].'\n"
            + "4. **Check for LOSERS that share the same attributes as your would-be ADDs.** If a "
            + "white solid won in one brand but white solids lost in two others, that's a "
            + "negative-transfer signal — that pattern is brand-specific, not transferable.\n"
            + "5. **Apply SEASON FIT for the target booking season.** Target is "
            + f"{target_season.upper()}. Fall/Winter favors heavier fabrics, deeper grounds "
            + "(navy, wine, hunter, charcoal), warmer prints. Spring/Summer favors lighter "
            + "weights, brighter grounds (white, sky, sage, coral), fresher prints. A navy that "
            + "won in Spring still translates to Fall, but a coral that won in Spring does NOT "
            + "translate to Fall.\n"
            + "6. **Prefer pipeline styles (pipeline:YES) over revival styles (pipeline:NO) when "
            + "patterns are roughly equal.** Revivals require noticeably stronger patterns because "
            + "of manufacturing setup cost.\n"
            + "7. **HIT THE COUNT WITH DIVERSITY.** You MUST return exactly the requested count. "
            + "No more than 2 picks from the same color family in 'add'. No more than 3 picks "
            + "from the same brand. If your strong-pattern shortlist runs out, fill with "
            + "honest Trial-strength cross-brand transfers — the diversity rule overrides the "
            + "conviction rule when those two conflict.\n\n"
            + "## OUTPUT FORMAT\n"
            + "Respond as STRICT JSON only — no markdown, no fences, no prose outside the JSON. "
            + "Exact shape:\n"
            + "{\n"
            + '  "add": [\n'
            + "    {\n"
            + '      "style": "<base style code, must appear verbatim in ELIGIBLE STYLES>",\n'
            + '      "brand": "<brand abbreviation>",\n'
            + '      "recommended_color": "<color or print description>",\n'
            + '      "strength": "Strong" | "Medium" | "Trial",\n'
            + '      "rationale": "<2-4 sentences naming the pattern and the supporting evidence>",\n'
            + '      "referenced_sellers": ["<SKU>", "<SKU>"]\n'
            + "    }\n"
            + "  ],\n"
            + '  "drop": [ /* same shape; strength is "Cut" | "Reduce" | "Watch" */ ]\n'
            + "}\n"
            + "Sections that don't apply for this mode get an empty array. NO other keys, NO commentary."
        )

        # ── Per-brand parallel branch ──
        # When the user has specified per-brand counts, we run N parallel API calls
        # (one per brand) and concatenate the results. This produces better quality
        # than asking one call to balance a mix because each call has the full
        # attention budget for its single brand. Costs the same in total tokens.
        if per_brand_counts:
            from concurrent.futures import ThreadPoolExecutor, as_completed

            def _one_brand_call(brand_abbr, want_count):
                """Build inventory filtered to brand_abbr only, then run a recommend call.
                Returns (parsed_dict, usage, error_str_or_None).

                CRITICAL CHANGES from earlier version:
                  1. We rebuild the inventory list FRESH for this single brand instead of
                     slicing from the shared `inv` list. Why: the shared `inv` was capped
                     at 1500 with a "sort by ATS desc, pipeline first" key. Geoffrey Beene
                     revival styles (ATS=0) would lose every cap slot to NAUTICA / DKNY
                     pipeline styles, even when the user explicitly asked for GB. Building
                     fresh per-brand gives each brand its full eligibility list.
                  2. We use a much higher per-brand cap (1500) since one brand has at most
                     a few hundred styles total. Revivals now make it through.
                """
                try:
                    # Fresh inventory build scoped to this brand only — bypasses the
                    # global cap entirely so revivals aren't squeezed out.
                    brand_inv = _ai_build_inventory_context(
                        ats_source, [brand_abbr],
                        ats_only=ats_only, ats_min=ats_min, ats_scope=ats_scope
                    )
                    if not brand_inv:
                        return (None, {}, f"No eligible styles for brand {brand_abbr}")
                    # Sort: pipeline (current stock) first, then revivals — but DON'T cap
                    # aggressively. The model needs to see the revival styles to recommend them.
                    brand_inv.sort(key=lambda s: (
                        0 if s.get('in_current_pipeline') else 1,
                        -(s.get('total_ats', 0))
                    ))
                    brand_inv = brand_inv[:1500]   # generous per-brand cap
                    # Eligible-styles block: only this brand's styles
                    brand_inv_lines = ['## ELIGIBLE STYLES (only recommend from this list, verbatim style codes)']
                    brand_inv_lines.append(f"## Brand: {brand_abbr}  ({len(brand_inv)} styles)")
                    for s in brand_inv:
                        brand_inv_lines.append(
                            f"  {s['style']} | {s.get('brand_abbr','')} | {s.get('color','') or '-'} | "
                            f"{s.get('fabric','') or '-'} | {s.get('fit','') or '-'} | "
                            f"WH:{s.get('total_warehouse',0)} INC:{s.get('incoming',0)} "
                            f"ATS:{s.get('total_ats',0)} pipeline:{('YES' if s.get('in_current_pipeline') else 'NO')}"
                        )
                    brand_inv_block = '\n'.join(brand_inv_lines)
                    # Reconstruct the entire user prompt with want_count embedded.
                    # Selling block, recurrence block, mode_instr, and the reasoning
                    # method are all unchanged across brands — only count + eligible
                    # styles differ. Note the diversity rule is RELAXED here ("no more
                    # than 3 from same brand") since EVERY pick is from one brand by
                    # definition; we tell the model explicitly.
                    booking_year = datetime.utcnow().year + (1 if datetime.utcnow().month >= 7 else 0)
                    brand_user_prompt = (
                        f"# CUSTOMER: {customer}\n"
                        f"# BOOKING FOR: {target_season.upper()} {booking_year}\n"
                        f"# MODE: {mode}\n"
                        f"# BRAND FOCUS: {brand_abbr} (this entire call is scoped to ONE brand)\n"
                        f"# REQUESTED COUNT: EXACTLY {want_count} {'add' if mode != 'drop_only' else 'drop'}"
                        + (f" AND EXACTLY {want_count} drop" if mode == 'both' else "")
                        + f" recommendation(s) for {brand_abbr} — return EXACTLY {want_count}, no more, no less.\n"
                        f"# Strength levels (ADD): {strength_labels_add}\n"
                        f"# Strength levels (DROP): {strength_labels_drop}\n\n"
                        f"## HISTORICAL SELLING DATA (chronological)\n"
                        f"{selling_block}\n\n"
                        + (f"{recurrence_block}\n\n" if recurrence_block else "")
                        + f"{brand_inv_block}\n\n"
                        + f"## TASK\n"
                        + f"Generate exactly {want_count} {brand_abbr} recommendation(s). "
                        + "All picks come from the ELIGIBLE STYLES list above (which is "
                        + f"already scoped to {brand_abbr} only — DO NOT recommend any other brand).\n\n"
                        + "## REASONING METHOD (apply explicitly, in order):\n"
                        + "1. **Identify PATTERNS at the ATTRIBUTE level** — color, print type, "
                        + "fabric, fit. Tag each winning pattern with its dominant attribute(s).\n"
                        + "2. **Cross-brand transfer is GOLD**: when a non-" + brand_abbr + " "
                        + "style won with a transferable attribute (a print style, fabric weight, "
                        + "or fit silhouette), look for a " + brand_abbr + " style with the same "
                        + "attribute profile. The rationale MUST say 'this pattern won in Brand X, "
                        + f"transferring to {brand_abbr} because [reason].'\n"
                        + "3. **Check for LOSERS that share the same attributes as your would-be "
                        + "ADDs**. If a white solid won in one brand but white solids lost in two "
                        + "others, that's a negative-transfer signal.\n"
                        + f"4. **Apply SEASON FIT for {target_season.upper()}**. Fall/Winter "
                        + "favors heavier fabrics + deeper grounds. Spring/Summer favors lighter "
                        + "weights + brighter grounds.\n"
                        + "5. **Prefer pipeline styles (pipeline:YES) over revivals (pipeline:NO) "
                        + "when patterns are roughly equal.**\n"
                        + f"6. **HIT THE COUNT — EXACTLY {want_count}.** This call is scoped to one "
                        + f"brand. The diversity rule (max 3 per brand) does NOT apply here, since "
                        + f"every pick is {brand_abbr} by design. Spread across COLOR/FABRIC/FIT "
                        + "diversity instead — no more than 2 picks from the same color family. "
                        + "If your high-conviction shortlist runs out, fill with honest "
                        + f"Trial-strength picks. NEVER undershoot {want_count}.\n\n"
                        + "## OUTPUT FORMAT\n"
                        + "Use the submit_recommendations tool. EXACTLY "
                        + f"{want_count} entries in 'add'" + (f" and EXACTLY {want_count} in 'drop'"
                          if mode == 'both' else " (or in 'drop' if mode is drop_only)") + ". "
                        + "Each pick MUST be from the ELIGIBLE STYLES list verbatim."
                    )
                    # Per-brand system prompt — strip the "max 3 from same brand" rule
                    # since by definition this call is single-brand, and STRONGLY re-emphasize
                    # the exact count.
                    brand_system_prompt = (
                        system_prompt
                        + f"\n\n## THIS SPECIFIC CALL\n"
                        + f"You are working on ONE brand: {brand_abbr}. "
                        + f"You MUST return EXACTLY {want_count} 'add' recommendations"
                        + (f" AND EXACTLY {want_count} 'drop' recommendations" if mode == 'both' else "")
                        + ". The 'max 3 from same brand' diversity rule from the general "
                        + "rules above DOES NOT apply on this call — every pick will be "
                        + f"{brand_abbr}. Focus diversity on COLOR/FABRIC/FIT instead "
                        + "(max 2 from same color family). If you can't find enough strong-"
                        + f"conviction picks, fill the remainder with Trial-strength bets — but "
                        + f"NEVER return fewer than {want_count} items."
                    )
                    resp = _anthropic_post_with_retry({
                            'model': AI_OPUS_MODEL,
                            'max_tokens': 16000,
                            'system': brand_system_prompt,
                            'messages': [{'role': 'user', 'content': brand_user_prompt}],
                            'tools': [{
                                'name': 'submit_recommendations',
                                'description': 'Submit booking add and drop recommendations.',
                                'input_schema': {
                                    'type': 'object',
                                    'properties': {
                                        'add': {'type': 'array', 'items': {'type': 'object',
                                            'properties': {
                                                'style': {'type': 'string'},
                                                'brand': {'type': 'string'},
                                                'recommended_color': {'type': 'string'},
                                                'strength': {'type': 'string', 'enum': ['Strong','Medium','Trial']},
                                                'rationale': {'type': 'string'},
                                                'referenced_sellers': {'type':'array','items':{'type':'string'}}
                                            },
                                            'required': ['style','brand','recommended_color','strength','rationale','referenced_sellers']
                                        }},
                                        'drop': {'type': 'array', 'items': {'type': 'object',
                                            'properties': {
                                                'style': {'type': 'string'},
                                                'brand': {'type': 'string'},
                                                'recommended_color': {'type': 'string'},
                                                'strength': {'type': 'string', 'enum': ['Cut','Reduce','Watch']},
                                                'rationale': {'type': 'string'},
                                                'referenced_sellers': {'type':'array','items':{'type':'string'}}
                                            },
                                            'required': ['style','brand','recommended_color','strength','rationale','referenced_sellers']
                                        }},
                                    },
                                    'required': ['add', 'drop']
                                }
                            }],
                            'tool_choice': {'type': 'tool', 'name': 'submit_recommendations'},
                        }, timeout=180)
                    if resp.status_code != 200:
                        return (None, {}, f"{brand_abbr}: HTTP {resp.status_code} — {resp.text[:200]}")
                    body_b = resp.json()
                    parsed_b = None
                    for block in body_b.get('content', []):
                        if block.get('type') == 'tool_use' and block.get('name') == 'submit_recommendations':
                            parsed_b = block.get('input') or {}
                            break
                    if parsed_b is None:
                        return (None, body_b.get('usage', {}), f"{brand_abbr}: model did not call tool")
                    return (parsed_b, body_b.get('usage', {}), None)
                except Exception as e:
                    return (None, {}, f"{brand_abbr}: {type(e).__name__}: {str(e)[:200]}")

            all_add = []
            all_drop = []
            brand_errors = []
            total_in = 0
            total_out = 0
            with ThreadPoolExecutor(max_workers=min(8, len(per_brand_counts))) as ex:
                futures = {ex.submit(_one_brand_call, b, c): b for b, c in per_brand_counts.items()}
                for f in as_completed(futures):
                    parsed_b, usage_b, err = f.result()
                    if err:
                        brand_errors.append(err)
                        continue
                    total_in  += usage_b.get('input_tokens', 0)
                    total_out += usage_b.get('output_tokens', 0)
                    all_add.extend(parsed_b.get('add', []))
                    all_drop.extend(parsed_b.get('drop', []))

            # ── Validate + enrich ──
            # We NEVER drop a recommendation. The resolver's lookup chain (eligible
            # → full catalog → design fuzzy match → SKU-derived synthesis) guarantees
            # every style code gets metadata. Styles the AI proposed that don't exist
            # anywhere as exact keys are still shown — flagged as 'synthesized' so the
            # frontend can label them "new style suggestion".
            #
            # This fixes the symptom where asking for 10 picks returned 6 + "4 skipped"
            # — the model was returning 10, our validator was dropping 4 that didn't
            # match exactly, leaving the user shortchanged. Now all 10 surface.
            eligible_styles = {s['style']: s for s in inv}
            full_catalog, design_index = _build_full_catalog_index()
            def _filter_brand(recs):
                kept, synthesized = [], []
                for r in (recs or []):
                    style = (r.get('style') or '').upper().strip()
                    if not style:
                        continue
                    inv_data, source = _resolve_style_metadata(
                        style, eligible_styles, full_catalog, design_index)
                    r['style'] = inv_data['style']
                    r['_inventory'] = {
                        'brand_abbr':       inv_data['brand_abbr'],
                        'brand_full':       inv_data['brand_full'] or BRAND_FULL_NAMES.get(inv_data['brand_abbr'], inv_data['brand_abbr']),
                        'color':            inv_data.get('color',''),
                        'fabric':           inv_data.get('fabric',''),
                        'fit':              inv_data.get('fit',''),
                        'total_warehouse':  inv_data.get('total_warehouse', 0),
                        'incoming':         inv_data.get('incoming', 0),
                        'total_ats':        inv_data.get('total_ats', 0),
                        'in_current_pipeline': inv_data.get('in_current_pipeline', True),
                        '_synthesized':     bool(inv_data.get('_synthesized', False)),
                        '_lookup_source':   source,
                    }
                    kept.append(r)
                    if source == 'synthesized':
                        synthesized.append(style)
                return kept, synthesized
            add_recs, add_synthesized = _filter_brand(all_add)
            drop_recs, drop_synthesized = _filter_brand(all_drop)
            cost = _ai_estimate_cost(total_in, total_out)
            return jsonify({
                'recommendations': {'add': add_recs, 'drop': drop_recs},
                'usage': {'input_tokens': total_in, 'output_tokens': total_out},
                'cost_usd': round(cost, 4),
                'model': AI_OPUS_MODEL,
                # Renamed from invalid_styles_dropped — these are NOT dropped, they're
                # shown but flagged as "new style suggestions" (style code didn't match
                # any exact catalog entry, so metadata was derived from the SKU itself).
                'invalid_styles_dropped': [],  # kept empty for frontend backward-compat
                'synthesized_styles': add_synthesized + drop_synthesized,
                'per_brand_mode': True,
                'per_brand_counts': per_brand_counts,
                'per_brand_errors': brand_errors,
                'context': {
                    'customer': customer,
                    'season': season,
                    'target_season': target_season,
                    'sheets_used': [s['name'] for s in sheets],
                    'inventory_styles_offered': len(inv),
                    'brands_requested': list(per_brand_counts.keys()),
                }
            })

        # ── Call Claude Opus (with auto-retry on transient 529/503/429) ──
        api_resp = _anthropic_post_with_retry({
                'model': AI_OPUS_MODEL,
                'max_tokens': 16000,
                'system': system_prompt,
                'messages': [{'role': 'user', 'content': user_prompt}],
                # ── Tool use forces structured output ──
                # Instead of "please return JSON" (which is fragile — model can wrap
                # in code fences, add prose, embed JS comments from our prompt example),
                # we define a tool with a strict JSON schema and force the model to
                # call it. The model's response then comes back as a tool_use block
                # with `input` already parsed into a dict — no string parsing required.
                'tools': [{
                    'name': 'submit_recommendations',
                    'description': 'Submit booking add and drop recommendations.',
                    'input_schema': {
                        'type': 'object',
                        'properties': {
                            'add': {
                                'type': 'array',
                                'description': 'Styles to add/book for next season.',
                                'items': {
                                    'type': 'object',
                                    'properties': {
                                        'style': {'type': 'string', 'description': 'Base style code from ELIGIBLE STYLES list, verbatim'},
                                        'brand': {'type': 'string'},
                                        'recommended_color': {'type': 'string'},
                                        'strength': {'type': 'string', 'enum': ['Strong', 'Medium', 'Trial']},
                                        'rationale': {'type': 'string', 'description': '2-4 sentences naming the pattern and evidence'},
                                        'referenced_sellers': {
                                            'type': 'array',
                                            'items': {'type': 'string'},
                                            'description': '1-4 SKUs from selling data that justify this pick'
                                        }
                                    },
                                    'required': ['style', 'brand', 'recommended_color', 'strength', 'rationale', 'referenced_sellers']
                                }
                            },
                            'drop': {
                                'type': 'array',
                                'description': 'Styles to drop/reduce from current bookings.',
                                'items': {
                                    'type': 'object',
                                    'properties': {
                                        'style': {'type': 'string'},
                                        'brand': {'type': 'string'},
                                        'recommended_color': {'type': 'string'},
                                        'strength': {'type': 'string', 'enum': ['Cut', 'Reduce', 'Watch']},
                                        'rationale': {'type': 'string'},
                                        'referenced_sellers': {
                                            'type': 'array',
                                            'items': {'type': 'string'}
                                        }
                                    },
                                    'required': ['style', 'brand', 'recommended_color', 'strength', 'rationale', 'referenced_sellers']
                                }
                            }
                        },
                        'required': ['add', 'drop']
                    }
                }],
                'tool_choice': {'type': 'tool', 'name': 'submit_recommendations'},
            }, timeout=180)
        if api_resp.status_code != 200:
            return jsonify({
                'error': f'Anthropic API returned {api_resp.status_code}',
                'detail': api_resp.text[:500]
            }), 502

        body = api_resp.json()
        # ── Extract the tool_use block ──
        # With tool_choice forcing our tool, the model's response is a single
        # tool_use content block whose `input` is the already-parsed JSON dict.
        # No string parsing, no fence stripping, no JSON decode errors.
        parsed = None
        for block in body.get('content', []):
            if block.get('type') == 'tool_use' and block.get('name') == 'submit_recommendations':
                parsed = block.get('input') or {}
                break
        if parsed is None:
            # Diagnostic fallback: include the raw content so we can see what came back.
            return jsonify({
                'error': 'Model did not call the recommendations tool',
                'raw_response': json.dumps(body.get('content', []))[:2000]
            }), 502

        # ── Validate: every recommended style must exist in the eligible inventory list ──
        # ── Validate + enrich (single-call path) ──
        # Same NEVER-DROP guarantee as the per-brand path: every AI suggestion gets
        # shown. Lookup chain progressively widens (eligible → full catalog → design
        # fuzzy → SKU-derived). See _resolve_style_metadata for details.
        eligible_styles = {s['style']: s for s in inv}
        full_catalog, design_index = _build_full_catalog_index()
        def _filter(recs):
            kept, synthesized = [], []
            for r in (recs or []):
                style = (r.get('style') or '').upper().strip()
                if not style:
                    continue
                inv_data, source = _resolve_style_metadata(
                    style, eligible_styles, full_catalog, design_index)
                r['style'] = inv_data['style']
                r['_inventory'] = {
                    'brand_abbr':       inv_data['brand_abbr'],
                    'brand_full':       inv_data['brand_full'] or BRAND_FULL_NAMES.get(inv_data['brand_abbr'], inv_data['brand_abbr']),
                    'color':            inv_data.get('color',''),
                    'fabric':           inv_data.get('fabric',''),
                    'fit':              inv_data.get('fit',''),
                    'total_warehouse':  inv_data.get('total_warehouse', 0),
                    'incoming':         inv_data.get('incoming', 0),
                    'total_ats':        inv_data.get('total_ats', 0),
                    'in_current_pipeline': inv_data.get('in_current_pipeline', True),
                    '_synthesized':     bool(inv_data.get('_synthesized', False)),
                    '_lookup_source':   source,
                }
                kept.append(r)
                if source == 'synthesized':
                    synthesized.append(style)
            return kept, synthesized

        add_recs, add_synthesized = _filter(parsed.get('add', []))
        drop_recs, drop_synthesized = _filter(parsed.get('drop', []))

        usage = body.get('usage', {})
        in_t  = usage.get('input_tokens', 0)
        out_t = usage.get('output_tokens', 0)
        cost  = _ai_estimate_cost(in_t, out_t)

        return jsonify({
            'recommendations': {'add': add_recs, 'drop': drop_recs},
            'usage': {'input_tokens': in_t, 'output_tokens': out_t},
            'cost_usd': round(cost, 4),
            'model': AI_OPUS_MODEL,
            # Empty for backward compatibility — nothing gets "dropped" anymore.
            'invalid_styles_dropped': [],
            # Synthesized styles: AI suggested codes that didn't match any catalog
            # entry; metadata was derived from the SKU code structure. Still shown
            # to the user — they decide if these new combinations are useful bets.
            'synthesized_styles': add_synthesized + drop_synthesized,
            'context': {
                'customer': customer,
                'season': season,
                'target_season': target_season,
                'sheets_used': [s['name'] for s in sheets],
                'inventory_styles_offered': len(inv),
            }
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


# ─────────────────────────────────────────────────────────────────────────────
# AI OPEN-ORDER PREDICTIONS  (Phase 2b — separate function from Add/Drop)
# Reads Ross's actual open orders from the open-orders-api service, then asks
# Opus to predict — for each open style — whether it will likely sell Best /
# Okay / Worst at Ross based on the historical selling-recap data.
# This is a forecasting tool, not a recommendation tool: every open order gets
# a verdict + reasoning. Nothing is suggested for "drop" — that's deliberate
# (you wanted predictions, not action items).
# ─────────────────────────────────────────────────────────────────────────────

# URL of the separate open-orders service. Matches what the frontend uses.
OPEN_ORDERS_API_URL = os.environ.get('OPEN_ORDERS_API_URL',
                                     'https://open-orders-api.onrender.com')

def _oo_api_headers():
    """Machine credential for calls to the open-orders service. Shares the
    INVENTORY_API_KEY value; harmless while that service's reads are open."""
    return {'X-Api-Key': INVENTORY_API_KEY} if INVENTORY_API_KEY else {}

def _fetch_open_orders():
    """Hit the open-orders-api /api/orders endpoint and return the raw list.
    Returns [] on any error so the caller can surface a clean message."""
    try:
        resp = http_requests.get(f"{OPEN_ORDERS_API_URL}/api/orders", headers=_oo_api_headers(), timeout=30)
        if resp.status_code != 200:
            print(f"[OpenOrders] /api/orders returned {resp.status_code}", flush=True)
            return []
        return resp.json().get('orders', []) or []
    except Exception as e:
        print(f"[OpenOrders] fetch failed: {e}", flush=True)
        return []

def _ross_open_orders(orders):
    """Filter the raw orders list to Ross-only entries.
    Ross can show up under several customer values: 'ROSS', 'ROSS STORES', 'RO'
    (the 2-char code), or 'RM' (Ross with size UPC). We match all of them.
    We also fall back to checking the style prefix in case the customer field
    is blank but the style code's first 2 chars are RO/RM."""
    ROSS_CUSTOMERS = {'ROSS', 'ROSS STORES', 'RO', 'RM', 'ROSS (WITH SIZE UPC)'}
    ROSS_PREFIXES  = {'RO', 'RM'}
    out = []
    for o in orders:
        cust = str(o.get('customer', '') or '').upper().strip()
        style = str(o.get('style', '') or o.get('baseStyle', '') or '').upper().strip()
        base = style.split('-')[0] if style else ''
        prefix = base[:2] if len(base) >= 2 else ''
        if cust in ROSS_CUSTOMERS or prefix in ROSS_PREFIXES:
            out.append(o)
    return out

def _aggregate_orders_by_style(ross_orders):
    """Many POs can hit the same base style. Aggregate into one row per style
    so the AI sees one prediction target per style rather than 17 line items
    for the same SKU. Sums qty + dollar value; collects POs and customers.

    Ship-date handling: an open order has a startDate (when it ships out) and a
    cancelDate (when the retailer will refuse late shipments). For each style we
    track the EARLIEST startDate across all its POs — that's when the first goods
    will land at the retailer. Used downstream by the AI to match the right
    historical selling season."""
    by_style = {}
    for o in ross_orders:
        style = str(o.get('style', '') or o.get('baseStyle', '') or '').upper().split('-')[0]
        if not style:
            continue
        try:
            open_qty = int(o.get('openQty') or 0)
        except (ValueError, TypeError):
            open_qty = 0
        try:
            pick_qty = int(o.get('pickQty') or 0)
        except (ValueError, TypeError):
            pick_qty = 0
        try:
            open_val = float(o.get('openValue') or 0)
        except (ValueError, TypeError):
            open_val = 0
        try:
            pick_val = float(o.get('pickValue') or 0)
        except (ValueError, TypeError):
            pick_val = 0
        ship_dt = _ai_parse_order_ship_date(o)   # datetime or None
        rec = by_style.setdefault(style, {
            'style': style,
            'pos': set(),
            'customer_variants': set(),
            'total_qty': 0,
            'total_value': 0.0,
            'earliest_ship': None,
            'latest_ship': None,
        })
        # The open-orders API returns the PO identifier under several different keys
        # depending on the upstream feed — the frontend already handles this with a
        # multi-key fallback (poOf). Mirror that here so the export doesn't end up
        # with blank PO columns whenever the API happens to use 'poNumber' instead of 'po'.
        po_val = (o.get('po') or o.get('poNumber') or o.get('customerPO')
                  or o.get('orderNo') or o.get('ctrlNo') or '')
        if po_val:
            rec['pos'].add(str(po_val).strip())
        if o.get('customer'):
            rec['customer_variants'].add(str(o['customer']))
        rec['total_qty']   += open_qty + pick_qty
        rec['total_value'] += open_val + pick_val
        if ship_dt:
            if rec['earliest_ship'] is None or ship_dt < rec['earliest_ship']:
                rec['earliest_ship'] = ship_dt
            if rec['latest_ship'] is None or ship_dt > rec['latest_ship']:
                rec['latest_ship'] = ship_dt
    # Serialize sets to sorted lists, attach derived season info for the prompt
    for r in by_style.values():
        r['pos'] = sorted(r['pos'])
        r['customer_variants'] = sorted(r['customer_variants'])
        r['ship_season'] = _ai_floor_season_from_ship_date(r['earliest_ship']) if r['earliest_ship'] else 'unknown'
        r['ship_date_str'] = r['earliest_ship'].strftime('%Y-%m-%d') if r['earliest_ship'] else 'no_date'
    return list(by_style.values())

def _enrich_open_orders_with_inventory(open_styles):
    """For each open-order style, look up color/fabric/fit from the style overrides
    DB and current inventory feed. The AI needs that metadata to apply the
    selling-pattern logic.

    Lookup chain for each style:
      1. Live inventoryData (current stock feed) — direct base-style key
      2. Live inventoryData — any size-variant match (size suffix stripped both sides)
      3. style_overrides DB — direct lookup
      4. style_overrides DB — case-insensitive scan as fallback
      5. SKU-derived signals — brand abbr from positions 2-4, fit code via
         _py_extract_fit_code, fabric code from positions 4-6, category via
         _py_get_item_category. ALL styles will get at least these, since they're
         derived purely from the SKU string.

    The result: NO style is ever skipped for "missing color/fabric data". Worst
    case the AI gets a style with empty color but a derived fit/fabric/category
    from the SKU code, which is still enough to apply pattern logic.
    """
    with _inv_lock:
        items_snap = list(_inventory['items'])
    with _overrides_lock:
        overrides_snap = dict(_style_overrides)

    # Build inventoryData lookups: direct + size-stripped + case-folded
    inv_by_style = {}
    for it in items_snap:
        base = (it.get('sku') or '').split('-')[0].upper()
        if base and base not in inv_by_style:
            inv_by_style[base] = it
    # Case-insensitive override map so a style code with unexpected casing still hits
    overrides_ci = { (k or '').upper(): v for k, v in overrides_snap.items() if isinstance(v, dict) }

    # SKU brand-code → full brand abbr map. Mirrors the frontend SKU_BRAND_CODE_MAP
    # so e.g. 'DK' → 'DKNY', 'CH' → 'CHAPS', 'GB' → 'BEENE'. Without this we'd just
    # show the 2-char code which is uninformative.
    sku_brand_code_map = {
        'NA': 'NAUTICA', 'DK': 'DKNY', 'EB': 'EB', 'RB': 'REEBOK', 'VC': 'VINCE',
        'BE': 'BEN', 'US': 'USPA', 'CH': 'CHAPS', 'LB': 'LUCKY', 'JN': 'JNY',
        'GB': 'BEENE', 'NM': 'NICOLE', 'SH': 'SHAQ', 'TA': 'TAYION', 'MS': 'STRAHAN',
        'VD': 'VD', 'VR': 'VERSA', 'CK': 'CHEROKEE', 'AC': 'AMERICA', 'BL': 'BLO',
        'D9': 'DN', 'KL': 'KL', 'RG': 'RG', 'NE': 'NE',
        'NT': 'NAUTICA',  # Nautica overflow serials past 999
    }

    for s in open_styles:
        base = (s['style'] or '').upper()
        if not base:
            continue

        # 1. Direct inventory hit
        inv = inv_by_style.get(base, {}) or {}
        # 2. Override DB (direct + case-insensitive fallback)
        ov = overrides_snap.get(base) or overrides_ci.get(base) or {}
        if not isinstance(ov, dict):
            ov = {}

        # Brand: prefer live inventory, then override 'brand', then SKU-derived
        brand_abbr = (inv.get('brand_abbr') or inv.get('brand') or ov.get('brand') or '').upper()
        if not brand_abbr and len(base) >= 4:
            # Brand code lives at positions 2-3 of the base SKU (after the 2-char customer code).
            # Example: TJDKTS006 → customer=TJ, brand=DK → DKNY
            brand_code = base[2:4]
            brand_abbr = sku_brand_code_map.get(brand_code, brand_code)
        s['brand_abbr'] = brand_abbr
        s['brand_full'] = _normalize_brand(
            inv.get('brand_full', '') or BRAND_FULL_NAMES.get(brand_abbr, brand_abbr)
        )

        # Color/fabric/fit: prefer override (cleanest data), then live inventory,
        # then SKU-derived for fit/fabric (color can't be derived from SKU codes).
        s['color']  = (ov.get('color')  or inv.get('color')        or '').strip()
        s['fabric'] = (ov.get('fabric') or inv.get('fabrication')  or '').strip()
        s['fit']    = (ov.get('fit')    or inv.get('fit')          or '').strip()

        # Fallback fit: extract from SKU positions per _py_extract_fit_code rules.
        # Example: ROCHYD166SLD → fit code 'SL' (slim sleeve / long sleeve marker).
        if not s['fit']:
            try:
                fit_code = _py_extract_fit_code(s['style'])
                # Display default: junk codes read as 'RF' (old behavior) rather
                # than feeding a raw serial fragment into the AI prompt/UI.
                if fit_code and fit_code not in _PY_ALL_FIT_CODES:
                    fit_code = 'RF'
                if fit_code:
                    # Friendly labels for the common fit codes — gives the AI
                    # something readable rather than just a 2-letter abbrev.
                    fit_friendly = {
                        'SL': 'Slim/Long Sleeve', 'RF': 'Regular Fit',
                        'TF': 'Trim Fit', 'MF': 'Modern Fit',
                        'BT': 'Big & Tall', 'BB': 'Big & Tall', 'TT': 'Tall',
                        'SS': 'Short Sleeve', 'SR': 'Short Sleeve Regular',
                        'SB': 'Short Sleeve B&T', 'ST': 'Short Sleeve Tall',
                    }
                    s['fit'] = fit_friendly.get(fit_code, fit_code)
            except Exception:
                pass

        # Fallback fabric: derive from SKU positions 4-6 (fabric code). These codes
        # are documented in _PY_SPORTSWEAR_FABRICS and _PY_YM_FABRIC_CODES — they tell
        # us whether it's a polo (PH/PJ/PO/PW), tee (TH), heather (HE), woven dress
        # shirt, etc. Better than empty.
        if not s['fabric'] and len(base) >= 6:
            fab_code = base[4:6]
            fab_friendly = {
                'PH': 'Pique (Polo)', 'PJ': 'Pique Jersey', 'PL': 'Pique LS',
                'PO': 'Pique Open', 'PW': 'Pique Woven', 'TH': 'Tee Heather',
                'HE': 'Heather', 'KN': 'Knit', 'WT': 'Woven Tee',
                'SD': 'Soft Dobby', 'SF': 'Soft Fabric',
                'YD': 'Yarn Dye', 'PT': 'Print/Pattern', 'TS': 'TC Stretch',
                'PK': 'Polyester Knit', 'FT': 'Flax Stretch',
                'BC': 'Carpenter Bottom', 'BR': 'Ripstop Bottom',
                'BH': 'Heavy Bottom', 'BA': 'Pinstripe Bottom',
            }
            if fab_code in fab_friendly:
                s['fabric'] = fab_friendly[fab_code]
            elif fab_code:
                s['fabric'] = fab_code   # at minimum, expose the code

        # Category — derived from SKU position rules; tells AI if it's pants/
        # sportswear/short-sleeve/long-sleeve/young men/big&tall. Always available.
        try:
            s['category'] = _py_get_item_category(s['style'], brand_abbr) or ''
        except Exception:
            s['category'] = ''

        # Source flag — tells the AI (and the UI) how much real metadata we had
        # vs how much was inferred from the SKU code. Useful for confidence.
        if ov.get('color') or inv.get('color'):
            s['_metadata_source'] = 'override' if ov.get('color') else 'inventory'
        elif s['fabric'] or s['fit']:
            s['_metadata_source'] = 'sku_derived'
        else:
            s['_metadata_source'] = 'minimal'

    return open_styles

@app.route('/ai-open-order-predictions/preview', methods=['GET', 'POST', 'OPTIONS'])
def ai_open_order_predictions_preview():
    """Returns the list of Ross open-order styles WITHOUT calling the AI.
    Used to populate the cost estimate and let the UI show 'will analyze N styles'
    before the user clicks the expensive Generate button."""
    if request.method == 'OPTIONS':
        return ('', 204)
    try:
        raw = _fetch_open_orders()
        ross = _ross_open_orders(raw)
        styles = _aggregate_orders_by_style(ross)
        styles = _enrich_open_orders_with_inventory(styles)
        # Filter out styles with no useful metadata — they'd be unhelpful for the AI
        # AND inflate cost. Still surface a count so the user knows.
        # No filtering — every style gets analyzed. The enricher always fills in at least
        # category/fit/fabric from SKU codes, so the AI always has something to reason from.
        analyzable = styles
        skipped_no_metadata_list = []
        # Token estimate (rough): ~80 chars per style row in the prompt, plus
        # the selling-data block (variable). We compute a generous in/out so the
        # UI shows a high-water mark.
        return jsonify({
            'total_open_orders': len(ross),
            'unique_styles': len(styles),
            'analyzable_styles': len(analyzable),
            'skipped_no_metadata': len(styles) - len(analyzable),
            'sample': [{'style': s['style'], 'brand': s['brand_abbr'],
                        'color': s['color'], 'qty': s['total_qty'],
                        'value': round(s['total_value'], 2),
                        'pos': s['pos'][:3]} for s in analyzable[:5]],
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/ai-open-order-predictions/estimate', methods=['POST', 'OPTIONS'])
def ai_open_order_predictions_estimate():
    """Cost estimate for the open-order prediction call."""
    if request.method == 'OPTIONS':
        return ('', 204)
    if not ANTHROPIC_API_KEY:
        return jsonify({'error': 'ANTHROPIC_API_KEY not configured'}), 500
    try:
        req = request.get_json(silent=True) or {}
        customer       = req.get('customer', 'Ross')
        season         = (req.get('season') or 'all').lower()
        sheet_filter   = req.get('sheet') or None

        sheets = _ai_build_selling_context(customer, season, sheet_filter)
        raw = _fetch_open_orders()
        ross = _ross_open_orders(raw)
        styles = _aggregate_orders_by_style(ross)
        styles = _enrich_open_orders_with_inventory(styles)
        # No filtering — every style gets analyzed. The enricher always fills in at least
        # category/fit/fabric from SKU codes, so the AI always has something to reason from.
        analyzable = styles
        skipped_no_metadata_list = []

        # Char-count heuristic same as the recommendations estimate.
        selling_chars = sum(
            len(r.get('sku','')) + len(r.get('description','')) + 20
            for sh in sheets for r in sh['rows']
        ) + 300
        open_chars = sum(len(s.get('color','')) + len(s.get('fabric','')) + 40 for s in analyzable)
        in_toks = (selling_chars + open_chars + 2000) // 4   # +2000 for system+instructions
        # Output: ~150 tokens per style (verdict + 2-3 sentence reason + refs)
        out_toks = max(500, len(analyzable) * 150)
        cost = _ai_estimate_cost(in_toks, out_toks)

        return jsonify({
            'input_tokens_est': in_toks,
            'output_tokens_est': out_toks,
            'cost_usd_est': round(cost, 3),
            'analyzable_styles': len(analyzable),
            'unique_open_styles': len(styles),
            'skipped_no_metadata': len(styles) - len(analyzable),
            'selling_sheets': len(sheets),
            'sheets_included': [s['name'] for s in sheets],
            'model': AI_OPUS_MODEL,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/ai-open-order-predictions', methods=['POST', 'OPTIONS'])
def ai_open_order_predictions():
    """Predict selling performance for every Ross open-order style.

    Request body:
      customer:  'Ross' (str) — which customer's selling data to read against
      season:    'fall'|'winter'|'spring'|'summer'|'all' — historical filter
      sheet:     optional specific sheet name override

    Returns: { predictions: [...], skipped: [...], cost_usd, usage, ... }
    Each prediction: { style, brand, color, qty, value, pos: [...],
                       verdict: 'likely_best'|'likely_okay'|'likely_worst',
                       confidence: 'high'|'medium'|'low',
                       rationale, referenced_sellers }
    """
    if request.method == 'OPTIONS':
        return ('', 204)
    if not ANTHROPIC_API_KEY:
        return jsonify({'error': 'ANTHROPIC_API_KEY not configured on server'}), 500
    try:
        req = request.get_json(silent=True) or {}
        customer     = req.get('customer', 'Ross')
        season       = (req.get('season') or 'all').lower()
        sheet_filter = req.get('sheet') or None

        sheets = _ai_build_selling_context(customer, season, sheet_filter)
        if not sheets:
            return jsonify({'error': 'No selling data matched the season/sheet filter'}), 404

        raw = _fetch_open_orders()
        if not raw:
            return jsonify({'error': 'Could not fetch open orders from open-orders-api '
                           '(service may be down or starting up)'}), 503
        ross = _ross_open_orders(raw)
        styles = _aggregate_orders_by_style(ross)
        styles = _enrich_open_orders_with_inventory(styles)
        # No filtering — every style gets analyzed. The enricher always fills in at least
        # category/fit/fabric from SKU codes, so the AI always has something to reason from.
        analyzable = styles
        skipped_no_metadata_list = []
        skipped = []  # kept as empty list for backward-compat response shape
        if not analyzable:
            return jsonify({'error': f'Found {len(ross)} Ross open orders but none had '
                           f'parseable style codes.'}), 404

        # ── Serialize selling data + recurrence table (same structure as Add/Drop endpoint) ──
        from collections import defaultdict
        selling_lines = []
        for sh in sheets:
            selling_lines.append(f"\n## Sheet: {sh['name']}  (season: {sh['season']})")
            for r in sh['rows']:
                lbl = (r.get('label') or 'unknown').upper()
                selling_lines.append(
                    f"  [{lbl}] {r.get('sku','')} | {r.get('brand','')} | {r.get('description','')[:120]}"
                )
        selling_block = '\n'.join(selling_lines)

        sku_track = defaultdict(lambda: {'best':0, 'okay':0, 'worst':0, 'desc':'', 'brand':''})
        for sh in sheets:
            for r in sh['rows']:
                sku = (r.get('sku') or '').upper()
                if not sku:
                    continue
                lbl = r.get('label')
                if lbl in ('best','okay','worst'):
                    sku_track[sku][lbl] += 1
                if r.get('description') and not sku_track[sku]['desc']:
                    sku_track[sku]['desc'] = r['description'][:100]
                if r.get('brand') and not sku_track[sku]['brand']:
                    sku_track[sku]['brand'] = r['brand']
        ranked = sorted(sku_track.items(),
                        key=lambda kv: (-(kv[1]['best']+kv[1]['okay']+kv[1]['worst']),
                                        -(kv[1]['best'] - kv[1]['worst'])))
        recurrence_lines = ['## RECURRENCE TABLE (same SKU across multiple sheets)']
        for sku, t in ranked:
            appearances = t['best']+t['okay']+t['worst']
            if appearances < 2:
                continue
            recurrence_lines.append(
                f"  {sku} | {t['brand']} | BEST×{t['best']} OKAY×{t['okay']} WORST×{t['worst']} | {t['desc']}"
            )
        recurrence_block = '\n'.join(recurrence_lines) if len(recurrence_lines) > 1 else ''

        # ── SKU SELLING WINDOWS ──
        # The user's selling-window model: a label on sheet X applies from sheet X's date
        # until the next sheet that mentions that SKU. This is the AI's primary input
        # for season-aware prediction — for any ship date we can ask "what label was
        # in effect for this SKU/family during the same season last year?"
        sku_windows = _ai_build_sku_windows(sheets)
        window_lines = ['## SKU SELLING WINDOWS  (each label is in effect FROM its date until the NEXT report mentions the SKU)',
                        '## Format: SKU | brand | window_1 → window_2 → … (last window is open-ended)']
        # Sort by total appearances desc — same as recurrence — so the model reads strongest signals first
        windowed_sorted = sorted(sku_windows.items(), key=lambda kv: -len(kv[1]))
        for sku, ws in windowed_sorted:
            if len(ws) < 1:
                continue
            brand = sku_track.get(sku, {}).get('brand', '')
            window_strs = ' → '.join(_ai_format_window(w) for w in ws)
            window_lines.append(f"  {sku} | {brand} | {window_strs}")
        windows_block = '\n'.join(window_lines) if len(window_lines) > 2 else ''

        # ── Open-orders block ──
        # Each row now includes the EARLIEST ship date + the season it ships into,
        # so the AI can match historical windows to the order's ship timing.
        open_lines = ['## ROSS OPEN ORDERS (must produce exactly one verdict per row, no skipping)',
                      '## Fields: style | brand | color | fabric | fit | category | qty | value | SHIP_DATE | SHIP_SEASON | DATA_SOURCE | POs',
                      '## SHIP_SEASON is the season the goods will land on the floor — match historical windows from this season heavily.',
                      '## SHIP_SEASON = "unknown" means no ship date was provided; use full history without season weighting for that style.',
                      '## DATA_SOURCE tells you how much real metadata we have:',
                      '##   "override"      = clean color/fabric from our style database (most reliable)',
                      '##   "inventory"     = pulled from current inventory feed (still good)',
                      '##   "sku_derived"   = no color stored, but fit/fabric inferred from SKU position codes (lower confidence)',
                      '##   "minimal"       = only the SKU itself + derived brand/category (lowest confidence — predict from category + brand patterns)']
        for s in analyzable:
            open_lines.append(
                f"  {s['style']} | {s.get('brand_abbr','')} | {s.get('color','') or '(no color stored)'} | "
                f"{s.get('fabric','') or '(no fabric)'} | {s.get('fit','') or '(no fit)'} | "
                f"{s.get('category','')} | "
                f"qty:{s['total_qty']} | ${s['total_value']:.0f} | "
                f"SHIP:{s.get('ship_date_str','no_date')} | "
                f"SEASON:{s.get('ship_season','unknown')} | "
                f"SRC:{s.get('_metadata_source','minimal')} | "
                f"PO:{','.join(s['pos'][:4])}{'+more' if len(s['pos']) > 4 else ''}"
            )
        open_block = '\n'.join(open_lines)

        system_prompt = (
            "You are an elite senior apparel buyer with 20 years of experience reading retail "
            "selling data. Today's job is FORECASTING — not recommending.\n\n"
            "You will be given (1) Ross's historical selling recap data, (2) per-SKU selling "
            "WINDOWS that show when each label was in effect over time, and (3) the list of styles "
            "Ross has on open order with their SHIP DATE and SHIP SEASON. For EVERY open-order "
            "style, predict whether it will likely sell BEST, OKAY, or WORST when it hits the floor.\n\n"
            "CORE MENTAL MODEL — read carefully:\n"
            "• A label on a recap sheet (BEST/OKAY/WORST) applies to a TIME WINDOW, not a moment. "
            "  The window starts on the sheet's date and ENDS when the next report mentions that SKU.\n"
            "• A SKU labeled BEST on Nov 10 and WORST on Dec 22 was a BEST seller from Nov 10 to "
            "  Dec 22, then became a WORST seller starting Dec 22. Both windows are real — neither "
            "  invalidates the other.\n"
            "• Selling is SEASONAL. A style that won during Fall windows is much more likely to win "
            "  in next year's Fall than in next year's Spring. A style that won in Spring is much "
            "  more likely to win in next year's Spring than in Fall.\n"
            "• Each open order ships at a specific time (SHIP_DATE / SHIP_SEASON). The goods will be "
            "  on the retailer's floor during the SHIP SEASON. Therefore the BEST evidence for that "
            "  order's likely performance comes from HISTORICAL WINDOWS THAT FALL IN THE SAME SEASON.\n\n"
            "Hard rules:\n"
            "1. Produce exactly one verdict per open-order style. Never skip a row. Never invent rows.\n"
            "2. Every verdict needs a rationale that NAMES THE SEASON-MATCHING REASONING. If the ship "
            "season is Fall, explain what you saw in Fall windows of the historical data, not just "
            "overall history.\n"
            "3. Every verdict needs `referenced_sellers` — 1 to 4 SKUs from the historical selling "
            "data that justify your call. Prefer SKUs whose WINNING WINDOW season matches the open "
            "order's ship season.\n"
            "4. Weight evidence with this hierarchy (most important first):\n"
            "   a. Same SKU in the windows table — if the open-order style itself has a history, "
            "      its most recent same-season window is the strongest single signal.\n"
            "   b. Same-season pattern matches — other SKUs whose winning/losing windows fall in "
            "      the same season as the open order's ship season AND whose color/fabric/fit "
            "      profile is similar.\n"
            "   c. Cross-season pattern matches — same color/fabric/fit family that won/lost in a "
            "      different season. Use as secondary context only.\n"
            "5. If SHIP_SEASON is 'unknown' for a style, that means no ship date was provided. Use "
            "   the full history without season weighting and lower confidence by one level.\n"
            "6. DATA_SOURCE affects how you reason, NOT whether you predict. Every style gets a "
            "   verdict, even DATA_SOURCE='minimal' ones. For minimal/sku_derived styles:\n"
            "   • Use BRAND + CATEGORY + FIT patterns from the historical data (e.g. 'DKNY slim-fit "
            "     long-sleeve fancies in Fall windows consistently won').\n"
            "   • The SKU code itself encodes information — first 2 chars = customer, chars 2-4 = "
            "     brand, chars 4-6 = fabric code (yarn dye=YD, pique=PH, dobby=SD, etc.).\n"
            "   • Lower confidence by one level relative to the same call made with full color "
            "     data — 'minimal' rarely earns 'high' confidence.\n"
            "7. Calibrate confidence honestly. 'high' = rich metadata + same-season evidence from "
            "   2+ matching patterns or same-SKU history. 'medium' = some same-season evidence OR "
            "   strong cross-season patterns. 'low' = thin or absent same-season evidence, or "
            "   minimal metadata.\n"
            "8. You are PREDICTING SELLING, not recommending action. Do not suggest cancellations, "
            "   reductions, or replacements. Just forecast.\n"
        )

        user_prompt = (
            f"# CUSTOMER: {customer}\n"
            f"# TASK: Predict selling performance for {len(analyzable)} open-order styles "
            f"using SEASON-AWARE reasoning from time-windowed selling history.\n\n"
            f"## HISTORICAL SELLING DATA (chronological)\n"
            f"{selling_block}\n\n"
            + (f"{recurrence_block}\n\n" if recurrence_block else "")
            + (f"{windows_block}\n\n" if windows_block else "")
            + f"{open_block}\n\n"
            + "## REASONING METHOD (apply in order, every time):\n"
            + "1. For each open-order style, FIRST identify its SHIP_SEASON from the order row. "
            + "This is the season the goods will be on Ross's floor. ALL season-matching logic "
            + "centers on this date.\n"
            + "2. Check if the open-order style appears in the SKU SELLING WINDOWS table. If yes, "
            + "look at its most recent same-season window. That is your strongest evidence. "
            + "(Example: open order ships Oct 15 → Fall season → look for that SKU's history in "
            + "any prior Fall window.)\n"
            + "3. If the style isn't in the windows table, find OTHER SKUs whose winning or losing "
            + "windows fall in the SAME SEASON as the ship season AND whose color/fabric/fit "
            + "profile is similar to the open-order style. These are your primary pattern signals.\n"
            + "4. Distinguish coincidence from pattern. ONE matching seller is not a pattern. A "
            + "pattern needs ≥2 same-season matches OR same-SKU multi-window history.\n"
            + "5. Use cross-season patterns (same color/fabric won in a different season) as "
            + "SECONDARY context only — never as the primary basis for a 'high' confidence verdict.\n"
            + "6. Cross-reference negatives. Don't predict likely_best on a navy plaid because one "
            + "navy plaid won in Fall if two other navy plaids LOST in Fall the same year.\n"
            + "7. Brand specificity matters. A pattern that won in DKNY does not automatically "
            + "transfer to Chaps. Same-brand evidence > cross-brand evidence.\n"
            + "8. For styles with SHIP_SEASON='unknown' (no ship date provided), use the full "
            + "history without season weighting — and lower confidence by one level since you "
            + "can't time-match.\n\n"
            + "## OUTPUT FORMAT\n"
            + "Respond as STRICT JSON only — no markdown, no fences, no prose outside the JSON. "
            + "Exact shape:\n"
            + "{\n"
            + '  "predictions": [\n'
            + "    {\n"
            + '      "style": "<base style from the OPEN ORDERS list, verbatim>",\n'
            + '      "verdict": "likely_best" | "likely_okay" | "likely_worst",\n'
            + '      "confidence": "high" | "medium" | "low",\n'
            + '      "rationale": "<2-4 sentences. MUST mention the ship season and what same-season historical evidence drove the call.>",\n'
            + '      "referenced_sellers": ["<SKU>", "<SKU>"]\n'
            + "    }\n"
            + "  ]\n"
            + "}\n"
            + f"You MUST return exactly {len(analyzable)} entries — one per open-order style. "
            + "NO other keys, NO commentary outside the JSON."
        )

        # ── Batched parallel API calls ──
        # 295 open-order styles × ~200 tokens/prediction = ~60k output tokens needed.
        # That blows past max_tokens=16k, which is what made the previous single-call
        # version silently return zero predictions. We batch into groups of ~50 styles
        # each (~10k output tokens per call, comfortably under 16k), run in parallel,
        # and use prompt caching so the heavy historical context (selling data + windows
        # + recurrence table) is paid for once and read cheaply on subsequent batches.
        BATCH_SIZE = 50
        MAX_WORKERS = 6
        batches = [analyzable[i:i+BATCH_SIZE]
                   for i in range(0, len(analyzable), BATCH_SIZE)]

        # The user_prompt was built above as one string. For caching we need to split
        # it into a cacheable historical-context block and a per-batch open-orders block.
        # We rebuild it here using the pieces already computed earlier in this function.
        cacheable_context = (
            f"# CUSTOMER: {customer}\n"
            f"# TASK: Predict selling performance for open-order styles "
            f"using SEASON-AWARE reasoning from time-windowed selling history.\n\n"
            f"## HISTORICAL SELLING DATA (chronological)\n"
            f"{selling_block}\n\n"
            + (f"{recurrence_block}\n\n" if recurrence_block else "")
            + (f"{windows_block}\n\n" if windows_block else "")
        )
        # The per-batch tail (open orders for THIS batch + reasoning method + output spec).
        # Built per-batch since open_block differs per batch.
        reasoning_tail = (
            "## REASONING METHOD (apply in order, every time):\n"
            "1. For each open-order style, FIRST identify its SHIP_SEASON from the order row. "
            "This is the season the goods will be on Ross's floor.\n"
            "2. Check if the open-order style appears in the SKU SELLING WINDOWS table. If yes, "
            "look at its most recent same-season window. That is your strongest evidence.\n"
            "3. If the style isn't in the windows table, find OTHER SKUs whose winning or losing "
            "windows fall in the SAME SEASON as the ship season AND whose color/fabric/fit "
            "profile is similar to the open-order style.\n"
            "4. Distinguish coincidence from pattern. ONE matching seller is not a pattern.\n"
            "5. Use cross-season patterns (same color/fabric won in a different season) as "
            "SECONDARY context only — never as the primary basis for a 'high' confidence verdict.\n"
            "6. Cross-reference negatives.\n"
            "7. Brand specificity matters. Same-brand evidence > cross-brand evidence.\n"
            "8. For styles with SHIP_SEASON='unknown' (no ship date provided), use the full "
            "history without season weighting — and lower confidence by one level.\n\n"
            "Produce exactly ONE prediction per open-order style listed above. No skipping. "
            "No invented styles. The `style` field MUST match the open-order row verbatim."
        )

        def _build_batch_open_block(batch_styles):
            lines = ['## ROSS OPEN ORDERS (must produce exactly one verdict per row, no skipping)',
                     '## Fields: style | brand | color | fabric | fit | category | qty | value | SHIP_DATE | SHIP_SEASON | DATA_SOURCE | POs']
            for s in batch_styles:
                lines.append(
                    f"  {s['style']} | {s.get('brand_abbr','')} | {s.get('color','') or '(no color stored)'} | "
                    f"{s.get('fabric','') or '(no fabric)'} | {s.get('fit','') or '(no fit)'} | "
                    f"{s.get('category','')} | "
                    f"qty:{s['total_qty']} | ${s['total_value']:.0f} | "
                    f"SHIP:{s.get('ship_date_str','no_date')} | "
                    f"SEASON:{s.get('ship_season','unknown')} | "
                    f"SRC:{s.get('_metadata_source','minimal')} | "
                    f"PO:{','.join(s['pos'][:4])}{'+more' if len(s['pos']) > 4 else ''}"
                )
            return '\n'.join(lines)

        def _call_one_batch(batch_styles, batch_idx):
            """Run one prediction API call for a subset of styles.
            Returns (parsed_dict, usage_dict, stop_reason, error_str_or_None)."""
            open_block_batch = _build_batch_open_block(batch_styles)
            batch_user_content = [
                # Cacheable historical context — same for every batch, paid for once
                # and then read from cache at 10% of the input price on subsequent batches.
                # cache_control: ephemeral expires after 5 minutes of inactivity.
                {'type': 'text', 'text': cacheable_context,
                 'cache_control': {'type': 'ephemeral'}},
                # Per-batch tail — different for each batch (the open orders to predict on).
                {'type': 'text',
                 'text': open_block_batch + "\n\n" + reasoning_tail}
            ]
            try:
                resp = _anthropic_post_with_retry({
                        'model': AI_OPUS_MODEL,
                        'max_tokens': 16000,
                        'system': [{'type': 'text', 'text': system_prompt,
                                    'cache_control': {'type': 'ephemeral'}}],
                        'messages': [{'role': 'user', 'content': batch_user_content}],
                        'tools': [{
                            'name': 'submit_predictions',
                            'description': 'Submit selling predictions for every open-order style in this batch.',
                            'input_schema': {
                                'type': 'object',
                                'properties': {
                                    'predictions': {
                                        'type': 'array',
                                        'description': 'One prediction per open-order style in this batch.',
                                        'items': {
                                            'type': 'object',
                                            'properties': {
                                                'style': {'type': 'string'},
                                                'verdict': {'type': 'string', 'enum': ['likely_best', 'likely_okay', 'likely_worst']},
                                                'confidence': {'type': 'string', 'enum': ['high', 'medium', 'low']},
                                                'rationale': {'type': 'string'},
                                                'referenced_sellers': {
                                                    'type': 'array',
                                                    'items': {'type': 'string'}
                                                }
                                            },
                                            'required': ['style', 'verdict', 'confidence', 'rationale', 'referenced_sellers']
                                        }
                                    }
                                },
                                'required': ['predictions']
                            }
                        }],
                        'tool_choice': {'type': 'tool', 'name': 'submit_predictions'},
                    }, timeout=180)
                if resp.status_code != 200:
                    return (None, {}, '', f"Batch {batch_idx}: HTTP {resp.status_code} — {resp.text[:200]}")
                body_b = resp.json()
                parsed_b = None
                for block in body_b.get('content', []):
                    if block.get('type') == 'tool_use' and block.get('name') == 'submit_predictions':
                        parsed_b = block.get('input') or {}
                        break
                stop_reason_b = body_b.get('stop_reason', '')
                usage_b = body_b.get('usage', {})
                if parsed_b is None:
                    return (None, usage_b, stop_reason_b,
                            f"Batch {batch_idx}: model did not call submit_predictions tool")
                return (parsed_b, usage_b, stop_reason_b, None)
            except Exception as e:
                return (None, {}, '', f"Batch {batch_idx}: {type(e).__name__}: {str(e)[:200]}")

        # ── Fire all batches in parallel ──
        from concurrent.futures import ThreadPoolExecutor, as_completed
        all_preds = []
        batch_errors = []
        stop_reasons = []
        total_usage = {
            'input_tokens': 0, 'output_tokens': 0,
            'cache_read_input_tokens': 0, 'cache_creation_input_tokens': 0,
        }
        with ThreadPoolExecutor(max_workers=min(MAX_WORKERS, len(batches))) as ex:
            future_to_idx = {ex.submit(_call_one_batch, b, i): i for i, b in enumerate(batches)}
            for f in as_completed(future_to_idx):
                parsed_b, usage_b, stop_reason_b, err = f.result()
                if err:
                    batch_errors.append(err)
                    continue
                if stop_reason_b:
                    stop_reasons.append(stop_reason_b)
                for k in total_usage:
                    total_usage[k] += usage_b.get(k, 0)
                all_preds.extend(parsed_b.get('predictions') or [])

        if not all_preds and batch_errors:
            return jsonify({
                'error': 'All prediction batches failed',
                'detail': '; '.join(batch_errors)[:1000],
            }), 502

        # ── Validate + enrich predictions ──
        by_style = {s['style']: s for s in analyzable}
        enriched = []
        invalid = []
        seen_styles = set()
        for p in all_preds:
            style = (p.get('style') or '').upper().strip()
            # Defensive: strip any accidental size suffix (e.g. "ROCHYD166SLD-M" → "ROCHYD166SLD")
            if style not in by_style and '-' in style:
                style = style.split('-')[0]
            if style not in by_style:
                invalid.append(style)
                continue
            if style in seen_styles:
                continue
            seen_styles.add(style)
            order = by_style[style]
            enriched.append({
                'style':              style,
                'verdict':            p.get('verdict') or 'likely_okay',
                'confidence':         p.get('confidence') or 'low',
                'rationale':          p.get('rationale') or '',
                'referenced_sellers': p.get('referenced_sellers') or [],
                'brand_abbr':         order.get('brand_abbr', ''),
                'brand_full':         order.get('brand_full', ''),
                'color':              order.get('color', ''),
                'fabric':             order.get('fabric', ''),
                'fit':                order.get('fit', ''),
                'total_qty':          order['total_qty'],
                'total_value':        order['total_value'],
                'pos':                order['pos'],
                'ship_date':          order.get('ship_date_str', 'no_date'),
                'ship_season':        order.get('ship_season', 'unknown'),
            })

        missing = [s['style'] for s in analyzable if s['style'] not in seen_styles]

        counts = {'likely_best': 0, 'likely_okay': 0, 'likely_worst': 0}
        for e in enriched:
            counts[e['verdict']] = counts.get(e['verdict'], 0) + 1

        usage = total_usage
        # Cost calc — cache reads cost 0.1x normal input rate
        regular_in = usage.get('input_tokens', 0) + usage.get('cache_creation_input_tokens', 0)
        cache_in   = usage.get('cache_read_input_tokens', 0)
        cost = _ai_estimate_cost(regular_in, usage.get('output_tokens', 0)) + \
               (cache_in / 1_000_000) * AI_OPUS_INPUT_PER_MTOK * 0.1

        # ── Pre-warm image cache for these styles in the background ──
        # By the time the user reads the predictions and clicks "Export to Excel",
        # we want every image already in _img_cache. Without this, the first export
        # fetches images one-by-one and can take 30-90s. With pre-warming kicked off
        # right after the AI call returns, the cache is usually populated by the
        # time the export button is clicked → export becomes near-instant.
        try:
            prewarm_items = [
                {
                    'sku':        e['style'],
                    'brand_abbr': e.get('brand_abbr', ''),
                    'brand':      e.get('brand_abbr', ''),
                }
                for e in enriched
            ]
            def _prewarm():
                try:
                    download_images_for_items(prewarm_items, S3_PHOTOS_URL, use_cache=True)
                    print(f"  [prewarm] cached images for {len(prewarm_items)} prediction styles", flush=True)
                except Exception as pwe:
                    print(f"  [prewarm] failed: {pwe}", flush=True)
            threading.Thread(target=_prewarm, daemon=True).start()
        except Exception as e:
            print(f"  [prewarm] kickoff failed: {e}", flush=True)

        return jsonify({
            'predictions': enriched,
            'counts': counts,
            'usage': usage,
            'cost_usd': round(cost, 4),
            'model': AI_OPUS_MODEL,
            'skipped_no_metadata': [s['style'] for s in skipped],
            'invalid_styles_returned': invalid,
            'missing_from_response': missing,
            # Batching diagnostics — surfaced for debug + UI status messaging
            'batches_total':       len(batches),
            'batches_succeeded':   len(batches) - len(batch_errors),
            'batches_failed':      len(batch_errors),
            'batch_errors':        batch_errors,
            'stop_reasons':        list(set(stop_reasons)),
            'context': {
                'customer': customer,
                'season': season,
                'sheets_used': [s['name'] for s in sheets],
                'open_orders_analyzed': len(enriched),
                'total_open_styles': len(styles),
                'batch_size': BATCH_SIZE,
            }
        })

    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT PREDICTIONS TO EXCEL  (with embedded images in column A, one sheet per brand)
# Reuses the same image pipeline as the catalog export so images appear identically.
# ─────────────────────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────────────────────
# EXPORT AI SUGGESTIONS  (Add/Drop recommendations → Excel, one tab per brand)
# Layout: one workbook, one Summary tab + one tab per brand. Within each brand
# tab, Adds come first (with strength Strong/Medium/Trial), then Drops (Cut/
# Reduce/Watch). Image in column A, same pattern as the catalog export.
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/export-suggestions', methods=['POST', 'OPTIONS'])
def export_suggestions():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        req = request.get_json() or {}
        recs = req.get('recommendations') or {}
        fname = req.get('filename', 'AI-Booking-Suggestions')
        adds = recs.get('add') or []
        drops = recs.get('drop') or []
        if not adds and not drops:
            return jsonify({'error': 'No recommendations to export'}), 400

        # Build a unified record list with direction tag so we can group by brand.
        # Each record carries its origin direction so the brand tab can interleave them.
        unified = []
        for r in adds:
            unified.append({**r, '_direction': 'add'})
        for r in drops:
            unified.append({**r, '_direction': 'drop'})

        # Group by brand (canonical name). Recommendations from per-brand calls have
        # _inventory.brand_full reliably set; legacy recs may use the top-level
        # 'brand' field. Normalize both into one canonical key per brand.
        by_brand = {}
        for r in unified:
            inv = r.get('_inventory') or {}
            raw = inv.get('brand_full') or inv.get('brand_abbr') or r.get('brand') or 'Unknown'
            brand = _normalize_brand(raw) or 'Unknown'
            by_brand.setdefault(brand, []).append(r)
        brands_sorted = sorted(by_brand.keys())

        # Fetch images for every style across all brands
        items_for_img = [
            {
                'sku':        r.get('style', ''),
                'brand_abbr': (r.get('_inventory') or {}).get('brand_abbr', '') or r.get('brand', ''),
                'brand':      (r.get('_inventory') or {}).get('brand_abbr', '') or r.get('brand', ''),
            }
            for r in unified
        ]
        all_imgs = download_images_for_items(items_for_img, S3_PHOTOS_URL, use_cache=True)
        img_by_style = {}
        for idx, data in all_imgs.items():
            if idx < len(unified):
                style = (unified[idx].get('style') or '').upper()
                if style and style not in img_by_style:
                    img_by_style[style] = data

        # Build workbook
        buf = BytesIO()
        wb = xlsxwriter.Workbook(buf, {'in_memory': True, 'strings_to_formulas': False})
        wb.set_properties({'title': 'AI Booking Suggestions',
                           'author': 'Versa Inventory System'})

        fmt_header = wb.add_format({'bold': True, 'bg_color': '#1f2937', 'font_color': 'white',
                                    'align': 'center', 'valign': 'vcenter', 'border': 1, 'font_size': 11})
        fmt_text = wb.add_format({'valign': 'vcenter', 'border': 1, 'font_size': 10})
        fmt_text_wrap = wb.add_format({'valign': 'vcenter', 'border': 1, 'font_size': 10, 'text_wrap': True})
        fmt_num = wb.add_format({'valign': 'vcenter', 'border': 1, 'font_size': 10,
                                 'num_format': '#,##0', 'align': 'right'})
        # Direction-colored cells
        fmt_add = wb.add_format({'valign': 'vcenter', 'border': 1, 'font_size': 10, 'bold': True,
                                 'bg_color': '#ede9fe', 'font_color': '#5b21b6', 'align': 'center'})
        fmt_drop = wb.add_format({'valign': 'vcenter', 'border': 1, 'font_size': 10, 'bold': True,
                                  'bg_color': '#fee2e2', 'font_color': '#991b1b', 'align': 'center'})
        fmt_strong = wb.add_format({'valign': 'vcenter', 'border': 1, 'font_size': 10, 'bold': True,
                                    'bg_color': '#dcfce7', 'font_color': '#166534', 'align': 'center'})
        fmt_medium = wb.add_format({'valign': 'vcenter', 'border': 1, 'font_size': 10, 'bold': True,
                                    'bg_color': '#dbeafe', 'font_color': '#1e3a8a', 'align': 'center'})
        fmt_trial = wb.add_format({'valign': 'vcenter', 'border': 1, 'font_size': 10, 'bold': True,
                                   'bg_color': '#fef3c7', 'font_color': '#92400e', 'align': 'center'})
        strength_fmt = {'Strong': fmt_strong, 'Medium': fmt_medium, 'Trial': fmt_trial,
                        'Cut': fmt_drop, 'Reduce': fmt_drop, 'Watch': fmt_trial}

        # Summary sheet
        ws_sum = wb.add_worksheet('Summary')
        sum_h = ['Brand', 'Total', 'Adds', 'Drops', 'Strong Adds', 'Cut Drops']
        for c, h in enumerate(sum_h):
            ws_sum.write(0, c, h, fmt_header)
        ws_sum.set_row(0, 22)
        for c, w in enumerate([28, 10, 10, 10, 14, 14]):
            ws_sum.set_column(c, c, w)
        for r, brand in enumerate(brands_sorted, start=1):
            recs_b = by_brand[brand]
            a = [x for x in recs_b if x['_direction'] == 'add']
            d = [x for x in recs_b if x['_direction'] == 'drop']
            strong = sum(1 for x in a if x.get('strength') == 'Strong')
            cut = sum(1 for x in d if x.get('strength') == 'Cut')
            ws_sum.write(r, 0, brand, fmt_text)
            ws_sum.write(r, 1, len(recs_b), fmt_num)
            ws_sum.write(r, 2, len(a), fmt_num)
            ws_sum.write(r, 3, len(d), fmt_num)
            ws_sum.write(r, 4, strong, fmt_num)
            ws_sum.write(r, 5, cut, fmt_num)

        # Per-brand tabs
        IMAGE_COL_WIDTH = COL_WIDTH_UNITS
        IMAGE_ROW_HEIGHT = 112.5
        headers = ['Image', 'Direction', 'Style', 'Color', 'Strength', 'Rationale',
                   'Referenced Sellers', 'Total ATS', 'Warehouse', 'Incoming']
        col_widths = [IMAGE_COL_WIDTH, 11, 16, 22, 12, 70, 32, 11, 11, 11]

        for brand in brands_sorted:
            safe = re.sub(r'[\\/*?\[\]:]', '', brand)[:31] or 'Brand'
            ws = wb.add_worksheet(safe)
            for c, h in enumerate(headers):
                ws.write(0, c, h, fmt_header)
            ws.set_row(0, 22)
            for c, w in enumerate(col_widths):
                ws.set_column(c, c, w)
            # Sort: Adds first (Strong → Medium → Trial), then Drops (Cut → Reduce → Watch)
            order_dir = {'add': 0, 'drop': 1}
            order_str = {'Strong': 0, 'Medium': 1, 'Trial': 2, 'Cut': 0, 'Reduce': 1, 'Watch': 2}
            recs_b = sorted(by_brand[brand], key=lambda r: (
                order_dir.get(r['_direction'], 9),
                order_str.get(r.get('strength'), 9),
                r.get('style', '')
            ))
            for r_idx, rec in enumerate(recs_b, start=1):
                ws.set_row(r_idx, IMAGE_ROW_HEIGHT)
                style = (rec.get('style') or '').upper()
                inv = rec.get('_inventory') or {}
                img_data = img_by_style.get(style)
                if img_data and img_data.get('image_data'):
                    try:
                        bio = img_data['image_data']
                        if hasattr(bio, 'seek'): bio.seek(0)
                        ratio = (TARGET_W - _IMG_CELL_PAD) / TARGET_W
                        ws.insert_image(r_idx, 0, 'img.png', {
                            'image_data': bio,
                            'x_scale':  (img_data.get('x_scale') or 1) * ratio,
                            'y_scale':  (img_data.get('y_scale') or 1) * ratio,
                            'x_offset': (img_data.get('x_offset') or 0) * ratio + _IMG_CELL_PAD / 2,
                            'y_offset': (img_data.get('y_offset') or 0) * ratio + _IMG_CELL_PAD / 2,
                            'object_position': 1,
                        })
                    except Exception as e:
                        ws.write(r_idx, 0, f'err: {type(e).__name__}', fmt_text)
                else:
                    ws.write(r_idx, 0, 'No Image', fmt_text)
                direction = rec.get('_direction', 'add')
                ws.write(r_idx, 1, '⬆ ADD' if direction == 'add' else '⬇ DROP',
                         fmt_add if direction == 'add' else fmt_drop)
                ws.write(r_idx, 2, style, fmt_text)
                ws.write(r_idx, 3, rec.get('recommended_color', '') or inv.get('color', '') or '', fmt_text)
                str_v = rec.get('strength', '')
                ws.write(r_idx, 4, str_v, strength_fmt.get(str_v, fmt_text))
                ws.write(r_idx, 5, rec.get('rationale', '') or '', fmt_text_wrap)
                ws.write(r_idx, 6, ', '.join(rec.get('referenced_sellers', []) or []), fmt_text)
                ws.write(r_idx, 7, int(inv.get('total_ats', 0) or 0), fmt_num)
                ws.write(r_idx, 8, int(inv.get('total_warehouse', 0) or 0), fmt_num)
                ws.write(r_idx, 9, int(inv.get('incoming', 0) or 0), fmt_num)
            ws.freeze_panes(1, 0)

        wb.close()
        buf.seek(0)
        ts = datetime.utcnow().strftime('%Y-%m-%d')
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{fname}_{ts}.xlsx"
        )
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


# ─────────────────────────────────────────────────────────────────────────────
# EXPORT BOOKING PLAN  (the user's saved-to-plan items → Excel, one tab per brand)
# Same brand-grouped layout as suggestions, but pulls from the plan items list
# the frontend sends. Includes qty (which is user-set, ignoring ATS).
# ─────────────────────────────────────────────────────────────────────────────
@app.route('/export-rec-plan', methods=['POST', 'OPTIONS'])
def export_rec_plan():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        req = request.get_json() or {}
        items = req.get('items') or []
        if not items:
            return jsonify({'error': 'Plan is empty'}), 400

        # Group by brand
        by_brand = {}
        for it in items:
            raw = it.get('brand') or it.get('brand_abbr') or 'Unknown'
            brand = _normalize_brand(raw) or 'Unknown'
            by_brand.setdefault(brand, []).append(it)
        brands_sorted = sorted(by_brand.keys())

        # Image fetch
        items_for_img = [
            {
                'sku':        it.get('style', ''),
                'brand_abbr': it.get('brand_abbr', '') or it.get('brand', ''),
                'brand':      it.get('brand_abbr', '') or it.get('brand', ''),
            }
            for it in items
        ]
        all_imgs = download_images_for_items(items_for_img, S3_PHOTOS_URL, use_cache=True)
        img_by_style = {}
        for idx, data in all_imgs.items():
            if idx < len(items):
                style = (items[idx].get('style') or '').upper()
                if style and style not in img_by_style:
                    img_by_style[style] = data

        buf = BytesIO()
        wb = xlsxwriter.Workbook(buf, {'in_memory': True, 'strings_to_formulas': False})
        wb.set_properties({'title': 'Booking Plan', 'author': 'Versa Inventory System'})

        fmt_header = wb.add_format({'bold': True, 'bg_color': '#1f2937', 'font_color': 'white',
                                    'align': 'center', 'valign': 'vcenter', 'border': 1, 'font_size': 11})
        fmt_text = wb.add_format({'valign': 'vcenter', 'border': 1, 'font_size': 10})
        fmt_text_wrap = wb.add_format({'valign': 'vcenter', 'border': 1, 'font_size': 10, 'text_wrap': True})
        fmt_num = wb.add_format({'valign': 'vcenter', 'border': 1, 'font_size': 10,
                                 'num_format': '#,##0', 'align': 'right'})
        fmt_add = wb.add_format({'valign': 'vcenter', 'border': 1, 'font_size': 10, 'bold': True,
                                 'bg_color': '#ede9fe', 'font_color': '#5b21b6', 'align': 'center'})
        fmt_drop = wb.add_format({'valign': 'vcenter', 'border': 1, 'font_size': 10, 'bold': True,
                                  'bg_color': '#fee2e2', 'font_color': '#991b1b', 'align': 'center'})

        # Summary
        ws_sum = wb.add_worksheet('Summary')
        sum_h = ['Brand', 'Styles', 'Add Qty', 'Drop Qty', 'Total Qty']
        for c, h in enumerate(sum_h):
            ws_sum.write(0, c, h, fmt_header)
        ws_sum.set_row(0, 22)
        for c, w in enumerate([28, 12, 12, 12, 12]):
            ws_sum.set_column(c, c, w)
        grand = {'styles': 0, 'add_qty': 0, 'drop_qty': 0}
        for r, brand in enumerate(brands_sorted, start=1):
            its = by_brand[brand]
            adds = [x for x in its if x.get('direction') == 'add']
            drops = [x for x in its if x.get('direction') == 'drop']
            add_q = sum(int(x.get('qty', 0) or 0) for x in adds)
            drop_q = sum(int(x.get('qty', 0) or 0) for x in drops)
            ws_sum.write(r, 0, brand, fmt_text)
            ws_sum.write(r, 1, len(its), fmt_num)
            ws_sum.write(r, 2, add_q, fmt_num)
            ws_sum.write(r, 3, drop_q, fmt_num)
            ws_sum.write(r, 4, add_q + drop_q, fmt_num)
            grand['styles'] += len(its); grand['add_qty'] += add_q; grand['drop_qty'] += drop_q

        # Per-brand
        IMAGE_COL_WIDTH = COL_WIDTH_UNITS
        IMAGE_ROW_HEIGHT = 112.5
        headers = ['Image', 'Direction', 'Style', 'Color', 'Strength', 'Qty', 'Rationale', 'Added']
        col_widths = [IMAGE_COL_WIDTH, 11, 16, 22, 12, 10, 60, 18]
        for brand in brands_sorted:
            safe = re.sub(r'[\\/*?\[\]:]', '', brand)[:31] or 'Brand'
            ws = wb.add_worksheet(safe)
            for c, h in enumerate(headers):
                ws.write(0, c, h, fmt_header)
            ws.set_row(0, 22)
            for c, w in enumerate(col_widths):
                ws.set_column(c, c, w)
            order_dir = {'add': 0, 'drop': 1}
            its = sorted(by_brand[brand], key=lambda i: (
                order_dir.get(i.get('direction', 'add'), 9),
                i.get('style', '')
            ))
            for r_idx, it in enumerate(its, start=1):
                ws.set_row(r_idx, IMAGE_ROW_HEIGHT)
                style = (it.get('style') or '').upper()
                img_data = img_by_style.get(style)
                if img_data and img_data.get('image_data'):
                    try:
                        bio = img_data['image_data']
                        if hasattr(bio, 'seek'): bio.seek(0)
                        ratio = (TARGET_W - _IMG_CELL_PAD) / TARGET_W
                        ws.insert_image(r_idx, 0, 'img.png', {
                            'image_data': bio,
                            'x_scale':  (img_data.get('x_scale') or 1) * ratio,
                            'y_scale':  (img_data.get('y_scale') or 1) * ratio,
                            'x_offset': (img_data.get('x_offset') or 0) * ratio + _IMG_CELL_PAD / 2,
                            'y_offset': (img_data.get('y_offset') or 0) * ratio + _IMG_CELL_PAD / 2,
                            'object_position': 1,
                        })
                    except Exception as e:
                        ws.write(r_idx, 0, f'err: {type(e).__name__}', fmt_text)
                else:
                    ws.write(r_idx, 0, 'No Image', fmt_text)
                direction = it.get('direction', 'add')
                ws.write(r_idx, 1, '⬆ ADD' if direction == 'add' else '⬇ DROP',
                         fmt_add if direction == 'add' else fmt_drop)
                ws.write(r_idx, 2, style, fmt_text)
                ws.write(r_idx, 3, it.get('color', '') or '', fmt_text)
                ws.write(r_idx, 4, it.get('strength', '') or '', fmt_text)
                ws.write(r_idx, 5, int(it.get('qty', 0) or 0), fmt_num)
                ws.write(r_idx, 6, it.get('rationale', '') or '', fmt_text_wrap)
                ws.write(r_idx, 7, (it.get('added_at', '') or '')[:10], fmt_text)
            ws.freeze_panes(1, 0)

        wb.close()
        buf.seek(0)
        ts = datetime.utcnow().strftime('%Y-%m-%d')
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"Booking-Plan_{ts}.xlsx"
        )
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


@app.route('/export-predictions', methods=['POST', 'OPTIONS'])
def export_predictions():
    """Build an xlsx workbook from a prediction-result payload.

    Request body:
      predictions: [...]   the enriched predictions list (passed back from /ai-open-order-predictions)
      filename:    str     base filename (without extension or date)

    Layout:
      • Summary sheet first (totals by brand)
      • One sheet per brand, sorted alphabetically
      • Column A holds the embedded image, sized to match catalog exports
      • Subsequent columns: Style, Color, Fabric, Fit, Ship Date, Ship Season,
        Verdict, Confidence, Rationale (wide), Referenced Sellers, Qty, Value, POs
    """
    if request.method == 'OPTIONS':
        return '', 204
    try:
        req = request.get_json() or {}
        preds = req.get('predictions') or []
        fname = req.get('filename', 'Ross-Open-Order-Predictions')
        if not preds:
            return jsonify({'error': 'No predictions in request body'}), 400

        # Group predictions by brand. Normalize so "BEENE", "GEOFFREY BEENE",
        # "GB", and "Geoffrey Beene" all collapse to a single canonical key
        # ("Geoffrey Beene") — one sheet per brand, not three.
        by_brand = {}
        for p in preds:
            brand = _normalize_brand(
                p.get('brand_full') or p.get('brand_abbr') or 'Unknown'
            ) or 'Unknown'
            by_brand.setdefault(brand, []).append(p)
        brands_sorted = sorted(by_brand.keys())

        # Image fetch — uses the same cached downloader as catalog exports. We feed it
        # a list of "item-like" dicts (only sku is required for the cache key).
        # Build "item-like" dicts for the image fetcher. brand_abbr is REQUIRED here —
        # without it the resolver can't use the brand-folder fallback path (where most
        # shirt images actually live in S3) and falls through to placeholder. We pass
        # both forms of brand because the resolver checks both keys.
        items_for_img = [
            {
                'sku':        p.get('style', ''),
                'brand_abbr': p.get('brand_abbr', '') or p.get('brand', ''),
                'brand':      p.get('brand_abbr', '') or p.get('brand', ''),
            }
            for p in preds
        ]
        all_imgs = download_images_for_items(items_for_img, S3_PHOTOS_URL, use_cache=True)
        # Build a style → image-bytes map so we can look up by style in each brand sheet.
        # download_images_for_items returns {original_index: image_data}, so reverse it
        # to a style-keyed dict for sheet-by-sheet lookup.
        img_by_style = {}
        for idx, img_data in all_imgs.items():
            if idx < len(preds):
                style = (preds[idx].get('style') or '').upper()
                if style and style not in img_by_style:
                    img_by_style[style] = img_data

        # Build the workbook.
        buf = BytesIO()
        wb = xlsxwriter.Workbook(buf, {'in_memory': True, 'strings_to_formulas': False})
        wb.set_properties({'title': 'Versa Open Order Predictions',
                           'author': 'Versa Inventory System'})

        # Reusable cell formats. Defined once on the workbook (xlsxwriter constraint —
        # formats are workbook-level objects shared across sheets).
        fmt_header = wb.add_format({
            'bold': True, 'bg_color': '#1f2937', 'font_color': 'white',
            'align': 'center', 'valign': 'vcenter', 'border': 1, 'font_size': 11,
        })
        fmt_text = wb.add_format({'valign': 'vcenter', 'border': 1, 'font_size': 10})
        fmt_text_wrap = wb.add_format({
            'valign': 'vcenter', 'border': 1, 'font_size': 10, 'text_wrap': True
        })
        fmt_num = wb.add_format({
            'valign': 'vcenter', 'border': 1, 'font_size': 10,
            'num_format': '#,##0', 'align': 'right'
        })
        fmt_money = wb.add_format({
            'valign': 'vcenter', 'border': 1, 'font_size': 10,
            'num_format': '$#,##0.00', 'align': 'right'
        })
        # Verdict-colored cells — green/yellow/red tints to match the on-screen pills
        fmt_best = wb.add_format({
            'valign': 'vcenter', 'border': 1, 'font_size': 10, 'bold': True,
            'bg_color': '#dcfce7', 'font_color': '#166534', 'align': 'center'
        })
        fmt_okay = wb.add_format({
            'valign': 'vcenter', 'border': 1, 'font_size': 10, 'bold': True,
            'bg_color': '#fef9c3', 'font_color': '#854d0e', 'align': 'center'
        })
        fmt_worst = wb.add_format({
            'valign': 'vcenter', 'border': 1, 'font_size': 10, 'bold': True,
            'bg_color': '#fee2e2', 'font_color': '#991b1b', 'align': 'center'
        })
        verdict_fmt_map = {
            'likely_best': (fmt_best,  'Likely Best'),
            'likely_okay': (fmt_okay,  'Likely Okay'),
            'likely_worst':(fmt_worst, 'Likely Worst'),
        }

        # Summary sheet — totals by brand.
        ws_sum = wb.add_worksheet('Summary')
        sum_headers = ['Brand', 'Total Styles', 'Likely Best', 'Likely Okay',
                       'Likely Worst', 'Total Open Qty', 'Total Open Value (USD)']
        for c, h in enumerate(sum_headers):
            ws_sum.write(0, c, h, fmt_header)
        ws_sum.set_row(0, 22)
        col_widths_sum = [28, 13, 13, 13, 13, 18, 22]
        for c, w in enumerate(col_widths_sum):
            ws_sum.set_column(c, c, w)
        grand = {'styles': 0, 'best': 0, 'okay': 0, 'worst': 0, 'qty': 0, 'value': 0.0}
        for r, brand in enumerate(brands_sorted, start=1):
            bps = by_brand[brand]
            best = sum(1 for p in bps if p.get('verdict') == 'likely_best')
            okay = sum(1 for p in bps if p.get('verdict') == 'likely_okay')
            worst = sum(1 for p in bps if p.get('verdict') == 'likely_worst')
            qty = sum(int(p.get('total_qty', 0) or 0) for p in bps)
            value = sum(float(p.get('total_value', 0) or 0) for p in bps)
            ws_sum.write(r, 0, brand, fmt_text)
            ws_sum.write(r, 1, len(bps), fmt_num)
            ws_sum.write(r, 2, best, fmt_num)
            ws_sum.write(r, 3, okay, fmt_num)
            ws_sum.write(r, 4, worst, fmt_num)
            ws_sum.write(r, 5, qty, fmt_num)
            ws_sum.write(r, 6, value, fmt_money)
            grand['styles'] += len(bps); grand['best'] += best
            grand['okay'] += okay;       grand['worst'] += worst
            grand['qty'] += qty;         grand['value'] += value
        # Bold TOTAL row at the bottom
        fmt_total_text = wb.add_format({'bold': True, 'valign': 'vcenter',
                                        'border': 1, 'font_size': 10, 'bg_color': '#f3f4f6'})
        fmt_total_num = wb.add_format({'bold': True, 'valign': 'vcenter',
                                       'border': 1, 'font_size': 10, 'num_format': '#,##0',
                                       'align': 'right', 'bg_color': '#f3f4f6'})
        fmt_total_money = wb.add_format({'bold': True, 'valign': 'vcenter',
                                         'border': 1, 'font_size': 10, 'num_format': '$#,##0.00',
                                         'align': 'right', 'bg_color': '#f3f4f6'})
        total_row = len(brands_sorted) + 1
        ws_sum.write(total_row, 0, 'TOTAL', fmt_total_text)
        ws_sum.write(total_row, 1, grand['styles'], fmt_total_num)
        ws_sum.write(total_row, 2, grand['best'], fmt_total_num)
        ws_sum.write(total_row, 3, grand['okay'], fmt_total_num)
        ws_sum.write(total_row, 4, grand['worst'], fmt_total_num)
        ws_sum.write(total_row, 5, grand['qty'], fmt_total_num)
        ws_sum.write(total_row, 6, grand['value'], fmt_total_money)

        # Per-brand sheets — image in column A, then the data columns.
        # IMAGE_COL_WIDTH and IMAGE_ROW_HEIGHT MATCH the catalog export exactly so the
        # embedded thumbnails are visually consistent. download_images_for_items returns
        # image dicts already pre-scaled for a 150x150 catalog cell — we use the same
        # helper (_padded_image_opts) that the catalog uses so the result is identical.
        IMAGE_COL_WIDTH = COL_WIDTH_UNITS   # 22 chars — same as catalog
        IMAGE_ROW_HEIGHT = 112.5             # points — same as catalog set_default_row

        headers = ['Image', 'Style', 'Color', 'Fabric', 'Fit', 'Ship Date',
                   'Ship Season', 'Verdict', 'Confidence', 'Rationale',
                   'Referenced Sellers', 'Open Qty', 'Open Value (USD)', 'PO Numbers']
        col_widths = [IMAGE_COL_WIDTH, 16, 20, 18, 16, 12, 13, 14, 12,
                      70, 32, 11, 16, 28]

        verdict_sort_rank = {'likely_best': 0, 'likely_okay': 1, 'likely_worst': 2}
        confidence_sort_rank = {'high': 0, 'medium': 1, 'low': 2}

        for brand in brands_sorted:
            # Sanitize the sheet name — Excel forbids \ / ? * [ ] : and caps at 31 chars.
            safe = re.sub(r'[\\/*?\[\]:]', '', brand)[:31] or 'Brand'
            ws = wb.add_worksheet(safe)

            for c, h in enumerate(headers):
                ws.write(0, c, h, fmt_header)
            ws.set_row(0, 22)
            for c, w in enumerate(col_widths):
                ws.set_column(c, c, w)

            # Sort within the brand: verdict (best→worst), confidence (high→low), value desc
            bps = sorted(by_brand[brand], key=lambda p: (
                verdict_sort_rank.get(p.get('verdict'), 3),
                confidence_sort_rank.get(p.get('confidence'), 3),
                -(p.get('total_value') or 0),
            ))

            for r_idx, p in enumerate(bps, start=1):
                ws.set_row(r_idx, IMAGE_ROW_HEIGHT)
                style = (p.get('style') or '').upper()

                # Column A: image. Two failure modes possible:
                #   1) img_data is None — image fetch failed or didn't run → "No Image" cell
                #   2) img_data exists but insert_image throws (rare — usually means corrupted
                #      bytes or missing required keys in the dict). We log the actual exception
                #      and write a short error tag in the cell so the rest of the sheet still
                #      builds successfully.
                # Defensive: build the insert-options dict ourselves rather than relying on
                # _padded_image_opts, which assumes every expected key is present. Open-order
                # styles that hit the brand-folder fallback path sometimes return image dicts
                # with partial keys.
                img_data = img_by_style.get(style)
                if img_data and img_data.get('image_data'):
                    try:
                        bio = img_data['image_data']
                        if hasattr(bio, 'seek'):
                            bio.seek(0)
                        ratio = (TARGET_W - _IMG_CELL_PAD) / TARGET_W
                        insert_opts = {
                            'image_data': bio,
                            'x_scale':  (img_data.get('x_scale') or 1) * ratio,
                            'y_scale':  (img_data.get('y_scale') or 1) * ratio,
                            'x_offset': (img_data.get('x_offset') or 0) * ratio + _IMG_CELL_PAD / 2,
                            'y_offset': (img_data.get('y_offset') or 0) * ratio + _IMG_CELL_PAD / 2,
                            'object_position': 1,
                            'url': img_data.get('url', '') or '',
                        }
                        ws.insert_image(r_idx, 0, 'img.png', insert_opts)
                    except Exception as e:
                        # Log full exception for server diagnostics
                        import traceback as _tb
                        print(f"  [export-predictions] image insert failed for {style}: "
                              f"{type(e).__name__}: {e}", flush=True)
                        _tb.print_exc()
                        # Short readable tag in the cell — concise so the row still fits
                        ws.write(r_idx, 0, f'err: {type(e).__name__}', fmt_text)
                else:
                    ws.write(r_idx, 0, 'No Image', fmt_text)

                ws.write(r_idx, 1, style, fmt_text)
                ws.write(r_idx, 2, p.get('color', '') or '', fmt_text)
                ws.write(r_idx, 3, p.get('fabric', '') or '', fmt_text)
                ws.write(r_idx, 4, p.get('fit', '') or '', fmt_text)
                ws.write(r_idx, 5, p.get('ship_date', '') or '', fmt_text)
                ws.write(r_idx, 6, (p.get('ship_season', '') or '').title(), fmt_text)
                # Verdict cell uses the colored format matching its bucket
                v = p.get('verdict', 'likely_okay')
                vfmt, vlabel = verdict_fmt_map.get(v, (fmt_okay, v))
                ws.write(r_idx, 7, vlabel, vfmt)
                ws.write(r_idx, 8, (p.get('confidence', '') or '').upper(), fmt_text)
                # Rationale — wide column with text wrap so it's actually readable
                ws.write(r_idx, 9, p.get('rationale', '') or '', fmt_text_wrap)
                ws.write(r_idx, 10, ', '.join(p.get('referenced_sellers', []) or []), fmt_text)
                try:
                    ws.write(r_idx, 11, int(p.get('total_qty', 0) or 0), fmt_num)
                except (ValueError, TypeError):
                    ws.write(r_idx, 11, 0, fmt_num)
                try:
                    ws.write(r_idx, 12, float(p.get('total_value', 0) or 0), fmt_money)
                except (ValueError, TypeError):
                    ws.write(r_idx, 12, 0.0, fmt_money)
                ws.write(r_idx, 13, ', '.join(p.get('pos', []) or []), fmt_text)

            # Freeze the header row so it stays visible while scrolling
            ws.freeze_panes(1, 0)

        wb.close()
        buf.seek(0)
        ts = datetime.utcnow().strftime('%Y-%m-%d')
        return send_file(
            buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"{fname}_{ts}.xlsx"
        )
    except Exception as e:
        import traceback
        return jsonify({'error': str(e), 'trace': traceback.format_exc()}), 500


# ─────────────────────────────────────────────────────────────────────────────
# FACTORY VIEW  (Open Order to PO)
# Sanitized data bundle for the Versa-Docs "Open Order to PO" factory view.
# Factory users (garment factories, some behind China's firewall) see which of
# their production orders have open customer PO demand.
# SECURITY RULE: factories must NEVER see dollar amounts. Every list in the
# response is built field-by-field from an explicit whitelist — never copy an
# upstream dict and delete keys, so a new upstream field can never flow
# through. Col-H shipmentNo (admin-only) is likewise excluded, matching the
# export rule elsewhere in this file.
# ─────────────────────────────────────────────────────────────────────────────

# Transit constants for the factory page's arrival math. THIS PAGE ONLY
# (David, Jul 21 2026): sail time is a flat ETD + 35 days — no pants split —
# and a port-dated row keeps its ATD + 10 arrival. The office platforms keep
# their own 45/55 rule; nothing here feeds them.
FACTORY_VIEW_TRANSIT = {'default_days': 35, 'pants_days': 35}

# Module-level cache for the FULL open-orders list (all customers, unfiltered).
# Deliberately separate from _fetch_open_orders() above — that one is uncached
# and its callers Ross-filter it for the AI prediction endpoints. Do not merge.
_all_open_orders_cache = {'data': None, 'fetched_at': 0, 'fail_until': 0}
_all_open_orders_lock = threading.Lock()
_ALL_OPEN_ORDERS_TTL = 600  # 10 minutes
_UPSTREAM_FAIL_BACKOFF = 120  # after a failure, don't retry for 2 minutes

# FOB pickup customers (route by ETD, never US warehouse). The open-orders app
# keeps the live admin-editable list at /api/fob-customers; this seed matches
# its built-in default and is served when that fetch fails, so the factory view
# always has a workable list. Proxied here so factory browsers only ever need
# one reachable host (onrender.com is blocked in China).
_FOB_CUSTOMERS_SEED = ['CENT1', 'GLOB', 'BFL', 'TJXAU', 'TJXUK', 'HALF', 'MULT', 'MULT1']
_fob_customers_cache = {'data': None, 'fetched_at': 0}
_fob_customers_lock = threading.Lock()


def _fetch_fob_customers():
    """Live FOB customer codes from the open-orders service (10-min cache),
    falling back to the seed list when unreachable."""
    now = time.time()
    with _fob_customers_lock:
        if (_fob_customers_cache['data'] is not None
                and now - _fob_customers_cache['fetched_at'] < _ALL_OPEN_ORDERS_TTL):
            return list(_fob_customers_cache['data'])
    try:
        resp = http_requests.get(f"{OPEN_ORDERS_API_URL}/api/fob-customers", headers=_oo_api_headers(), timeout=10)
        if resp.status_code != 200:
            raise RuntimeError(f"HTTP {resp.status_code}")
        data = resp.json()
        # Response shape: {'fobCustomers': [...]} — null when the admin list
        # was never saved, in which case the seed applies.
        codes = data.get('fobCustomers')
        if not isinstance(codes, list) or not codes:
            raise RuntimeError('list unset upstream')
        codes = [str(c).strip().upper() for c in codes if str(c).strip()]
        with _fob_customers_lock:
            _fob_customers_cache['data'] = codes
            _fob_customers_cache['fetched_at'] = time.time()
        return list(codes)
    except Exception as e:
        print(f"[FactoryView] fob-customers fetch failed ({e}) — using seed list", flush=True)
        with _fob_customers_lock:
            stale = _fob_customers_cache['data']
            return list(stale) if stale else list(_FOB_CUSTOMERS_SEED)


def _fetch_all_open_orders():
    """Fetch the full open-orders list from the open-orders-api service.

    Returns (orders, ok):
      orders — list of raw order dicts (same /api/orders response shape that
               _fetch_open_orders parses)
      ok     — True when the list is fresh (cache hit inside TTL, or a
               successful fetch); False when the upstream fetch failed and we
               are serving a stale cache (or nothing at all).
    Cached module-level with a 10-minute TTL so factory-view polling doesn't
    hammer the open-orders service."""
    now = time.time()
    with _all_open_orders_lock:
        if (_all_open_orders_cache['data'] is not None
                and now - _all_open_orders_cache['fetched_at'] < _ALL_OPEN_ORDERS_TTL):
            return list(_all_open_orders_cache['data']), True
        # After a failure, serve what we have without re-attempting for a
        # while: with a single sync worker, retrying a dead upstream on every
        # request would block the whole API for the timeout each time.
        if now < _all_open_orders_cache['fail_until']:
            stale = _all_open_orders_cache['data']
            return (list(stale) if stale else []), False

    try:
        resp = http_requests.get(f"{OPEN_ORDERS_API_URL}/api/orders", headers=_oo_api_headers(), timeout=15)
        if resp.status_code != 200:
            raise RuntimeError(f"HTTP {resp.status_code}")
        orders = resp.json().get('orders', []) or []
        # An empty list is a bad payload, not a real 'no open orders' state —
        # don't cache it as fresh for the next 10 minutes.
        if not orders:
            raise RuntimeError('empty orders payload')
        with _all_open_orders_lock:
            _all_open_orders_cache['data'] = orders
            _all_open_orders_cache['fetched_at'] = time.time()
            _all_open_orders_cache['fail_until'] = 0
        return list(orders), True
    except Exception as e:
        print(f"[FactoryView] open-orders fetch failed: {e}", flush=True)
        with _all_open_orders_lock:
            _all_open_orders_cache['fail_until'] = time.time() + _UPSTREAM_FAIL_BACKOFF
            stale = _all_open_orders_cache['data']
            return (list(stale) if stale else []), False


# ── Caller identity (Versa-Docs Supabase session) ────────────────────────────
# The factory a caller may see is derived from their OWN profile row, never
# from the query string: a factory account is pinned to its prefix even if it
# asks for another one. The user's access token is used for both calls, so no
# service key lives here — RLS lets a signed-in user read their own profile.
VERSA_DOCS_SUPABASE_URL = os.environ.get(
    'VERSA_DOCS_SUPABASE_URL', 'https://api.versa-docs.com')
VERSA_DOCS_SUPABASE_ANON = os.environ.get(
    'VERSA_DOCS_SUPABASE_ANON', 'sb_publishable_zKJbBUqAFcqZzuaOhGWL0A_miIpnHio')
_identity_cache = {}          # access_token -> (expires_at, profile dict)
_identity_lock = threading.Lock()
_IDENTITY_TTL = 300           # 5 minutes


def _caller_identity(token):
    """Resolve a Versa-Docs access token to {'role','factory_prefix'}.

    Returns None when the token is missing/invalid/unverifiable. Cached
    briefly so polling doesn't hit Supabase on every request."""
    if not token:
        return None
    now = time.time()
    with _identity_lock:
        hit = _identity_cache.get(token)
        if hit and hit[0] > now:
            return hit[1]
    hdr = {'apikey': VERSA_DOCS_SUPABASE_ANON, 'Authorization': f'Bearer {token}'}
    try:
        u = http_requests.get(f'{VERSA_DOCS_SUPABASE_URL}/auth/v1/user',
                              headers=hdr, timeout=10)
        if u.status_code != 200:
            return None
        uid = (u.json() or {}).get('id')
        if not uid:
            return None
        p = http_requests.get(
            f'{VERSA_DOCS_SUPABASE_URL}/rest/v1/profiles'
            f'?select=role,factory_prefix&id=eq.{uid}',
            headers=hdr, timeout=10)
        if p.status_code != 200:
            return None
        rows = p.json() or []
        if not rows:
            return None
        prof = {'role': (rows[0].get('role') or '').strip().lower(),
                'factory_prefix': (rows[0].get('factory_prefix') or '').strip().upper()}
        with _identity_lock:
            _identity_cache[token] = (now + _IDENTITY_TTL, prof)
            if len(_identity_cache) > 500:      # bound the cache
                for k in [k for k, v in _identity_cache.items() if v[0] <= now]:
                    _identity_cache.pop(k, None)
        return prof
    except Exception as e:
        print(f"[FactoryView] identity check failed: {e}", flush=True)
        return None


@app.route('/factory-view', methods=['GET', 'OPTIONS'])
def factory_view():
    """Consumed by Versa-Docs 'Open Order to PO' factory view. Whitelisted —
    no price/value fields may ever be added here.

    GET /factory-view?factory=<CODE>   (Authorization: Bearer <supabase token>)
      CODE — TF/NB/PC/DP/FR/NK (case-insensitive) for a single factory, or
             ALL for office mode (every row, no external_supply masking).

    AUTHORIZATION: the caller's Versa-Docs session decides what they get.
    A 'factory' profile is FORCED to its own prefix regardless of the query
    param; staff may request any factory or ALL. No valid session → 401.

    Bundle: the factory's production ledger rows, masked external supply for
    the same SKUs (routing engine needs the full supply picture per SKU),
    matching open orders / APO / VW allocations, and summed warehouse
    inventory. Explicitly forbidden: sales price, open/pick dollar value,
    unit price, control #, warehouse code, order notes, shipment # —
    no dollar amounts, no admin metadata, ever."""
    if request.method == 'OPTIONS':
        return '', 204

    # ── Who is asking? ──
    auth = request.headers.get('Authorization', '')
    token = auth[7:].strip() if auth.lower().startswith('bearer ') else ''
    ident = _caller_identity(token)
    if not ident:
        return jsonify({'error': 'Sign in to Versa Docs to view this data.'}), 401

    requested = str(request.args.get('factory', '') or '').strip().upper()
    valid_codes = sorted(FACTORY_NAMES.keys())

    if ident['role'] == 'factory':
        # Pinned to their own factory — the query param is ignored entirely.
        code = ident['factory_prefix']
        if code not in FACTORY_NAMES:
            return jsonify({'error': 'Your account is not linked to a factory yet.'}), 403
    elif ident['role'] == 'staff':
        code = requested or 'ALL'
        if code != 'ALL' and code not in FACTORY_NAMES:
            return jsonify({
                'error': f"Invalid factory code '{code}'. "
                         f"Valid codes: {', '.join(valid_codes)} or ALL"
            }), 400
    else:
        return jsonify({'error': 'Your account does not have access to this data.'}), 403

    all_mode = (code == 'ALL')
    factory_name = 'All factories' if all_mode else FACTORY_NAMES[code]

    # ── Production ledger (Dropbox-backed cache, self-TTL'd) ──
    ledger = load_production_from_dropbox()

    productions = []
    sku_set = set()
    for row in ledger:
        prod_ref = str(row.get('production', '') or '').strip()
        if not all_mode and _factory_label(prod_ref) != code:
            continue
        style = str(row.get('style', '') or '').strip().upper()
        if style:
            sku_set.add(style)
        # Field-by-field whitelist — never dict-copy the ledger row
        # (shipmentNo / col H is admin-only and must not leak here).
        # ETD is the RAW ledger date (etd_raw): factories see what the ledger
        # says, never the G-27 backdate. Port-dated rows may therefore carry
        # arrival without etd when column F was blank.
        productions.append({
            'production': prod_ref,
            'poName': row.get('poName', ''),
            'style': style,
            'units': row.get('units', 0),
            'brand': row.get('brand', ''),
            'etd': row.get('etd_raw') if row.get('port_dated') else row.get('etd'),
            'arrival': row.get('arrival'),
            'port_dated': bool(row.get('port_dated')),
            'fob_flag': bool(row.get('fob_flag')),
            'fob_note': row.get('fob_note', ''),
        })

    # ── External supply (single-factory mode only) ──
    # The routing engine needs the FULL supply per SKU to work out which open
    # orders this factory's production actually covers. Rows from OTHER
    # factories for the same SKUs are included but MASKED: a stable "EXT-n"
    # ref (one per distinct external production #, stable within this
    # response) replaces the production #, and poName / brand are omitted
    # entirely — factories must not see each other's order details.
    external_supply = []
    if not all_mode:
        ext_ref_map = {}  # real production # → masked "EXT-n" ref
        for row in ledger:
            prod_ref = str(row.get('production', '') or '').strip()
            if _factory_label(prod_ref) == code:
                continue
            style = str(row.get('style', '') or '').strip().upper()
            if not style or style not in sku_set:
                continue
            if prod_ref not in ext_ref_map:
                ext_ref_map[prod_ref] = f"EXT-{len(ext_ref_map) + 1}"
            external_supply.append({
                'ref': ext_ref_map[prod_ref],
                'style': style,
                'units': row.get('units', 0),
                'etd': row.get('etd_raw') if row.get('port_dated') else row.get('etd'),
                'arrival': row.get('arrival'),
                'port_dated': bool(row.get('port_dated')),
                'fob_flag': bool(row.get('fob_flag')),
            })

    # ── Open orders (exact-SKU uppercase match on the style field) ──
    raw_orders, orders_ok = _fetch_all_open_orders()
    orders = []
    for o in raw_orders:
        style = str(o.get('style', '') or '').strip().upper()
        if style not in sku_set:
            continue
        try:
            open_qty = int(o.get('openQty') or 0)
        except (ValueError, TypeError):
            open_qty = 0
        try:
            pick_qty = int(o.get('pickQty') or 0)
        except (ValueError, TypeError):
            pick_qty = 0
        # WHITELIST — built field-by-field on purpose so a new upstream field
        # can never flow through. The upstream rows also carry sales-price /
        # open-value / pick-value dollar fields, the control #, warehouse
        # code, and notes — none of those may EVER be added here.
        orders.append({
            'style': style,
            'customer': o.get('customer', ''),
            'customerFull': o.get('customerFull', ''),
            'orderNo': o.get('orderNo', ''),
            'startDate': o.get('startDate', ''),
            'cancelDate': o.get('cancelDate', ''),
            'openQty': open_qty,
            'pickQty': pick_qty,
            'isPipeline': bool(o.get('isPipeline')),
        })

    # ── APO allocations (in-memory cache, hourly Dropbox sync) ──
    with _apo_lock:
        apo_rows = list(_apo_data)
    apo = []
    for row in apo_rows:
        style = str(row.get('style', '') or '').strip().upper()
        if style not in sku_set:
            continue
        # 'po' is intentionally not served — the page never displays an APO's
        # internal booking reference to a factory.
        apo.append({
            'style': style,
            'customer': row.get('customer', ''),
            'qty': row.get('qty', 0),
        })

    # ── VW allocations (S3 sheet + manual entries — same merge as /allocations) ──
    try:
        s3_alloc = load_allocation_from_s3()
    except Exception as e:
        print(f"[FactoryView] VW allocation load failed: {e}", flush=True)
        s3_alloc = []
    with _manual_alloc_lock:
        manual_alloc = list(_manual_allocations)
    vw = []
    for row in s3_alloc + manual_alloc:
        sku = str(row.get('sku', '') or '').strip().upper()
        if sku not in sku_set:
            continue
        vw.append({
            'sku': sku,
            'customer': row.get('customer', ''),
            'qty': row.get('qty', 0),
        })

    # ── Inventory (warehouse TOTAL only — per-warehouse breakdown withheld) ──
    with _inv_lock:
        inv_items = list(_inventory['items'])
    # MERGE duplicate feed rows per exact SKU (BUGBSA002SLS bug, Aug 13 2026):
    # the ATS feed can carry a real stock row PLUS a zero-stock row that only
    # holds the commitment. Consumers of this bundle (weekly factory emails,
    # orders2po) key inventory by SKU last-row-wins, so the duplicate wiped the
    # warehouse and routed in-stock POs onto far-future productions as "late".
    # Merge rule: warehouses/incoming SUM; committed/allocated are style-level
    # figures duplicated onto each row — keep the largest magnitude, never sum.
    inv_by_sku = {}
    for item in inv_items:
        sku = str(item.get('sku', '') or '').strip().upper()
        if sku not in sku_set:
            continue
        try:
            nj_qty = int(item.get('nj', 0) or 0)
            warehouse = (int(item.get('jtw', 0) or 0) + int(item.get('tr', 0) or 0)
                         + int(item.get('dcw', 0) or 0) + int(item.get('qa', 0) or 0)
                         + nj_qty)
        except (ValueError, TypeError):
            warehouse, nj_qty = 0, 0
        incoming = int(item.get('incoming', 0) or 0)
        committed = int(item.get('committed', 0) or 0)
        allocated = int(item.get('allocated', 0) or 0)
        m = inv_by_sku.get(sku)
        if m is None:
            # 'nj' rides along separately (still no per-warehouse breakdown of
            # the four factory-served warehouses): consumers need it to EXCLUDE
            # NJ from landed-batch suppression checks — 3PL stock never arrives
            # through a production line (audit fix).
            inv_by_sku[sku] = {'sku': sku, 'warehouse': warehouse, 'nj': nj_qty,
                               'incoming': incoming,
                               'committed': committed, 'allocated': allocated}
        else:
            m['warehouse'] += warehouse
            m['nj'] = m.get('nj', 0) + nj_qty
            m['incoming'] += incoming
            if abs(committed) > abs(m['committed']):
                m['committed'] = committed
            if abs(allocated) > abs(m['allocated']):
                m['allocated'] = allocated
    inventory_rows = list(inv_by_sku.values())

    # ── Suppression overrides (S3-backed exemption list, same source as
    # GET /suppression-overrides) — the frontend's arrival-suppression rule
    # needs these so an exempted SKU stays visible, matching the office view. ──
    load_suppression_overrides_from_s3()
    with _suppression_overrides_lock:
        # Scoped to this factory's SKUs — the full list is a catalog-wide
        # inventory of style codes and isn't a factory's business.
        suppression_overrides = [s for s in _suppression_overrides
                                 if str(s or '').strip().upper() in sku_set]

    return jsonify({
        'factory': {'code': code, 'name': factory_name},
        'generated_at': datetime.utcnow().isoformat() + 'Z',
        'transit': dict(FACTORY_VIEW_TRANSIT),
        'productions': productions,
        'external_supply': external_supply,
        'orders': orders,
        'apo': apo,
        'vw': vw,
        'inventory': inventory_rows,
        'suppression_overrides': suppression_overrides,
        'fob_customers': _fetch_fob_customers(),
        'orders_ok': orders_ok,
        # False during a cold start before the first inventory/ledger sync —
        # the page warns instead of implying the factory has no work.
        'inventory_ok': bool(inv_items),
        'ledger_ok': bool(ledger),
    })


# ─────────────────────────────────────────────────────────────────────────────
# FACTORY REPORTS  (Excel builder for the Open Order to PO page)
# The page computes the rows (it owns the routing engine); this builds the
# workbook, embedding the same 150x150 product thumbnails the catalog exports
# use. Same authorization rule as /factory-view: a factory account can only
# ever export its own factory, whatever the payload claims.
# NO price, value or margin field may ever be written here.
# ─────────────────────────────────────────────────────────────────────────────

# Header + status vocabulary per report language. English falls through to the
# key itself, so a missing translation degrades to English rather than blank.
_FR_T = {
    'zh': {
        'Report': '报告', 'Summary': '汇总', 'Image': '图片', 'Production #': '生产单号',
        'PO Name': '生产单名称', 'Factory': '工厂', 'Style': '款号', 'Brand': '品牌',
        'Units in production': '生产件数', 'Ex-Factory': '出厂日期', 'Arrival': '到仓日期',
        'Customer': '客户', 'Customer PO': '客户订单号', 'Units needed': '需求件数',
        'Start': '起始日期', 'Cancel': '截止日期', 'Status': '状态', 'Issue': '问题说明',
        'ON TIME': '准时', 'TIGHT': '紧张', 'LATE': '延误', 'SHORT': '短缺',
        'SHORT + LATE': '短缺且延误', 'NO ORDER': '暂无订单', 'APO': '预留', 'HOLD': '保留',
        'BULK': '期货订单', 'FOB PICKUP': 'FOB提货', 'ESTIMATE': '估算',
        'Open Order to PO': '订单对应生产单', 'Generated': '生成时间',
        'Productions': '生产订单', 'Style lines': '款式行', 'Units': '件数',
        'Late lines': '延误款', 'Short lines': '短缺款', 'Tight lines': '紧张款',
        'Report type': '报告类型', 'Language': '语言', 'Filters': '筛选条件',
        'Slack (days)': '可延误天数', 'Cumulative units': '累计件数', 'Ship by': '最迟出厂日',
        'days late': '天延误', 'units short': '件短缺', 'port-dated': '港口日期',
        'Arrives after the cancel date': '在截止日期之后到仓',
        'Arrives after the start date but before cancel': '在起始日之后、截止日之前到仓',
        'Not enough units in production to cover this order': '生产数量不足以覆盖此订单',
        'Short and arriving late': '数量短缺且到仓延误',
        'No customer order linked yet': '暂无关联客户订单',
        'Arrives in time': '按时到仓',
        'Allocation booking (APO)': '预留分配 (APO)',
        'Estimated allocation - not reconciled': '估算分配 — 未核对',
        'All factories': '全部工厂', 'Problem styles only': '仅问题款式',
        'Full line sheet': '完整款式表', 'By customer': '按客户',
    },
    'bn': {
        'Report': 'রিপোর্ট', 'Summary': 'সারসংক্ষেপ', 'Image': 'ছবি', 'Production #': 'প্রোডাকশন #',
        'PO Name': 'PO নাম', 'Factory': 'ফ্যাক্টরি', 'Style': 'স্টাইল', 'Brand': 'ব্র্যান্ড',
        'Units in production': 'প্রোডাকশন ইউনিট', 'Ex-Factory': 'এক্স-ফ্যাক্টরি', 'Arrival': 'পৌঁছানোর তারিখ',
        'Customer': 'কাস্টমার', 'Customer PO': 'কাস্টমার PO', 'Units needed': 'প্রয়োজনীয় ইউনিট',
        'Start': 'শুরু', 'Cancel': 'ক্যানসেল', 'Status': 'স্ট্যাটাস', 'Issue': 'সমস্যা',
        'ON TIME': 'সময়মতো', 'TIGHT': 'টাইট', 'LATE': 'লেট', 'SHORT': 'শর্ট',
        'SHORT + LATE': 'শর্ট + লেট', 'NO ORDER': 'অর্ডার নেই', 'APO': 'APO', 'HOLD': 'হোল্ড',
        'BULK': 'বাল্ক', 'FOB PICKUP': 'FOB পিকআপ', 'ESTIMATE': 'আনুমানিক',
        'Open Order to PO': 'ওপেন অর্ডার → PO', 'Generated': 'তৈরি হয়েছে',
        'Productions': 'প্রোডাকশন', 'Style lines': 'স্টাইল লাইন', 'Units': 'ইউনিট',
        'Late lines': 'লেট লাইন', 'Short lines': 'শর্ট লাইন', 'Tight lines': 'টাইট লাইন',
        'Report type': 'রিপোর্টের ধরন', 'Language': 'ভাষা', 'Filters': 'ফিল্টার',
        'Slack (days)': 'অতিরিক্ত দিন', 'Cumulative units': 'ক্রমযোজিত ইউনিট', 'Ship by': 'শিপ ডেডলাইন',
        'days late': 'দিন দেরি', 'units short': 'ইউনিট কম', 'port-dated': 'পোর্ট-তারিখ',
        'Arrives after the cancel date': 'ক্যানসেল তারিখের পরে পৌঁছায়',
        'Arrives after the start date but before cancel': 'শুরুর পরে কিন্তু ক্যানসেলের আগে পৌঁছায়',
        'Not enough units in production to cover this order': 'এই অর্ডারের জন্য যথেষ্ট ইউনিট নেই',
        'Short and arriving late': 'কম এবং দেরিতে পৌঁছাচ্ছে',
        'No customer order linked yet': 'এখনও কোনো কাস্টমার অর্ডার নেই',
        'Arrives in time': 'সময়মতো পৌঁছায়',
        'Allocation booking (APO)': 'অ্যালোকেশন বুকিং (APO)',
        'Estimated allocation - not reconciled': 'আনুমানিক বরাদ্দ — মিলানো হয়নি',
        'All factories': 'সব ফ্যাক্টরি', 'Problem styles only': 'শুধু সমস্যার স্টাইল',
        'Full line sheet': 'সম্পূর্ণ লাইন শিট', 'By customer': 'কাস্টমার অনুযায়ী',
    },
    'ko': {
        'Report': '리포트', 'Summary': '요약', 'Image': '이미지', 'Production #': '생산 번호',
        'PO Name': 'PO 이름', 'Factory': '공장', 'Style': '스타일', 'Brand': '브랜드',
        'Units in production': '생산 수량', 'Ex-Factory': '출고일', 'Arrival': '도착일',
        'Customer': '고객', 'Customer PO': '고객 PO', 'Units needed': '필요 수량',
        'Start': '시작일', 'Cancel': '마감일', 'Status': '상태', 'Issue': '문제',
        'ON TIME': '정상', 'TIGHT': '빠듯', 'LATE': '지연', 'SHORT': '부족',
        'SHORT + LATE': '부족 + 지연', 'NO ORDER': '주문 없음', 'APO': 'APO', 'HOLD': '홀드',
        'BULK': '벌크', 'FOB PICKUP': 'FOB 인수', 'ESTIMATE': '추정',
        'Open Order to PO': '오픈 오더 → PO', 'Generated': '생성',
        'Productions': '생산 오더', 'Style lines': '스타일 라인', 'Units': '수량',
        'Late lines': '지연 라인', 'Short lines': '부족 라인', 'Tight lines': '빠듯한 라인',
        'Report type': '리포트 유형', 'Language': '언어', 'Filters': '필터',
        'Slack (days)': '여유 일수', 'Cumulative units': '누적 수량', 'Ship by': '선적 기한',
        'days late': '일 지연', 'units short': '수량 부족', 'port-dated': '항구 일정',
        'Arrives after the cancel date': '마감일 이후 도착',
        'Arrives after the start date but before cancel': '시작일 이후, 마감일 이전 도착',
        'Not enough units in production to cover this order': '이 주문을 채울 생산 수량 부족',
        'Short and arriving late': '수량 부족 및 지연 도착',
        'No customer order linked yet': '연결된 고객 주문 없음',
        'Arrives in time': '기한 내 도착',
        'Allocation booking (APO)': '할당 예약 (APO)',
        'Estimated allocation - not reconciled': '추정 배분 — 대조되지 않음',
        'All factories': '전체 공장', 'Problem styles only': '문제 스타일만',
        'Full line sheet': '전체 라인시트', 'By customer': '고객별',
    },
    'vi': {
        'Report': 'Báo cáo', 'Summary': 'Tổng hợp', 'Image': 'Hình', 'Production #': 'Số PO sản xuất',
        'PO Name': 'Tên PO', 'Factory': 'Nhà máy', 'Style': 'Mã hàng', 'Brand': 'Thương hiệu',
        'Units in production': 'SL sản xuất', 'Ex-Factory': 'Ngày xuất xưởng', 'Arrival': 'Ngày đến kho',
        'Customer': 'Khách hàng', 'Customer PO': 'PO khách hàng', 'Units needed': 'SL cần',
        'Start': 'Ngày bắt đầu', 'Cancel': 'Ngày hủy', 'Status': 'Trạng thái', 'Issue': 'Vấn đề',
        'ON TIME': 'ĐÚNG HẠN', 'TIGHT': 'SÁT HẠN', 'LATE': 'TRỄ', 'SHORT': 'THIẾU',
        'SHORT + LATE': 'THIẾU + TRỄ', 'NO ORDER': 'CHƯA CÓ ĐƠN', 'APO': 'APO', 'HOLD': 'GIỮ',
        'BULK': 'BULK', 'FOB PICKUP': 'FOB nhận tại xưởng', 'ESTIMATE': 'Ước tính',
        'Open Order to PO': 'Đơn hàng → PO', 'Generated': 'Tạo lúc',
        'Productions': 'Đơn sản xuất', 'Style lines': 'Dòng mã hàng', 'Units': 'Số lượng',
        'Late lines': 'Dòng trễ', 'Short lines': 'Dòng thiếu', 'Tight lines': 'Dòng sát hạn',
        'Report type': 'Loại báo cáo', 'Language': 'Ngôn ngữ', 'Filters': 'Bộ lọc',
        'Slack (days)': 'Số ngày dư', 'Cumulative units': 'SL lũy kế', 'Ship by': 'Hạn xuất xưởng',
        'days late': 'ngày trễ', 'units short': 'thiếu', 'port-dated': 'theo lịch cảng',
        'Arrives after the cancel date': 'Đến sau ngày hủy',
        'Arrives after the start date but before cancel': 'Đến sau ngày bắt đầu, trước ngày hủy',
        'Not enough units in production to cover this order': 'Không đủ số lượng sản xuất cho đơn này',
        'Short and arriving late': 'Thiếu và đến trễ',
        'No customer order linked yet': 'Chưa có đơn khách hàng',
        'Arrives in time': 'Đến đúng hạn',
        'Allocation booking (APO)': 'Đặt chỗ phân bổ (APO)',
        'Estimated allocation - not reconciled': 'Phân bổ ước tính — chưa đối chiếu',
        'All factories': 'Tất cả nhà máy', 'Problem styles only': 'Chỉ mã có vấn đề',
        'Full line sheet': 'Line sheet đầy đủ', 'By customer': 'Theo khách hàng',
    },
}


def _frt(lang, s):
    """Translate a report string; unknown language or key falls back to English."""
    return _FR_T.get(lang, {}).get(s, s)


_FR_STATUS_LABEL = {
    'ok': 'ON TIME', 'confirm': 'TIGHT', 'late': 'LATE', 'short': 'SHORT',
    'short_late': 'SHORT + LATE', 'none': 'NO ORDER', 'apo': 'APO', 'hold': 'HOLD',
}
_FR_STATUS_REASON = {
    'late': 'Arrives after the cancel date',
    'confirm': 'Arrives after the start date but before cancel',
    'short': 'Not enough units in production to cover this order',
    'short_late': 'Short and arriving late',
    'none': 'No customer order linked yet',
    'ok': 'Arrives in time',
    'apo': 'Allocation booking (APO)',
    'hold': 'Allocation booking (APO)',
}
# Severity drives sort order, tab colour and the summary counts.
_FR_SEVERITY = {'short_late': 0, 'late': 1, 'short': 2, 'confirm': 3,
                'ok': 4, 'apo': 5, 'hold': 5, 'none': 6}


def _fr_sheet_name(raw, used):
    """Excel sheet names: <=31 chars, no \\ / ? * [ ] :, unique within the book."""
    safe = re.sub(r'[\\/*?\[\]:]', '-', str(raw or '')).strip() or 'Sheet'
    safe = safe[:31]
    if safe not in used:
        used.add(safe)
        return safe
    for n in range(2, 200):
        suffix = f'~{n}'
        cand = safe[:31 - len(suffix)] + suffix
        if cand not in used:
            used.add(cand)
            return cand
    used.add(safe[:28] + '~xx')
    return safe[:28] + '~xx'


def _factory_report_build(req, rows, code, lang, split, want_images, job=None):
    """Build the Open Order to PO workbook. Runs in a background thread for
    big reports (a full all-factories build with ~1,600 embedded swatches
    takes minutes — far past the 100s edge / 120s worker limits that killed
    synchronous builds). `job`, when given, receives stage updates.

    Returns (xlsx_bytes, download_name). All authorization and scope
    enforcement happens in the route BEFORE this is called."""
    T = lambda s: _frt(lang, s)
    factory_name = T('All factories') if code == 'ALL' else FACTORY_NAMES.get(code, code)
    t0 = time.time()
    def _stage(s):
        if job is not None:
            job['stage'] = s
            if job.get('jid'):   # keep the S3 ground truth's stage current too
                _report_job_s3_put_status(job['jid'], {
                    'status': 'building', 'stage': s, 'created': job['created']})

    # ── Images: one fetch per unique style, reused across every sheet ──
    img_by_style = {}
    if want_images:
        _stage('images')
        try:
            uniq = []
            seen_styles = set()
            for r in rows:
                s = str(r.get('style') or '').upper()
                if s and s not in seen_styles:
                    seen_styles.add(s)
                    # brand_abbr feeds the brand-folder fallback in the image
                    # resolution chain (extract_image_code); the ledger's brand
                    # column is already the abbreviation.
                    b = str(r.get('brand') or '').strip()
                    uniq.append({'sku': s, 'brand': b, 'brand_abbr': b})
            if uniq:
                fetched = download_images_for_items(uniq, S3_PHOTOS_URL, use_cache=True)
                for i, item in enumerate(uniq):
                    if i in fetched:
                        img_by_style[item['sku']] = fetched[i]
            print(f"[FactoryReport] images: {len(img_by_style)}/{len(uniq)} in "
                  f"{time.time() - t0:.1f}s", flush=True)
        except Exception as e:
            print(f"[FactoryReport] image fetch failed: {e}", flush=True)

    _stage('workbook')
    buf = BytesIO()
    wb = xlsxwriter.Workbook(buf, {'in_memory': True, 'strings_to_formulas': False})

    # CJK/Bengali need a font with the right coverage or Excel shows boxes.
    base_font = {'zh': 'Microsoft YaHei', 'ko': 'Malgun Gothic',
                 'bn': 'Nirmala UI', 'vi': 'Segoe UI'}.get(lang, 'Calibri')

    f_title = wb.add_format({'bold': True, 'font_size': 18, 'font_name': base_font,
                             'font_color': '#16181D'})
    f_sub = wb.add_format({'font_size': 11, 'font_name': base_font, 'font_color': '#5B6068'})
    f_kpi_n = wb.add_format({'bold': True, 'font_size': 20, 'font_name': base_font,
                             'align': 'center', 'font_color': '#16181D'})
    f_kpi_l = wb.add_format({'font_size': 9, 'font_name': base_font, 'align': 'center',
                             'font_color': '#5B6068'})
    f_hdr = wb.add_format({'bold': True, 'font_size': 10, 'font_name': base_font,
                           'bg_color': '#16181D', 'font_color': '#FFFFFF',
                           'align': 'center', 'valign': 'vcenter', 'text_wrap': True,
                           'border': 1, 'border_color': '#16181D'})
    f_cell = wb.add_format({'font_size': 10, 'font_name': base_font, 'valign': 'vcenter',
                            'border': 1, 'border_color': '#E7E8EC', 'text_wrap': True})
    f_num = wb.add_format({'font_size': 10, 'font_name': base_font, 'valign': 'vcenter',
                           'border': 1, 'border_color': '#E7E8EC', 'num_format': '#,##0',
                           'align': 'right'})
    f_mono = wb.add_format({'font_size': 10, 'font_name': 'Consolas', 'valign': 'vcenter',
                            'border': 1, 'border_color': '#E7E8EC'})
    f_date = wb.add_format({'font_size': 10, 'font_name': base_font, 'valign': 'vcenter',
                            'border': 1, 'border_color': '#E7E8EC', 'align': 'center'})
    f_dateval = wb.add_format({'font_size': 10, 'font_name': base_font, 'valign': 'vcenter',
                               'border': 1, 'border_color': '#E7E8EC', 'align': 'center',
                               'num_format': 'yyyy-mm-dd'})
    # Status colours mirror the web page and stay legible printed in mono.
    _STATUS_FMT_SPEC = {
        'late':       ('#FEF3F2', '#B42318'), 'short_late': ('#FEE4E2', '#912018'),
        'short':      ('#FFF8E6', '#9A6700'), 'confirm':    ('#FFF6ED', '#B54708'),
        'ok':         ('#EAF6EE', '#0A7D3C'), 'apo':        ('#EFF8FF', '#175CD3'),
        'hold':       ('#EFF8FF', '#175CD3'), 'none':       ('#F7F8FA', '#5B6068'),
    }
    f_status = {k: wb.add_format({'font_size': 10, 'font_name': base_font, 'bold': True,
                                  'align': 'center', 'valign': 'vcenter', 'border': 1,
                                  'border_color': '#E7E8EC', 'bg_color': bg, 'font_color': fg})
                for k, (bg, fg) in _STATUS_FMT_SPEC.items()}

    # "Slack" is how many days this production can slip before the tightest
    # customer PO on it misses its cancel date — the single most actionable
    # number a factory can have. In tab-per-production mode the sheet is a
    # cut-split ticket, so it also carries a running unit total.
    per_production = (split == 'production')
    HEADERS = ['Image', 'Production #', 'PO Name', 'Factory', 'Style', 'Brand',
               'Units in production', 'Ex-Factory', 'Arrival', 'Slack (days)',
               'Customer', 'Customer PO', 'Units needed']
    WIDTHS = [COL_WIDTH_UNITS, 16, 26, 9, 18, 16, 13, 13, 13, 11, 22, 16, 12]
    if per_production:
        HEADERS += ['Cumulative units']
        WIDTHS += [13]
    HEADERS += ['Start', 'Cancel', 'Ship by', 'Status', 'Issue']
    WIDTHS += [12, 12, 12, 15, 42]
    if not want_images:
        HEADERS = HEADERS[1:]
        WIDTHS = WIDTHS[1:]
    NUM_COLS = {'Units in production', 'Units needed', 'Cumulative units', 'Slack (days)'}
    DATE_COLS = {'Ex-Factory', 'Arrival', 'Start', 'Cancel', 'Ship by'}
    MONO_COLS = {'Production #', 'Style', 'Customer PO'}

    def _as_date(v):
        """'YYYY-MM-DD' -> datetime so Excel sorts and filters it as a real
        date. Anything else is written through as text."""
        try:
            return datetime.strptime(str(v)[:10], '%Y-%m-%d')
        except Exception:
            return None

    def _issue_text(r):
        st = r.get('status') or 'ok'
        base = T(_FR_STATUS_REASON.get(st, ''))
        bits = []
        if st in ('late', 'short_late') and r.get('days'):
            bits.append(f"{int(r['days'])} {T('days late')}")
        if st in ('short', 'short_late') and r.get('gap'):
            bits.append(f"{int(r['gap'])} {T('units short')}")
        if r.get('fobPickup'):
            bits.append(T('FOB PICKUP'))
        if r.get('bulk'):
            bits.append(T('BULK'))
        if r.get('portDated'):
            bits.append(T('port-dated'))
        if r.get('est'):
            bits.append(T('ESTIMATE'))
        return base + (' · ' + ' · '.join(bits) if bits else '')

    def _write_sheet(ws, sheet_rows):
        for c, h in enumerate(HEADERS):
            ws.write(0, c, T(h), f_hdr)
        ws.set_row(0, 30)
        for c, w in enumerate(WIDTHS):
            ws.set_column(c, c, w)
        ws.freeze_panes(1, 0)
        ws.autofilter(0, 0, max(len(sheet_rows), 1), len(HEADERS) - 1)
        # Print setup: landscape, one page wide, header row repeated.
        ws.set_landscape()
        ws.fit_to_pages(1, 0)
        ws.repeat_rows(0)
        ws.set_margins(0.3, 0.3, 0.4, 0.4)
        ws.set_footer('&L' + str(T('Open Order to PO')) + ' · ' + str(factory_name) + '&R&P/&N')

        running = 0
        for i, r in enumerate(sheet_rows):
            row = i + 1
            if want_images:
                ws.set_row(row, 112.5)
            st = r.get('status') or 'ok'
            # Cumulative units answers "if only part of the cut is finished by
            # ex-factory, how far down this list can we ship?"
            if per_production and st not in ('none',):
                try:
                    running += int(r.get('unitsNeeded') or 0)
                except (TypeError, ValueError):
                    pass
            vals = {
                'Production #': r.get('production') or '',
                'PO Name': r.get('poName') or '',
                'Factory': (r.get('factoryCode') or _factory_label(r.get('production')) or ''),
                'Style': r.get('style') or '',
                'Brand': r.get('brand') or '',
                'Units in production': r.get('units') or 0,
                'Ex-Factory': r.get('etd') or '',
                'Arrival': r.get('arrival') or '',
                'Slack (days)': r.get('slackDays') if r.get('slackDays') != '' else '',
                'Customer': r.get('customer') or '',
                'Customer PO': r.get('orderNo') or '',
                'Units needed': r.get('unitsNeeded') or '',
                'Cumulative units': running if per_production else '',
                'Start': r.get('start') or '',
                'Cancel': r.get('cancel') or '',
                'Ship by': r.get('shipBy') or '',
                'Status': T(_FR_STATUS_LABEL.get(st, st)),
                'Issue': _issue_text(r),
            }
            for c, h in enumerate(HEADERS):
                if h == 'Image':
                    continue
                v = vals.get(h, '')
                if h == 'Status':
                    ws.write(row, c, v, f_status.get(st, f_cell))
                elif h in DATE_COLS:
                    d = _as_date(v)
                    if d:
                        ws.write_datetime(row, c, d, f_dateval)
                    else:
                        ws.write(row, c, v or '', f_date)
                elif h in NUM_COLS:
                    ws.write(row, c, v if v != '' else '', f_num)
                elif h in MONO_COLS:
                    ws.write(row, c, v, f_mono)
                else:
                    ws.write(row, c, v, f_cell)
            if want_images:
                img = img_by_style.get(str(r.get('style') or '').upper())
                if img and img.get('image_data'):
                    try:
                        bio = img['image_data']
                        if hasattr(bio, 'seek'):
                            bio.seek(0)
                        ws.insert_image(row, 0, 'img.png', _padded_image_opts(img))
                    except Exception:
                        ws.write(row, 0, '', f_cell)
                else:
                    ws.write(row, 0, '', f_cell)

    # ── Summary / cover sheet ──
    counts = {}
    for r in rows:
        st = r.get('status') or 'ok'
        counts[st] = counts.get(st, 0) + 1
    total_units = sum(int(r.get('units') or 0) for r in rows)
    n_prod = len({r.get('production') for r in rows if r.get('production')})

    ws_sum = wb.add_worksheet(_fr_sheet_name(T('Summary'), set()))
    ws_sum.hide_gridlines(2)
    ws_sum.set_column(0, 0, 3)
    ws_sum.set_column(1, 6, 20)
    ws_sum.write(1, 1, req.get('title') or T('Open Order to PO'), f_title)
    ws_sum.write(2, 1, f"{T('Factory')}: {factory_name}", f_sub)
    ws_sum.write(3, 1, f"{T('Generated')}: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC", f_sub)
    if req.get('filters'):
        ws_sum.write(4, 1, f"{T('Filters')}: {req.get('filters')}", f_sub)
    kpis = [(T('Productions'), n_prod), (T('Style lines'), len(rows)), (T('Units'), total_units),
            (T('Late lines'), counts.get('late', 0) + counts.get('short_late', 0)),
            (T('Short lines'), counts.get('short', 0) + counts.get('short_late', 0)),
            (T('Tight lines'), counts.get('confirm', 0))]
    for i, (label, val) in enumerate(kpis):
        ws_sum.write(6, 1 + i, val, f_kpi_n)
        ws_sum.write(7, 1 + i, label, f_kpi_l)

    # Legend — so anyone opening the file cold knows what the colours mean
    # without being told, and so a mono printout is still readable.
    f_leg = wb.add_format({'font_size': 10, 'font_name': base_font, 'font_color': '#5B6068'})
    ws_sum.write(10, 1, T('Status'), wb.add_format({'bold': True, 'font_name': base_font,
                                                    'font_size': 11}))
    LEGEND = [('late', 'Arrives after the cancel date'),
              ('short_late', 'Short and arriving late'),
              ('short', 'Not enough units in production to cover this order'),
              ('confirm', 'Arrives after the start date but before cancel'),
              ('ok', 'Arrives in time'), ('none', 'No customer order linked yet')]
    for i, (stk, desc) in enumerate(LEGEND):
        ws_sum.write(11 + i, 1, T(_FR_STATUS_LABEL.get(stk, stk)), f_status.get(stk, f_cell))
        ws_sum.write(11 + i, 2, T(desc), f_leg)
    ws_sum.set_landscape()

    # ── Data sheets ──
    sev = lambda r: _FR_SEVERITY.get(r.get('status') or 'ok', 9)
    used_names = {ws_sum.get_name()}
    if split == 'none':
        ws = wb.add_worksheet(_fr_sheet_name(T('Report'), used_names))
        _write_sheet(ws, sorted(rows, key=lambda r: (sev(r), str(r.get('production') or ''),
                                                     str(r.get('style') or ''))))
    else:
        key = (lambda r: r.get('production') or '-') if split == 'production' \
            else (lambda r: r.get('customer') or '-')
        groups = {}
        for r in rows:
            groups.setdefault(key(r), []).append(r)
        # Worst groups first so the urgent tabs sit nearest the summary.
        ordered = sorted(groups.items(),
                         key=lambda kv: (min((sev(r) for r in kv[1]), default=9), str(kv[0])))
        shown = ordered[:180]           # Excel tolerates more, humans do not
        links = []
        for name, grp in shown:
            ws = wb.add_worksheet(_fr_sheet_name(name, used_names))
            worst = min((sev(r) for r in grp), default=9)
            ws.set_tab_color({0: '#B42318', 1: '#B42318', 2: '#9A6700',
                              3: '#B54708', 4: '#0A7D3C'}.get(worst, '#98A2B3'))
            # Tab-per-production doubles as a cut-split ticket: order by the
            # customer deadline, so the top of the sheet ships first.
            if per_production:
                grp_sorted = sorted(grp, key=lambda r: (str(r.get('cancel') or '9999-12-31'),
                                                        sev(r), str(r.get('style') or '')))
            else:
                grp_sorted = sorted(grp, key=lambda r: (sev(r), str(r.get('style') or '')))
            _write_sheet(ws, grp_sorted)
            links.append((ws.get_name(), name, len(grp), worst))
        # Contents list on the cover, hyperlinked to each tab.
        f_link = wb.add_format({'font_size': 10, 'font_name': base_font,
                                'font_color': '#1D4ED8', 'underline': 1})
        base_row = 12 + len(LEGEND)
        ws_sum.write(base_row, 1, T('Report'), wb.add_format({'bold': True, 'font_name': base_font,
                                                              'font_size': 11}))
        for i, (sheet, label, n, worst) in enumerate(links):
            rr = base_row + 1 + i
            ws_sum.write_url(rr, 1, f"internal:'{sheet}'!A1", f_link, str(label))
            ws_sum.write(rr, 2, n, f_leg)
        if len(ordered) > len(shown):
            ws_sum.write(base_row + 1 + len(links), 1,
                         f"+{len(ordered) - len(shown)} more (narrow the filters)", f_leg)

    wb.close()
    buf.seek(0)
    data = buf.read()
    print(f"[FactoryReport] built {code} lang={lang} split={split} rows={len(rows)} "
          f"images={len(img_by_style)} -> {len(data) // 1024}KB in {time.time() - t0:.1f}s",
          flush=True)
    stamp = datetime.utcnow().strftime('%Y-%m-%d')
    return data, f"open-order-to-po_{code}_{stamp}.xlsx"


# Job store: in-memory fast path PLUS S3 ground truth. This service runs as
# multiple processes/instances in practice (learned the hard way with the
# overrides extract-images job, Jul 2026): the POST that creates a job and the
# poll that checks it routinely land on different processes, and a purely
# in-memory store answers "expired" to a job that is building fine elsewhere.
# Status and the finished workbook therefore live in S3; memory only
# short-circuits the same-process case.
_report_jobs = {}
_report_jobs_lock = threading.Lock()
_REPORT_JOB_TTL = 1800
_REPORT_S3_PREFIX = 'inventory/factory_report_jobs/'


def _report_job_s3_put_status(jid, d):
    try:
        get_s3().put_object(Bucket=S3_BUCKET, Key=f'{_REPORT_S3_PREFIX}{jid}.status.json',
                            Body=json.dumps(d).encode('utf-8'),
                            ContentType='application/json')
    except Exception as e:
        print(f"[FactoryReport] S3 status write failed for {jid}: {e}", flush=True)


def _report_job_s3_get_status(jid):
    try:
        resp = get_s3().get_object(Bucket=S3_BUCKET,
                                   Key=f'{_REPORT_S3_PREFIX}{jid}.status.json')
        return json.loads(resp['Body'].read().decode('utf-8'))
    except Exception:
        return None


def _report_job_s3_put_file(jid, data):
    get_s3().put_object(Bucket=S3_BUCKET, Key=f'{_REPORT_S3_PREFIX}{jid}.xlsx', Body=data,
                        ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def _report_job_s3_get_file(jid):
    try:
        resp = get_s3().get_object(Bucket=S3_BUCKET, Key=f'{_REPORT_S3_PREFIX}{jid}.xlsx')
        return resp['Body'].read()
    except Exception:
        return None


def _prune_report_jobs():
    now = time.time()
    with _report_jobs_lock:
        for jid in [j for j, v in _report_jobs.items() if now - v['created'] > _REPORT_JOB_TTL]:
            _report_jobs.pop(jid, None)
    # Best-effort S3 cleanup of finished/abandoned jobs older than 2 hours.
    try:
        s3 = get_s3()
        listed = s3.list_objects_v2(Bucket=S3_BUCKET, Prefix=_REPORT_S3_PREFIX, MaxKeys=200)
        from datetime import timezone as _tz
        cutoff = datetime.now(_tz.utc).timestamp() - 7200
        stale = [o['Key'] for o in listed.get('Contents', [])
                 if o['LastModified'].timestamp() < cutoff]
        if stale:
            s3.delete_objects(Bucket=S3_BUCKET,
                              Delete={'Objects': [{'Key': k} for k in stale[:100]]})
    except Exception:
        pass


@app.route('/factory-report', methods=['POST', 'OPTIONS'])
def factory_report():
    """Excel report builder for the Versa-Docs Open Order to PO page.

    Asynchronous: big builds (all factories, ~1,600 embedded swatches) take
    minutes, far past the edge/worker timeouts — so this returns {'job': id}
    immediately and a background thread builds the workbook. Poll
    GET /factory-report/job/<id>, then GET /factory-report/job/<id>/download.
    Payloads without 'async' keep the old synchronous behaviour for small
    reports and backwards compatibility.

    POST body: factory, lang, title, split, images, filters, async,
    rows=[{production, poName, factoryCode, style, brand, units, etd,
    arrival, portDated, fob, customer, orderNo, unitsNeeded, start, cancel,
    shipBy, status, days, gap, bulk, fobPickup, est}].
    Whitelisted by construction — no price or value field is ever read.

    AUTHORIZATION: same rule as /factory-view — a factory account is pinned
    to its own prefix regardless of the payload; staff may build any."""
    if request.method == 'OPTIONS':
        return '', 204

    auth = request.headers.get('Authorization', '')
    token = auth[7:].strip() if auth.lower().startswith('bearer ') else ''
    ident = _caller_identity(token)
    if not ident:
        return jsonify({'error': 'Sign in to Versa Docs to build a report.'}), 401

    try:
        req = request.get_json(force=True) or {}
    except Exception:
        return jsonify({'error': 'Invalid JSON body'}), 400

    rows = req.get('rows') or []
    if not isinstance(rows, list) or not rows:
        return jsonify({'error': 'No rows to export'}), 400
    if len(rows) > 20000:
        return jsonify({'error': 'Report too large — narrow the filters'}), 413

    lang = str(req.get('lang') or 'en').lower()
    if lang not in ('en', 'zh', 'bn', 'ko', 'vi'):
        lang = 'en'
    split = str(req.get('split') or 'none').lower()
    if split not in ('none', 'production', 'customer'):
        split = 'none'
    want_images = bool(req.get('images', True))

    # ── Scope enforcement (server-side, same rule as /factory-view) ──
    requested = str(req.get('factory') or 'ALL').strip().upper()
    if ident['role'] == 'factory':
        code = ident['factory_prefix']
        if code not in FACTORY_NAMES:
            return jsonify({'error': 'Your account is not linked to a factory yet.'}), 403
        # Never trust client-side filtering for another factory's rows.
        rows = [r for r in rows
                if str(r.get('factoryCode') or _factory_label(r.get('production'))).upper() == code]
        if not rows:
            return jsonify({'error': 'No rows for your factory'}), 400
    elif ident['role'] == 'staff':
        code = requested if (requested == 'ALL' or requested in FACTORY_NAMES) else 'ALL'
    else:
        return jsonify({'error': 'Your account does not have access to this data.'}), 403

    if not req.get('async'):
        # Legacy synchronous path — fine for small no-image pulls.
        data, fname = _factory_report_build(req, rows, code, lang, split, want_images)
        return send_file(BytesIO(data), as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    _prune_report_jobs()
    with _report_jobs_lock:
        building = sum(1 for v in _report_jobs.values() if v['status'] == 'building')
    if building >= 3:
        return jsonify({'error': 'Too many reports building right now — try again in a minute.'}), 429

    jid = uuid.uuid4().hex
    job = {'jid': jid, 'status': 'building', 'stage': 'queued', 'created': time.time(),
           'data': None, 'fname': None, 'error': None}
    with _report_jobs_lock:
        _report_jobs[jid] = job
    # Ground truth BEFORE returning the id — the next poll may hit another
    # process, and it must find the job there.
    _report_job_s3_put_status(jid, {'status': 'building', 'stage': 'queued',
                                    'created': job['created']})

    def _run():
        try:
            data, fname = _factory_report_build(req, rows, code, lang, split,
                                                want_images, job=job)
            job['data'] = data
            job['fname'] = fname
            _report_job_s3_put_file(jid, data)
            job['status'] = 'done'
            _report_job_s3_put_status(jid, {'status': 'done', 'stage': 'done',
                                            'created': job['created'], 'fname': fname})
        except Exception as e:
            import traceback as _tb
            _tb.print_exc()
            job['error'] = f'{type(e).__name__}: {e}'
            job['status'] = 'error'
            _report_job_s3_put_status(jid, {'status': 'error', 'stage': 'error',
                                            'created': job['created'],
                                            'error': job['error']})

    threading.Thread(target=_run, daemon=True).start()
    return jsonify({'job': jid})


@app.route('/factory-report/job/<jid>', methods=['GET', 'OPTIONS'])
def factory_report_job(jid):
    if request.method == 'OPTIONS':
        return '', 204
    auth = request.headers.get('Authorization', '')
    token = auth[7:].strip() if auth.lower().startswith('bearer ') else ''
    if not _caller_identity(token):
        return jsonify({'error': 'Sign in to Versa Docs.'}), 401
    if not re.fullmatch(r'[0-9a-f]{32}', jid or ''):
        return jsonify({'error': 'Bad job id.'}), 400
    with _report_jobs_lock:
        job = _report_jobs.get(jid)
    if job:
        return jsonify({'status': job['status'], 'stage': job['stage'], 'error': job['error']})
    # Not ours — another process may own it. S3 is the ground truth.
    s3job = _report_job_s3_get_status(jid)
    if not s3job:
        return jsonify({'error': 'Report expired or the service restarted — build it again.'}), 404
    # A build whose process died mid-way stays 'building' in S3 forever —
    # surface that honestly instead of letting the page poll into its timeout.
    if s3job.get('status') == 'building' and time.time() - (s3job.get('created') or 0) > 900:
        return jsonify({'status': 'error', 'stage': 'stale',
                        'error': 'The build was interrupted — build it again.'})
    return jsonify({'status': s3job.get('status'), 'stage': s3job.get('stage'),
                    'error': s3job.get('error')})


@app.route('/factory-report/job/<jid>/download', methods=['GET', 'OPTIONS'])
def factory_report_download(jid):
    if request.method == 'OPTIONS':
        return '', 204
    auth = request.headers.get('Authorization', '')
    token = auth[7:].strip() if auth.lower().startswith('bearer ') else ''
    if not _caller_identity(token):
        return jsonify({'error': 'Sign in to Versa Docs.'}), 401
    if not re.fullmatch(r'[0-9a-f]{32}', jid or ''):
        return jsonify({'error': 'Bad job id.'}), 400
    with _report_jobs_lock:
        job = _report_jobs.get(jid)
    if job and job['status'] == 'done' and job['data']:
        return send_file(BytesIO(job['data']), as_attachment=True, download_name=job['fname'],
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # Cross-process download: the workbook lives in S3.
    s3job = _report_job_s3_get_status(jid)
    if s3job and s3job.get('status') == 'done':
        data = _report_job_s3_get_file(jid)
        if data:
            return send_file(BytesIO(data), as_attachment=True,
                             download_name=s3job.get('fname') or 'report.xlsx',
                             mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    return jsonify({'error': 'Report not ready.'}), 404


# ============================================================
# AI AGENT — server-side tool loop for the platform chatbot
# ============================================================
# The old /api/ai-proxy is a one-shot passthrough: the model gets a big
# snapshot and answers blind. This endpoint runs a real agentic loop —
# the model calls tools that query the LIVE in-memory data (inventory,
# production ledger, APO, open orders) and can build line sheet files,
# iterating until it has a grounded answer. The frontend chat keeps its
# existing response contract: the final assistant text passes through in
# an Anthropic-shaped envelope, so the client parser is unchanged.
#
# Uses the official anthropic SDK (lazy import so a missing package can
# never kill boot — the endpoint 503s instead; lesson of the packaging
# deploy failure). Model: claude-opus-5 (thinking on by default there;
# effort medium keeps chat latency sane — override via env).

_ai_agent_sdk_client = None


def _ai_agent_client():
    global _ai_agent_sdk_client
    if _ai_agent_sdk_client is None:
        import anthropic  # lazy: boot survives a missing package
        key = os.environ.get('ANTHROPIC_API_KEY') or os.environ.get('CLAUDE_API_KEY')
        if not key:
            raise RuntimeError('ANTHROPIC_API_KEY not configured')
        _ai_agent_sdk_client = anthropic.Anthropic(api_key=key)
    return _ai_agent_sdk_client


AI_AGENT_MODEL = os.environ.get('AI_AGENT_MODEL', 'claude-opus-5')
AI_AGENT_EFFORT = os.environ.get('AI_AGENT_EFFORT', 'medium')
_AI_AGENT_MAX_ITER = int(os.environ.get('AI_AGENT_MAX_ITER', '8'))
_AI_AGENT_WALL_SECONDS = int(os.environ.get('AI_AGENT_WALL_SECONDS', '110'))
_AI_AGENT_SELF_URL = os.environ.get('RENDER_EXTERNAL_URL', 'https://versa-inventory-api.onrender.com').rstrip('/')


def _ai_agent_base(sku):
    return str(sku or '').split('-')[0].strip().upper()


def _ai_agent_agg_inventory():
    """Aggregate the per-size feed rows to one record per base style.
    Physical quantities SUM across size rows; committed/allocated are
    style-level NEGATIVE figures duplicated onto each size row, so keep
    the largest magnitude — never sum (the BUGBSA002SLS lesson)."""
    with _inv_lock:
        items = list(_inventory.get('items') or [])
    agg = {}
    for it in items:
        base = _ai_agent_base(it.get('sku'))
        if not base:
            continue
        r = agg.get(base)
        if r is None:
            r = {'style': base, 'brand_abbr': (it.get('brand_abbr') or it.get('brand') or '').upper(),
                 'jtw': 0, 'tr': 0, 'dcw': 0, 'qa': 0, 'nj': 0, 'incoming': 0, 'total_ats': 0,
                 'committed': 0, 'allocated': 0, 'size_rows': 0}
            agg[base] = r
        for k in ('jtw', 'tr', 'dcw', 'qa', 'nj', 'incoming', 'total_ats'):
            try:
                r[k] += int(it.get(k) or 0)
            except (TypeError, ValueError):
                pass
        for k in ('committed', 'allocated'):
            try:
                v = int(it.get(k) or 0)
            except (TypeError, ValueError):
                v = 0
            if abs(v) > abs(r[k]):
                r[k] = v
        r['size_rows'] += 1
    for r in agg.values():
        r['total_warehouse'] = r['jtw'] + r['tr'] + r['dcw'] + r['qa'] + r.get('nj', 0)
    return agg


def _ai_agent_prod_for(base, prods=None):
    """Production ledger rows for one base style with effective arrivals."""
    if prods is None:
        with _production_lock:
            prods = list(_production_data)
    base = base.upper()
    out = []
    for p in prods:
        if _ai_agent_base(p.get('style')) != base:
            continue
        pants = False
        try:
            pants = _py_is_bottom(base)
        except Exception:
            pass
        arr = None
        try:
            arr = _apo_prod_arrival(p, pants)
        except Exception:
            pass
        out.append({
            'production': p.get('production'), 'po_name': p.get('poName'),
            'units': p.get('units'), 'etd': p.get('etd'),
            'arrival': str(arr) if arr else ('FOB/' + str(p.get('fob_note') or 'TBD') if p.get('fob_flag') else 'TBD'),
            'port_dated': bool(p.get('port_dated')),
        })
    return out


def _ai_agent_filter(params):
    """Shared filter over aggregated inventory. Returns (rows, prods_snapshot).
    Rows are enriched with color/fit/fabrication lazily AFTER filtering."""
    agg = _ai_agent_agg_inventory()
    with _production_lock:
        prods = list(_production_data)

    brands = params.get('brands') or ([params['brand']] if params.get('brand') else [])
    brand_keys = set()
    for b in brands:
        bu = str(b or '').strip().upper()
        if not bu:
            continue
        brand_keys.add(bu)
        for abbr, full in (BRAND_FULL_NAMES or {}).items():
            if str(full).upper() == bu:
                brand_keys.add(str(abbr).upper())
    category = (params.get('category') or '').strip().lower() or None
    fabric_codes = {str(f).strip().upper() for f in (params.get('fabric_codes') or []) if str(f).strip()}
    color_q = (params.get('color') or '').strip().lower() or None
    search = (params.get('search') or '').strip().upper() or None
    stock = (params.get('stock') or 'any').strip().lower()
    min_units = int(params.get('min_units') or 0)
    arrive_before = _apo_parse_date(params.get('arrive_before')) if params.get('arrive_before') else None
    arrive_after = _apo_parse_date(params.get('arrive_after')) if params.get('arrive_after') else None

    # Pre-index qualifying production arrivals when a date window is set
    date_window = arrive_before or arrive_after
    qual_units = {}
    if date_window:
        for p in prods:
            b = _ai_agent_base(p.get('style'))
            if not b:
                continue
            try:
                arr = _apo_prod_arrival(p, _py_is_bottom(b))
            except Exception:
                arr = None
            if not arr:
                continue
            if arrive_before and not (arr <= arrive_before):
                continue
            if arrive_after and not (arr >= arrive_after):
                continue
            qual_units[b] = qual_units.get(b, 0) + int(p.get('units') or 0)

    rows = []
    for base, r in agg.items():
        wh = r['total_warehouse']
        inc = r['incoming']
        if stock == 'warehouse' and wh <= 0:
            continue
        if stock == 'overseas' and inc <= 0:
            continue
        if stock == 'any' and wh <= 0 and inc <= 0:
            continue
        if brand_keys and r['brand_abbr'] not in brand_keys:
            continue
        if search and search not in base:
            continue
        if fabric_codes and (len(base) < 6 or base[4:6] not in fabric_codes):
            continue
        if category:
            try:
                if category in ('blazer', 'blazers'):
                    if not _py_is_blazer(base):
                        continue
                elif not _py_matches_category(base, r['brand_abbr'], category):
                    continue
            except Exception:
                continue
        if date_window and base not in qual_units:
            continue
        basis = wh if stock == 'warehouse' else (inc if stock == 'overseas' else max(r['total_ats'], wh + inc))
        if min_units and basis < min_units:
            continue
        if date_window:
            r = dict(r)
            r['qualifying_po_units'] = qual_units.get(base, 0)
        rows.append(r)

    sort = (params.get('sort') or 'total_ats').strip().lower()
    key = {'warehouse': 'total_warehouse', 'incoming': 'incoming'}.get(sort, 'total_ats')
    rows.sort(key=lambda x: x.get(key, 0), reverse=True)
    return rows, prods


def _ai_agent_enrich(r):
    base = r['style']
    ab = r['brand_abbr']
    try:
        color = _apo_style_color(base, ab) or ''
    except Exception:
        color = ''
    try:
        fab = _apo_fabrication(base) or ''
    except Exception:
        fab = ''
    try:
        fit = _apo_fit_label(base) or ''
    except Exception:
        fit = ''
    return color, fit, fab


def _ai_tool_query_inventory(params):
    rows, _prods = _ai_agent_filter(params)
    color_q = (params.get('color') or '').strip().lower() or None
    limit = max(1, min(int(params.get('limit') or 30), 200))
    out_rows = []
    tot_wh = tot_inc = tot_ats = 0
    matched = 0
    for r in rows:
        color, fit, fab = _ai_agent_enrich(r)
        if color_q:
            try:
                bucket = _apo_classify_color(color, r['brand_abbr'])
            except Exception:
                bucket = ''
            cl = color.lower()
            if color_q in ('solid', 'solids'):
                if bucket == 'fancies':
                    continue
            elif color_q in ('fancy', 'fancies'):
                if bucket != 'fancies':
                    continue
            # blue/indigo also match the navy bucket (it holds the whole blue family)
            elif color_q in ('blue', 'indigo', 'navy/blue', 'navy blue') and bucket == 'navy':
                pass
            elif color_q not in cl and color_q != bucket:
                continue
        matched += 1
        tot_wh += r['total_warehouse']
        tot_inc += r['incoming']
        tot_ats += r['total_ats']
        if len(out_rows) < limit:
            row = {'style': r['style'], 'brand': r['brand_abbr'], 'color': color, 'fit': fit,
                   'fabrication': fab, 'warehouse': r['total_warehouse'], 'incoming': r['incoming'],
                   'total_ats': r['total_ats'], 'committed_units': abs(r['committed']),
                   'allocated_units': abs(r['allocated'])}
            if 'qualifying_po_units' in r:
                row['qualifying_po_units'] = r['qualifying_po_units']
            out_rows.append(row)
    return {'matched_styles': matched, 'total_warehouse_units': tot_wh, 'total_incoming_units': tot_inc,
            'total_ats_units': tot_ats, 'rows': out_rows, 'truncated': matched > len(out_rows),
            'note': 'quantities aggregate all size rows per base style; committed/allocated shown as positive magnitudes'}


def _ai_tool_style_detail(params):
    base = _ai_agent_base(params.get('style'))
    if not base:
        return {'error': 'style required'}
    agg = _ai_agent_agg_inventory()
    r = agg.get(base)
    with _inv_lock:
        size_rows = [dict(it) for it in (_inventory.get('items') or []) if _ai_agent_base(it.get('sku')) == base]
    if not r and not size_rows:
        # Style not in the current feed — still report classification + production
        try:
            ab = _apo_style_brand(base) or ''
        except Exception:
            ab = ''
        r = {'style': base, 'brand_abbr': ab, 'jtw': 0, 'tr': 0, 'dcw': 0, 'qa': 0, 'nj': 0, 'incoming': 0,
             'total_ats': 0, 'committed': 0, 'allocated': 0, 'total_warehouse': 0, 'size_rows': 0}
    color, fit, fab = _ai_agent_enrich(r)
    try:
        category = _py_get_item_category(base, r['brand_abbr'])
    except Exception:
        category = ''
    prods = _ai_agent_prod_for(base)
    with _apo_lock:
        apo = [{'customer': a.get('customer'), 'qty': a.get('qty'), 'po': a.get('po')}
               for a in (_apo_data or []) if _ai_agent_base(a.get('style')) == base]
    orders_out = []
    try:
        orders, ok = _fetch_all_open_orders()
        if ok:
            for o in orders:
                if _ai_agent_base(o.get('style') or o.get('baseStyle')) != base:
                    continue
                if not (o.get('openQty') or o.get('pickQty')):
                    continue
                orders_out.append({'customer': o.get('customerFull') or o.get('customer'),
                                   'order_no': o.get('orderNo'), 'open_qty': o.get('openQty'),
                                   'pick_qty': o.get('pickQty'), 'start': o.get('startDate'),
                                   'cancel': o.get('cancelDate')})
    except Exception:
        pass
    return {'style': base, 'brand': r['brand_abbr'], 'color': color, 'fit': fit, 'fabrication': fab,
            'category': category, 'warehouse': {'jtw': r['jtw'], 'tr': r['tr'], 'dcw': r['dcw'], 'qa': r['qa'], 'nj': r.get('nj', 0),
            'total': r['total_warehouse']}, 'incoming': r['incoming'], 'total_ats': r['total_ats'],
            'committed_units': abs(r['committed']), 'allocated_units': abs(r['allocated']),
            'size_variants': [{'sku': s.get('sku'), 'warehouse': (s.get('jtw') or 0) + (s.get('tr') or 0) + (s.get('dcw') or 0) + (s.get('qa') or 0) + (s.get('nj') or 0),
                               'incoming': s.get('incoming'), 'total_ats': s.get('total_ats')} for s in size_rows[:40]],
            'production_orders': prods[:20], 'apo_allocations': apo[:20], 'open_orders': orders_out[:20],
            'image_url': f'{_AI_AGENT_SELF_URL}/image/{base}'}


def _ai_tool_brand_summary(params):
    agg = _ai_agent_agg_inventory()
    brand = (params.get('brand') or '').strip().upper()
    if brand:
        for abbr, full in (BRAND_FULL_NAMES or {}).items():
            if str(full).upper() == brand:
                brand = str(abbr).upper()
                break
    by_brand = {}
    for r in agg.values():
        if r['total_warehouse'] <= 0 and r['incoming'] <= 0:
            continue
        b = r['brand_abbr'] or '?'
        s = by_brand.setdefault(b, {'brand': b, 'styles': 0, 'warehouse_units': 0, 'incoming_units': 0, 'total_ats': 0})
        s['styles'] += 1
        s['warehouse_units'] += r['total_warehouse']
        s['incoming_units'] += r['incoming']
        s['total_ats'] += r['total_ats']
    if not brand:
        rows = sorted(by_brand.values(), key=lambda x: x['total_ats'], reverse=True)
        return {'brands': rows}
    fabs, cats = {}, {}
    for r in agg.values():
        if r['brand_abbr'] != brand or (r['total_warehouse'] <= 0 and r['incoming'] <= 0):
            continue
        base = r['style']
        try:
            fab = _apo_fabrication(base) or 'Unknown'
        except Exception:
            fab = 'Unknown'
        code = base[4:6] if len(base) >= 6 else '??'
        f = fabs.setdefault(code, {'fabric_code': code, 'fabrication': fab, 'styles': 0, 'warehouse_units': 0, 'incoming_units': 0})
        f['styles'] += 1
        f['warehouse_units'] += r['total_warehouse']
        f['incoming_units'] += r['incoming']
        try:
            cat = _py_get_item_category(base, brand)
        except Exception:
            cat = 'unknown'
        c = cats.setdefault(cat, {'category': cat, 'styles': 0, 'units': 0})
        c['styles'] += 1
        c['units'] += r['total_warehouse'] + r['incoming']
    return {'summary': by_brand.get(brand) or {'brand': brand, 'styles': 0},
            'fabrications': sorted(fabs.values(), key=lambda x: x['warehouse_units'] + x['incoming_units'], reverse=True),
            'categories': sorted(cats.values(), key=lambda x: x['units'], reverse=True)}


def _ai_tool_open_orders(params):
    orders, ok = _fetch_all_open_orders()
    if not ok and not orders:
        return {'error': 'open orders feed unavailable right now'}
    cust_q = (params.get('customer') or '').strip().lower() or None
    style_q = _ai_agent_base(params.get('style')) if params.get('style') else None
    limit = max(1, min(int(params.get('limit') or 30), 200))
    rows, tot_qty, tot_val, matched = [], 0, 0.0, 0
    for o in orders:
        open_qty = int(o.get('openQty') or 0) + int(o.get('pickQty') or 0)
        if open_qty <= 0:
            continue
        cust = str(o.get('customerFull') or o.get('customer') or '')
        if cust_q and cust_q not in cust.lower() and cust_q not in str(o.get('customer') or '').lower():
            continue
        if style_q and _ai_agent_base(o.get('style') or o.get('baseStyle')) != style_q:
            continue
        matched += 1
        tot_qty += open_qty
        try:
            tot_val += float(o.get('openValue') or 0) + float(o.get('pickValue') or 0)
        except (TypeError, ValueError):
            pass
        if len(rows) < limit:
            rows.append({'customer': cust, 'order_no': o.get('orderNo'), 'style': o.get('style'),
                         'open_qty': open_qty, 'start': o.get('startDate'), 'cancel': o.get('cancelDate'),
                         'is_bulk': bool(o.get('isPipeline'))})
    return {'matched_orders': matched, 'total_open_units': tot_qty, 'total_open_value': round(tot_val, 2),
            'rows': rows, 'truncated': matched > len(rows)}


# ── Past orders (received-PO history) — proxied from the open-orders service ──
# The open-orders API snapshots the order book daily (since 2026-06-03) and
# rolls every PO that leaves the book into a per-customer archive
# (/api/po-history). This lets the AI agent + MCP connector query that history.
_po_hist_proxy_cache = {'summary': None, 'summary_at': 0.0, 'accounts': {}, 'accounts_at': {},
                        'fail_until': {}}
_po_hist_proxy_lock = threading.Lock()
_PO_HIST_PROXY_TTL = 900       # 15 min — the archive gains at most one new day per day
_PO_HIST_FAIL_BACKOFF = 120    # after a failed fetch, don't re-hit that key for 2 min


def _fetch_po_history(account=None):
    """Fetch the received-PO archive (summary, or one account's POs) from the
    open-orders service, cached module-level. Returns (payload_or_None, ok)."""
    now = time.time()
    key = (str(account).strip().upper() if account else None) or None
    with _po_hist_proxy_lock:
        if key is None:
            cached, at = _po_hist_proxy_cache['summary'], _po_hist_proxy_cache['summary_at']
        else:
            cached = _po_hist_proxy_cache['accounts'].get(key)
            at = _po_hist_proxy_cache['accounts_at'].get(key, 0.0)
        fail_until = _po_hist_proxy_cache['fail_until'].get(key or '__summary__', 0.0)
    if cached is not None and now - at < _PO_HIST_PROXY_TTL:
        return cached, True
    if now < fail_until:
        # Recent failure — serve whatever we have without re-hitting a dead
        # upstream (mirrors _fetch_all_open_orders' backoff for the sync worker).
        return cached, cached is not None
    try:
        resp = http_requests.get(f"{OPEN_ORDERS_API_URL}/api/po-history", headers=_oo_api_headers(),
                                 params=({'account': key} if key else None), timeout=30)
        if resp.status_code != 200:
            raise RuntimeError(f"HTTP {resp.status_code}")
        data = resp.json()
        if not data.get('ready'):
            return data, False   # index still building server-side — don't cache
        with _po_hist_proxy_lock:
            if key is None:
                _po_hist_proxy_cache['summary'] = data
                _po_hist_proxy_cache['summary_at'] = time.time()
            else:
                _po_hist_proxy_cache['accounts'][key] = data
                _po_hist_proxy_cache['accounts_at'][key] = time.time()
        return data, True
    except Exception as e:
        print(f"[PastOrders] po-history fetch failed ({key or 'summary'}): {e}", flush=True)
        with _po_hist_proxy_lock:
            _po_hist_proxy_cache['fail_until'][key or '__summary__'] = time.time() + _PO_HIST_FAIL_BACKOFF
        return cached, cached is not None


def _past_po_row(p, include_lines=True):
    row = {'customer': p.get('customer'), 'customer_full': p.get('customerFull'),
           'order_no': p.get('orderNo'),
           'status': 'shipped' if p.get('shipped') else 'cancelled_or_pulled',
           'ship_start': p.get('start') or None, 'cancel': p.get('cancel') or None,
           'left_book': p.get('lastSeen') or None,
           'units': p.get('units'), 'value': p.get('value'),
           # peak booked state — a partially-shipped PO's true size (the plain
           # units/value fields hold only the last-day residual)
           'units_peak': p.get('unitsPeak'), 'value_peak': p.get('valuePeak')}
    if include_lines:
        row['lines'] = [{'style': l.get('style'), 'qty': l.get('qty'),
                         'price': l.get('price'), 'value': l.get('value')}
                        for l in (p.get('lines') or [])]
    return row


def _ai_tool_past_orders(params):
    summary, ok = _fetch_po_history()
    if summary is None:
        return {'error': 'past-orders archive unavailable right now'}
    if not summary.get('ready'):
        prog = summary.get('progress') or {}
        return {'building': True,
                'note': (f"the archive index is still building "
                         f"({prog.get('done', 0)}/{prog.get('total', 0)} days) — retry in a minute")}
    accounts = [a for a in (summary.get('accounts') or [])
                if (a.get('poCount') or 0) > 0 or (a.get('cancelledCount') or 0) > 0]
    cust_q = (params.get('customer') or '').strip() or None
    style_q = (params.get('style') or '').strip().upper() or None
    status_q = (params.get('status') or 'all').strip().lower()
    after_q = (params.get('received_after') or '').strip() or None
    before_q = (params.get('received_before') or '').strip() or None
    limit = max(1, min(int(params.get('limit') or 30), 200))

    def _keep(p):
        if status_q == 'shipped' and not p.get('shipped'):
            return False
        if status_q in ('cancelled', 'pulled', 'cancelled_or_pulled') and p.get('shipped'):
            return False
        seen = p.get('lastSeen') or ''
        if after_q and seen < after_q:
            return False
        if before_q and seen > before_q:
            return False
        return True

    # ── No filters: per-customer summary ──
    if not cust_q and not style_q:
        ignored = [k for k, v in (('status', status_q != 'all'), ('received_after', bool(after_q)),
                                  ('received_before', bool(before_q))) if v]
        extra = {'filters_ignored': ignored, 'note2': 'status/date filters apply only with a customer or style'} if ignored else {}
        return {**extra, 'history_start': '2026-06-03', 'latest_date': summary.get('latestDate'),
                'note': ("'received' = the day a PO left the open-order book (shipped-out approximation, "
                         "accurate to ~1 day). cancelled/pulled = left before its ship window opened."),
                'totals': summary.get('totals'),
                'customers': [{'customer': a.get('customer'), 'customer_full': a.get('customerFull'),
                               'received_pos': a.get('poCount'), 'units': a.get('units'),
                               'value': a.get('value'),
                               'cancelled_pos': a.get('cancelledCount'),
                               'cancelled_value': a.get('cancelledValue'),
                               'first_received': a.get('firstReceived') or None,
                               'last_received': a.get('lastReceived') or None}
                              for a in accounts]}

    # ── Customer filter: resolve to ONE account code ──
    if cust_q:
        cl = cust_q.lower()
        matches = [a for a in accounts
                   if cl == str(a.get('customer') or '').lower()
                   or cl in str(a.get('customerFull') or '').lower()]
        if not matches:
            return {'error': f"no past-order customer matches '{cust_q}'",
                    'customers': [a.get('customerFull') or a.get('customer') for a in accounts]}
        if len(matches) > 1 and not any(cl == str(a.get('customer') or '').lower() for a in matches):
            return {'ambiguous_customer': [{'customer': a.get('customer'),
                                            'customer_full': a.get('customerFull')} for a in matches]}
        acct = next((a for a in matches if cl == str(a.get('customer') or '').lower()), matches[0])
        if not acct.get('customer'):
            return {'error': 'that account has no customer code in the archive'}
        data, _ok = _fetch_po_history(acct.get('customer'))
        if data is None or not data.get('ready'):
            return {'error': f"couldn't load history for {acct.get('customer')} right now"}
        pos = [p for p in (data.get('pos') or []) if _keep(p)]
        if style_q:
            pos = [p for p in pos
                   if any(style_q in str(l.get('style') or '').upper() for l in (p.get('lines') or []))]
        pos.sort(key=lambda p: str(p.get('lastSeen') or ''), reverse=True)
        if style_q:
            # Totals count only the MATCHING style lines, not whole POs.
            tot_units = sum(l.get('qty') or 0 for p in pos for l in (p.get('lines') or [])
                            if style_q in str(l.get('style') or '').upper())
            tot_val = round(sum(l.get('value') or 0 for p in pos for l in (p.get('lines') or [])
                                if style_q in str(l.get('style') or '').upper()), 2)
        else:
            # Shipped POs count at their PEAK booked size (the last-day residual
            # understates partial shipments); pulled POs keep their final state.
            tot_units = sum(((p.get('unitsPeak') or p.get('units') or 0) if p.get('shipped')
                             else (p.get('units') or 0)) for p in pos)
            tot_val = round(sum(((p.get('valuePeak') or p.get('value') or 0) if p.get('shipped')
                                 else (p.get('value') or 0)) for p in pos), 2)
        out = {'customer': acct.get('customer'), 'customer_full': acct.get('customerFull'),
               'matched_pos': len(pos), 'total_units': tot_units, 'total_value': tot_val,
               'rows': [_past_po_row(p) for p in pos[:limit]],
               'truncated': len(pos) > limit}
        if style_q:
            out['note'] = f"total_units/total_value count only lines matching '{style_q}'; rows show full POs"
        return out

    # ── Style search across every customer ──
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as ex:
        fetched = list(ex.map(lambda a: _fetch_po_history(a.get('customer')), accounts))
    hits = []
    partial = False
    for a, (data, _ok) in zip(accounts, fetched):
        if data is None or not data.get('ready'):
            partial = True
            continue
        for p in (data.get('pos') or []):
            if not _keep(p):
                continue
            for l in (p.get('lines') or []):
                if style_q in str(l.get('style') or '').upper():
                    hits.append({'customer': a.get('customer'), 'customer_full': a.get('customerFull'),
                                 'order_no': p.get('orderNo'),
                                 'status': 'shipped' if p.get('shipped') else 'cancelled_or_pulled',
                                 'left_book': p.get('lastSeen') or None, 'style': l.get('style'),
                                 'qty': l.get('qty'), 'price': l.get('price'), 'value': l.get('value')})
    hits.sort(key=lambda h: str(h.get('left_book') or ''), reverse=True)
    tot_qty = sum(h.get('qty') or 0 for h in hits)
    tot_val = round(sum(h.get('value') or 0 for h in hits), 2)
    out = {'style_query': style_q, 'matched_lines': len(hits), 'total_units': tot_qty,
           'total_value': tot_val, 'rows': hits[:limit], 'truncated': len(hits) > limit}
    if partial:
        out['note'] = 'some customers failed to load — results may be partial'
    return out


def _sales_resolve_customer(cust_q):
    """Resolve a customer NAME or code to the exact invoice account code.
    Upstream matches codes exactly, so 'Burlington' must become 'BURL' here.
    Uses the sales summary's code list + the po-history accounts' full names.
    Returns (code, None) on success or (None, error_dict) on failure."""
    resp = http_requests.get(f"{OPEN_ORDERS_API_URL}/api/sales-history", headers=_oo_api_headers(), timeout=45)
    data = resp.json()
    if not data.get('ready'):
        return None, {'error': 'sales history backfill not available yet'}
    codes = [str(c.get('customer') or '') for c in (data.get('customers') or [])]
    cu = cust_q.strip().upper()
    if cu in codes:
        return cu, None
    # full-name resolution via the po-history account directory
    name_map = {}
    hist, _ok = _fetch_po_history()
    for a in ((hist or {}).get('accounts') or []):
        name_map[str(a.get('customer') or '').upper()] = str(a.get('customerFull') or '')
    matches = [c for c in codes
               if cu in c.upper() or (name_map.get(c.upper()) and cu in name_map[c.upper()].upper())]
    if not matches:
        return None, {'error': f"no invoiced customer matches '{cust_q}'",
                      'customers': [(name_map.get(c.upper()) and f"{c} ({name_map[c.upper()]})") or c
                                    for c in codes]}
    if len(matches) > 1:
        return None, {'ambiguous_customer': [{'customer': c,
                                              'customer_full': name_map.get(c.upper()) or c}
                                             for c in matches]}
    return matches[0], None


def _ai_tool_sales_history(params):
    """Invoice-level sales history (A2000 InvoiceReport backfill on the
    open-orders service): real invoice dates + shipped quantities, Nov 2019 →
    the last ingest. Proxied live; the upstream serves prebuilt aggregates."""
    cust_q = (params.get('customer') or '').strip()
    style_q = (params.get('style') or '').strip()
    month_q = (params.get('month') or '').strip()
    from_q = (params.get('from') or '').strip()
    to_q = (params.get('to') or '').strip()
    limit = max(1, min(int(params.get('limit') or 30), 200))
    try:
        if style_q:
            code = None
            if cust_q:
                code, err = _sales_resolve_customer(cust_q)
                if err:
                    return err
            resp = http_requests.get(f"{OPEN_ORDERS_API_URL}/api/sales-history", headers=_oo_api_headers(),
                                     params={'style': style_q}, timeout=45)
            data = resp.json()
            if not data.get('ready'):
                return {'error': 'sales history backfill not available yet'}
            rows = data.get('rows') or []
            out = {'style_query': data.get('style_query')}
            if code:
                rows = [r for r in rows if str(r.get('customer') or '').upper() == code]
                out['customer'] = code
                out['matched_qty'] = sum(r.get('qty') or 0 for r in rows)
                out['matched_value'] = round(sum(r.get('value') or 0 for r in rows), 2)
                if data.get('truncated'):
                    out['note'] = ('upstream capped the search at its newest 500 lines — older matches '
                                   'for this customer may be missing; totals cover the returned lines only')
            else:
                out['matched_qty'] = data.get('matched_qty')
                out['matched_value'] = data.get('matched_value')
            out['rows'] = rows[:limit]
            out['truncated'] = bool(data.get('truncated')) or len(rows) > limit
            return out
        if cust_q:
            code, err = _sales_resolve_customer(cust_q)
            if err:
                return err
            qp = {'account': code}
            if month_q:
                qp['month'] = month_q
            elif from_q or to_q:
                if from_q:
                    qp['from'] = from_q
                if to_q:
                    qp['to'] = to_q
            resp = http_requests.get(f"{OPEN_ORDERS_API_URL}/api/sales-history", params=qp, headers=_oo_api_headers(), timeout=45)
            data = resp.json()
            if not data.get('ready'):
                return {'error': 'sales history backfill not available yet'}
            if 'pos' in data:   # ranged query → PO rollups with style lines
                pos = data.get('pos') or []
                return {'customer': data.get('account'), 'from': data.get('from'), 'to': data.get('to'),
                        'po_count': data.get('poCount'), 'units': data.get('units'),
                        'value': data.get('value'), 'rows': pos[:limit],
                        'truncated': bool(data.get('truncated')) or len(pos) > limit}
            return {'customer': data.get('account'), 'invoice_row_count': data.get('rows'),
                    'po_count': data.get('poCount'), 'years': data.get('years'),
                    'months': data.get('months'), 'divisions': data.get('divisions'),
                    'top_styles': (data.get('topStyles') or [])[:30],
                    'recent_pos': (data.get('posRecent') or [])[:limit],
                    'note': 'divisions: OB = bulk, DS = dropship'}
        resp = http_requests.get(f"{OPEN_ORDERS_API_URL}/api/sales-history", headers=_oo_api_headers(), timeout=45)
        data = resp.json()
        if not data.get('ready'):
            return {'error': 'sales history backfill not available yet'}
        return {'source': data.get('source'), 'totals': data.get('totals'),
                'years': data.get('years'),
                'customers': [{'customer': c.get('customer'), 'units': c.get('units'),
                               'value': c.get('value'), 'invoices': c.get('invoiceCount'),
                               'first': c.get('firstInv'), 'last': c.get('lastInv')}
                              for c in (data.get('customers') or [])[:60]],
                'note': ('invoice ground truth (real invoice dates + shipped quantities). '
                         'Use past_orders_lookup for POs newer than source.to (book-tracked).')}
    except Exception as e:
        return {'error': f'sales history unavailable: {e}'}


def _sisc(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def _sfsc(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _ai_tool_build_sales_sheet(params):
    """Build a downloadable Excel SALES SHEET with photos for one customer:
    invoiced shipments (real invoice dates), book-tracked shipped-awaiting-
    invoice POs (peak sizes), and current open hard POs. Uploaded to S3 like
    build_line_sheet; returns a public download_url."""
    cust_q = (params.get('customer') or '').strip()
    if not cust_q:
        return {'error': 'customer required'}
    code, err = _sales_resolve_customer(cust_q)
    if err:
        return err
    # full display name from the po-history directory
    full_name = code
    hist_sum, _ok = _fetch_po_history()
    for a in ((hist_sum or {}).get('accounts') or []):
        if str(a.get('customer') or '').upper() == code:
            full_name = a.get('customerFull') or code
            break

    # ── date bounds: from/to win; year / month are conveniences ──
    from_q = (params.get('from') or '').strip()
    to_q = (params.get('to') or '').strip()
    year_q = (str(params.get('year') or '')).strip()
    month_q = (params.get('month') or '').strip()
    if not from_q and not to_q:
        if month_q:
            from_q, to_q = month_q + '-01', month_q + '-31'
        elif year_q:
            from_q, to_q = year_q + '-01-01', year_q + '-12-31'
    lo = from_q or '0000-00-00'
    hi = to_q or '9999-99-99'
    period_label = (f"{from_q or 'start of history'} → {to_q or 'today'}"
                    if (from_q or to_q) else 'Full history')

    include = params.get('include') or ['invoiced', 'shipped_pending', 'on_order']
    if isinstance(include, str):
        include = [include]
    include = {str(x).strip().lower() for x in include}
    group_by = (params.get('group_by') or 'po').strip().lower()
    notes = []

    # ── 1. invoiced PO rollups with style lines ──
    inv_pos, inv_po_set = [], set()
    if 'invoiced' in include:
        try:
            resp = http_requests.get(f"{OPEN_ORDERS_API_URL}/api/sales-history", headers=_oo_api_headers(),
                                     params={'account': code, 'from': lo, 'to': hi}, timeout=60)
            data = resp.json()
            if data.get('ready'):
                inv_pos = data.get('pos') or []
                inv_po_set = {p.get('po') for p in inv_pos}
                if data.get('truncated'):
                    notes.append('invoiced list truncated at 1500 POs — narrow the date range for full coverage')
            else:
                notes.append('invoice backfill unavailable — invoiced tab skipped')
        except Exception as e:
            notes.append(f'invoiced data failed to load: {e}')

    # ── 2. book-tracked shipped POs not yet invoiced (peak sizes) ──
    pend_pos = []
    if 'shipped_pending' in include:
        data, _ok2 = _fetch_po_history(code)
        if data and data.get('ready'):
            for p in (data.get('pos') or []):
                if not p.get('shipped') or p.get('orderNo') in inv_po_set:
                    continue
                seen = p.get('lastSeen') or ''
                if seen < lo or seen > hi:
                    continue
                pend_pos.append(p)
        elif data is None:
            notes.append('book-tracked archive unavailable — awaiting-invoice tab skipped')

    # ── 3. current open hard POs (no bulks/allocations; not date-filtered) ──
    open_pos = []
    if 'on_order' in include:
        orders, ok3 = _fetch_all_open_orders()
        agg = {}
        for o in orders:
            if o.get('isPipeline'):
                continue
            if str(o.get('customer') or '').upper() != code:
                continue
            qty = _sisc(o.get('openQty')) + _sisc(o.get('pickQty'))
            if qty <= 0:
                continue
            po = str(o.get('orderNo') or '?')
            e = agg.setdefault(po, {'po': po, 'start': '', 'cancel': '', 'units': 0,
                                    'value': 0.0, 'lines': {}})
            e['units'] += qty
            e['value'] += _sfsc(o.get('openValue')) + _sfsc(o.get('pickValue'))
            st = (o.get('startDate') or '')[:10]
            cn = (o.get('cancelDate') or '')[:10]
            if st and (not e['start'] or st < e['start']):
                e['start'] = st
            if cn and (not e['cancel'] or cn > e['cancel']):
                e['cancel'] = cn
            sk = str(o.get('style') or '').strip()
            ln = e['lines'].setdefault(sk.upper(), {'style': sk, 'qty': 0, 'value': 0.0})
            ln['qty'] += qty
            ln['value'] += _sfsc(o.get('openValue')) + _sfsc(o.get('pickValue'))
        open_pos = sorted(agg.values(), key=lambda p: p['start'] or '9999')
        if not ok3 and not open_pos:
            notes.append('open-orders feed unavailable — on-order tab may be incomplete')

    if not inv_pos and not pend_pos and not open_pos:
        return {'error': f'nothing found for {full_name} ({code}) in {period_label} with include={sorted(include)}'}

    # ── flatten into per-tab line rows: (style, color, po, d1, d2, extra, qty, price, value) ──
    def _line_rows_invoiced():
        rows = []
        for p in inv_pos:
            for l in (p.get('lines') or []):
                q = _sisc(l.get('qty'))
                v = _sfsc(l.get('value'))
                rows.append({'style': l.get('style') or '', 'color': l.get('color') or '',
                             'po': p.get('po') or '', 'div': 'Dropship' if p.get('div') == 'DS' else 'Wholesale',
                             'first': p.get('firstInv') or '', 'last': p.get('lastInv') or '',
                             'invs': p.get('invoices') or 1,
                             'qty': q, 'price': round(v / q, 2) if q else 0, 'value': round(v, 2)})
        if group_by == 'month':
            rows.sort(key=lambda r: (r['last'][:7], r['po'], -r['qty']), reverse=True)
        else:
            rows.sort(key=lambda r: (r['last'], r['po'], -r['qty']), reverse=True)
        return rows

    def _line_rows_pending():
        rows = []
        for p in pend_pos:
            lines = p.get('linesPeak') or p.get('lines') or []
            for l in lines:
                q = _sisc(l.get('qty'))
                v = _sfsc(l.get('value'))
                rows.append({'style': l.get('style') or '', 'color': '',
                             'po': p.get('orderNo') or '', 'div': '',
                             'first': p.get('start') or '', 'last': p.get('lastSeen') or '',
                             'invs': '',
                             'qty': q, 'price': round(v / q, 2) if q else _sfsc(l.get('price')), 'value': round(v, 2)})
        rows.sort(key=lambda r: (r['last'], r['po']), reverse=True)
        return rows

    def _line_rows_open():
        rows = []
        for p in open_pos:
            for l in sorted(p['lines'].values(), key=lambda x: -x['qty']):
                q = l['qty']
                v = l['value']
                rows.append({'style': l['style'], 'color': '',
                             'po': p['po'], 'div': '',
                             'first': p['start'], 'last': p['cancel'], 'invs': '',
                             'qty': q, 'price': round(v / q, 2) if q else 0, 'value': round(v, 2)})
        return rows

    tabs = []
    if inv_pos:
        tabs.append(('Invoiced', _line_rows_invoiced(),
                     ['Image', 'Style', 'Color', 'PO #', 'Division', 'First Invoiced', 'Last Invoiced', 'Qty', 'Unit Price', 'Value'],
                     lambda r: [r['style'], r['color'], r['po'], r['div'], r['first'], r['last'], r['qty'], r['price'], r['value']]))
    if pend_pos:
        tabs.append(('Shipped - Awaiting Invoice', _line_rows_pending(),
                     ['Image', 'Style', 'PO #', 'Window Start', 'Left Book', 'Qty (peak booked)', 'Est. Price', 'Est. Value'],
                     lambda r: [r['style'], r['po'], r['first'], r['last'], r['qty'], r['price'], r['value']]))
    if open_pos:
        tabs.append(('On Order', _line_rows_open(),
                     ['Image', 'Style', 'PO #', 'Window Start', 'Cancel', 'Qty', 'Unit Price', 'Value'],
                     lambda r: [r['style'], r['po'], r['first'], r['last'], r['qty'], r['price'], r['value']]))

    ROW_CAP = 3000
    buf = BytesIO()
    wb = xlsxwriter.Workbook(buf, {'in_memory': True, 'strings_to_formulas': False})
    f_title = wb.add_format({'bold': True, 'font_size': 16, 'font_name': 'Calibri'})
    f_sub = wb.add_format({'font_size': 10, 'font_color': '#666666', 'font_name': 'Calibri'})
    f_head = wb.add_format({'bold': True, 'font_color': '#FFFFFF', 'bg_color': '#6D28D9',
                            'border': 1, 'align': 'center', 'valign': 'vcenter', 'font_name': 'Calibri'})
    f_cell = wb.add_format({'border': 1, 'valign': 'vcenter', 'font_name': 'Calibri'})
    f_cell2 = wb.add_format({'border': 1, 'valign': 'vcenter', 'bg_color': '#F5F3FF', 'font_name': 'Calibri'})
    f_num = wb.add_format({'border': 1, 'valign': 'vcenter', 'num_format': '#,##0', 'font_name': 'Calibri'})
    f_num2 = wb.add_format({'border': 1, 'valign': 'vcenter', 'num_format': '#,##0', 'bg_color': '#F5F3FF', 'font_name': 'Calibri'})
    f_money = wb.add_format({'border': 1, 'valign': 'vcenter', 'num_format': '$#,##0.00', 'font_name': 'Calibri'})
    f_money2 = wb.add_format({'border': 1, 'valign': 'vcenter', 'num_format': '$#,##0.00', 'bg_color': '#F5F3FF', 'font_name': 'Calibri'})
    f_bold = wb.add_format({'bold': True, 'font_name': 'Calibri'})

    # Summary tab
    ws = wb.add_worksheet('Summary')
    ws.set_column(0, 0, 30)
    ws.set_column(1, 1, 46)
    ws.write(0, 0, f'{full_name} — Sales Sheet', f_title)
    ws.write(1, 0, f'Period: {period_label} · generated {datetime.now().strftime("%b %d, %Y %H:%M")}', f_sub)
    srow = 3
    for name, rows, _h, _g in tabs:
        units = sum(r['qty'] for r in rows)
        value = sum(r['value'] for r in rows)
        pos_ct = len({r['po'] for r in rows})
        ws.write(srow, 0, name, f_bold)
        ws.write(srow, 1, f'{pos_ct:,} POs · {units:,} units · ${value:,.2f}')
        srow += 1
    srow += 1
    for n in ['Invoiced = A2000 invoice records (real invoice dates and shipped quantities).',
              'Awaiting invoice = left the open-order book after its window opened; quantities are the PEAK booked size.',
              'On order = current open hard POs (bulks and allocations excluded), regardless of the date filter.'] + notes:
        ws.write(srow, 0, n, f_sub)
        srow += 1

    for name, rows, headers, getter in tabs:
        wsx = wb.add_worksheet(name[:31])
        truncated = len(rows) > ROW_CAP
        rows = rows[:ROW_CAP]
        wsx.set_column(0, 0, 14)
        wsx.set_column(1, 1, 18)
        for ci in range(2, len(headers)):
            wsx.set_column(ci, ci, 16)
        wsx.set_row(0, 22)
        for ci, h in enumerate(headers):
            wsx.write(0, ci, h, f_head)
        wsx.freeze_panes(1, 0)
        # Photos only for modern-taxonomy SKUs — legacy invoice styles (NAU-10395SS,
        # SHAQ-22-106…) would collapse to bogus bases and could attach wrong images.
        def _img_sku(style):
            base = get_base_style(str(style or '').upper())
            return base if re.match(r'^[A-Z]{6}\d{3}[A-Z]{2,3}$', base) else ''
        img_items = [{'sku': _img_sku(r['style']), 'brand_abbr': ''} for r in rows]
        try:
            for it in img_items:
                try:
                    it['brand_abbr'] = _apo_style_brand(it['sku']) or ''
                except Exception:
                    pass
            images = download_images_for_items(img_items, S3_PHOTOS_URL, use_cache=True)
        except Exception:
            images = {}
        for ri, r in enumerate(rows):
            row = ri + 1
            wsx.set_row(row, 60)
            cf = f_cell2 if ri % 2 else f_cell
            nf = f_num2 if ri % 2 else f_num
            mf = f_money2 if ri % 2 else f_money
            vals = getter(r)
            for ci, v in enumerate(vals):
                col = ci + 1
                h = headers[col]
                if h in ('Qty', 'Qty (peak booked)'):
                    wsx.write(row, col, v, nf)
                elif h in ('Unit Price', 'Value', 'Est. Price', 'Est. Value'):
                    wsx.write(row, col, v, mf)
                else:
                    wsx.write(row, col, v, cf)
            img = images.get(ri)
            if img:
                try:
                    wsx.insert_image(row, 0, 'img.png', _padded_image_opts(img))
                except Exception:
                    wsx.write(row, 0, '', cf)
            else:
                wsx.write(row, 0, '', cf)
        if truncated:
            notes.append(f'{name}: only the first {ROW_CAP:,} lines are in the sheet')

    wb.close()
    buf.seek(0)
    fname = re.sub(r'[^A-Za-z0-9 _.-]+', '', str(params.get('filename') or f'{full_name} Sales Sheet')).strip() or 'Sales Sheet'
    key = f"claude uploaded/{fname} {datetime.now().strftime('%Y-%m-%d %H%M%S')}.xlsx"
    get_s3().put_object(Bucket=S3_BUCKET, Key=key, Body=buf.read(),
                        ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    from urllib.parse import quote
    url = f"https://{S3_BUCKET}.s3.us-east-2.amazonaws.com/{quote(key)}"
    out = {'download_url': url, 'customer': code, 'customer_full': full_name,
           'period': period_label,
           'tabs': [{'tab': n, 'lines': min(len(r), ROW_CAP), 'pos': len({x['po'] for x in r}),
                     'units': sum(x['qty'] for x in r), 'value': round(sum(x['value'] for x in r), 2)}
                    for n, r, _h, _g in tabs],
           'note': 'give the user this link as a clickable download'}
    if notes:
        out['warnings'] = notes
    return out


def _ai_tool_build_line_sheet(params):
    tabs_in = params.get('tabs') or []
    if not isinstance(tabs_in, list) or not tabs_in:
        return {'error': 'tabs required: a list of filter objects, one per tab'}
    customer_view = bool(params.get('customer_view'))
    tabs_out = []
    summary = []
    with _production_lock:
        prods_all = list(_production_data)
    prod_by_base = {}
    for p in prods_all:
        prod_by_base.setdefault(_ai_agent_base(p.get('style')), []).append(p)
    agg_cache = None
    for ti, t in enumerate(tabs_in[:4]):
        skus_req = []
        seen_req = set()
        for s in (t.get('skus') or []):
            su = str(s or '').strip().upper()
            # Dedupe on the RESOLVED base (resolution below also keys on base) so a
            # base and its size-suffixed twin can't render the same style twice.
            sb = _ai_agent_base(su)
            if su and sb and sb not in seen_req:
                seen_req.add(sb)
                skus_req.append(su)
        missing_req = []
        if skus_req:
            # Explicit style list: exactly these bases in this order, no other
            # filters second-guess the caller's curation.
            if agg_cache is None:
                agg_cache = _ai_agent_agg_inventory()
            rows = []
            for su in skus_req:
                r = agg_cache.get(_ai_agent_base(su))
                if r is not None:
                    rows.append(r)
                else:
                    missing_req.append(su)
            color_q = None
        else:
            rows, _ = _ai_agent_filter(t)
            color_q = (t.get('color') or '').strip().lower() or None
        items = []
        over_cap = rows[300:]
        for r in rows[:300]:
            color, fit, fab = _ai_agent_enrich(r)
            if color_q:
                try:
                    bucket = _apo_classify_color(color, r['brand_abbr'])
                except Exception:
                    bucket = ''
                cl = color.lower()
                if color_q in ('solid', 'solids'):
                    if bucket == 'fancies':
                        continue
                elif color_q in ('fancy', 'fancies'):
                    if bucket != 'fancies':
                        continue
                # blue/indigo also match the navy bucket (it holds the whole blue family)
                elif color_q in ('blue', 'indigo', 'navy/blue', 'navy blue') and bucket == 'navy':
                    pass
                elif color_q not in cl and color_q != bucket:
                    continue
            base = r['style']
            # Nearest production dates for delivery columns
            etd_s = arr_s = po_ref = ''
            best = None
            for p in prod_by_base.get(base, []):
                try:
                    arr = _apo_prod_arrival(p, _py_is_bottom(base))
                except Exception:
                    arr = None
                if arr and (best is None or arr < best[0]):
                    best = (arr, p)
            if best:
                arr_s = str(best[0])
                etd_s = str(best[1].get('etd') or '')
                po_ref = str(best[1].get('production') or '')
            wh_names = [n for n, k in (('JTW', 'jtw'), ('TR', 'tr'), ('DCW', 'dcw'), ('QA', 'qa'), ('NJ', 'nj')) if r[k] > 0]
            item = {'sku': base, 'brand_abbr': r['brand_abbr'],
                    'brand_full': (BRAND_FULL_NAMES or {}).get(r['brand_abbr'], r['brand_abbr']),
                    'color': color, 'fit': fit, 'fabric_code': base[4:6] if len(base) >= 6 else '',
                    'fabrication': fab, 'delivery': 'ATS' if r['total_warehouse'] > 0 else (arr_s or 'Overseas'),
                    'total_ats': r['total_ats'], 'jtw': r['jtw'], 'tr': r['tr'], 'dcw': r['dcw'], 'qa': r['qa'],
                    'nj': r.get('nj', 0),
                    'incoming': r['incoming'], 'total_warehouse': r['total_warehouse'],
                    'committed': r['committed'], 'allocated': r['allocated'],
                    'warehouse': ', '.join(wh_names) or '—',
                    'ex_factory': etd_s, 'arrival': arr_s, 'po_ref': po_ref}
            items.append(item)
        if not items:
            entry = {'tab': t.get('title') or f'Tab {ti + 1}', 'styles': 0, 'skipped': True}
            if missing_req:
                entry['skus_not_found'] = missing_req
            summary.append(entry)
            continue
        try:
            # Returns annotated COPIES — assigning is what makes Override Tool size
            # packs and fit/customer-narrowed prepack rules reach the size charts.
            items = _annotate_items_for_prepack(items)
        except Exception:
            pass
        stock = (t.get('stock') or 'any').strip().lower()
        view_mode = 'incoming' if stock == 'overseas' else ('ats' if stock == 'warehouse' else 'all')
        name = t.get('title') or f"{(t.get('brand') or (t.get('brands') or ['Styles'])[0])} {view_mode}"
        tabs_out.append({'brand_name': str(name), 'tab_name': str(name), 'items': items,
                         'view_mode': view_mode, 'flow_mode': False, 'keep_order': True})
        entry = {'tab': str(name), 'styles': len(items)}
        if missing_req:
            entry['skus_not_found'] = missing_req
        if over_cap:
            # Never let a silent cap read as full coverage: name the dropped styles
            # on the curated path so the model can move them to another tab or warn.
            entry['truncated'] = True
            entry['styles_dropped_over_300_cap'] = ([r['style'] for r in over_cap]
                                                    if skus_req else len(over_cap))
        summary.append(entry)
    if not tabs_out:
        return {'error': 'no styles matched any tab filters', 'tabs': summary}
    try:
        pd_snap = _fresh_prepack_defaults()
    except Exception:
        pd_snap = None
    xl = build_multi_brand_excel(tabs_out, S3_PHOTOS_URL, catalog_mode=customer_view,
                                 view_mode='all', flow_mode=False, prepack_defaults=pd_snap)
    fname = re.sub(r'[^A-Za-z0-9 _.-]+', '', str(params.get('filename') or 'AI Line Sheet')).strip() or 'AI Line Sheet'
    key = f"claude uploaded/{fname} {datetime.now().strftime('%Y-%m-%d %H%M%S')}.xlsx"
    get_s3().put_object(Bucket=S3_BUCKET, Key=key, Body=xl,
                        ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    from urllib.parse import quote
    url = f"https://{S3_BUCKET}.s3.us-east-2.amazonaws.com/{quote(key)}"
    return {'download_url': url, 'tabs': summary, 'customer_view': customer_view,
            'note': 'give the user this link as a clickable download'}


_AI_AGENT_TOOLS = [
    {'name': 'query_inventory',
     'description': ("Query LIVE inventory aggregated per base style. Filters: brands (abbr like NAUTICA or full name), "
                     "category (long_sleeve|short_sleeve|pants|sportswear|big_tall|young_men|accessories|blazers), "
                     "fabric_codes (2-letter SKU codes), color (color word, or buckets: solids/fancies/white/black/navy — "
                     "the navy bucket covers the whole blue family incl. blue/indigo, and any color name containing "
                     "'dobby' counts as a SOLID even with stripe/check words), "
                     "search (substring of style #), stock (any|warehouse|overseas), min_units, "
                     "arrive_before/arrive_after (YYYY-MM-DD, filters styles with production arriving in that window and "
                     "reports qualifying_po_units). Totals cover ALL matches even when rows are truncated. "
                     "Use for ANY quantity/availability question."),
     'input_schema': {'type': 'object', 'properties': {
         'brands': {'type': 'array', 'items': {'type': 'string'}},
         'category': {'type': 'string'}, 'fabric_codes': {'type': 'array', 'items': {'type': 'string'}},
         'color': {'type': 'string'}, 'search': {'type': 'string'},
         'stock': {'type': 'string', 'enum': ['any', 'warehouse', 'overseas']},
         'min_units': {'type': 'integer'}, 'arrive_before': {'type': 'string'}, 'arrive_after': {'type': 'string'},
         'sort': {'type': 'string', 'enum': ['total_ats', 'warehouse', 'incoming']},
         'limit': {'type': 'integer'}}}},
    {'name': 'style_detail',
     'description': ('Everything about ONE style (base style # or full SKU): per-warehouse stock, size variants, '
                     'incoming production orders with arrival dates, APO allocations, open customer orders, '
                     'color/fit/fabrication/category, image link. Use when asked about a specific style number.'),
     'input_schema': {'type': 'object', 'properties': {'style': {'type': 'string'}}, 'required': ['style']}},
    {'name': 'brand_summary',
     'description': ('Without brand: every brand with style counts and unit totals. With brand: that brand plus its '
                     'fabrication rollup (codes, descriptions, units) and category rollup.'),
     'input_schema': {'type': 'object', 'properties': {'brand': {'type': 'string'}}}},
    {'name': 'open_orders_lookup',
     'description': ('Open customer orders (A2000 + bulks) with quantities, dollars, and ship windows. '
                     'Filter by customer name and/or style #.'),
     'input_schema': {'type': 'object', 'properties': {
         'customer': {'type': 'string'}, 'style': {'type': 'string'}, 'limit': {'type': 'integer'}}}},
    {'name': 'past_orders_lookup',
     'description': ('CLOSED order history (past orders): every customer PO that already LEFT the open-order '
                     'book, archived daily since 2026-06-03. status separates shipped vs cancelled_or_pulled '
                     '(left the book before its ship window opened); left_book = the day it shipped out '
                     '(accurate to about a day). Rows carry units (last-day residual) AND units_peak '
                     '(true booked size — use units_peak for shipped POs; totals already do). '
                     'No args = per-customer summary. customer = that '
                     "customer's past POs with style lines. style = search a style # (substring) across "
                     'every customer\'s history — a cold call can take up to a minute while per-customer '
                     'archives load. received_after/received_before filter on left_book (YYYY-MM-DD). '
                     'Use open_orders_lookup for orders still open.'),
     'input_schema': {'type': 'object', 'properties': {
         'customer': {'type': 'string'}, 'style': {'type': 'string'},
         'status': {'type': 'string', 'enum': ['all', 'shipped', 'cancelled']},
         'received_after': {'type': 'string'}, 'received_before': {'type': 'string'},
         'limit': {'type': 'integer'}}}},
    {'name': 'sales_history_lookup',
     'description': ('INVOICED sales history (ground truth): A2000 invoice records with real invoice dates '
                     'and actually-shipped quantities, Nov 2019 through the last backfill ingest. '
                     'No args = company summary (per-customer lifetime + per-year totals). '
                     'customer = that customer\'s years/months/top styles/recent POs; add month (YYYY-MM) '
                     'or from/to (YYYY-MM-DD) for that period\'s POs WITH style lines. '
                     'style = search a style # (substring) across all invoices. '
                     'For POs newer than the backfill, use past_orders_lookup (book-tracked, updated daily).'),
     'input_schema': {'type': 'object', 'properties': {
         'customer': {'type': 'string'}, 'style': {'type': 'string'},
         'month': {'type': 'string'}, 'from': {'type': 'string'}, 'to': {'type': 'string'},
         'limit': {'type': 'integer'}}}},
    {'name': 'build_sales_sheet',
     'description': ('Build a downloadable Excel SALES SHEET with photos for ONE customer\'s past selling '
                     'and pipeline: an Invoiced tab (real invoice dates + shipped quantities per PO and '
                     'style), a Shipped-Awaiting-Invoice tab (book-tracked POs at peak booked size), and '
                     'an On Order tab (current open hard POs with ship windows; bulks excluded). '
                     'customer accepts a name or code (Winners, Burlington, ROSS…). Date filters: from/to '
                     '(YYYY-MM-DD), or year (2026), or month (YYYY-MM) — applies to invoiced + awaiting; '
                     'On Order always shows the current book. include narrows the tabs '
                     "(['invoiced','shipped_pending','on_order']); group_by 'po' (default) or 'month' sets "
                     'the sort. Takes up to a minute. Present the returned download_url to the user as a '
                     'clickable link. Use this whenever the user wants past selling as a line sheet / '
                     'document rather than numbers in chat.'),
     'input_schema': {'type': 'object', 'properties': {
         'customer': {'type': 'string'},
         'from': {'type': 'string'}, 'to': {'type': 'string'},
         'year': {'type': 'string'}, 'month': {'type': 'string'},
         'include': {'type': 'array', 'items': {'type': 'string',
                     'enum': ['invoiced', 'shipped_pending', 'on_order']}},
         'group_by': {'type': 'string', 'enum': ['po', 'month']},
         'filename': {'type': 'string'}},
         'required': ['customer']}},
    {'name': 'build_line_sheet',
     'description': ('Build a real multi-tab Excel line sheet with embedded photos and return a download URL. '
                     'tabs = list of filter objects (same filters as query_inventory, plus title) — OR give a tab '
                     'an explicit skus list (base style #s, e.g. hand-picked from query_inventory results) for an '
                     'exactly-curated tab in your order; skus overrides every other filter on that tab. '
                     'customer_view=true for customer-facing columns (no committed/allocated), false for full admin. '
                     'Max 4 tabs, 300 styles per tab. Takes up to a minute. Present the returned download_url '
                     'to the user as a clickable link.'),
     'input_schema': {'type': 'object', 'properties': {
         'tabs': {'type': 'array', 'items': {'type': 'object', 'properties': {
             'title': {'type': 'string'}, 'brand': {'type': 'string'},
             'brands': {'type': 'array', 'items': {'type': 'string'}},
             'skus': {'type': 'array', 'items': {'type': 'string'}},
             'category': {'type': 'string'}, 'fabric_codes': {'type': 'array', 'items': {'type': 'string'}},
             'color': {'type': 'string'}, 'stock': {'type': 'string', 'enum': ['any', 'warehouse', 'overseas']},
             'min_units': {'type': 'integer'}, 'arrive_before': {'type': 'string'}, 'arrive_after': {'type': 'string'},
             'limit': {'type': 'integer'}}}},
         'customer_view': {'type': 'boolean'}, 'filename': {'type': 'string'}},
         'required': ['tabs']}},
]

_AI_AGENT_TOOL_FNS = {
    'query_inventory': _ai_tool_query_inventory,
    'style_detail': _ai_tool_style_detail,
    'brand_summary': _ai_tool_brand_summary,
    'open_orders_lookup': _ai_tool_open_orders,
    'past_orders_lookup': _ai_tool_past_orders,
    'sales_history_lookup': _ai_tool_sales_history,
    'build_sales_sheet': _ai_tool_build_sales_sheet,
    'build_line_sheet': _ai_tool_build_line_sheet,
}

_AI_AGENT_TOOL_GUIDANCE = """
LIVE DATA TOOLS
You have server-side tools that query the live inventory database directly. They are fresher and more precise than any snapshot in this prompt. Use them for EVERY question about quantities, styles, availability, fabrications, colors, arrivals, customer orders, or dollar values. Never estimate from the snapshot when a tool can answer; run the tool. Chain tools when needed (e.g. query_inventory to find styles, style_detail to drill in). Quantities from tools are per base style with all size rows aggregated; committed/allocated come back as positive magnitudes.
build_line_sheet creates a real Excel file with photos and returns download_url. When you use it, put the link in your final message as <a href="URL" target="_blank">Download the line sheet</a>.
CURATED SELECTIONS: when a request needs styles no single filter expresses (e.g. several specific colors in one tab), query_inventory FIRST, pick the exact styles from the results yourself, then pass them as an explicit style list — build_line_sheet tabs[].skus, or the saveLineSheetViews action's skus param. Never ask the user to paste style numbers you can look up.
UI actions (navigate, filters, saveLineSheetViews, etc.) still work exactly as documented; use tools for DATA and actions for controlling the UI. Never emit a queryInventory action — it is retired; its output rendered only in the user's browser and never came back to you. After your tools finish, respond in the required JSON format.
The message field renders as raw HTML in the chat bubble. Format with HTML only: <b>, <br>, &bull; lists, <a> links. NEVER markdown (**bold**, ##, tables) — it shows as literal asterisks.
"""

_AI_AGENT_DEFAULT_SYSTEM = (
    'You are the Versa Group inventory assistant with live data tools. Answer questions about inventory, '
    'production, and orders using the tools. Be concise and concrete; cite real numbers from tool results.'
    + _AI_AGENT_TOOL_GUIDANCE)


@app.route('/api/ai-agent', methods=['POST', 'OPTIONS'])
def api_ai_agent():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        client = _ai_agent_client()
    except Exception as e:
        return jsonify({'error': f'AI agent unavailable: {e}'}), 503
    body = request.get_json(silent=True) or {}
    messages = body.get('messages') or []
    convo = [{'role': m.get('role'), 'content': m.get('content')}
             for m in messages if m.get('role') in ('user', 'assistant') and m.get('content')]
    # Anthropic requires the first message to be role 'user' — clients that window
    # their history can start on an assistant turn, which 400s the whole request.
    while convo and convo[0]['role'] != 'user':
        convo.pop(0)
    if not convo:
        return jsonify({'error': 'messages required'}), 400
    model = body.get('model') or AI_AGENT_MODEL
    max_tokens = min(int(body.get('max_tokens') or 8192), 16000)
    system_static = str(body.get('system_static') or '')
    system_dynamic = str(body.get('system_dynamic') or '')
    if system_static:
        static_text = system_static + '\n' + _AI_AGENT_TOOL_GUIDANCE
    else:
        static_text = _AI_AGENT_DEFAULT_SYSTEM
    system = [{'type': 'text', 'text': static_text, 'cache_control': {'type': 'ephemeral'}}]
    if system_dynamic:
        system.append({'type': 'text', 'text': system_dynamic})

    started = time.time()
    iterations = 0
    tools_used = []
    usage_tot = {'input_tokens': 0, 'output_tokens': 0, 'cache_read_input_tokens': 0, 'cache_creation_input_tokens': 0}
    force_final = False
    final_text = ''
    try:
        while True:
            iterations += 1
            kwargs = {}
            if force_final:
                kwargs['tool_choice'] = {'type': 'none'}
            resp = client.with_options(timeout=90.0).messages.create(
                model=model, max_tokens=max_tokens, system=system, messages=convo,
                tools=_AI_AGENT_TOOLS, output_config={'effort': AI_AGENT_EFFORT}, **kwargs)
            u = getattr(resp, 'usage', None)
            if u:
                for k in usage_tot:
                    usage_tot[k] += int(getattr(u, k, 0) or 0)
            if resp.stop_reason == 'refusal':
                final_text = json.dumps({'message': 'I can\'t help with that request.', 'actions': []})
                break
            if resp.stop_reason == 'pause_turn':
                convo.append({'role': 'assistant', 'content': resp.content})
                continue
            tool_uses = [b for b in resp.content if getattr(b, 'type', '') == 'tool_use']
            if not tool_uses:
                final_text = ''.join(getattr(b, 'text', '') for b in resp.content if getattr(b, 'type', '') == 'text')
                break
            convo.append({'role': 'assistant', 'content': resp.content})
            results = []
            for tu in tool_uses:
                tools_used.append(tu.name)
                fn = _AI_AGENT_TOOL_FNS.get(tu.name)
                try:
                    if fn is None:
                        raise ValueError(f'unknown tool {tu.name}')
                    out = fn(tu.input or {})
                    content = json.dumps(out, default=str)
                    if len(content) > 60000:
                        content = content[:60000] + '... [truncated]'
                    results.append({'type': 'tool_result', 'tool_use_id': tu.id, 'content': content})
                except Exception as te:
                    results.append({'type': 'tool_result', 'tool_use_id': tu.id,
                                    'content': f'Tool error: {te}', 'is_error': True})
            convo.append({'role': 'user', 'content': results})
            if iterations >= _AI_AGENT_MAX_ITER or (time.time() - started) > _AI_AGENT_WALL_SECONDS:
                force_final = True
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'AI agent failed: {e}'}), 502

    print(f"[AI-Agent] {iterations} iterations, tools={tools_used}, "
          f"in={usage_tot['input_tokens']} cached={usage_tot['cache_read_input_tokens']} "
          f"out={usage_tot['output_tokens']}, {time.time() - started:.1f}s")
    return jsonify({'content': [{'type': 'text', 'text': final_text}], 'model': model,
                    'stop_reason': 'end_turn', 'usage': usage_tot,
                    'agent': {'iterations': iterations, 'tools': tools_used,
                              'elapsed_seconds': round(time.time() - started, 1)}})


# ============================================================
# MCP CONNECTOR — the same agent tools for claude.ai
# ============================================================
# Serves the same agent tools as /api/ai-agent, over MCP.
# Implements the Model Context Protocol's Streamable HTTP transport as a
# plain Flask route (JSON-RPC 2.0 over POST, single JSON responses — the
# spec allows servers to skip SSE). David adds the URL as a custom
# connector in claude.ai settings and full Claude can query the live
# platform from any chat, including mobile. Stateless: no sessions.
# Auth: if the MCP_TOKEN env var is set, the token must appear in the
# URL path (/mcp/<token>); unset = open, matching this API's posture.

_MCP_PROTOCOL_VERSION = '2025-06-18'

_MCP_INSTRUCTIONS = """Versa Group live inventory connector. Versa is a men's apparel manufacturer/reseller (Nautica, DKNY, Chaps, U.S. Polo Assn., Von Dutch, and ~20 more brands).
Conventions: styles are SKUs shaped [CUSTOMER 2][BRAND 2][FABRIC 2][SERIAL 3][FIT 2][COLLAR 1]; a base style is the part before any -size suffix. Warehouse = stock on hand now (JTW/TR/DCW/QA/NJ). Incoming/overseas = on production order. Total ATS = available to sell. committed/allocated come back as positive magnitudes already deducted from ATS. NT is Nautica overflow (NT 201 and NA 201 are DIFFERENT styles).
Use query_inventory for any quantity/availability question (totals cover all matches even when rows truncate), style_detail for one style, brand_summary for rollups, open_orders_lookup for customer demand and dollars, past_orders_lookup for CLOSED order history (what each customer already received or cancelled, daily archive since Jun 2026), sales_history_lookup for INVOICED sales going back to Nov 2019 (real invoice dates and shipped quantities — the ground truth for what shipped), build_sales_sheet to turn one customer's past selling + open orders into a downloadable Excel sales sheet with photos, build_line_sheet to produce a current-inventory Excel with photos (both return a download URL; each takes up to a minute)."""


def _mcp_rpc_result(rid, result):
    return jsonify({'jsonrpc': '2.0', 'id': rid, 'result': result})


def _mcp_rpc_error(rid, code, message):
    return jsonify({'jsonrpc': '2.0', 'id': rid, 'error': {'code': code, 'message': message}})


def _mcp_handle(payload):
    if not isinstance(payload, dict):
        return _mcp_rpc_error(None, -32600, 'batch requests not supported'), 400
    method = payload.get('method')
    rid = payload.get('id')
    params = payload.get('params') or {}
    # Notifications (no id) are acknowledged with 202 and no body
    if rid is None:
        return ('', 202)
    if method == 'initialize':
        return _mcp_rpc_result(rid, {
            'protocolVersion': _MCP_PROTOCOL_VERSION,
            'capabilities': {'tools': {}},
            'serverInfo': {'name': 'versa-inventory', 'version': '1.0.0'},
            'instructions': _MCP_INSTRUCTIONS,
        })
    if method == 'ping':
        return _mcp_rpc_result(rid, {})
    if method == 'tools/list':
        tools = [{'name': t['name'], 'description': t['description'],
                  'inputSchema': t['input_schema']} for t in _AI_AGENT_TOOLS]
        return _mcp_rpc_result(rid, {'tools': tools})
    if method == 'tools/call':
        name = params.get('name')
        args = params.get('arguments') or {}
        fn = _AI_AGENT_TOOL_FNS.get(name)
        if fn is None:
            return _mcp_rpc_error(rid, -32602, f'unknown tool: {name}')
        try:
            out = fn(args)
            text = json.dumps(out, default=str)
            if len(text) > 90000:
                text = text[:90000] + '... [truncated]'
            return _mcp_rpc_result(rid, {'content': [{'type': 'text', 'text': text}],
                                         'isError': bool(isinstance(out, dict) and out.get('error'))})
        except Exception as e:
            import traceback
            traceback.print_exc()
            return _mcp_rpc_result(rid, {'content': [{'type': 'text', 'text': f'Tool error: {e}'}],
                                         'isError': True})
    return _mcp_rpc_error(rid, -32601, f'method not found: {method}')


@app.route('/mcp', methods=['POST', 'GET', 'DELETE', 'OPTIONS'])
@app.route('/mcp/<token>', methods=['POST', 'GET', 'DELETE', 'OPTIONS'])
def mcp_endpoint(token=None):
    if request.method == 'OPTIONS':
        return '', 204
    required = os.environ.get('MCP_TOKEN')
    if required and token != required:
        return jsonify({'error': 'unauthorized'}), 401
    if request.method != 'POST':
        # No server-initiated SSE stream; stateless single-response server.
        return jsonify({'error': 'POST JSON-RPC messages to this endpoint'}), 405
    payload = request.get_json(silent=True)
    if payload is None:
        return _mcp_rpc_error(None, -32700, 'parse error'), 400
    result = _mcp_handle(payload)
    return result


# Register swatch card extractor routes (/api/ai-proxy, /api/swatch/commit, /api/swatch/history)
from swatch_extractor import register_swatch_routes
register_swatch_routes(app, get_s3, S3_BUCKET)


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)

