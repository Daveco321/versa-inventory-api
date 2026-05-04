"""
SWATCH CARD EXTRACTOR — BACKEND

Frontend handles the Anthropic vision call directly via /api/ai-proxy.
This module handles things that need server-side credentials:

  POST /api/swatch/commit              — APPEND-ONLY color map write to S3
  GET  /api/swatch/history             — last N color map upload events
  POST /api/ai-proxy                   — forwards Claude API calls server-side

  POST /api/swatch/extract-images      — render PDF, detect & crop swatches,
                                         return base64 thumbnails for review
  POST /api/swatch/check-image-dupes   — given SKUs, check which already have
                                         images in STYLE OVERRIDES
  POST /api/swatch/commit-images       — upload approved swatch JPEGs to S3

DESIGN GUARANTEE: Color map writes are APPEND-ONLY. Image uploads write
to <SKU>.jpg in S3 STYLE OVERRIDES — overwrites only happen when the user
explicitly confirms in the UI.

Wire-up in app.py (one line near the bottom, after `app = Flask(__name__)`
and after `get_s3()` / `S3_BUCKET` are defined):

    from swatch_extractor import register_swatch_routes
    register_swatch_routes(app, get_s3, S3_BUCKET)

Environment variables (all optional — sane defaults):
    ANTHROPIC_API_KEY            — required for /api/ai-proxy
    S3_COLOR_MAP_KEY             — default "Inventory Colors Data/style_color_map.xlsx"
    S3_SWATCH_HISTORY_KEY        — default "inventory/swatch_upload_history.json"
    S3_STYLE_OVERRIDES_PREFIX    — default "ALL INVENTORY Photos/STYLE OVERRIDES"
    S3_IMAGE_HISTORY_KEY         — default "inventory/swatch_image_upload_history.json"
"""

import os
import io
import json
import base64
import threading
from datetime import datetime, timezone

from flask import request, jsonify
import openpyxl
import requests as http_requests

# Imports for image extraction. Wrapped so the module still loads even if
# these aren't installed yet — the image endpoints will return 503 instead
# of crashing the whole module.
try:
    import fitz  # PyMuPDF — pure Python, no system deps
    import cv2
    import numpy as np
    from PIL import Image
    _IMAGE_DEPS_OK = True
    _IMAGE_DEPS_ERR = None
except ImportError as _e:
    _IMAGE_DEPS_OK = False
    _IMAGE_DEPS_ERR = str(_e)


# ─────────────────────────────────────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────────────────────────────────────
S3_COLOR_MAP_KEY      = os.environ.get('S3_COLOR_MAP_KEY',
                                       'Inventory Colors Data/style_color_map.xlsx')
S3_SWATCH_HISTORY_KEY = os.environ.get('S3_SWATCH_HISTORY_KEY',
                                       'inventory/swatch_upload_history.json')

# Image extraction config — STYLE OVERRIDES is the folder the frontend reads
# images from when a card needs a custom image (see S3_EXACT_MATCH_URL in
# index.html). New swatch images get written as <SKU>.jpg in this folder.
S3_STYLE_OVERRIDES_PREFIX = os.environ.get('S3_STYLE_OVERRIDES_PREFIX',
                                       'ALL INVENTORY Photos/STYLE OVERRIDES').rstrip('/')
S3_IMAGE_HISTORY_KEY      = os.environ.get('S3_IMAGE_HISTORY_KEY',
                                       'inventory/swatch_image_upload_history.json')

# Image extraction parameters
PDF_RENDER_DPI            = 300        # render quality for swatch detection
SWATCH_OUTPUT_SIZE        = 700        # final JPEG dimensions (px, square)
SWATCH_OUTPUT_QUALITY     = 90         # JPEG quality
SWATCH_INSET_PCT          = 0.08       # crop inset to remove zigzag edges (8%)
SWATCH_THUMB_SIZE         = 200        # thumbnail size for review UI

# Anthropic API key — pulled from Render environment (NEVER hardcoded).
# The /api/ai-proxy endpoint uses this to forward requests on behalf of the
# browser, so the key never leaves the server.
ANTHROPIC_API_KEY     = os.environ.get('ANTHROPIC_API_KEY', '')
ANTHROPIC_API_URL     = 'https://api.anthropic.com/v1/messages'
ANTHROPIC_VERSION     = '2023-06-01'

# Sheet name the frontend looks for first (loadColorMapFromS3 in index.html)
COLOR_MAP_SHEET_NAME = 'Color Map'

# Serializes concurrent commits within a single worker.
_commit_lock = threading.Lock()


# ─────────────────────────────────────────────────────────────────────────────
# COLOR MAP READ / APPEND
# ─────────────────────────────────────────────────────────────────────────────
def _download_color_map(get_s3, s3_bucket):
    """Download the master color map xlsx → openpyxl Workbook + ETag + size."""
    s3 = get_s3()
    resp = s3.get_object(Bucket=s3_bucket, Key=S3_COLOR_MAP_KEY)
    raw = resp['Body'].read()
    etag = resp.get('ETag', '').strip('"')
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    return wb, etag, len(raw)


def _resolve_color_map_sheet(wb):
    """Match frontend logic: prefer 'Color Map' sheet, else first sheet."""
    if COLOR_MAP_SHEET_NAME in wb.sheetnames:
        return wb[COLOR_MAP_SHEET_NAME]
    return wb[wb.sheetnames[0]]


def _append_rows(ws, rows):
    """Append rows at the bottom. NEVER touches existing rows.

    Reads the header row to figure out which column is Key and which is
    Color_Description, then writes new values into those exact columns
    (leaving any middle columns blank). This is critical because the
    actual sheet may have extra columns like Brand_Prefix, Style_Number,
    Brand_Name between Key and Color_Description — writing positionally
    would put the color into the wrong column and the frontend would
    silently skip the row.
    """
    # Empty sheet: write minimal headers and append flush-left
    if ws.max_row < 1:
        ws.append(['Key', 'Color_Description'])
        for sku, color in rows:
            ws.append([sku, color])
        return

    # Read header row, find the columns we care about (case-insensitive,
    # whitespace-tolerant). Frontend's reader accepts Key OR Style_Number
    # OR Style_Num as the SKU column — match that priority order.
    headers = [str(c.value or '').strip() for c in ws[1]]
    headers_lower = [h.lower() for h in headers]

    def find_col(*candidates):
        for cand in candidates:
            if cand.lower() in headers_lower:
                return headers_lower.index(cand.lower())  # 0-indexed
        return None

    key_col   = find_col('Key', 'Style_Number', 'Style_Num')
    color_col = find_col('Color_Description')
    n_cols    = max(len(headers), (key_col or 0) + 1, (color_col or 0) + 1)

    # Defensive fallback: if the sheet has no recognizable schema, append
    # in the original 2-column form rather than refusing to write.
    if key_col is None and color_col is None:
        for sku, color in rows:
            ws.append([sku, color])
        return

    # If only one of the two was found, default the other to a sensible position
    if key_col is None:    key_col = 0
    if color_col is None:  color_col = max(1, n_cols - 1)

    for sku, color in rows:
        new_row = [None] * n_cols
        new_row[key_col]   = sku
        new_row[color_col] = color
        ws.append(new_row)


def _upload_color_map(get_s3, s3_bucket, wb) -> str:
    buf = io.BytesIO()
    wb.save(buf)
    body = buf.getvalue()
    s3 = get_s3()
    resp = s3.put_object(
        Bucket=s3_bucket,
        Key=S3_COLOR_MAP_KEY,
        Body=body,
        ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    new_etag = resp.get('ETag', '').strip('"')
    print(f"  [swatch] ✓ Uploaded color map to s3://{s3_bucket}/{S3_COLOR_MAP_KEY} "
          f"({len(body):,} bytes, ETag {new_etag[:8]})", flush=True)
    return new_etag


# ─────────────────────────────────────────────────────────────────────────────
# AUDIT LOG
# ─────────────────────────────────────────────────────────────────────────────
def _append_audit_entry(get_s3, s3_bucket, entry):
    """Append an entry to the JSON audit log. Best-effort; failure is logged
    but does not abort the commit."""
    s3 = get_s3()
    history = []
    try:
        resp = s3.get_object(Bucket=s3_bucket, Key=S3_SWATCH_HISTORY_KEY)
        history = json.loads(resp['Body'].read())
        if not isinstance(history, list):
            history = []
    except s3.exceptions.NoSuchKey:
        pass
    except Exception as e:
        print(f"  [swatch] ⚠ Could not read existing audit log: {e}", flush=True)

    history.append(entry)
    history = history[-1000:]  # cap log size

    try:
        s3.put_object(
            Bucket=s3_bucket,
            Key=S3_SWATCH_HISTORY_KEY,
            Body=json.dumps(history, indent=2).encode('utf-8'),
            ContentType='application/json',
        )
    except Exception as e:
        print(f"  [swatch] ⚠ Failed to write audit log: {e}", flush=True)


# ─────────────────────────────────────────────────────────────────────────────
# IMAGE EXTRACTION — PDF → swatch crops
# ─────────────────────────────────────────────────────────────────────────────
def _render_pdf_pages(pdf_bytes, dpi=PDF_RENDER_DPI):
    """Render every page of a PDF to a list of PIL.Image objects.

    Uses PyMuPDF — pure Python, no Poppler / system dependencies. Each page
    is rendered at the configured DPI for reliable contour detection.
    """
    doc = fitz.open(stream=pdf_bytes, filetype='pdf')
    images = []
    for page in doc:
        mat = fitz.Matrix(dpi / 72, dpi / 72)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        img = Image.frombytes('RGB', (pix.width, pix.height), pix.samples)
        images.append(img)
    doc.close()
    return images


def _detect_swatch_frames(pil_img):
    """Find swatch bounding boxes on a swatch-card page.

    Strategy: every swatch on the source cards is wrapped in a thin gray
    rectangular frame. Detecting that frame via Canny edges + rectangular
    contour analysis is reliable and deterministic — same input → same output
    every time, no AI guessing required.

    Returns list of (x, y, w, h) tuples in image pixel coordinates, sorted
    in reading order (top-to-bottom, then left-to-right).
    """
    np_img = np.array(pil_img)
    gray = cv2.cvtColor(np_img, cv2.COLOR_RGB2GRAY)
    H, W = gray.shape

    edges = cv2.Canny(gray, 30, 100)
    kernel = np.ones((3, 3), np.uint8)
    dilated = cv2.dilate(edges, kernel, iterations=1)
    contours, _ = cv2.findContours(dilated, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)

    candidates = []
    for c in contours:
        x, y, w, h = cv2.boundingRect(c)
        # Aspect ratio: swatch frames are square (within ±15%)
        ar = w / h if h > 0 else 0
        if not (0.85 <= ar <= 1.15):
            continue
        # Size: 18%-35% of page width.
        # Lower bound excludes buttons (typically 8-12% wide) and other
        # small trim photos. Upper bound excludes full garment renders.
        if w < W * 0.18 or w > W * 0.35:
            continue
        # Position: must be in the body region (header/footer excluded)
        if y < H * 0.08 or (y + h) > H * 0.94:
            continue
        # Polygon approximation should yield a rectangle (4-8 vertices,
        # allowing for slightly imperfect corners)
        peri = cv2.arcLength(c, True)
        approx = cv2.approxPolyDP(c, 0.04 * peri, True)
        if not (4 <= len(approx) <= 8):
            continue
        candidates.append((x, y, w, h))

    # Non-maximum suppression: when multiple contours describe the same frame,
    # keep the largest. Two boxes are duplicates if their centers are within
    # 30% of average side length.
    candidates.sort(key=lambda c: -c[2] * c[3])
    kept = []
    for cand in candidates:
        x1, y1, w1, h1 = cand
        cx1, cy1 = x1 + w1 / 2, y1 + h1 / 2
        is_dup = False
        for x2, y2, w2, h2 in kept:
            cx2, cy2 = x2 + w2 / 2, y2 + h2 / 2
            avg = (w1 + h1 + w2 + h2) / 4
            if abs(cx1 - cx2) < avg * 0.3 and abs(cy1 - cy2) < avg * 0.3:
                is_dup = True
                break
        if not is_dup:
            kept.append(cand)

    if not kept:
        return []
    # Sort reading order: top-to-bottom (rows), then left-to-right within a row
    avg_h = sum(h for _, _, _, h in kept) / len(kept)
    kept.sort(key=lambda c: (c[1] // (avg_h * 0.5), c[0]))
    return kept


def _crop_swatch(pil_img, bbox, output_size=SWATCH_OUTPUT_SIZE,
                 inset_pct=SWATCH_INSET_PCT):
    """Crop a swatch from its detected bbox, inset to remove zigzag edges,
    output as a clean square image."""
    x, y, w, h = bbox
    inset_x = int(w * inset_pct)
    inset_y = int(h * inset_pct)
    cropped = pil_img.crop((x + inset_x, y + inset_y,
                            x + w - inset_x, y + h - inset_y))
    cropped = cropped.resize((output_size, output_size), Image.LANCZOS)
    return cropped


def _image_to_jpeg_bytes(pil_img, quality=SWATCH_OUTPUT_QUALITY):
    """Encode a PIL image as JPEG bytes."""
    buf = io.BytesIO()
    if pil_img.mode != 'RGB':
        pil_img = pil_img.convert('RGB')
    pil_img.save(buf, 'JPEG', quality=quality, optimize=True)
    return buf.getvalue()


def _bytes_to_data_url(jpeg_bytes):
    """Encode JPEG bytes as a data: URL for inline preview in the review UI."""
    b64 = base64.b64encode(jpeg_bytes).decode('ascii')
    return f"data:image/jpeg;base64,{b64}"


def _check_image_exists(s3, s3_bucket, sku):
    """Check whether <SKU>.jpg or <SKU>.png already exists in STYLE OVERRIDES.

    Returns the existing extension (e.g. 'jpg') if found, else None.
    Uses head_object — fast, no body download.
    """
    sku = sku.strip().upper()
    for ext in ('jpg', 'JPG', 'jpeg', 'JPEG', 'png', 'PNG'):
        key = f"{S3_STYLE_OVERRIDES_PREFIX}/{sku}.{ext}"
        try:
            s3.head_object(Bucket=s3_bucket, Key=key)
            return ext
        except Exception:
            continue
    return None


def _upload_swatch_image(s3, s3_bucket, sku, jpeg_bytes):
    """Upload a single swatch JPEG to S3 STYLE OVERRIDES as <SKU>.jpg.

    Always writes .jpg (lowercase) per existing convention. Returns the
    full S3 key written.
    """
    sku = sku.strip().upper()
    key = f"{S3_STYLE_OVERRIDES_PREFIX}/{sku}.jpg"
    s3.put_object(
        Bucket=s3_bucket,
        Key=key,
        Body=jpeg_bytes,
        ContentType='image/jpeg',
        CacheControl='public, max-age=3600',
    )
    return key


def _append_image_history(get_s3, s3_bucket, entry):
    """Append an entry to the image-upload audit log. Best-effort."""
    s3 = get_s3()
    history = []
    try:
        resp = s3.get_object(Bucket=s3_bucket, Key=S3_IMAGE_HISTORY_KEY)
        history = json.loads(resp['Body'].read())
        if not isinstance(history, list):
            history = []
    except s3.exceptions.NoSuchKey:
        pass
    except Exception as e:
        print(f"  [swatch-img] ⚠ Could not read existing image audit log: {e}", flush=True)

    history.append(entry)
    history = history[-1000:]

    try:
        s3.put_object(
            Bucket=s3_bucket,
            Key=S3_IMAGE_HISTORY_KEY,
            Body=json.dumps(history, indent=2).encode('utf-8'),
            ContentType='application/json',
        )
    except Exception as e:
        print(f"  [swatch-img] ⚠ Failed to write image audit log: {e}", flush=True)


# ─────────────────────────────────────────────────────────────────────────────
# ROUTE REGISTRATION
# ─────────────────────────────────────────────────────────────────────────────
def register_swatch_routes(app, get_s3, s3_bucket):
    """Register swatch routes on a Flask app.

    Args:
        app: Flask application instance
        get_s3: callable returning a boto3 S3 client (your existing helper)
        s3_bucket: bucket name (e.g. S3_BUCKET)
    """

    @app.route('/api/swatch/commit', methods=['POST'])
    def swatch_commit():
        """Reviewed rows in → APPEND-ONLY to style_color_map.xlsx in S3."""
        body = request.get_json(silent=True) or {}
        rows = body.get('rows') or []
        if not rows:
            return jsonify({'error': 'No rows provided'}), 400

        # Sanitize: drop empties, strip whitespace, dedupe within payload only
        clean_rows = []
        seen_in_payload = set()
        for r in rows:
            sku = (r.get('sku') or '').strip()
            color = (r.get('color_description') or '').strip()
            if not sku or not color:
                continue
            sig = (sku.upper(), color)
            if sig in seen_in_payload:
                continue
            seen_in_payload.add(sig)
            clean_rows.append((sku, color))

        if not clean_rows:
            return jsonify({'error': 'No valid rows after cleaning'}), 400

        with _commit_lock:
            try:
                wb, prev_etag, _ = _download_color_map(get_s3, s3_bucket)
                ws = _resolve_color_map_sheet(wb)
                row_count_before = ws.max_row

                _append_rows(ws, clean_rows)
                new_etag = _upload_color_map(get_s3, s3_bucket, wb)

            except Exception as e:
                print(f"  [swatch] ✗ Commit failed: {e}", flush=True)
                return jsonify({'error': f'Commit failed: {e}'}), 500

        entry = {
            'timestamp':        datetime.now(timezone.utc).isoformat(),
            'filename':         body.get('filename', ''),
            'po_ref':           body.get('po_ref', ''),
            'delivery_date':    body.get('delivery_date', ''),
            'fabrication':      body.get('fabrication', ''),
            'rows_appended':    len(clean_rows),
            'prev_etag':        prev_etag,
            'new_etag':         new_etag,
            'row_count_before': row_count_before,
            'row_count_after':  row_count_before + len(clean_rows),
            'rows':             [{'sku': s, 'color': c} for s, c in clean_rows],
        }
        try:
            _append_audit_entry(get_s3, s3_bucket, entry)
        except Exception as e:
            print(f"  [swatch] ⚠ Audit log error (non-fatal): {e}", flush=True)

        return jsonify({
            'success':           True,
            'rows_appended':     len(clean_rows),
            'row_count_before':  row_count_before,
            'row_count_after':   row_count_before + len(clean_rows),
            'prev_etag':         prev_etag,
            'new_etag':          new_etag,
        })

    @app.route('/api/ai-proxy', methods=['POST'])
    def ai_proxy():
        """Server-side proxy for Anthropic API calls. The browser sends the
        same body it would send directly to api.anthropic.com; we attach the
        key from the Render environment and forward.
        """
        if not ANTHROPIC_API_KEY:
            return jsonify({
                'error': 'ANTHROPIC_API_KEY not configured on server. '
                         'Set it in Render → Environment → Environment Variables.'
            }), 503

        body = request.get_json(silent=True)
        if not body or not isinstance(body, dict):
            return jsonify({'error': 'Request body must be JSON'}), 400

        # Whitelist the fields we forward — protects against the browser
        # sneaking in fields we don't expect, and keeps the proxy minimal.
        allowed = {'model', 'messages', 'system', 'max_tokens',
                   'temperature', 'top_p', 'top_k', 'stop_sequences', 'tools'}
        forward = {k: v for k, v in body.items() if k in allowed}
        if 'model' not in forward or 'messages' not in forward:
            return jsonify({'error': 'Required fields: model, messages'}), 400
        forward.setdefault('max_tokens', 4096)

        try:
            resp = http_requests.post(
                ANTHROPIC_API_URL,
                headers={
                    'x-api-key': ANTHROPIC_API_KEY,
                    'anthropic-version': ANTHROPIC_VERSION,
                    'content-type': 'application/json',
                },
                json=forward,
                timeout=180,
            )
        except http_requests.exceptions.Timeout:
            return jsonify({'error': 'Anthropic API timed out'}), 504
        except Exception as e:
            print(f"  [ai-proxy] ✗ Upstream error: {e}", flush=True)
            return jsonify({'error': f'Upstream error: {e}'}), 502

        try:
            return jsonify(resp.json()), resp.status_code
        except ValueError:
            return jsonify({'error': f'Anthropic returned non-JSON ({resp.status_code})'}), 502

    @app.route('/api/swatch/history', methods=['GET'])
    def swatch_history():
        """Return last N audit log entries (most recent first)."""
        limit = int(request.args.get('limit', 25))
        try:
            s3 = get_s3()
            resp = s3.get_object(Bucket=s3_bucket, Key=S3_SWATCH_HISTORY_KEY)
            history = json.loads(resp['Body'].read())
            if not isinstance(history, list):
                history = []
        except Exception:
            history = []
        return jsonify({'history': history[-limit:][::-1]})

    # ─────────────────────────────────────────────────────────────────────
    # IMAGE EXTRACTION ENDPOINTS
    # ─────────────────────────────────────────────────────────────────────

    @app.route('/api/swatch/extract-images', methods=['POST'])
    def swatch_extract_images():
        """Render an uploaded PDF, detect swatches, return base64 thumbnails.

        Multipart upload with field "file" containing a PDF.
        Returns:
        {
          "swatches": [
            {
              "page": 1,
              "index": 0,                 // index within the page (reading order)
              "bbox": [x, y, w, h],       // pixel coords for re-cropping later
              "page_dims": [W, H],        // full page dimensions
              "thumbnail_data_url": "data:image/jpeg;base64,...",
              "full_image_data_url": "data:image/jpeg;base64,..."  // 700×700
            }
          ]
        }
        """
        if not _IMAGE_DEPS_OK:
            return jsonify({
                'error': f'Image extraction unavailable: missing dependency ({_IMAGE_DEPS_ERR}). '
                         f'Install pymupdf, opencv-python-headless, Pillow.'
            }), 503

        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded (expected multipart field "file")'}), 400
        f = request.files['file']
        if not f.filename:
            return jsonify({'error': 'Empty filename'}), 400
        pdf_bytes = f.read()
        if not pdf_bytes:
            return jsonify({'error': 'Uploaded file is empty'}), 400

        # Guard: only PDFs are supported for image extraction (image uploads
        # would just be a single swatch the user could upload directly).
        mime = (f.mimetype or '').lower()
        if mime != 'application/pdf' and not f.filename.lower().endswith('.pdf'):
            return jsonify({'error': 'Image extraction requires a PDF file'}), 400

        try:
            pages = _render_pdf_pages(pdf_bytes)
        except Exception as e:
            print(f"  [swatch-img] ✗ PDF render failed: {e}", flush=True)
            return jsonify({'error': f'PDF render failed: {e}'}), 500

        results = []
        for page_idx, page_img in enumerate(pages):
            try:
                bboxes = _detect_swatch_frames(page_img)
            except Exception as e:
                print(f"  [swatch-img] ⚠ Detection failed on page {page_idx+1}: {e}", flush=True)
                continue

            for swatch_idx, bbox in enumerate(bboxes):
                try:
                    # Full-resolution clean swatch
                    full = _crop_swatch(page_img, bbox)
                    full_jpeg = _image_to_jpeg_bytes(full)

                    # Thumbnail for the review UI
                    thumb = full.copy()
                    thumb.thumbnail((SWATCH_THUMB_SIZE, SWATCH_THUMB_SIZE), Image.LANCZOS)
                    thumb_jpeg = _image_to_jpeg_bytes(thumb, quality=80)

                    results.append({
                        'page': page_idx + 1,
                        'index': swatch_idx,
                        'bbox': list(bbox),
                        'page_dims': [page_img.width, page_img.height],
                        'thumbnail_data_url': _bytes_to_data_url(thumb_jpeg),
                        'full_image_data_url': _bytes_to_data_url(full_jpeg),
                    })
                except Exception as e:
                    print(f"  [swatch-img] ⚠ Crop failed page {page_idx+1} idx {swatch_idx}: {e}",
                          flush=True)

        print(f"  [swatch-img] ✓ Extracted {len(results)} swatch images from "
              f"{f.filename} ({len(pages)} pages)", flush=True)
        return jsonify({'swatches': results})

    @app.route('/api/swatch/check-image-dupes', methods=['POST'])
    def swatch_check_image_dupes():
        """Given a list of SKUs, return which already have images in STYLE OVERRIDES.

        Body: {"skus": ["JNAW263SSP", "JNAW111SSP", ...]}
        Returns: {"existing": {"JNAW263SSP": "jpg", "JNAW111SSP": null, ...}}

        For each SKU, the value is the existing extension (e.g. "jpg") if a
        file exists, or null if not. The frontend uses this to render
        per-row "✓ New" or "⚠ Existing" badges in the review UI.
        """
        body = request.get_json(silent=True) or {}
        skus = body.get('skus') or []
        if not isinstance(skus, list):
            return jsonify({'error': 'skus must be an array'}), 400

        s3 = get_s3()
        result = {}
        for sku in skus:
            if not sku or not isinstance(sku, str):
                continue
            try:
                result[sku.strip().upper()] = _check_image_exists(s3, s3_bucket, sku)
            except Exception as e:
                print(f"  [swatch-img] ⚠ check-dupe failed for {sku}: {e}", flush=True)
                result[sku.strip().upper()] = None
        return jsonify({'existing': result})

    @app.route('/api/swatch/commit-images', methods=['POST'])
    def swatch_commit_images():
        """Upload approved swatch images to S3 STYLE OVERRIDES.

        Body: {
          "filename": "JONES_SS_9_25-_R.pdf",
          "images": [
            {
              "sku": "JNAW263SSP",
              "image_data_url": "data:image/jpeg;base64,...",
              "overwrite": true   // user confirmed they want to overwrite an existing file
            }
          ]
        }
        Returns: {"uploaded": [...], "skipped": [...], "errors": [...]}
        """
        body = request.get_json(silent=True) or {}
        images = body.get('images') or []
        if not images:
            return jsonify({'error': 'No images provided'}), 400

        s3 = get_s3()
        uploaded, skipped, errors = [], [], []

        for img in images:
            sku = (img.get('sku') or '').strip().upper()
            data_url = img.get('image_data_url') or ''
            overwrite = bool(img.get('overwrite'))

            if not sku or not data_url:
                errors.append({'sku': sku, 'error': 'Missing sku or image_data_url'})
                continue

            # Decode the data URL to raw JPEG bytes
            try:
                if ',' not in data_url:
                    raise ValueError('Malformed data URL')
                jpeg_bytes = base64.b64decode(data_url.split(',', 1)[1])
            except Exception as e:
                errors.append({'sku': sku, 'error': f'Decode failed: {e}'})
                continue

            # Duplicate check: skip unless explicitly overwriting
            try:
                existing = _check_image_exists(s3, s3_bucket, sku)
            except Exception as e:
                errors.append({'sku': sku, 'error': f'Duplicate check failed: {e}'})
                continue

            if existing and not overwrite:
                skipped.append({'sku': sku, 'reason': f'Already exists as .{existing}'})
                continue

            try:
                key = _upload_swatch_image(s3, s3_bucket, sku, jpeg_bytes)
                uploaded.append({
                    'sku':       sku,
                    'key':       key,
                    'overwrote': bool(existing),
                    'bytes':     len(jpeg_bytes),
                })
                print(f"  [swatch-img] ✓ Uploaded {sku} → s3://{s3_bucket}/{key} "
                      f"({len(jpeg_bytes):,} bytes{' [OVERWRITE]' if existing else ''})",
                      flush=True)
            except Exception as e:
                errors.append({'sku': sku, 'error': f'Upload failed: {e}'})

        # Audit log entry
        if uploaded:
            try:
                _append_image_history(get_s3, s3_bucket, {
                    'timestamp': datetime.now(timezone.utc).isoformat(),
                    'filename':  body.get('filename', ''),
                    'uploaded':  uploaded,
                    'skipped':   skipped,
                    'errors':    errors,
                })
            except Exception as e:
                print(f"  [swatch-img] ⚠ Audit log error (non-fatal): {e}", flush=True)

        return jsonify({
            'uploaded': uploaded,
            'skipped':  skipped,
            'errors':   errors,
        })

    print("  ✓ Swatch routes registered: /api/swatch/commit, /api/swatch/history, "
          "/api/ai-proxy, /api/swatch/extract-images, /api/swatch/check-image-dupes, "
          "/api/swatch/commit-images", flush=True)
    if not ANTHROPIC_API_KEY:
        print("  ⚠ ANTHROPIC_API_KEY env var not set — /api/ai-proxy will return 503 until configured",
              flush=True)
    if not _IMAGE_DEPS_OK:
        print(f"  ⚠ Image deps missing ({_IMAGE_DEPS_ERR}). "
              f"Add 'pymupdf', 'opencv-python-headless', 'Pillow' to requirements.txt "
              f"to enable swatch image extraction.", flush=True)
